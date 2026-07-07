from datetime import datetime

import pandas as pd
from openpyxl import Workbook

from codebase.archive.outlook_mapping_maintenance_workflow import (
    _apply_subtotal_overrides_to_sheet,
    _archive_workbook,
    _build_stale_subtotal_override_rows,
    _compute_leap_subtotals,
    _mapping_style_sector_path_from_export_segments,
    _split_allowed_subtotal_mismatches,
    _split_allowed_crosswalk_conflicts,
    _split_allowed_many_to_many,
)
from codebase.mapping_tools.build_energy_balance_relationships import (
    _split_allowed_duplicate_source_pairs,
    _split_allowed_duplicate_target_pairs,
)


def _write_exception_workbook(path, sheet_name: str, rows: list[dict[str, object]]) -> None:
    pd.DataFrame(rows).to_excel(path, sheet_name=sheet_name, index=False)


def test_archive_workbook_writes_unique_timestamped_copy(tmp_path) -> None:
    workbook_path = tmp_path / "outlook_mappings_master.xlsx"
    archive_dir = tmp_path / "archive"
    workbook_path.write_bytes(b"workbook bytes")
    timestamp = datetime(2026, 6, 26, 14, 5, 6)

    first_archive = _archive_workbook(
        workbook_path,
        archive_dir=archive_dir,
        timestamp=timestamp,
    )
    second_archive = _archive_workbook(
        workbook_path,
        archive_dir=archive_dir,
        timestamp=timestamp,
    )

    assert first_archive.name == "outlook_mappings_master.maintenance_run_20260626_140506.xlsx"
    assert second_archive.name == "outlook_mappings_master.maintenance_run_20260626_140506_2.xlsx"
    assert first_archive.read_bytes() == b"workbook bytes"
    assert second_archive.read_bytes() == b"workbook bytes"


def test_split_allowed_many_to_many_keeps_placeholder_rows_out_of_conflicts(tmp_path) -> None:
    many_to_many = pd.DataFrame(
        [
            {
                "sheet": "leap_combined_ninth",
                "leap_sector_name_full_path": "Electricity Generation",
                "raw_leap_fuel_name": "Solar nonspecified",
                "ninth_sector": "09_01_electricity_plants",
                "ninth_fuel": "12_solar_unallocated",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "leap_combined_ninth",
                "leap_sector_name_full_path": "Electricity Generation",
                "raw_leap_fuel_name": "Solar nonspecified",
                "ninth_sector": "09_01_electricity_plants",
                "ninth_fuel": "12_x_other_solar",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "leap_combined_ninth",
                "leap_sector_name_full_path": "Electricity interim/Electricity interim",
                "raw_leap_fuel_name": "Solar nonspecified",
                "ninth_sector": "09_01_electricity_plants",
                "ninth_fuel": "12_solar_unallocated",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "leap_combined_ninth",
                "leap_sector_name_full_path": "Electricity interim/Electricity interim",
                "raw_leap_fuel_name": "Solar nonspecified",
                "ninth_sector": "09_01_electricity_plants",
                "ninth_fuel": "12_x_other_solar",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "ninth_pairs_to_esto_pairs",
                "esto_flow": "09.01.01 Electricity plants",
                "esto_product": "15.05 Other biomass",
                "9th_sector": "09_01_electricity_plants",
                "9th_fuel": "15_05_other_biomass",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "ninth_pairs_to_esto_pairs",
                "esto_flow": "09.01.02 CHP plants",
                "esto_product": "15.05 Other biomass",
                "9th_sector": "09_02_chp_plants",
                "9th_fuel": "15_05_other_biomass",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "ninth_pairs_to_esto_pairs",
                "esto_flow": "09.02.02 CHP plants",
                "esto_product": "15.05 Other biomass",
                "9th_sector": "09_02_chp_plants",
                "9th_fuel": "15_05_other_biomass",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "ninth_pairs_to_esto_pairs",
                "esto_flow": "09.01.02 CHP plants",
                "esto_product": "15.05 Other biomass",
                "9th_sector": "09_02_chp_plants",
                "9th_fuel": "15_solid_biomass_unallocated",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "ninth_pairs_to_esto_pairs",
                "esto_flow": "09.02.02 CHP plants",
                "esto_product": "15.05 Other biomass",
                "9th_sector": "09_02_chp_plants",
                "9th_fuel": "15_solid_biomass_unallocated",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "ninth_pairs_to_esto_pairs",
                "esto_flow": "09.01.02 CHP plants",
                "esto_product": "16.09 Other sources",
                "9th_sector": "09_02_chp_plants",
                "9th_fuel": "16_others_unallocated",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "ninth_pairs_to_esto_pairs",
                "esto_flow": "09.01.03 Heat plants",
                "esto_product": "15.05 Other biomass",
                "9th_sector": "09_x_heat_plants",
                "9th_fuel": "15_05_other_biomass",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "ninth_pairs_to_esto_pairs",
                "esto_flow": "09.01.03 Heat plants",
                "esto_product": "16.09 Other sources",
                "9th_sector": "09_x_heat_plants",
                "9th_fuel": "16_09_other_sources",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "ninth_pairs_to_esto_pairs",
                "esto_flow": "18.01 MAP electricity plants",
                "esto_product": "15.05 Other biomass",
                "9th_sector": "18_01_electricity_plants",
                "9th_fuel": "15_05_other_biomass",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "ninth_pairs_to_esto_pairs",
                "esto_flow": "18.02 MAP CHP plants",
                "esto_product": "15.05 Other biomass",
                "9th_sector": "18_01_electricity_plants",
                "9th_fuel": "15_05_other_biomass",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "ninth_pairs_to_esto_pairs",
                "esto_flow": "18.01 MAP electricity plants",
                "esto_product": "15.05 Other biomass",
                "9th_sector": "18_01_electricity_plants",
                "9th_fuel": "15_solid_biomass_unallocated",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "ninth_pairs_to_esto_pairs",
                "esto_flow": "18.02 MAP CHP plants",
                "esto_product": "15.05 Other biomass",
                "9th_sector": "18_01_electricity_plants",
                "9th_fuel": "15_solid_biomass_unallocated",
                "cardinality": "many_to_many",
            },
            {
                "sheet": "leap_combined_esto",
                "leap_sector_name_full_path": "Other loss and own use/Liquefaction and regasification plants",
                "raw_leap_fuel_name": "Electricity",
                "esto_flow": "09.06 Gas processing plants",
                "esto_product": "17 Electricity",
                "cardinality": "many_to_many",
            },
        ]
    ).fillna("")

    exception_path = tmp_path / "mapping_issue_exception_sets.xlsx"
    exception_rows = many_to_many.iloc[:16].copy()
    exception_rows.insert(0, "enabled", True)
    exception_rows["notes"] = "Reviewed manually"
    _write_exception_workbook(
        exception_path,
        "many_to_many_allowed",
        exception_rows.to_dict("records"),
    )

    conflicts, allowed = _split_allowed_many_to_many(
        many_to_many,
        exception_workbook_path=exception_path,
    )

    assert len(allowed) == 16
    assert set(allowed["many_to_many_review_status"]) == {"allowed"}
    assert set(allowed["many_to_many_review_reason"]) == {"Reviewed manually"}
    assert len(conflicts) == 1
    assert conflicts.loc[0, "sheet"] == "leap_combined_esto"
    assert "many_to_many_review_status" not in conflicts.columns


def test_split_allowed_crosswalk_conflicts_uses_manual_workbook(tmp_path) -> None:
    crosswalk_conflicts = pd.DataFrame(
        [
            {
                "leap_sector_name_full_path": "Transfers",
                "raw_leap_fuel_name": "Petroleum coke",
                "ninth_sector": "08_transfers",
                "ninth_fuel": "07_x_other_petroleum_products",
                "implied_esto_targets": (
                    "08 Transfers || 07.12 White spirit SBP | "
                    "08 Transfers || 07.17 Other products"
                ),
                "active_esto_targets": (
                    "08 Transfers || 07.16 Petroleum coke | "
                    "08 Transfers || 07_x_other_petroleum_products"
                ),
                "conflict_reason": "implied_esto_target_not_active_for_leap_source",
                "conflict_classification": "target_mismatch_review",
            },
            {
                "leap_sector_name_full_path": "Other loss and own use/Liquefaction and regasification plants",
                "raw_leap_fuel_name": "Electricity",
                "ninth_sector": "09_06_gas_processing_plants",
                "ninth_fuel": "17_electricity",
                "implied_esto_targets": (
                    "09.06 Gas processing plants || 17 Electricity | "
                    "09.06.02 Liquefaction/regasification plants || 17 Electricity"
                ),
                "active_esto_targets": (
                    "10.01.03 Liquefaction/regasification plants || 17 Electricity"
                ),
                "conflict_reason": "implied_esto_target_not_active_for_leap_source",
                "conflict_classification": "target_mismatch_review",
            },
        ]
    )

    exception_path = tmp_path / "mapping_issue_exception_sets.xlsx"
    exception_row = crosswalk_conflicts.iloc[[0]].copy()
    exception_row.insert(0, "enabled", True)
    exception_row["notes"] = "Reviewed rollup category overlap"
    _write_exception_workbook(
        exception_path,
        "crosswalk_allowed",
        exception_row.to_dict("records"),
    )

    conflicts, allowed = _split_allowed_crosswalk_conflicts(
        crosswalk_conflicts,
        exception_workbook_path=exception_path,
    )

    assert len(allowed) == 1
    assert allowed.loc[0, "leap_sector_name_full_path"] == "Transfers"
    assert allowed.loc[0, "crosswalk_review_status"] == "allowed"
    assert allowed.loc[0, "crosswalk_review_reason"] == "Reviewed rollup category overlap"
    assert len(conflicts) == 1
    assert conflicts.loc[0, "leap_sector_name_full_path"].startswith("Other loss and own use")
    assert "crosswalk_review_status" not in conflicts.columns


def test_split_allowed_subtotal_mismatches_uses_manual_allowlist(tmp_path) -> None:
    subtotal_mismatches = pd.DataFrame(
        [
            {
                "leap_sector_name_full_path": "Total Transformation",
                "raw_leap_fuel_name": "Coal products",
                "esto_flow": "09 Total transformation sector",
                "esto_product": "02 Coal products",
                "leap_is_subtotal": "False",
                "esto_pair_is_subtotal": "True",
                "mismatch_reason": "leaf_source_maps_to_aggregate_target_and_more_specific_target_exists",
                "sheet": "leap_combined_esto",
            },
            {
                "leap_sector_name_full_path": "Transfers",
                "raw_leap_fuel_name": "Petroleum coke",
                "esto_flow": "08 Transfers",
                "esto_product": "07.16 Petroleum coke",
                "leap_is_subtotal": "False",
                "esto_pair_is_subtotal": "True",
                "mismatch_reason": "leaf_source_maps_to_aggregate_target_and_more_specific_target_exists",
                "sheet": "leap_combined_esto",
            },
        ]
    )
    exception_path = tmp_path / "mapping_issue_exception_sets.xlsx"
    allowed_row = subtotal_mismatches.iloc[[0]].copy()
    allowed_row.insert(0, "enabled", True)
    allowed_row["notes"] = "Reviewed manually"
    _write_exception_workbook(
        exception_path,
        "subtotal_mismatch_allowed",
        allowed_row.to_dict("records"),
    )

    needs_review, allowed = _split_allowed_subtotal_mismatches(
        subtotal_mismatches,
        exception_workbook_path=exception_path,
    )

    assert len(allowed) == 1
    assert allowed.loc[0, "subtotal_mismatch_review_status"] == "allowed"
    assert allowed.loc[0, "subtotal_mismatch_review_reason"] == "Reviewed manually"
    assert len(needs_review) == 1
    assert needs_review.loc[0, "leap_sector_name_full_path"] == "Transfers"


def test_split_allowed_duplicate_source_pairs_uses_manual_allowlist(tmp_path) -> None:
    duplicate_source_pairs = pd.DataFrame(
        [
            {
                "source_flow": "Total final consumption",
                "source_product": "Anthracite",
                "included_row_count": 3,
                "target_pair_count": 3,
                "target_pairs": "A | B | C",
                "target_flows": "A | B | C",
                "source_rows": "1|2|3",
                "cardinality": "many_to_many",
                "relationship_types": "direct_or_existing_mapping",
                "relationship_levels": "total|parent",
                "qa_status": "review_expected",
                "qa_severity": "info",
                "qa_reason": "source-to-target duplication is expected by cardinality or total/subtotal metadata.",
                "expected_duplicate": True,
            },
            {
                "source_flow": "Industry",
                "source_product": "Bagasse",
                "included_row_count": 2,
                "target_pair_count": 2,
                "target_pairs": "D | E",
                "target_flows": "D | E",
                "source_rows": "4|5",
                "cardinality": "many_to_many",
                "relationship_types": "direct_or_existing_mapping",
                "relationship_levels": "parent",
                "qa_status": "review_expected",
                "qa_severity": "info",
                "qa_reason": "source-to-target duplication is expected by cardinality or total/subtotal metadata.",
                "expected_duplicate": True,
            },
        ]
    )

    exception_path = tmp_path / "mapping_issue_exception_sets.xlsx"
    exception_row = duplicate_source_pairs.iloc[[0]].copy()
    exception_row.insert(0, "enabled", True)
    exception_row["notes"] = "Reviewed manually"
    _write_exception_workbook(
        exception_path,
        "leap_dup_source_allowed",
        exception_row[["enabled", "source_flow", "source_product", "notes"]].to_dict("records"),
    )

    conflicts, allowed = _split_allowed_duplicate_source_pairs(
        duplicate_source_pairs,
        exception_workbook_path=exception_path,
    )

    assert len(allowed) == 1
    assert allowed.loc[0, "duplicate_source_review_status"] == "allowed"
    assert allowed.loc[0, "duplicate_source_review_reason"] == "Reviewed manually"
    assert len(conflicts) == 1
    assert conflicts.loc[0, "source_flow"] == "Industry"


def test_split_allowed_duplicate_target_pairs_uses_manual_allowlist(tmp_path) -> None:
    duplicate_target_pairs = pd.DataFrame(
        [
            {
                "target_flow": "09.06 Gas processing plants",
                "target_product": "17 Electricity",
                "included_row_count": 2,
                "source_pair_count": 2,
                "source_pairs": "A | B",
                "source_rows": "1|2",
                "cardinality": "many_to_many",
                "relationship_types": "direct_or_existing_mapping",
                "relationship_levels": "parent",
                "target_flows": "09.06 Gas processing plants",
                "qa_status": "review_expected",
                "qa_severity": "info",
                "qa_reason": "target-to-source duplication is expected by cardinality or total/subtotal metadata.",
                "expected_duplicate": True,
            },
            {
                "target_flow": "14 Industry sector",
                "target_product": "15 Bagasse",
                "included_row_count": 2,
                "source_pair_count": 2,
                "source_pairs": "C | D",
                "source_rows": "3|4",
                "cardinality": "many_to_many",
                "relationship_types": "direct_or_existing_mapping",
                "relationship_levels": "parent",
                "target_flows": "14 Industry sector",
                "qa_status": "review_expected",
                "qa_severity": "info",
                "qa_reason": "target-to-source duplication is expected by cardinality or total/subtotal metadata.",
                "expected_duplicate": True,
            },
        ]
    )

    exception_path = tmp_path / "mapping_issue_exception_sets.xlsx"
    exception_row = duplicate_target_pairs.iloc[[0]].copy()
    exception_row.insert(0, "enabled", True)
    exception_row["notes"] = "Reviewed manually"
    _write_exception_workbook(
        exception_path,
        "leap_dup_target_allowed",
        exception_row[["enabled", "target_flow", "target_product", "notes"]].to_dict("records"),
    )

    conflicts, allowed = _split_allowed_duplicate_target_pairs(
        duplicate_target_pairs,
        exception_workbook_path=exception_path,
    )

    assert len(allowed) == 1
    assert allowed.loc[0, "duplicate_target_review_status"] == "allowed"
    assert allowed.loc[0, "duplicate_target_review_reason"] == "Reviewed manually"
    assert len(conflicts) == 1
    assert conflicts.loc[0, "target_flow"] == "14 Industry sector"


def test_mapping_style_sector_path_from_export_segments_matches_workbook_paths() -> None:
    fuel_names = {"Electricity", "Natural gas", "Other bituminous coal"}

    assert (
        _mapping_style_sector_path_from_export_segments(
            ["Transformation", "Oil Refining", "Processes", "Oil Refining", "Feedstock Fuels", "Natural gas"],
            fuel_names,
        )
        == "Oil Refining/Oil Refining"
    )
    assert (
        _mapping_style_sector_path_from_export_segments(
            ["Transformation", "Electricity Generation", "Output Fuels", "Electricity"],
            fuel_names,
        )
        == "Electricity Generation"
    )
    assert (
        _mapping_style_sector_path_from_export_segments(
            ["Demand", "Industry", "Manufacturing", "Iron and steel", "Other bituminous coal"],
            fuel_names,
        )
        == "Industry/Manufacturing/Iron and steel"
    )
    assert _mapping_style_sector_path_from_export_segments(["Key", "Macro", "GDP"], fuel_names) == ""


def test_full_export_paths_make_power_parent_branches_subtotals() -> None:
    fuel_names = {"Electricity", "Natural gas", "Wind"}
    export_branch_segments = [
        ["Transformation", "Electricity Generation"],
        ["Transformation", "Electricity Generation", "Output Fuels", "Electricity"],
        ["Transformation", "Electricity Generation", "Processes", "Gas"],
        ["Transformation", "Electricity Generation", "Processes", "Gas", "Feedstock Fuels", "Natural gas"],
        ["Transformation", "Electricity Generation", "Processes", "Wind"],
        ["Transformation", "Electricity Generation", "Processes", "Wind", "Feedstock Fuels", "Wind"],
    ]
    export_paths = {
        _mapping_style_sector_path_from_export_segments(segments, fuel_names)
        for segments in export_branch_segments
    }
    export_paths.discard("")

    assert "Electricity Generation" in _compute_leap_subtotals(export_paths)


def test_reviewed_subtotal_overrides_are_idempotent() -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.title = "leap_combined_esto"
    ws.append([
        "leap_sector_name_full_path",
        "raw_leap_fuel_name",
        "esto_flow",
        "esto_product",
        "leap_is_subtotal",
        "esto_pair_is_subtotal",
    ])
    ws.append(["Industry", "Electricity", "14 Industry", "17 Electricity", False, False])
    overrides = {
        (
            "leap_combined_esto",
            ("Industry", "Electricity", "14 Industry", "17 Electricity"),
            "leap_is_subtotal",
        ): True,
        (
            "leap_combined_esto",
            ("Industry", "Electricity", "14 Industry", "17 Electricity"),
            "esto_pair_is_subtotal",
        ): True,
    }

    assert _apply_subtotal_overrides_to_sheet(ws, overrides) == 2
    first_values = (ws["E2"].value, ws["F2"].value)
    assert _apply_subtotal_overrides_to_sheet(ws, overrides) == 2
    second_values = (ws["E2"].value, ws["F2"].value)

    assert first_values == (True, True)
    assert second_values == first_values


def test_stale_subtotal_overrides_are_reported() -> None:
    mapping_frames = {
        "leap_combined_esto": pd.DataFrame([{
            "leap_sector_name_full_path": "Industry",
            "raw_leap_fuel_name": "Electricity",
            "esto_flow": "14 Industry",
            "esto_product": "17 Electricity",
        }]),
        "leap_combined_ninth": pd.DataFrame(columns=[
            "leap_sector_name_full_path", "raw_leap_fuel_name", "ninth_sector", "ninth_fuel"
        ]),
        "ninth_pairs_to_esto_pairs": pd.DataFrame(columns=[
            "9th_sector", "9th_fuel", "esto_flow", "esto_product"
        ]),
    }
    override_df = pd.DataFrame([
        {
            "enabled": True,
            "sheet": "leap_combined_esto",
            "leap_sector_name_full_path": "Industry",
            "raw_leap_fuel_name": "Electricity",
            "esto_flow": "14 Industry",
            "esto_product": "17 Electricity",
            "leap_is_subtotal": True,
            "esto_pair_is_subtotal": True,
        },
        {
            "enabled": True,
            "sheet": "leap_combined_esto",
            "leap_sector_name_full_path": "Deleted branch",
            "raw_leap_fuel_name": "Electricity",
            "esto_flow": "14 Industry",
            "esto_product": "17 Electricity",
            "leap_is_subtotal": True,
            "esto_pair_is_subtotal": True,
        },
    ])

    stale = _build_stale_subtotal_override_rows(override_df, mapping_frames)

    assert len(stale) == 1
    assert stale.loc[0, "leap_sector_name_full_path"] == "Deleted branch"
    assert stale.loc[0, "stale_reason"] == "mapping_key_not_found"

