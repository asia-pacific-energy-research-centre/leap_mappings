"""Regression tests for Python-only separate-axis workbook generation."""

from __future__ import annotations

import csv
import inspect
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from codebase import separate_axis_mapping_workbooks_builder as production_builder
from codebase.separate_axis_mapping_gap_review_workbook_builder import (
    build_gap_review_workbook,
)
from codebase.separate_axis_mapping_refresh_workflow import (
    run_separate_axis_mapping_refresh,
)


def _write_csv(path: Path, rows: list[list[object]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as file:
        csv.writer(file).writerows(rows)


def _save_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> None:
    workbook = Workbook()
    workbook.remove(workbook.active)
    for name, rows in sheets.items():
        sheet = workbook.create_sheet(name)
        for row in rows:
            sheet.append(row)
    workbook.save(path)


def test_refresh_public_api_has_no_node_runtime_parameter() -> None:
    parameters = inspect.signature(run_separate_axis_mapping_refresh).parameters
    assert "node_executable" not in parameters


def test_gap_review_builder_creates_six_sheet_workbook(tmp_path: Path) -> None:
    data_root = tmp_path / "data"
    preview_root = tmp_path / "previews"
    output = tmp_path / "gap_review.xlsx"
    _write_csv(data_root / "summary.csv", [["section", "metric", "value"], ["A", "rows", 1]])
    _write_csv(
        data_root / "missing_mappings.csv",
        [["review_queue", "source_flow", "active"], ["boundary_policy_review", "Flow", "true"]],
    )
    for filename in (
        "exact_subtotal_differences.csv",
        "master_subtotal_review.csv",
        "incomplete_current_rows.csv",
    ):
        _write_csv(data_root / filename, [["reason", "subtotal"], ["Review", "false"]])

    built = build_gap_review_workbook(data_root, output, preview_root)

    workbook = load_workbook(built)
    assert workbook.sheetnames == [
        "README",
        "Summary",
        "Missing mappings",
        "Exact subtotal differences",
        "Master subtotal review",
        "Incomplete current rows",
    ]
    assert workbook["Missing mappings"]["C2"].value == "TRUE"
    assert len(list(preview_root.glob("*.png"))) == 6


def test_production_builder_preserves_contract_and_literal_booleans(
    tmp_path: Path,
    monkeypatch: object,
) -> None:
    output_root = tmp_path / "workbooks"
    data_root = output_root / "data"
    previews = output_root / "previews"
    canonical = tmp_path / "canonical.xlsx"
    editable = tmp_path / "editable.xlsx"
    pairs = tmp_path / "pairs.xlsx"
    candidate = tmp_path / "candidate.xlsx"
    compiled_headers = ["leap_sector", "leap_fuel", "leap_is_subtotal"]
    canonical_sheets = {
        "Guide": [["old guide"]],
        "compiled": [compiled_headers, ["Old", "Old", False]],
        "leap_rollup_rules": [["from", "to"], ["A", "B"]],
        "esto_rollup_rules": [["from", "to"], ["C", "D"]],
        "ninth_rollup_rules": [["from", "to"], ["E", "F"]],
        "rollup_label_overrides": [["from", "to"]],
        "preserved": [["value"], [42]],
    }
    _save_workbook(canonical, canonical_sheets)
    editable_sheets = {
        "leap_sector_to_esto": [
            ["leap_sector", "esto_flow", "esto_dataset_scope"],
            ["Road", "15.02 Road", "ESTO"],
            ["Road", "15.02 Road", "ESTO_EXTENDED"],
        ],
        "exceptions": [production_builder.EXCEPTION_HEADERS, production_builder.INITIAL_EXCEPTION_ROW],
        "leap_rollup_rules": [["from", "to"], ["A", "B"]],
        "esto_rollup_rules": [["from", "to"], ["C", "D"]],
        "ninth_rollup_rules": [["from", "to"], ["E", "F"]],
        "rollup_label_overrides": [["from", "to"]],
    }
    _save_workbook(editable, editable_sheets)
    _write_csv(
        data_root / "editable.csv",
        [["leap_sector", "esto_flow", "esto_dataset_scope"], ["Road", "15.02 Road", "BOTH"]],
    )
    _write_csv(
        data_root / "pairs.csv",
        [["leap_sector", "leap_fuel", "exists_in_dataset", "pair_is_subtotal"], ["Road", "Gas", "true", "false"]],
    )
    _write_csv(
        data_root / "compiled.csv",
        [compiled_headers, ["Road", "Gas", "true"]],
    )
    manifest = {
        "prototype_status": "Test contract.",
        "historical_boundary_year": 2023,
        "canonical_master_path": str(canonical),
        "editable_axis_workbook_path": str(editable),
        "generated_pair_workbook_path": str(pairs),
        "generated_master_workbook_path": str(candidate),
        "editable_sources": {"leap_sector_to_esto": "data/editable.csv"},
        "editable_counts": {"leap_sector_to_esto": 1},
        "pair_sources": {"LEAP key pairs": "data/pairs.csv"},
        "pair_counts": {"LEAP key pairs": 1},
        "compiled_sources": {"compiled": "data/compiled.csv"},
        "compiled_counts": {"compiled": 1},
        "compiled_columns": {"compiled": compiled_headers},
        "leap_registry_authority": "Test authority.",
    }
    manifest_path = output_root / "split_workbook_manifest.json"
    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest), encoding="utf-8")
    monkeypatch.setattr(production_builder, "OUTPUT_ROOT", output_root)
    monkeypatch.setattr(production_builder, "PREVIEW_ROOT", previews)
    monkeypatch.setattr(
        production_builder,
        "EDITABLE_DUPLICATE_AUDIT_PATH",
        output_root / "editable_duplicate_cleanup.json",
    )
    monkeypatch.setattr(
        production_builder,
        "REPLACEMENT_SUMMARY_PATH",
        output_root / "generated_master_replacement_summary.json",
    )
    monkeypatch.setattr(
        production_builder,
        "MAPPING_WORKFLOW_DIAGRAM_PATH",
        tmp_path / "missing_diagram.png",
    )

    result = production_builder.build_separate_axis_mapping_workbooks(
        promote_master=False,
        rebuild_editable_workbook=False,
        clean_editable_duplicates=True,
        split_manifest_path=manifest_path,
    )

    assert result["status"] == "candidate_generated_not_promoted"
    pair_workbook = load_workbook(pairs)
    assert pair_workbook["LEAP key pairs"]["C2"].value is True
    assert pair_workbook["LEAP key pairs"]["D2"].value is False
    candidate_workbook = load_workbook(candidate)
    assert candidate_workbook.sheetnames == list(canonical_sheets)
    assert candidate_workbook["compiled"]["C2"].value is True
    assert candidate_workbook["preserved"]["A2"].value == 42
    cleaned_editable = load_workbook(editable)
    assert cleaned_editable["leap_sector_to_esto"].max_row == 2
    assert cleaned_editable["leap_sector_to_esto"]["C2"].value == "BOTH"
    audit = json.loads((output_root / "editable_duplicate_cleanup.json").read_text())
    assert audit["duplicate_rows_removed"] == 1
