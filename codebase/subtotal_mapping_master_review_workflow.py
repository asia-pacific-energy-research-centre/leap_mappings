#%%
"""Verify mapping-workbook subtotal flags against their source hierarchies.

The mapping sheets are loaded as pandas DataFrames. ESTO, Ninth, LEAP, and
ESTO Extended pairs are classified by the canonical hierarchy adapters. The
workflow writes diagnostics plus a copied workbook with a CHANGED column; it
never modifies the source workbook or the exception workbook.
"""

from __future__ import annotations

from copy import copy
import json
from pathlib import Path
import shutil

import openpyxl
import pandas as pd

from codebase.mapping_issue_exceptions import (
    load_exception_sheet,
    matching_exception_row,
)
from codebase.mapping_tools.hierarchy_subtotal_adapters import (
    current_adapter_registry,
)
from codebase.mapping_tools.hierarchy_subtotal_contract import (
    build_contract_frames,
    load_contract,
    write_contract,
)
from codebase.mapping_tools.hierarchy_subtotal_review import (
    SIDE_CONFIGS,
    build_review_frames,
    write_review_csvs,
)


# --- Stable configuration ---

REPO_ROOT = Path(__file__).resolve().parents[1]
MAPPING_SHEETS = tuple(sorted({config["sheet"] for config in SIDE_CONFIGS}))
COMPLETE_HIERARCHY_STATUSES = {
    "complete_declared_code_list",
    "complete_declared_schema",
    "derived_declared_structure",
}
CHANGED_COLUMN_BY_SHEET = {
    "leap_combined_esto": 9,
    "leap_combined_ninth": 8,
    "ninth_pairs_to_esto_pairs": 9,
}


def _resolve(path_value: str | Path) -> Path:
    normalized = Path(str(path_value).replace("\\", "/"))
    return normalized if normalized.is_absolute() else REPO_ROOT / normalized


def _truthy(value: object) -> bool:
    if value is True:
        return True
    return str(value or "").strip().casefold() in {"true", "1", "yes", "y"}


def load_mapping_sheets_as_dataframes(
    workbook_path: str | Path,
) -> dict[str, pd.DataFrame]:
    """Load the three maintained mapping sheets once, preserving Excel indexes."""
    path = _resolve(workbook_path)
    frames = pd.read_excel(
        path,
        sheet_name=list(MAPPING_SHEETS),
        dtype=object,
    )
    missing = sorted(set(MAPPING_SHEETS).difference(frames))
    if missing:
        raise ValueError(f"Mapping workbook is missing sheets: {missing}")
    return frames


def build_source_dataset_contract(
    *,
    workbook_path: str | Path,
    exception_workbook_path: str | Path,
    contract_output_dir: str | Path,
    review_csv_dir: str | Path,
) -> tuple[
    dict[str, object],
    dict[str, pd.DataFrame],
    dict[str, pd.DataFrame],
    dict[str, pd.DataFrame],
]:
    """Build a raw/derived-source contract without Common ESTO output state."""
    workbook = _resolve(workbook_path)
    exceptions = _resolve(exception_workbook_path)
    contract_dir = _resolve(contract_output_dir)
    review_dir = _resolve(review_csv_dir)
    mapping_frames = load_mapping_sheets_as_dataframes(workbook)

    adapters = current_adapter_registry(
        REPO_ROOT,
        workbook,
        include_common_esto=False,
    )
    contract_frames, registry = build_contract_frames(adapters)
    input_paths = [
        workbook,
        exceptions,
        REPO_ROOT / "data" / "00APEC_2025_low_with_subtotals.csv",
        REPO_ROOT / "data" / "merged_file_energy_ALL_20251106.csv",
        REPO_ROOT / "data" / "temp" / "new leap rows.xlsx",
        REPO_ROOT / "results" / "tree_structure" / "esto_extended_tree.csv",
    ]
    manifest = write_contract(
        output_dir=contract_dir,
        frames=contract_frames,
        registry=registry,
        input_paths=input_paths,
        repo_root=REPO_ROOT,
        compatibility={
            "purpose": "mapping_subtotal_verification",
            "common_esto_dependency": "excluded",
        },
    )
    _, loaded_frames = load_contract(
        contract_dir,
        expected_build_id=str(manifest["build_id"]),
    )
    review_frames = build_review_frames(
        workbook,
        exceptions,
        loaded_frames,
        workbook_sheets=mapping_frames,
    )
    write_review_csvs(review_dir, review_frames)
    return manifest, loaded_frames, review_frames, mapping_frames


def _mapping_row(
    mapping_frames: dict[str, pd.DataFrame],
    sheet_name: str,
    excel_row: int,
) -> pd.Series:
    frame = mapping_frames[sheet_name]
    dataframe_index = excel_row - 2
    if dataframe_index not in frame.index:
        raise ValueError(f"{sheet_name} row {excel_row} is outside the DataFrame.")
    return frame.loc[dataframe_index]


def _label_exception_candidate(cell: pd.Series) -> pd.Series:
    dataset_id = str(cell["dataset_id"])
    draft_type = {
        "leap": "leap_pairs",
        "ninth": "ninth_pairs",
        "esto": "esto_pairs",
        "esto_extended": "esto_pairs",
    }.get(dataset_id, "")
    return pd.Series(
        {
            "draft_type": draft_type,
            "key_1": cell["axis_1_source_key"],
            "key_2": cell["axis_2_source_key"],
            "accepted_value": cell["current_value"],
            "proposed_value": cell["proposed_value"],
        }
    )


def attach_exception_evidence(
    *,
    affected_cells: pd.DataFrame,
    mapping_frames: dict[str, pd.DataFrame],
    exception_workbook_path: str | Path,
) -> pd.DataFrame:
    """Attach existing exception decisions without treating them as hierarchy."""
    output = affected_cells.copy()
    exception_path = _resolve(exception_workbook_path)
    exception_frames = {
        sheet: load_exception_sheet(sheet, workbook_path=exception_path)
        for sheet in [
            "subtotal_mismatch_allowed",
            "subtotal_label_exceptions",
            "subtotal_label_overrides",
        ]
    }
    records: list[dict[str, object]] = []
    for _, cell in output.iterrows():
        mapping_row = _mapping_row(
            mapping_frames,
            str(cell["mapping_sheet"]),
            int(cell["excel_row"]),
        ).copy()
        mapping_row["sheet"] = cell["mapping_sheet"]
        label_candidate = _label_exception_candidate(cell)

        mismatch = matching_exception_row(
            mapping_row,
            exception_frames["subtotal_mismatch_allowed"],
        )
        override = matching_exception_row(
            mapping_row,
            exception_frames["subtotal_label_overrides"],
        )
        label_exception = matching_exception_row(
            label_candidate,
            exception_frames["subtotal_label_exceptions"],
        )
        records.append(
            {
                "subtotal_mismatch_exception": mismatch is not None,
                "subtotal_mismatch_exception_notes": (
                    "" if mismatch is None else str(mismatch.get("notes", "") or "")
                ),
                "subtotal_label_override": override is not None,
                "subtotal_label_override_notes": (
                    "" if override is None else str(override.get("notes", "") or "")
                ),
                "subtotal_label_exception": label_exception is not None,
                "subtotal_label_exception_notes": (
                    ""
                    if label_exception is None
                    else str(label_exception.get("notes", "") or "")
                ),
            }
        )
    return pd.concat(
        [output.reset_index(drop=True), pd.DataFrame(records)],
        axis=1,
    )


def classify_suggestion_actions(affected_cells: pd.DataFrame) -> pd.DataFrame:
    """Classify which proposals are safe to apply to the copied workbook."""
    output = affected_cells.copy()
    complete = (
        output["every_node_resolved"].map(_truthy)
        & output["axis_1_hierarchy_status"].isin(COMPLETE_HIERARCHY_STATUSES)
        & output["axis_2_hierarchy_status"].isin(COMPLETE_HIERARCHY_STATUSES)
    )
    prior_label_decision = (
        output["subtotal_label_override"].map(_truthy)
        | output["subtotal_label_exception"].map(_truthy)
    )
    output["suggestion_action"] = "review_partial_or_unresolved_source"
    output.loc[
        complete & prior_label_decision,
        "suggestion_action",
    ] = "review_conflicts_with_prior_label_exception"
    output.loc[
        complete & ~prior_label_decision,
        "suggestion_action",
    ] = "apply_complete_source_evidence"
    return output


def _annotation(cell: pd.Series) -> str:
    current = str(_truthy(cell["current_value"]))
    proposed = str(_truthy(cell["proposed_value"]))
    evidence = (
        f"{cell['dataset_id']} original hierarchy: "
        f"{cell['axis_1_hierarchy_status']} + "
        f"{cell['axis_2_hierarchy_status']}"
    )
    action = str(cell["suggestion_action"])
    if action == "apply_complete_source_evidence":
        prefix = "UPDATED"
    elif action == "review_conflicts_with_prior_label_exception":
        prefix = "SUGGESTED — NOT APPLIED; PRIOR LABEL EXCEPTION"
    else:
        prefix = "REVIEW REQUIRED — NOT APPLIED"
    note = (
        f"{prefix}: {cell['flag_column']} {current} -> {proposed} "
        f"[{evidence}]"
    )
    if _truthy(cell.get("subtotal_mismatch_exception", False)):
        note += " [cross-dataset subtotal mismatch is already allowed]"
    return note


def _copy_cell_style(source: openpyxl.cell.cell.Cell, target: openpyxl.cell.cell.Cell) -> None:
    if source.has_style:
        target._style = copy(source._style)
    if source.number_format:
        target.number_format = source.number_format
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.protection = copy(source.protection)


def write_review_workbook(
    *,
    source_workbook_path: str | Path,
    output_workbook_path: str | Path,
    affected_cells: pd.DataFrame,
) -> dict[str, int]:
    """Copy the workbook, apply safe proposals, and add CHANGED annotations."""
    source = _resolve(source_workbook_path)
    output = _resolve(output_workbook_path)
    output.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source, output)
    workbook = openpyxl.load_workbook(output)

    updated_cells = 0
    annotated_rows = 0
    for sheet_name, sheet_cells in affected_cells.groupby("mapping_sheet"):
        worksheet = workbook[str(sheet_name)]
        headers = {
            str(cell.value).strip(): cell.column
            for cell in worksheet[1]
            if cell.value is not None and str(cell.value).strip()
        }
        changed_column = CHANGED_COLUMN_BY_SHEET[str(sheet_name)]
        if worksheet.cell(1, changed_column).value not in {None, "", "CHANGED"}:
            raise ValueError(
                f"{sheet_name} column {changed_column} is not available for CHANGED."
            )
        style_source_column = changed_column - 1
        _copy_cell_style(
            worksheet.cell(1, style_source_column),
            worksheet.cell(1, changed_column),
        )
        worksheet.cell(1, changed_column).value = "CHANGED"
        worksheet.column_dimensions[
            openpyxl.utils.get_column_letter(changed_column)
        ].width = 85

        for excel_row, row_cells in sheet_cells.groupby("excel_row"):
            annotations: list[str] = []
            for _, cell in row_cells.sort_values("flag_column").iterrows():
                flag_column = str(cell["flag_column"])
                flag_index = headers[flag_column]
                workbook_value = _truthy(worksheet.cell(int(excel_row), flag_index).value)
                expected_current = _truthy(cell["current_value"])
                if workbook_value != expected_current:
                    raise ValueError(
                        f"{sheet_name}!{flag_column}{excel_row} changed since verification."
                    )
                if cell["suggestion_action"] == "apply_complete_source_evidence":
                    worksheet.cell(int(excel_row), flag_index).value = _truthy(
                        cell["proposed_value"]
                    )
                    updated_cells += 1
                annotations.append(_annotation(cell))

            changed_cell = worksheet.cell(int(excel_row), changed_column)
            _copy_cell_style(
                worksheet.cell(int(excel_row), style_source_column),
                changed_cell,
            )
            changed_cell.value = " | ".join(annotations)
            changed_cell.alignment = copy(changed_cell.alignment)
            changed_cell.alignment = openpyxl.styles.Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True,
            )
            annotated_rows += 1

    workbook.save(output)
    return {
        "updated_cells": updated_cells,
        "annotated_rows": annotated_rows,
    }


def verify_exported_workbook(
    *,
    output_workbook_path: str | Path,
    affected_cells: pd.DataFrame,
) -> None:
    """Re-import the output as DataFrames and verify every affected cell."""
    output = _resolve(output_workbook_path)
    frames = pd.read_excel(
        output,
        sheet_name=list(MAPPING_SHEETS),
        dtype=object,
    )
    for _, cell in affected_cells.iterrows():
        sheet_name = str(cell["mapping_sheet"])
        dataframe_row = int(cell["excel_row"]) - 2
        row = frames[sheet_name].loc[dataframe_row]
        expected = (
            _truthy(cell["proposed_value"])
            if cell["suggestion_action"] == "apply_complete_source_evidence"
            else _truthy(cell["current_value"])
        )
        actual = _truthy(row[cell["flag_column"]])
        if actual != expected:
            raise ValueError(
                f"{sheet_name} row {cell['excel_row']} {cell['flag_column']} "
                "did not survive workbook export."
            )
        changed_value = str(row.get("CHANGED", "") or "")
        if not changed_value:
            raise ValueError(
                f"{sheet_name} row {cell['excel_row']} lost its CHANGED annotation."
            )


def run_subtotal_mapping_review(
    *,
    source_workbook_path: str | Path,
    exception_workbook_path: str | Path,
    output_dir: str | Path,
    output_workbook_name: str,
) -> dict[str, object]:
    """Run the reusable source-dataset subtotal verification system."""
    output = _resolve(output_dir)
    contract_dir = output / "contract"
    review_dir = output / "review_csv"
    output_workbook = output / output_workbook_name

    manifest, _, review_frames, mapping_frames = build_source_dataset_contract(
        workbook_path=source_workbook_path,
        exception_workbook_path=exception_workbook_path,
        contract_output_dir=contract_dir,
        review_csv_dir=review_dir,
    )
    affected = attach_exception_evidence(
        affected_cells=review_frames["affected_workbook_cells"],
        mapping_frames=mapping_frames,
        exception_workbook_path=exception_workbook_path,
    )
    affected = classify_suggestion_actions(affected)
    affected["changed_annotation"] = affected.apply(_annotation, axis=1)
    affected.to_csv(
        review_dir / "affected_workbook_cells_with_actions.csv",
        index=False,
        lineterminator="\n",
    )

    workbook_result = write_review_workbook(
        source_workbook_path=source_workbook_path,
        output_workbook_path=output_workbook,
        affected_cells=affected,
    )
    verify_exported_workbook(
        output_workbook_path=output_workbook,
        affected_cells=affected,
    )

    action_counts = {
        str(key): int(value)
        for key, value in affected["suggestion_action"].value_counts().items()
    }
    result = {
        "contract_build_id": manifest["build_id"],
        "source_workbook": str(_resolve(source_workbook_path)),
        "output_workbook": str(output_workbook),
        "affected_cells": int(len(affected)),
        "affected_rows": int(
            affected[["mapping_sheet", "excel_row"]].drop_duplicates().shape[0]
        ),
        "action_counts": action_counts,
        **workbook_result,
    }
    (output / "review_summary.json").write_text(
        json.dumps(result, indent=2),
        encoding="utf-8",
    )
    return result


# --- Notebook-friendly run block ---

RUN_REVIEW = False
SOURCE_WORKBOOK_PATH = (
    "outputs/subtotal_mapping_master_review_fad4223/"
    "outlook_mappings_master_todo_fad4223.xlsx"
)
EXCEPTION_WORKBOOK_PATH = "config/mapping_issue_exception_sets.xlsx"
OUTPUT_DIR = "outputs/subtotal_mapping_master_review_fad4223"
OUTPUT_WORKBOOK_NAME = (
    "outlook_mappings_master_todo_fad4223_subtotal_review.xlsx"
)


#%%
if __name__ == "__main__" and RUN_REVIEW:
    RESULT = run_subtotal_mapping_review(
        source_workbook_path=SOURCE_WORKBOOK_PATH,
        exception_workbook_path=EXCEPTION_WORKBOOK_PATH,
        output_dir=OUTPUT_DIR,
        output_workbook_name=OUTPUT_WORKBOOK_NAME,
    )
    print(json.dumps(RESULT, indent=2))


#%%
