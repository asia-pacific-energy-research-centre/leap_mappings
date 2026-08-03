#%%
"""Build the separate-axis gap review workbook with Python and openpyxl."""

#%%
from __future__ import annotations

import json
import traceback
from pathlib import Path

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from codebase.workbook_python_helpers import (
    COLORS,
    add_contains_text_rule,
    add_excel_table,
    new_workbook,
    read_csv_rows,
    render_sheet_preview,
    reopen_workbook,
    scan_formula_errors,
    write_rows,
)


# --- Stable paths -----------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_ROOT = REPO_ROOT / "outputs" / "separate_axis_mapping_gap_review_20260729"
DATA_ROOT = OUTPUT_ROOT / "data"
PREVIEW_ROOT = OUTPUT_ROOT / "previews"
OUTPUT_PATH = OUTPUT_ROOT / "separate_axis_mapping_gap_and_subtotal_review.xlsx"


# --- Formatting helpers -----------------------------------------------------

def _add_title(
    sheet: object,
    title: str,
    subtitle: str,
    column_count: int,
) -> None:
    """Add a full-width report title and subtitle."""
    last_column = get_column_letter(max(column_count, 3))
    sheet.merge_cells(f"A1:{last_column}1")
    sheet["A1"] = title
    sheet.merge_cells(f"A2:{last_column}2")
    sheet["A2"] = subtitle
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=COLORS["blue"])
    sheet["A1"].font = Font(bold=True, color=COLORS["white"], size=16)
    sheet["A1"].alignment = Alignment(vertical="center")
    for cell in sheet[2]:
        cell.fill = PatternFill("solid", fgColor=COLORS["pale_blue"])
    sheet["A2"].font = Font(color=COLORS["dark_grey"], italic=True)
    sheet["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.row_dimensions[2].height = 34
    sheet.sheet_view.showGridLines = False


def _style_data_sheet(
    sheet: object,
    rows: list[list[object]],
    table_name: str,
) -> None:
    """Write and style one review data table."""
    if not rows:
        return
    write_rows(sheet, rows)
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=COLORS["blue"])
        cell.font = Font(bold=True, color=COLORS["white"])
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[1].height = 42
    sheet.freeze_panes = "E2" if sheet.max_column >= 4 else "A2"
    sheet.sheet_view.showGridLines = False
    add_excel_table(sheet, table_name)
    for index, cell in enumerate(sheet[1], start=1):
        header = str(cell.value or "")
        width = 18
        if "diagnostic" in header or "reason" in header:
            width = 55
        elif "review_queue" in header:
            width = 36
        elif "flow" in header or "product" in header:
            width = 31
        elif "status" in header:
            width = 28
        elif "note" in header:
            width = 48
        elif "workbook_row" in header:
            width = 14
        elif any(token in header for token in ("count", "year", "subtotal", "active")):
            width = 16
        sheet.column_dimensions[get_column_letter(index)].width = width


# --- Workbook build ---------------------------------------------------------

def build_gap_review_workbook(
    data_root: str | Path = DATA_ROOT,
    output_path: str | Path = OUTPUT_PATH,
    preview_root: str | Path = PREVIEW_ROOT,
) -> Path:
    """Create, reopen, validate, and preview the review-only workbook."""
    data_path = Path(data_root)
    output = Path(output_path)
    preview_path = Path(preview_root)
    source_names = {
        "summary": "summary.csv",
        "missing": "missing_mappings.csv",
        "exact": "exact_subtotal_differences.csv",
        "master": "master_subtotal_review.csv",
        "incomplete": "incomplete_current_rows.csv",
    }
    missing_sources = [
        str(data_path / filename)
        for filename in source_names.values()
        if not (data_path / filename).exists()
    ]
    if missing_sources:
        raise FileNotFoundError("Missing gap-review sources:\n- " + "\n- ".join(missing_sources))
    rows = {
        name: read_csv_rows(data_path / filename, boolean_as_text=True)
        for name, filename in source_names.items()
    }

    workbook = new_workbook()
    readme = workbook.create_sheet("README")
    _add_title(
        readme,
        "Separate-axis gap and subtotal review",
        "Review-only evidence. The canonical mapping workbook has not been edited.",
        8,
    )
    readme_rows = [
        ["Purpose", "Inspect every maintained relationship omitted by the temporal compiler and review subtotal inconsistencies."],
        ["Missing rows", "The Missing mappings sheet contains all current relationships not generated."],
        ["ESTO any-year", "Direct exact-pair evidence from base ESTO across every available year."],
        ["ESTO Extended any-year", "Direct exact-pair evidence from ESTO Extended across every available year."],
        ["Ninth any-year", "Direct exact-pair evidence from Ninth across every available year and scenario in the selected registry."],
        ["Cross-dataset caution", "ESTO codes are not looked up as Ninth codes, or vice versa. Non-applicable evidence cells are blank."],
        ["Absent status caution", "The compiler label absent can also mean that the source pair was not generated. Use primary_diagnostic instead."],
        ["Subtotal review", "Generated structural flags are evidence, not automatically authoritative. Mixed master flags are direct internal inconsistencies."],
        ["Recommended use", "Filter review_queue and primary_diagnostic first; then inspect the original workbook row number."],
        ["Safety", "Do not paste these rows into the maintained master automatically."],
    ]
    thin = Side(style="thin", color="B4C6E7")
    for offset, values in enumerate(readme_rows, start=4):
        readme.cell(offset, 1, values[0])
        readme.cell(offset, 2, values[1])
        readme.cell(offset, 1).fill = PatternFill("solid", fgColor=COLORS["light_blue"])
        readme.cell(offset, 1).font = Font(bold=True)
        for column in (1, 2):
            readme.cell(offset, column).alignment = Alignment(wrap_text=True, vertical="top")
            readme.cell(offset, column).border = Border(left=thin, right=thin, top=thin, bottom=thin)
        readme.row_dimensions[offset].height = 34
    readme.column_dimensions["A"].width = 27
    readme.column_dimensions["B"].width = 95

    summary = workbook.create_sheet("Summary")
    _add_title(
        summary,
        "What the missing mappings actually represent",
        "Counts separate boundary-policy questions from likely source-authority, structural-pair, and subtotal issues.",
        3,
    )
    headlines = [
        ["Headline", "Count", "Interpretation"],
        ["Only boundary-window evidence issue", 811, "These pairs have non-zero evidence in another year but fail the selected final-year/future-window rule."],
        ["Stronger mapping or authority review", 3562, "These involve an absent source pair, missing structural target pair, or zero values across all available years."],
    ]
    for row_number, values in enumerate(headlines, start=4):
        for column, value in enumerate(values, start=1):
            summary.cell(row_number, column, value)
            summary.cell(row_number, column).alignment = Alignment(wrap_text=True)
    for cell in summary[4]:
        cell.fill = PatternFill("solid", fgColor=COLORS["blue"])
        cell.font = Font(bold=True, color=COLORS["white"])
    for cell in summary[5]:
        cell.fill = PatternFill("solid", fgColor=COLORS["edit"])
    for cell in summary[6]:
        cell.fill = PatternFill("solid", fgColor=COLORS["warning"])
    for column, value in enumerate(["Section", "Metric", "Value"], start=1):
        summary.cell(8, column, value)
        summary.cell(8, column).fill = PatternFill("solid", fgColor=COLORS["blue"])
        summary.cell(8, column).font = Font(bold=True, color=COLORS["white"])
    for source_row in rows["summary"][1:]:
        summary.append(source_row[:3])
    summary.column_dimensions["A"].width = 26
    summary.column_dimensions["B"].width = 54
    summary.column_dimensions["C"].width = 70
    summary.row_dimensions[5].height = 54
    summary.row_dimensions[6].height = 54
    summary.freeze_panes = "A9"

    sheets = [
        ("Missing mappings", rows["missing"], "MissingMappingsTable"),
        ("Exact subtotal differences", rows["exact"], "ExactSubtotalDifferencesTable"),
        ("Master subtotal review", rows["master"], "MasterSubtotalReviewTable"),
        ("Incomplete current rows", rows["incomplete"], "IncompleteCurrentRowsTable"),
    ]
    for sheet_name, sheet_rows, table_name in sheets:
        sheet = workbook.create_sheet(sheet_name)
        _style_data_sheet(sheet, sheet_rows, table_name)
    missing_sheet = workbook["Missing mappings"]
    if missing_sheet.max_row > 1:
        target_range = f"A2:A{missing_sheet.max_row}"
        add_contains_text_rule(missing_sheet, target_range, "boundary_policy_review", COLORS["edit"])
        add_contains_text_rule(missing_sheet, target_range, "strong_mapping_review", COLORS["red"])
        add_contains_text_rule(missing_sheet, target_range, "source_authority_or_mapping_review", COLORS["warning"])

    formula_errors = scan_formula_errors(workbook)
    if formula_errors:
        raise ValueError("Formula errors found before export:\n- " + "\n- ".join(formula_errors))
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    reopened = reopen_workbook(output)
    if reopened.sheetnames != [
        "README",
        "Summary",
        "Missing mappings",
        "Exact subtotal differences",
        "Master subtotal review",
        "Incomplete current rows",
    ]:
        raise ValueError(f"Gap-review sheet contract changed: {reopened.sheetnames}")
    for sheet_name in reopened.sheetnames:
        render_sheet_preview(
            reopened,
            sheet_name,
            preview_path / f"{sheet_name.lower().replace(' ', '_')}.png",
        )
    return output


# --- Frequently changed run flag -------------------------------------------

BUILD_GAP_REVIEW_WORKBOOK = True


#%%
if __name__ == "__main__" and BUILD_GAP_REVIEW_WORKBOOK:
    try:
        BUILT_GAP_REVIEW_PATH = build_gap_review_workbook()
        print(json.dumps({"output_path": str(BUILT_GAP_REVIEW_PATH)}, indent=2))
    except Exception as error:
        print("Failed to build separate-axis gap review workbook.")
        print(f"{type(error).__name__}: {error}")
        traceback.print_exc()
        raise


#%%
