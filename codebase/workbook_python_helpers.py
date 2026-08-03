#%%
"""Shared openpyxl helpers for Python-only mapping workbook generation."""

#%%
from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from PIL import Image, ImageDraw, ImageFont


COLORS = {
    "navy": "17365D",
    "blue": "1F4E78",
    "light_blue": "D9EAF7",
    "pale_blue": "EAF3F8",
    "edit": "FFF2CC",
    "edit_border": "D6B656",
    "generated": "E7E6E6",
    "warning": "FCE4D6",
    "warning_text": "9C0006",
    "green": "E2F0D9",
    "green_text": "375623",
    "red": "F4CCCC",
    "border": "B4C6D7",
    "dark_grey": "595959",
    "white": "FFFFFF",
}

BOOLEAN_HEADERS = {
    "exists_in_dataset",
    "active_in_final_esto_year",
    "active_after_final_esto_year",
    "eligible_for_compilation",
    "pair_is_subtotal",
    "leap_is_subtotal",
    "esto_pair_is_subtotal",
    "ninth_pair_is_subtotal",
    "duplicate_to_remove",
}

FORMULA_ERROR_PATTERN = re.compile(
    r"#(?:REF!|DIV/0!|VALUE!|NAME\?|N/A)", re.IGNORECASE
)
INTEGER_PATTERN = re.compile(r"^-?(?:0|[1-9]\d*)$")
DECIMAL_PATTERN = re.compile(r"^-?(?:0|[1-9]\d*)\.\d+$")


def new_workbook() -> Workbook:
    """Return an empty workbook without openpyxl's default sheet."""
    workbook = Workbook()
    workbook.remove(workbook.active)
    return workbook


def read_csv_rows(
    path: str | Path,
    boolean_headers: Iterable[str] = BOOLEAN_HEADERS,
    boolean_as_text: bool = False,
) -> list[list[object]]:
    """Read a CSV while keeping identifiers as text and normalising booleans."""
    with Path(path).open("r", encoding="utf-8-sig", newline="") as file:
        rows = list(csv.reader(file))
    if not rows:
        return []
    bool_headers = set(boolean_headers)
    headers = rows[0]
    output: list[list[object]] = [headers]
    for row in rows[1:]:
        converted: list[object] = []
        for index, value in enumerate(row):
            header = headers[index] if index < len(headers) else ""
            lowered = value.strip().lower()
            if boolean_as_text and lowered in {"true", "false"}:
                converted.append(lowered.upper())
            elif header in bool_headers and lowered in {"true", "false"}:
                converted.append(lowered == "true")
            elif INTEGER_PATTERN.fullmatch(value.strip()):
                converted.append(int(value))
            elif DECIMAL_PATTERN.fullmatch(value.strip()):
                converted.append(float(value))
            else:
                converted.append(value)
        output.append(converted)
    return output


def write_rows(sheet: object, rows: list[list[object]]) -> None:
    """Write a rectangular matrix to a worksheet."""
    for row in rows:
        sheet.append(row)


def style_title(
    sheet: object,
    title: str,
    banner_text: str,
    banner_fill: str,
    banner_font: str,
) -> None:
    """Apply the shared production-workbook title and banner."""
    sheet.sheet_view.showGridLines = False
    sheet.merge_cells("A1:H2")
    sheet["A1"] = title
    sheet["A1"].fill = PatternFill("solid", fgColor=COLORS["navy"])
    sheet["A1"].font = Font(
        name="Aptos Display", size=18, bold=True, color=COLORS["white"]
    )
    sheet["A1"].alignment = Alignment(vertical="center", horizontal="left")
    for row in sheet["A1:H2"]:
        for cell in row:
            cell.fill = PatternFill("solid", fgColor=COLORS["navy"])
    sheet.merge_cells("A4:H4")
    sheet["A4"] = banner_text
    sheet["A4"].fill = PatternFill("solid", fgColor=banner_fill)
    sheet["A4"].font = Font(
        name="Aptos", size=11, bold=True, color=banner_font
    )
    sheet["A4"].alignment = Alignment(vertical="center", wrap_text=True)
    sheet.row_dimensions[4].height = 36


def style_readme_table(
    sheet: object,
    start_row: int,
    rows: list[list[object]],
) -> None:
    """Write and style a two-column README table."""
    thin = Side(style="thin", color=COLORS["border"])
    for offset, row in enumerate(rows):
        row_number = start_row + offset
        sheet.cell(row_number, 1, row[0])
        sheet.cell(row_number, 2, row[1])
        for column in (1, 2):
            cell = sheet.cell(row_number, column)
            cell.font = Font(name="Aptos", size=10, bold=column == 1)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
        sheet.cell(row_number, 1).fill = PatternFill(
            "solid", fgColor=COLORS["light_blue"]
        )
    sheet.column_dimensions["A"].width = 35
    sheet.column_dimensions["B"].width = 95
    for column in range(3, 9):
        sheet.column_dimensions[get_column_letter(column)].width = 4


def style_data_sheet(sheet: object, editable: bool) -> None:
    """Apply the shared table-sheet presentation without form controls."""
    sheet.sheet_view.showGridLines = False
    sheet.freeze_panes = "A2"
    if sheet.max_row < 1 or sheet.max_column < 1:
        return
    header_fill = COLORS["blue"] if editable else COLORS["navy"]
    for cell in sheet[1]:
        cell.fill = PatternFill("solid", fgColor=header_fill)
        cell.font = Font(name="Aptos", size=10, bold=True, color=COLORS["white"])
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    sheet.row_dimensions[1].height = 34
    # Generated sheets can contain tens of thousands of rows. Their default
    # body style is already readable, so avoid creating several style objects
    # per cell. Editable sheets are small and keep their yellow input styling.
    if editable:
        hair = Side(style="hair", color=COLORS["edit_border"])
        editable_fill = PatternFill("solid", fgColor=COLORS["edit"])
        editable_font = Font(name="Aptos", size=9)
        editable_alignment = Alignment(vertical="top")
        editable_border = Border(bottom=hair)
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                cell.fill = editable_fill
                cell.font = editable_font
                cell.alignment = editable_alignment
                cell.border = editable_border
    headers = [str(cell.value or "") for cell in sheet[1]]
    for column, label in enumerate(headers, start=1):
        width = 38
        if any(
            token in label
            for token in (
                "scope",
                "status",
                "eligible",
                "active_",
                "exists_",
                "subtotal",
            )
        ):
            width = 42 if "status" in label else 23
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.auto_filter.ref = sheet.dimensions


def add_excel_table(sheet: object, table_name: str) -> None:
    """Add an Excel table when a sheet contains a header and data rows."""
    if sheet.max_row < 2 or sheet.max_column < 1:
        return
    table = Table(displayName=table_name, ref=sheet.dimensions)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)


def add_contains_text_rule(
    sheet: object,
    cell_range: str,
    text: str,
    fill_color: str,
) -> None:
    """Highlight cells containing a review status using an Excel formula."""
    start_cell = cell_range.split(":", maxsplit=1)[0]
    formula = [f'ISNUMBER(SEARCH("{text}",{start_cell}))']
    sheet.conditional_formatting.add(
        cell_range,
        FormulaRule(formula=formula, fill=PatternFill("solid", fgColor=fill_color)),
    )


def validate_literal_booleans(
    workbook: Workbook,
    sheet_names: Iterable[str],
) -> list[str]:
    """Require real Boolean cells in every known Boolean column."""
    problems: list[str] = []
    checked_columns: list[str] = []
    for sheet_name in sheet_names:
        if sheet_name not in workbook.sheetnames:
            problems.append(f"{sheet_name}: missing sheet")
            continue
        sheet = workbook[sheet_name]
        headers = [str(cell.value or "") for cell in sheet[1]]
        for column, header in enumerate(headers, start=1):
            if header not in BOOLEAN_HEADERS:
                continue
            checked_columns.append(f"{sheet_name}.{header}")
            for row in range(2, sheet.max_row + 1):
                value = sheet.cell(row, column).value
                if value is not None and not isinstance(value, bool):
                    problems.append(
                        f"{sheet_name}!R{row}C{column} ({header}) is "
                        f"{value!r}, not a Boolean"
                    )
                    if len(problems) >= 30:
                        break
    if problems:
        raise ValueError("Workbook Boolean validation failed:\n- " + "\n- ".join(problems))
    return checked_columns


def scan_formula_errors(workbook: Workbook) -> list[str]:
    """Return cells containing obvious stored formula-error strings."""
    errors: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and FORMULA_ERROR_PATTERN.search(cell.value):
                    errors.append(f"{sheet.title}!{cell.coordinate}: {cell.value}")
    return errors


def reopen_workbook(path: str | Path) -> Workbook:
    """Save-time validation helper with links retained and VBA disabled."""
    return load_workbook(Path(path), data_only=False, read_only=False)


def render_sheet_preview(
    workbook: Workbook,
    sheet_name: str,
    output_path: str | Path,
    max_rows: int = 20,
    max_columns: int = 14,
) -> Path:
    """Render a compact workbook preview using Python only.

    This is intentionally a QA view rather than an Excel rendering engine. It
    exposes sheet content, widths, fills, wrapping, and obvious clipping while
    keeping workbook generation independent of Excel, Node, and a browser.
    """
    sheet = workbook[sheet_name]
    row_count = min(sheet.max_row, max_rows)
    column_count = min(sheet.max_column, max_columns)
    font = ImageFont.load_default()
    widths = []
    for column in range(1, column_count + 1):
        configured = sheet.column_dimensions[get_column_letter(column)].width or 12
        widths.append(max(70, min(260, int(configured * 6.2))))
    heights = [max(24, int(sheet.row_dimensions[row].height or 24)) for row in range(1, row_count + 1)]
    image = Image.new("RGB", (sum(widths) + 1, sum(heights) + 1), "white")
    draw = ImageDraw.Draw(image)
    merged_anchors: dict[tuple[int, int], tuple[int, int]] = {}
    merged_children: set[tuple[int, int]] = set()
    for merged in sheet.merged_cells.ranges:
        if merged.min_row > row_count or merged.min_col > column_count:
            continue
        anchor = (merged.min_row, merged.min_col)
        merged_anchors[anchor] = (
            min(merged.max_row, row_count),
            min(merged.max_col, column_count),
        )
        for row in range(merged.min_row, min(merged.max_row, row_count) + 1):
            for column in range(merged.min_col, min(merged.max_col, column_count) + 1):
                if (row, column) != anchor:
                    merged_children.add((row, column))
    top = 0
    for row in range(1, row_count + 1):
        left = 0
        for column in range(1, column_count + 1):
            if (row, column) in merged_children:
                left += widths[column - 1]
                continue
            cell = sheet.cell(row, column)
            end_row, end_column = merged_anchors.get((row, column), (row, column))
            cell_width = sum(widths[column - 1 : end_column])
            cell_height = sum(heights[row - 1 : end_row])
            fill = cell.fill.fgColor.rgb if cell.fill.fill_type == "solid" else None
            if isinstance(fill, str) and len(fill) == 8:
                fill = f"#{fill[-6:]}"
            else:
                fill = "#FFFFFF"
            draw.rectangle(
                (left, top, left + cell_width, top + cell_height),
                fill=fill,
                outline="#B7C9D6",
            )
            value = "" if cell.value is None else str(cell.value)
            maximum_characters = max(8, cell_width // 7)
            if len(value) > maximum_characters:
                value = value[: maximum_characters - 1] + "…"
            text_color = "#FFFFFF" if fill.upper() in {"#17365D", "#1F4E78"} else "#222222"
            draw.text((left + 4, top + 4), value, fill=text_color, font=font)
            left += widths[column - 1]
        top += heights[row - 1]
    target = Path(output_path)
    target.parent.mkdir(parents=True, exist_ok=True)
    image.save(target)
    return target


#%%
