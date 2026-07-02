#%%
"""
Maintenance workflow for config/outlook_mappings_master.xlsx.

Stage 0 of the mapping pipeline.  Run this before Stage 1 whenever the
workbook or the source data CSVs have been updated.

What it does
------------
1. Populates subtotal flags in the three base mapping sheets:
     leap_is_subtotal        — True when the LEAP branch is a parent of
                               another active mapping row (within-sheet check)
     esto_pair_is_subtotal   — from data/00APEC_2025_low_with_subtotals.csv
     ninth_pair_is_subtotal  — from data/merged_file_energy_ALL_20251106.csv
                               (subtotal_layout OR subtotal_results)

2. Produces QA CSV outputs in results/maintenance/:
     cardinality_leap_esto.csv      — (LEAP source, ESTO target) cardinality
     cardinality_leap_ninth.csv     — (LEAP source, 9th target) cardinality
     cardinality_ninth_esto.csv     — (9th source, ESTO target) cardinality
     unmapped_nonzero_esto_pairs.csv        — ESTO (flow, product) pairs in data with
                                      no active mapping row
     unmapped_nonzero_ninth_pairs.csv       — 9th (sector, fuel) pairs in data with
                                      no active mapping row
     subtotal_mismatches.csv        — M6 rule: leaf source → aggregate target
                                      rows not present in the manual allowlist

3. Optionally writes paste-ready zero rows required by reviewed structural
   rules and non-zero Ninth mappings.  Source files are never edited
   automatically.

Usage:
    python codebase/outlook_mapping_maintenance_workflow.py
"""

from __future__ import annotations

import re
import shutil
import sys
from datetime import datetime
from pathlib import Path
from typing import Callable, Dict, Tuple

import pandas as pd
import openpyxl

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from codebase.mapping_issue_exceptions import EXCEPTION_WORKBOOK_PATH, split_allowed_rows

WORKBOOK_PATH = REPO_ROOT / "config" / "outlook_mappings_master.xlsx"
ARCHIVE_DIR = REPO_ROOT / "config" / "archive"
QA_DIR = REPO_ROOT / "results" / "maintenance"

ESTO_CSV_PATH = REPO_ROOT / "data" / "00APEC_2025_low_with_subtotals.csv"
ESTO_SOURCE_DATA_PATHS = [
    ESTO_CSV_PATH,
    REPO_ROOT / "data" / "00APEC_2024_low_with_subtotals.csv",
]
NINTH_CSV_PATH = REPO_ROOT / "data" / "merged_file_energy_ALL_20251106.csv"
FULL_MODEL_EXPORT_PATHS = [
    REPO_ROOT / "data" / "full model export.xlsx",
    REPO_ROOT.parent / "leap_initialisation" / "data" / "full model export.xlsx",
]
FULL_MODEL_EXPORT_SHEET = "Export"

# User toggle.  This only writes review/copy files under results/maintenance;
# it never modifies either ESTO source CSV.
GENERATE_MISSING_MAPPED_ESTO_ROWS = True
MISSING_MAPPED_ESTO_ROWS_DIR = QA_DIR / "missing_mapped_esto_rows"
SUBTOTAL_CHANGE_PREVIEW_PATH = QA_DIR / "subtotal_change_preview.xlsx"
SUBTOTAL_OVERRIDE_SHEET = "subtotal_label_overrides"
SUBTOTAL_OVERRIDE_STALE_PATH = QA_DIR / "subtotal_label_overrides_stale.csv"
APPLY_SUBTOTAL_CHANGES_TO_WORKBOOK = False
UNMAPPED_NONZERO_ESTO_ALLOWED_SHEET = "unmapped_esto_nonzero_allowed"
UNMAPPED_NONZERO_NINTH_ALLOWED_SHEET = "unmapped_ninth_nonzero_allowed"

SUBTOTAL_OVERRIDE_CONFIGS = {
    "leap_combined_esto": {
        "keys": ("leap_sector_name_full_path", "raw_leap_fuel_name", "esto_flow", "esto_product"),
        "subtotal_columns": ("leap_is_subtotal", "esto_pair_is_subtotal"),
    },
    "leap_combined_ninth": {
        "keys": ("leap_sector_name_full_path", "raw_leap_fuel_name", "ninth_sector", "ninth_fuel"),
        "subtotal_columns": ("leap_is_subtotal", "ninth_pair_is_subtotal"),
    },
    "ninth_pairs_to_esto_pairs": {
        "keys": ("9th_sector", "9th_fuel", "esto_flow", "esto_product"),
        "subtotal_columns": ("ninth_pair_is_subtotal", "esto_pair_is_subtotal"),
    },
}

# ── helpers ──────────────────────────────────────────────────────────────────

def _norm(value: object) -> str:
    """Normalize a cell value: strip, collapse internal whitespace."""
    text = " ".join(str(value or "").split())
    return "" if text.lower() in {"nan", "none"} else text


def _is_x(value: object) -> bool:
    """Return True if value is exactly the 9th-Outlook placeholder 'x'."""
    return str(value or "").strip() == "x"


def _truthy(value: object) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "t", "yes", "y", "on"}


def _subtotal_bool(value: object) -> bool | None:
    """Parse a subtotal value without treating the boolean False as blank."""
    if value is None or pd.isna(value):
        return None
    if isinstance(value, bool):
        return value
    text = str(value).strip().lower()
    if not text:
        return None
    if text in {"1", "true", "t", "yes", "y", "on"}:
        return True
    if text in {"0", "false", "f", "no", "n", "off"}:
        return False
    raise ValueError(f"Unexpected subtotal override value: {value!r}")


def _active_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Filter out rows where remove_row or duplicate_to_remove is truthy."""
    remove = df.get("remove_row", pd.Series(False, index=df.index)).map(_truthy)
    duplicate = df.get("duplicate_to_remove", pd.Series(False, index=df.index)).map(_truthy)
    return df[~(remove | duplicate)].copy()


def _year_columns(df: pd.DataFrame) -> list[str]:
    """Return columns that look like year/value columns in a balance table."""
    return [col for col in df.columns if re.fullmatch(r"\d{4}", str(col).strip())]


def _nonzero_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Return only rows with at least one non-zero value across year columns."""
    year_cols = _year_columns(df)
    if not year_cols:
        return df.iloc[0:0].copy()

    numeric_years = df[year_cols].apply(pd.to_numeric, errors="coerce").fillna(0)
    mask = numeric_years.abs().gt(0).any(axis=1)
    return df[mask].copy()


# ── subtotal lookups ─────────────────────────────────────────────────────────

def _build_esto_subtotal_lookup() -> Dict[Tuple[str, str], bool]:
    """
    Return a dict mapping (norm(esto_flow), norm(esto_product)) -> bool.

    True if any row in the ESTO CSV for that pair has is_subtotal=True.
    """
    df = pd.read_csv(ESTO_CSV_PATH)
    if "is_subtotal" not in df.columns:
        raise ValueError(f"is_subtotal column not found in {ESTO_CSV_PATH}")

    df = _nonzero_rows(df)
    df["_flow"] = df["flows"].fillna("").map(_norm)
    df["_product"] = df["products"].fillna("").map(_norm)
    df["_is_sub"] = df["is_subtotal"].fillna(False).map(_truthy)

    grouped = (
        df[df["_flow"].ne("") & df["_product"].ne("")]
        .groupby(["_flow", "_product"])["_is_sub"]
        .max()
    )
    return {(flow, product): bool(flag) for (flow, product), flag in grouped.items()}


def _build_ninth_subtotal_lookup() -> Dict[Tuple[str, str], bool]:
    """
    Return a dict mapping (norm(ninth_sector), norm(ninth_fuel)) -> bool.

    ninth_sector = deepest non-'x' sector level (sub4 -> sub3 -> sub2 -> sub1 -> sector).
    ninth_fuel   = deepest non-'x' fuel level (subfuels -> fuels).

    True if subtotal_layout OR subtotal_results is True for that pair.
    """
    df = pd.read_csv(NINTH_CSV_PATH)
    df = _nonzero_rows(df)

    sector_cols = ["sub4sectors", "sub3sectors", "sub2sectors", "sub1sectors", "sectors"]
    fuel_cols = ["subfuels", "fuels"]
    flag_cols = ["subtotal_layout", "subtotal_results"]

    for col in sector_cols + fuel_cols + flag_cols:
        if col not in df.columns:
            df[col] = ""

    for col in flag_cols:
        df[col] = df[col].fillna(False).map(_truthy)

    def _deepest_sector(row: pd.Series) -> str:
        for col in sector_cols:
            val = _norm(row[col])
            if val and not _is_x(val):
                return val
        return ""

    def _deepest_fuel(row: pd.Series) -> str:
        for col in fuel_cols:
            val = _norm(row[col])
            if val and not _is_x(val):
                return val
        return ""

    df["_sector"] = df.apply(_deepest_sector, axis=1)
    df["_fuel"] = df.apply(_deepest_fuel, axis=1)
    df["_is_sub"] = df["subtotal_layout"] | df["subtotal_results"]

    grouped = (
        df[df["_sector"].ne("") & df["_fuel"].ne("")]
        .groupby(["_sector", "_fuel"])["_is_sub"]
        .max()
    )
    return {(sector, fuel): bool(flag) for (sector, fuel), flag in grouped.items()}


def _compute_leap_subtotals(paths: set[str]) -> set[str]:
    """
    Return the subset of paths that are subtotals (i.e. have at least one
    child path in the active set).  A child path starts with 'parent/'.
    """
    subtotals: set[str] = set()
    sorted_paths = sorted(paths)
    for path in sorted_paths:
        prefix = path + "/"
        for other in sorted_paths:
            if other != path and other.startswith(prefix):
                subtotals.add(path)
                break
    return subtotals


def _find_full_model_export_path(
    candidate_paths: list[Path] | None = None,
) -> Path | None:
    """Return the first available full model export workbook path."""
    for path in candidate_paths or FULL_MODEL_EXPORT_PATHS:
        if path.exists():
            return path
    return None


def _normalize_export_branch_path(branch_path: object) -> list[str]:
    """Normalize a LEAP export Branch Path into clean path segments."""
    text = _norm(branch_path).replace("\\", "/")
    return [segment.strip() for segment in text.split("/") if segment.strip()]


def _mapping_style_sector_path_from_export_segments(
    segments: list[str],
    fuel_names: set[str],
) -> str:
    """
    Convert full-model Branch Path segments to mapping-style LEAP sector paths.

    Mapping sheets omit the LEAP root branch (Demand/Transformation) and omit
    technical transformation grouping nodes such as Processes, Output Fuels,
    Feedstock Fuels, and Auxiliary Fuels.
    """
    if not segments:
        return ""

    root = segments[0]
    if root not in {"Demand", "Transformation"}:
        return ""

    clean_segments = segments[1:]
    if not clean_segments:
        return ""

    technical_nodes = {"Processes", "Output Fuels", "Feedstock Fuels", "Auxiliary Fuels"}
    clean_segments = [segment for segment in clean_segments if segment not in technical_nodes]
    if clean_segments and clean_segments[-1] in fuel_names:
        clean_segments = clean_segments[:-1]

    return "/".join(clean_segments)


def _load_leap_paths_from_full_model_export(
    export_path: Path,
    fuel_names: set[str],
    sheet_name: str = FULL_MODEL_EXPORT_SHEET,
) -> set[str]:
    """Load mapping-style LEAP sector paths from the full model export workbook."""
    export_df = pd.read_excel(
        export_path,
        sheet_name=sheet_name,
        header=2,
        usecols=["Branch Path"],
        dtype=object,
    )
    paths: set[str] = set()
    for raw_path in export_df["Branch Path"].dropna().unique():
        segments = _normalize_export_branch_path(raw_path)
        path = _mapping_style_sector_path_from_export_segments(segments, fuel_names)
        if path:
            paths.add(path)
    return paths


def _compute_leap_subtotals_from_full_model_export(
    active_mapping_paths: set[str],
    fuel_names: set[str],
    export_path: Path | None = None,
) -> tuple[set[str], set[str], Path | None]:
    """
    Return subtotal paths from the full model export, limited to mapped paths.

    Returns (subtotal_paths, export_paths, used_export_path). If no export is
    available, all three values are suitable for falling back to mapping-only
    inference.
    """
    used_export_path = export_path or _find_full_model_export_path()
    if used_export_path is None:
        return set(), set(), None

    export_paths = _load_leap_paths_from_full_model_export(
        export_path=used_export_path,
        fuel_names=fuel_names,
    )
    export_subtotal_paths = _compute_leap_subtotals(export_paths)
    return export_subtotal_paths & active_mapping_paths, export_paths, used_export_path


# ── workbook helpers ──────────────────────────────────────────────────────────

def _archive_workbook(
    path: Path,
    archive_dir: Path = ARCHIVE_DIR,
    timestamp: datetime | None = None,
) -> Path:
    """
    Copy the workbook to the archive directory for this maintenance run.

    The filename includes seconds for readability and gets a numeric suffix if
    the workflow is run more than once in the same second.
    """
    archive_dir.mkdir(parents=True, exist_ok=True)
    ts = (timestamp or datetime.now()).strftime("%Y%m%d_%H%M%S")
    base_name = f"{path.stem}.maintenance_run_{ts}"
    dest = archive_dir / f"{base_name}{path.suffix}"
    suffix = 2
    while dest.exists():
        dest = archive_dir / f"{base_name}_{suffix}{path.suffix}"
        suffix += 1
    shutil.copy2(path, dest)
    print(f"Archived workbook copy: {dest}")
    return dest


def _col_index(ws, header_name: str) -> int | None:
    """Return the 1-based column index of header_name in row 1, or None."""
    for cell in ws[1]:
        if _norm(cell.value) == _norm(header_name):
            return cell.column
    return None


def _get_header_row(ws) -> list[str]:
    return [_norm(cell.value) for cell in ws[1]]


def _update_sheet_column(
    ws,
    col_name: str,
    key_col_names: list[str],
    lookup: Dict[Tuple[str, ...], bool],
) -> tuple[int, int, int]:
    """
    Write lookup results into col_name for each data row in ws.

    key_col_names: column headers whose normed values form the lookup key.
    Returns (updated, not_found, skipped_blank_key) counts.
    """
    target_col_idx = _col_index(ws, col_name)
    if target_col_idx is None:
        print(f"  WARNING: column '{col_name}' not found in sheet '{ws.title}' — skipping")
        return 0, 0, 0

    key_col_idxs = []
    for kname in key_col_names:
        idx = _col_index(ws, kname)
        if idx is None:
            print(f"  WARNING: key column '{kname}' not found in sheet '{ws.title}' — skipping")
            return 0, 0, 0
        key_col_idxs.append(idx)

    updated = not_found = skipped = 0
    for row in ws.iter_rows(min_row=2):
        key_parts = tuple(_norm(row[i - 1].value) for i in key_col_idxs)
        if any(part == "" for part in key_parts):
            skipped += 1
            continue
        flag = lookup.get(key_parts)
        if flag is None:
            not_found += 1
            row[target_col_idx - 1].value = None
        else:
            row[target_col_idx - 1].value = flag
            updated += 1

    return updated, not_found, skipped


def _update_sheet_leap_subtotals(ws, subtotal_paths: set[str]) -> tuple[int, int]:
    """Write leap_is_subtotal for each row based on precomputed subtotal_paths set."""
    target_col_idx = _col_index(ws, "leap_is_subtotal")
    if target_col_idx is None:
        print(f"  WARNING: column 'leap_is_subtotal' not found in sheet '{ws.title}' — skipping")
        return 0, 0

    path_col_idx = _col_index(ws, "leap_sector_name_full_path")
    if path_col_idx is None:
        print(f"  WARNING: column 'leap_sector_name_full_path' not found in sheet '{ws.title}' — skipping")
        return 0, 0

    updated = skipped = 0
    for row in ws.iter_rows(min_row=2):
        path = _norm(row[path_col_idx - 1].value)
        if not path:
            skipped += 1
            continue
        row[target_col_idx - 1].value = path in subtotal_paths
        updated += 1

    return updated, skipped


def _read_sheet_as_df(wb, sheet_name: str) -> pd.DataFrame:
    """Read a workbook sheet into a DataFrame using openpyxl (avoids re-read from disk)."""
    ws = wb[sheet_name]
    headers = [_norm(cell.value) for cell in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append([_norm(v) for v in row])
    return pd.DataFrame(rows, columns=headers)


def _subtotal_preview_value(value: object) -> bool | None:
    """Normalize stored subtotal values for preview comparisons."""
    text = _norm(value)
    if text == "":
        return None
    if _truthy(text):
        return True
    if text.lower() in {"false", "0", "no", "off"}:
        return False
    return None


def _build_subtotal_change_preview(
    df: pd.DataFrame,
    sheet_name: str,
    key_columns: list[str],
    change_specs: list[tuple[str, Callable[[pd.Series], object], str]],
) -> pd.DataFrame:
    """Return one row per subtotal cell whose proposed value differs from current."""
    rows: list[dict[str, object]] = []
    for row_number, (_, row) in enumerate(df.iterrows(), start=2):
        for change_column, proposed_getter, reason in change_specs:
            current_value = _subtotal_preview_value(row.get(change_column, ""))
            proposed_value = proposed_getter(row)
            proposed_value = _subtotal_preview_value(proposed_value)
            if current_value == proposed_value:
                continue
            preview_row: dict[str, object] = {
                "sheet_name": sheet_name,
                "excel_row_number": row_number,
                "change_column": change_column,
                "current_value": "" if current_value is None else current_value,
                "proposed_value": "" if proposed_value is None else proposed_value,
                "change_reason": reason,
            }
            for key_column in key_columns:
                preview_row[key_column] = row.get(key_column, "")
            rows.append(preview_row)
    return pd.DataFrame(rows)


def _build_subtotal_proposed_rows_preview(
    df: pd.DataFrame,
    change_specs: list[tuple[str, Callable[[pd.Series], object]]],
) -> pd.DataFrame:
    """Return only rows whose subtotal values would change, with proposed values applied."""
    rows: list[dict[str, object]] = []
    for _, row in df.iterrows():
        # String-backed pandas rows reject bool assignments on newer pandas.
        proposed_row = row.astype(object).copy()
        changed = False
        for change_column, proposed_getter in change_specs:
            proposed_value = _subtotal_preview_value(proposed_getter(row))
            current_value = _subtotal_preview_value(row.get(change_column, ""))
            if current_value != proposed_value:
                proposed_row[change_column] = "" if proposed_value is None else proposed_value
                changed = True
        if changed:
            rows.append(proposed_row.to_dict())
    return pd.DataFrame(rows, columns=list(df.columns))


def _load_subtotal_label_overrides(
    exception_workbook_path: Path = EXCEPTION_WORKBOOK_PATH,
) -> tuple[dict[tuple[str, tuple[str, ...], str], bool], pd.DataFrame]:
    """Load reviewed subtotal values keyed by sheet, mapping keys, and column."""
    try:
        override_df = pd.read_excel(
            exception_workbook_path,
            sheet_name=SUBTOTAL_OVERRIDE_SHEET,
            dtype=object,
        )
    except (FileNotFoundError, ValueError):
        return {}, pd.DataFrame()

    if "enabled" in override_df.columns:
        override_df = override_df[override_df["enabled"].map(_truthy)].copy()

    overrides: dict[tuple[str, tuple[str, ...], str], bool] = {}
    for row_number, (_, row) in enumerate(override_df.iterrows(), start=2):
        sheet_name = _norm(row.get("sheet", ""))
        config = SUBTOTAL_OVERRIDE_CONFIGS.get(sheet_name)
        if config is None:
            continue
        key = tuple(_norm(row.get(column, "")) for column in config["keys"])
        if all(not value for value in key):
            continue
        for subtotal_column in config["subtotal_columns"]:
            value = _subtotal_bool(row.get(subtotal_column))
            if value is None:
                continue
            override_key = (sheet_name, key, subtotal_column)
            previous = overrides.get(override_key)
            if previous is not None and previous != value:
                raise ValueError(
                    f"Conflicting {SUBTOTAL_OVERRIDE_SHEET} rows for {sheet_name} "
                    f"{key} {subtotal_column}; conflict at Excel row {row_number}."
                )
            overrides[override_key] = value
    return overrides, override_df


def _resolved_subtotal_value(
    sheet_name: str,
    row: pd.Series,
    subtotal_column: str,
    computed_value: object,
    overrides: dict[tuple[str, tuple[str, ...], str], bool],
) -> object:
    """Return a reviewed override when present, otherwise the computed value."""
    config = SUBTOTAL_OVERRIDE_CONFIGS[sheet_name]
    key = tuple(_norm(row.get(column, "")) for column in config["keys"])
    return overrides.get((sheet_name, key, subtotal_column), computed_value)


def _apply_subtotal_overrides_to_sheet(
    ws,
    overrides: dict[tuple[str, tuple[str, ...], str], bool],
) -> int:
    """Apply reviewed overrides after computed subtotal values are written."""
    config = SUBTOTAL_OVERRIDE_CONFIGS[ws.title]
    key_indexes = [_col_index(ws, column) for column in config["keys"]]
    subtotal_indexes = {
        column: _col_index(ws, column)
        for column in config["subtotal_columns"]
    }
    if any(index is None for index in key_indexes) or any(
        index is None for index in subtotal_indexes.values()
    ):
        raise ValueError(f"Missing override key or subtotal column in sheet {ws.title}")

    updated = 0
    for row in ws.iter_rows(min_row=2):
        key = tuple(_norm(row[index - 1].value) for index in key_indexes)
        for subtotal_column, subtotal_index in subtotal_indexes.items():
            override = overrides.get((ws.title, key, subtotal_column))
            if override is not None:
                row[subtotal_index - 1].value = override
                updated += 1
    return updated


def _build_stale_subtotal_override_rows(
    override_df: pd.DataFrame,
    mapping_frames: dict[str, pd.DataFrame],
) -> pd.DataFrame:
    """Return enabled overrides whose sheet/key no longer exists in the master."""
    if override_df.empty:
        return pd.DataFrame(columns=["override_excel_row", "sheet", "stale_reason"])

    mapping_keys = {}
    for sheet_name, config in SUBTOTAL_OVERRIDE_CONFIGS.items():
        frame = mapping_frames[sheet_name]
        mapping_keys[sheet_name] = {
            tuple(_norm(row.get(column, "")) for column in config["keys"])
            for _, row in frame.iterrows()
        }

    stale_rows: list[dict[str, object]] = []
    for row_number, (_, row) in enumerate(override_df.iterrows(), start=2):
        sheet_name = _norm(row.get("sheet", ""))
        config = SUBTOTAL_OVERRIDE_CONFIGS.get(sheet_name)
        reason = ""
        if config is None:
            reason = "unknown_mapping_sheet"
        else:
            key = tuple(_norm(row.get(column, "")) for column in config["keys"])
            if all(not value for value in key):
                reason = "blank_override_key"
            elif key not in mapping_keys[sheet_name]:
                reason = "mapping_key_not_found"
        if reason:
            stale_row = {"override_excel_row": row_number, "stale_reason": reason}
            stale_row.update(row.to_dict())
            stale_rows.append(stale_row)
    output_columns = ["override_excel_row", "stale_reason", *list(override_df.columns)]
    return pd.DataFrame(stale_rows, columns=output_columns)


# ── cardinality helpers ───────────────────────────────────────────────────────

def _mapping_cardinality(n_targets: int, n_sources: int) -> str:
    if n_targets <= 0 or n_sources <= 0:
        return ""
    if n_targets == 1 and n_sources == 1:
        return "one_to_one"
    if n_targets > 1 and n_sources == 1:
        return "one_to_many"
    if n_targets == 1 and n_sources > 1:
        return "many_to_one"
    return "many_to_many"


def _compute_cardinality(
    df: pd.DataFrame,
    source_flow_col: str,
    source_product_col: str,
    target_flow_col: str,
    target_product_col: str,
) -> pd.DataFrame:
    """Return a cardinality summary DataFrame for (source, target) pairs."""
    active = _active_rows(df)
    valid = active[
        active[source_flow_col].ne("") & active[source_product_col].ne("")
        & active[target_flow_col].ne("") & active[target_product_col].ne("")
    ].copy()
    if valid.empty:
        return pd.DataFrame(columns=[
            source_flow_col, source_product_col,
            target_flow_col, target_product_col,
            "n_targets_for_source", "n_sources_for_target", "cardinality",
        ])
    valid["_src"] = valid[source_flow_col] + "|||" + valid[source_product_col]
    valid["_tgt"] = valid[target_flow_col] + "|||" + valid[target_product_col]
    pairs = valid[["_src", "_tgt"]].drop_duplicates()
    src_target_count = pairs.groupby("_src")["_tgt"].nunique()
    tgt_source_count = pairs.groupby("_tgt")["_src"].nunique()

    rows = []
    for _, pair_row in pairs.iterrows():
        src, tgt = pair_row["_src"], pair_row["_tgt"]
        sf, sp = src.split("|||", 1)
        tf, tp = tgt.split("|||", 1)
        n_targets = int(src_target_count.get(src, 0))
        n_sources = int(tgt_source_count.get(tgt, 0))
        rows.append({
            source_flow_col: sf,
            source_product_col: sp,
            target_flow_col: tf,
            target_product_col: tp,
            "n_targets_for_source": n_targets,
            "n_sources_for_target": n_sources,
            "cardinality": _mapping_cardinality(n_targets, n_sources),
        })
    return pd.DataFrame(rows).sort_values([source_flow_col, source_product_col, target_flow_col, target_product_col])


# ── unmapped pair detection ───────────────────────────────────────────────────

def _unmapped_esto_pairs(
    mapping_dfs: list[pd.DataFrame],
    esto_subtotal_lookup: Dict[Tuple[str, str], bool],
) -> pd.DataFrame:
    """Find ESTO (flow, product) pairs in the data that have no active mapping row."""
    active_esto: set[tuple[str, str]] = set()
    for df in mapping_dfs:
        active = _active_rows(df)
        for col_flow, col_product in [("esto_flow", "esto_product")]:
            if col_flow in active.columns and col_product in active.columns:
                for _, row in active.iterrows():
                    f, p = _norm(row.get(col_flow, "")), _norm(row.get(col_product, ""))
                    if f and p:
                        active_esto.add((f, p))

    rows = []
    for (flow, product), is_sub in esto_subtotal_lookup.items():
        if (flow, product) not in active_esto:
            rows.append({"esto_flow": flow, "esto_product": product, "is_subtotal": is_sub})
    return pd.DataFrame(rows).sort_values(["esto_flow", "esto_product"]) if rows else pd.DataFrame(
        columns=["esto_flow", "esto_product", "is_subtotal"]
    )


def _unmapped_ninth_pairs(
    mapping_dfs: list[pd.DataFrame],
    ninth_subtotal_lookup: Dict[Tuple[str, str], bool],
) -> pd.DataFrame:
    """Find 9th (sector, fuel) pairs in the data that have no active mapping row."""
    active_ninth: set[tuple[str, str]] = set()
    for df in mapping_dfs:
        active = _active_rows(df)
        for sector_col, fuel_col in [
            ("ninth_sector", "ninth_fuel"),
            ("9th_sector", "9th_fuel"),
        ]:
            if sector_col in active.columns and fuel_col in active.columns:
                for _, row in active.iterrows():
                    s, f = _norm(row.get(sector_col, "")), _norm(row.get(fuel_col, ""))
                    if s and f:
                        active_ninth.add((s, f))

    rows = []
    for (sector, fuel), is_sub in ninth_subtotal_lookup.items():
        if (sector, fuel) not in active_ninth:
            rows.append({"ninth_sector": sector, "ninth_fuel": fuel, "is_subtotal": is_sub})
    return pd.DataFrame(rows).sort_values(["ninth_sector", "ninth_fuel"]) if rows else pd.DataFrame(
        columns=["ninth_sector", "ninth_fuel", "is_subtotal"]
    )


# ── M6: subtotal mismatch detection ──────────────────────────────────────────

def _subtotal_mismatches(
    df: pd.DataFrame,
    source_flow_col: str,
    source_product_col: str,
    target_flow_col: str,
    target_product_col: str,
    source_subtotal_col: str,
    target_subtotal_col: str,
) -> pd.DataFrame:
    """
    Flag rows where a leaf-level source maps to an aggregate target AND a more
    specific (non-subtotal) target also exists for a different source.

    Rule (M6): only flag when:
      - source is NOT a subtotal (leaf level)
      - target IS a subtotal (aggregate)
      - at least one other active row maps to a non-subtotal target at the same
        flow-level (same target_flow_col prefix)
    """
    active = _active_rows(df)
    for col in [source_flow_col, source_product_col, target_flow_col, target_product_col,
                source_subtotal_col, target_subtotal_col]:
        if col not in active.columns:
            active[col] = ""
        active[col] = active[col].fillna("").astype(str).str.strip()

    # Rows where leaf source → aggregate target
    candidates = active[
        (active[source_subtotal_col].str.lower().isin(["false", "0", "no", ""]))
        & (active[target_subtotal_col].str.lower().isin(["true", "1", "yes"]))
    ].copy()

    if candidates.empty:
        return pd.DataFrame()

    # Build set of target flows that also have non-subtotal targets active
    non_subtotal_targets = set(
        active.loc[
            active[target_subtotal_col].str.lower().isin(["false", "0", "no", ""]),
            target_flow_col,
        ]
    )

    mismatch_rows = candidates[
        candidates[target_flow_col].isin(non_subtotal_targets)
    ].copy()
    if mismatch_rows.empty:
        return mismatch_rows

    mismatch_rows["mismatch_reason"] = (
        "leaf_source_maps_to_aggregate_target_and_more_specific_target_exists"
    )
    return mismatch_rows[[
        source_flow_col, source_product_col,
        target_flow_col, target_product_col,
        source_subtotal_col, target_subtotal_col,
        "mismatch_reason",
    ]]


def _split_allowed_subtotal_mismatches(
    subtotal_mismatches: pd.DataFrame,
    exception_workbook_path: Path = EXCEPTION_WORKBOOK_PATH,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Split subtotal mismatches using the manual exception workbook."""
    return split_allowed_rows(
        subtotal_mismatches,
        sheet_name="subtotal_mismatch_allowed",
        status_column="subtotal_mismatch_review_status",
        reason_column="subtotal_mismatch_review_reason",
        workbook_path=exception_workbook_path,
    )


# â”€â”€ migrated legacy conflict checks â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€â”€

def _many_to_many_conflicts(
    cardinality_frames: list[tuple[str, pd.DataFrame]],
) -> pd.DataFrame:
    """Return active mapping pair cardinality rows that remain many-to-many."""
    records = []
    for sheet_name, frame in cardinality_frames:
        work = frame[frame["cardinality"].eq("many_to_many")].copy()
        if work.empty:
            continue
        work.insert(0, "sheet", sheet_name)
        records.append(work)
    if not records:
        return pd.DataFrame(columns=["sheet", "cardinality"])
    return pd.concat(records, ignore_index=True).fillna("")


def _split_allowed_many_to_many(
    many_to_many: pd.DataFrame,
    exception_workbook_path: Path = EXCEPTION_WORKBOOK_PATH,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Split all many-to-many rows using the manual exception workbook."""
    return split_allowed_rows(
        many_to_many,
        sheet_name="many_to_many_allowed",
        status_column="many_to_many_review_status",
        reason_column="many_to_many_review_reason",
        workbook_path=exception_workbook_path,
    )


def _leap_source_presence_conflicts(
    leap_esto_df: pd.DataFrame,
    leap_ninth_df: pd.DataFrame,
) -> pd.DataFrame:
    """Find active LEAP source pairs present in only one of the two LEAP mapping sheets."""
    source_cols = ["leap_sector_name_full_path", "raw_leap_fuel_name"]

    def _active_sources(frame: pd.DataFrame) -> pd.DataFrame:
        active = _active_rows(frame)
        for col in source_cols:
            if col not in active.columns:
                active[col] = ""
            active[col] = active[col].fillna("").astype(str).map(_norm)
        return active[active[source_cols].apply(lambda col: col.ne("")).all(axis=1)][source_cols].drop_duplicates()

    esto_sources = _active_sources(leap_esto_df)
    ninth_sources = _active_sources(leap_ninth_df)
    esto_sources["_in_leap_combined_esto"] = True
    ninth_sources["_in_leap_combined_ninth"] = True

    merged = esto_sources.merge(ninth_sources, on=source_cols, how="outer")
    merged["_in_leap_combined_esto"] = merged["_in_leap_combined_esto"].fillna(False).astype(bool)
    merged["_in_leap_combined_ninth"] = merged["_in_leap_combined_ninth"].fillna(False).astype(bool)
    conflicts = merged[merged["_in_leap_combined_esto"].ne(merged["_in_leap_combined_ninth"])].copy()
    if conflicts.empty:
        return pd.DataFrame(columns=[
            *source_cols,
            "presence_status",
            "in_leap_combined_esto",
            "in_leap_combined_ninth",
        ])
    conflicts["presence_status"] = conflicts.apply(
        lambda row: "active_in_leap_combined_esto_only"
        if row["_in_leap_combined_esto"]
        else "active_in_leap_combined_ninth_only",
        axis=1,
    )
    conflicts = conflicts.rename(columns={
        "_in_leap_combined_esto": "in_leap_combined_esto",
        "_in_leap_combined_ninth": "in_leap_combined_ninth",
    })
    return conflicts[[
        *source_cols,
        "presence_status",
        "in_leap_combined_esto",
        "in_leap_combined_ninth",
    ]].sort_values(source_cols).reset_index(drop=True)


def _split_allowed_leap_source_presence(
    conflicts: pd.DataFrame,
    exception_workbook_path: Path = EXCEPTION_WORKBOOK_PATH,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Split source-presence conflicts using the manual exception workbook."""
    return split_allowed_rows(
        conflicts,
        sheet_name="leap_source_presence_allowed",
        status_column="source_presence_review_status",
        reason_column="source_presence_review_reason",
        workbook_path=exception_workbook_path,
    )


def _leading_code_expression(label: object) -> str:
    """Return the leading ESTO-style code/range expression from a label."""
    text = _norm(label)
    match = re.match(r"^([0-9][0-9.,-]*)\b", text)
    return match.group(1) if match else ""


def _code_tuple(code: str) -> tuple[int, ...] | None:
    parts = str(code).split(".")
    if not parts or any(not part.isdigit() for part in parts):
        return None
    return tuple(int(part) for part in parts)


def _single_code_covers(active_code: str, implied_code: str) -> bool:
    """Return True if active_code exactly equals or is a parent of implied_code."""
    active_tuple = _code_tuple(active_code)
    implied_tuple = _code_tuple(implied_code)
    if active_tuple is None or implied_tuple is None:
        return False
    return implied_tuple[: len(active_tuple)] == active_tuple


def _range_code_covers(start_code: str, end_code: str, implied_code: str) -> bool:
    start_tuple = _code_tuple(start_code)
    end_tuple = _code_tuple(end_code)
    implied_tuple = _code_tuple(implied_code)
    if start_tuple is None or end_tuple is None or implied_tuple is None:
        return False
    if len(start_tuple) != len(end_tuple):
        return False
    if len(implied_tuple) < len(start_tuple):
        return False
    implied_prefix = implied_tuple[: len(start_tuple)]
    return start_tuple <= implied_prefix <= end_tuple


def _code_expression_covers(active_expression: str, implied_label: str) -> bool:
    """Return True if a compressed active code expression covers an implied label."""
    implied_code = _leading_code_expression(implied_label)
    if not implied_code:
        return False
    for token in [part.strip() for part in str(active_expression).split(",") if part.strip()]:
        if "-" in token:
            start_code, end_code = [part.strip() for part in token.split("-", 1)]
            if _range_code_covers(start_code, end_code, implied_code):
                return True
        elif _single_code_covers(token, implied_code):
            return True
    return False


def _target_covers(active_target: str, implied_target: str) -> bool:
    """Return True if an active ESTO target covers an implied ESTO target."""
    if " || " not in active_target or " || " not in implied_target:
        return False
    active_flow, active_product = active_target.split(" || ", 1)
    implied_flow, implied_product = implied_target.split(" || ", 1)
    return (
        _code_expression_covers(_leading_code_expression(active_flow), implied_flow)
        and _code_expression_covers(_leading_code_expression(active_product), implied_product)
    )


def _split_allowed_crosswalk_conflicts(
    crosswalk_conflicts: pd.DataFrame,
    exception_workbook_path: Path = EXCEPTION_WORKBOOK_PATH,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Split crosswalk conflicts using the manual exception workbook."""
    return split_allowed_rows(
        crosswalk_conflicts,
        sheet_name="crosswalk_allowed",
        status_column="crosswalk_review_status",
        reason_column="crosswalk_review_reason",
        workbook_path=exception_workbook_path,
    )


def _classify_crosswalk_conflict(row: pd.Series) -> str:
    """Classify whether a crosswalk conflict is likely expected aggregation or a real gap."""
    if row.get("conflict_reason") == "ninth_pair_missing_from_crosswalk":
        return "missing_crosswalk_mapping"
    implied_targets = [
        target.strip()
        for target in str(row.get("implied_esto_targets", "")).split(" | ")
        if target.strip()
    ]
    active_targets = [
        target.strip()
        for target in str(row.get("active_esto_targets", "")).split(" | ")
        if target.strip()
    ]
    if not implied_targets:
        return "missing_crosswalk_mapping"
    covered_count = sum(
        1 for implied_target in implied_targets
        if any(_target_covers(active_target, implied_target) for active_target in active_targets)
    )
    if covered_count == len(implied_targets):
        return "expected_combined_or_aggregate_target"
    if covered_count > 0:
        return "partial_combined_target_review"
    return "target_mismatch_review"


def _crosswalk_target_conflicts(
    leap_esto_df: pd.DataFrame,
    leap_ninth_df: pd.DataFrame,
    ninth_esto_df: pd.DataFrame,
) -> pd.DataFrame:
    """
    Find active LEAP->9th rows whose 9th->ESTO crosswalk target is not active
    for the same LEAP source in leap_combined_esto.
    """
    source_cols = ["leap_sector_name_full_path", "raw_leap_fuel_name"]
    esto_cols = ["esto_flow", "esto_product"]
    ninth_cols = ["ninth_sector", "ninth_fuel"]
    crosswalk_ninth_cols = ["9th_sector", "9th_fuel"]

    active_esto = _active_rows(leap_esto_df).copy()
    active_ninth = _active_rows(leap_ninth_df).copy()
    active_crosswalk = _active_rows(ninth_esto_df).copy()

    for frame, cols in [
        (active_esto, source_cols + esto_cols),
        (active_ninth, source_cols + ninth_cols),
        (active_crosswalk, crosswalk_ninth_cols + esto_cols),
    ]:
        for col in cols:
            if col not in frame.columns:
                frame[col] = ""
            frame[col] = frame[col].fillna("").astype(str).map(_norm)

    active_esto = active_esto[
        active_esto[source_cols + esto_cols].apply(lambda col: col.ne("")).all(axis=1)
    ].copy()
    active_ninth = active_ninth[
        active_ninth[source_cols + ninth_cols].apply(lambda col: col.ne("")).all(axis=1)
    ].copy()
    active_crosswalk = active_crosswalk[
        active_crosswalk[crosswalk_ninth_cols + esto_cols].apply(lambda col: col.ne("")).all(axis=1)
    ].copy()

    if active_esto.empty or active_ninth.empty or active_crosswalk.empty:
        return pd.DataFrame(columns=[
            *source_cols,
            *ninth_cols,
            "implied_esto_targets",
            "active_esto_targets",
            "conflict_reason",
            "conflict_classification",
        ])

    active_esto_pairs = set(
        zip(
            active_esto["leap_sector_name_full_path"],
            active_esto["raw_leap_fuel_name"],
            active_esto["esto_flow"],
            active_esto["esto_product"],
        )
    )
    active_esto_targets = (
        active_esto.assign(_target=active_esto["esto_flow"] + " || " + active_esto["esto_product"])
        .groupby(source_cols)["_target"]
        .agg(lambda values: " | ".join(sorted(set(values))))
        .reset_index()
        .rename(columns={"_target": "active_esto_targets"})
    )

    merged = active_ninth.merge(
        active_crosswalk,
        left_on=ninth_cols,
        right_on=crosswalk_ninth_cols,
        how="left",
    )
    merged["has_crosswalk_target"] = merged["esto_flow"].fillna("").astype(str).str.strip().ne("")
    merged["has_matching_active_esto_target"] = merged.apply(
        lambda row: (
            row["leap_sector_name_full_path"],
            row["raw_leap_fuel_name"],
            row.get("esto_flow", ""),
            row.get("esto_product", ""),
        )
        in active_esto_pairs,
        axis=1,
    )

    grouped = (
        merged.groupby(source_cols + ninth_cols, dropna=False)
        .agg(
            implied_esto_targets=(
                "esto_flow",
                lambda series: " | ".join(
                    sorted(
                        {
                            f"{flow} || {product}"
                            for flow, product in zip(
                                series.astype(str),
                                merged.loc[series.index, "esto_product"].astype(str),
                            )
                            if _norm(flow) and _norm(product)
                        }
                    )
                ),
            ),
            has_crosswalk_target=("has_crosswalk_target", "max"),
            has_matching_active_esto_target=("has_matching_active_esto_target", "max"),
        )
        .reset_index()
    )
    grouped = grouped.merge(active_esto_targets, on=source_cols, how="left")
    grouped["active_esto_targets"] = grouped["active_esto_targets"].fillna("")

    conflicts = grouped[
        grouped["active_esto_targets"].ne("")
        & (
            ~grouped["has_crosswalk_target"].astype(bool)
            | ~grouped["has_matching_active_esto_target"].astype(bool)
        )
    ].copy()
    if conflicts.empty:
        return pd.DataFrame(columns=[
            *source_cols,
            *ninth_cols,
            "implied_esto_targets",
            "active_esto_targets",
            "conflict_reason",
            "conflict_classification",
        ])

    conflicts["conflict_reason"] = conflicts.apply(
        lambda row: "ninth_pair_missing_from_crosswalk"
        if not bool(row["has_crosswalk_target"])
        else "implied_esto_target_not_active_for_leap_source",
        axis=1,
    )
    conflicts["conflict_classification"] = conflicts.apply(_classify_crosswalk_conflict, axis=1)
    return conflicts[[
        *source_cols,
        *ninth_cols,
        "implied_esto_targets",
        "active_esto_targets",
        "conflict_reason",
        "conflict_classification",
    ]].sort_values(source_cols + ninth_cols).reset_index(drop=True)


def _write_maintenance_summary(
    summary_path: Path,
    qa_dir: Path = QA_DIR,
    tree_dir: Path = REPO_ROOT / "results" / "tree_structure",
) -> pd.DataFrame:
    """Write a compact row-count summary for Stage 0 QA outputs."""
    output_specs = [
        ("maintenance", "cardinality_leap_esto.csv", "info"),
        ("maintenance", "cardinality_leap_ninth.csv", "info"),
        ("maintenance", "cardinality_ninth_esto.csv", "info"),
        ("maintenance", "many_to_many_allowed_matched.csv", "info"),
        ("maintenance", "many_to_many_conflicts.csv", "review"),
        ("maintenance", "leap_source_presence_conflicts.csv", "review"),
        ("maintenance", "crosswalk_target_conflicts_allowed_matched.csv", "info"),
        ("maintenance", "crosswalk_target_conflicts.csv", "review"),
        ("maintenance", "unmapped_nonzero_esto_pairs.csv", "review"),
        ("maintenance", "unmapped_nonzero_ninth_pairs.csv", "review"),
        ("maintenance", "subtotal_mismatches.csv", "review"),
        ("maintenance", "subtotal_mismatches_allowed_matched.csv", "info"),
        ("maintenance", "missing_mapped_esto_rows/missing_mapped_esto_rows_summary.csv", "review"),
        ("maintenance", "display_names_qa.csv", "review"),
        ("tree_structure", "esto_validation.csv", "validation"),
        ("tree_structure", "common_esto_validation.csv", "validation"),
        ("tree_structure", "common_esto_non_esto_parent_child_edges.csv", "review"),
    ]
    rows = []
    for output_area, file_name, output_type in output_specs:
        path = (qa_dir if output_area == "maintenance" else tree_dir) / file_name
        if path.exists():
            row_count = len(pd.read_csv(path))
            if output_type == "validation":
                status = "pass" if row_count == 0 else "fail"
            elif output_type == "info":
                status = "info"
            else:
                status = "review" if row_count else "empty"
        else:
            row_count = None
            status = "missing"
        rows.append({
            "output_area": output_area,
            "file_name": file_name,
            "output_type": output_type,
            "row_count": row_count,
            "status": status,
        })
    summary = pd.DataFrame(rows)
    summary_path.parent.mkdir(parents=True, exist_ok=True)
    summary.to_csv(summary_path, index=False)
    return summary


def _remove_stale_generated_exception_outputs() -> None:
    """Remove old generated allowed-output filenames that predate matched suffixes."""
    for file_name in [
        "many_to_many_allowed.csv",
        "crosswalk_target_conflicts_allowed.csv",
        "subtotal_mismatches_allowed.csv",
    ]:
        path = QA_DIR / file_name
        if path.exists():
            path.unlink()


# ── main ──────────────────────────────────────────────────────────────────────

def run(
    apply_subtotal_changes_to_workbook: bool = APPLY_SUBTOTAL_CHANGES_TO_WORKBOOK,
) -> None:

    if GENERATE_MISSING_MAPPED_ESTO_ROWS:
        print("Building paste-ready rows for mapped ESTO pairs missing from source data …")
        from codebase.mapping_tools.build_missing_mapped_esto_rows import (
            write_missing_mapped_esto_rows,
        )

        write_missing_mapped_esto_rows(
            esto_csv_paths=ESTO_SOURCE_DATA_PATHS,
            mapping_workbook_path=WORKBOOK_PATH,
            ninth_csv_path=NINTH_CSV_PATH,
            output_dir=MISSING_MAPPED_ESTO_ROWS_DIR,
        )
    else:
        print("Skipping missing mapped ESTO row generation (toggle is False).")

    print("Loading subtotal lookups …")
    esto_lookup = _build_esto_subtotal_lookup()
    ninth_lookup = _build_ninth_subtotal_lookup()
    print(f"  ESTO lookup: {len(esto_lookup):,} (flow, product) pairs")
    print(f"  9th lookup:  {len(ninth_lookup):,} (sector, fuel) pairs")

    print(f"\nOpening {WORKBOOK_PATH} …")
    wb = openpyxl.load_workbook(WORKBOOK_PATH)

    # ── Compute LEAP subtotal paths from both LEAP sheets combined ───────────
    print("\nComputing LEAP subtotals …")
    df_lcesto = _read_sheet_as_df(wb, "leap_combined_esto")
    df_lcninth = _read_sheet_as_df(wb, "leap_combined_ninth")
    df_nesto = _read_sheet_as_df(wb, "ninth_pairs_to_esto_pairs")
    subtotal_overrides, subtotal_override_df = _load_subtotal_label_overrides()
    mapping_frames = {
        "leap_combined_esto": df_lcesto,
        "leap_combined_ninth": df_lcninth,
        "ninth_pairs_to_esto_pairs": df_nesto,
    }
    stale_subtotal_overrides = _build_stale_subtotal_override_rows(
        subtotal_override_df,
        mapping_frames,
    )
    QA_DIR.mkdir(parents=True, exist_ok=True)
    stale_subtotal_overrides.to_csv(SUBTOTAL_OVERRIDE_STALE_PATH, index=False)
    print(
        f"  Reviewed subtotal overrides: {len(subtotal_override_df):,} rows; "
        f"stale: {len(stale_subtotal_overrides):,} -> {SUBTOTAL_OVERRIDE_STALE_PATH}"
    )
    active_esto_paths = set(
        _active_rows(df_lcesto)["leap_sector_name_full_path"].map(_norm)
        .loc[lambda s: s.ne("")]
    )
    active_ninth_paths = set(
        _active_rows(df_lcninth)["leap_sector_name_full_path"].map(_norm)
        .loc[lambda s: s.ne("")]
    )
    all_leap_paths = active_esto_paths | active_ninth_paths
    fuel_names = set(
        pd.concat([
            _active_rows(df_lcesto)["raw_leap_fuel_name"],
            _active_rows(df_lcninth)["raw_leap_fuel_name"],
        ], ignore_index=True)
        .map(_norm)
        .loc[lambda s: s.ne("")]
    )
    subtotal_paths, full_export_paths, full_export_path = _compute_leap_subtotals_from_full_model_export(
        active_mapping_paths=all_leap_paths,
        fuel_names=fuel_names,
    )
    mapping_inferred_subtotal_paths = _compute_leap_subtotals(all_leap_paths)
    if full_export_path is None:
        subtotal_paths = mapping_inferred_subtotal_paths
        print("  Full model export not found; using active mapping paths as fallback.")
    else:
        missing_from_export = all_leap_paths - full_export_paths
        fallback_subtotal_paths = mapping_inferred_subtotal_paths & missing_from_export
        subtotal_paths = subtotal_paths | fallback_subtotal_paths
        print(f"  Full model export: {full_export_path}")
        print(f"  Export-derived LEAP paths: {len(full_export_paths):,}")
        if missing_from_export:
            print(
                "  WARNING: "
                f"{len(missing_from_export):,} active mapping path(s) were not found in the "
                "full model export-derived path set; using mapping-sheet fallback for those paths."
            )
            print(f"  Fallback subtotal paths from mapping sheets: {len(fallback_subtotal_paths):,}")
    print(f"  Active LEAP paths: {len(all_leap_paths):,}  Subtotal paths: {len(subtotal_paths):,}")
    proposed_lcesto = _build_subtotal_proposed_rows_preview(
        df_lcesto,
        change_specs=[
            (
                "leap_is_subtotal",
                lambda row: _resolved_subtotal_value(
                    "leap_combined_esto",
                    row,
                    "leap_is_subtotal",
                    True if _norm(row.get("leap_sector_name_full_path", "")) in subtotal_paths else False if _norm(row.get("leap_sector_name_full_path", "")) else None,
                    subtotal_overrides,
                ),
            ),
            (
                "esto_pair_is_subtotal",
                lambda row: _resolved_subtotal_value(
                    "leap_combined_esto",
                    row,
                    "esto_pair_is_subtotal",
                    esto_lookup.get((_norm(row.get("esto_flow", "")), _norm(row.get("esto_product", "")))),
                    subtotal_overrides,
                ),
            ),
        ],
    )
    proposed_lcninth = _build_subtotal_proposed_rows_preview(
        df_lcninth,
        change_specs=[
            (
                "leap_is_subtotal",
                lambda row: _resolved_subtotal_value(
                    "leap_combined_ninth",
                    row,
                    "leap_is_subtotal",
                    True if _norm(row.get("leap_sector_name_full_path", "")) in subtotal_paths else False if _norm(row.get("leap_sector_name_full_path", "")) else None,
                    subtotal_overrides,
                ),
            ),
            (
                "ninth_pair_is_subtotal",
                lambda row: _resolved_subtotal_value(
                    "leap_combined_ninth",
                    row,
                    "ninth_pair_is_subtotal",
                    ninth_lookup.get((_norm(row.get("ninth_sector", "")), _norm(row.get("ninth_fuel", "")))),
                    subtotal_overrides,
                ),
            ),
        ],
    )
    proposed_nesto = _build_subtotal_proposed_rows_preview(
        df_nesto,
        change_specs=[
            (
                "ninth_pair_is_subtotal",
                lambda row: _resolved_subtotal_value(
                    "ninth_pairs_to_esto_pairs",
                    row,
                    "ninth_pair_is_subtotal",
                    ninth_lookup.get((_norm(row.get("9th_sector", "")), _norm(row.get("9th_fuel", "")))),
                    subtotal_overrides,
                ),
            ),
            (
                "esto_pair_is_subtotal",
                lambda row: _resolved_subtotal_value(
                    "ninth_pairs_to_esto_pairs",
                    row,
                    "esto_pair_is_subtotal",
                    esto_lookup.get((_norm(row.get("esto_flow", "")), _norm(row.get("esto_product", "")))),
                    subtotal_overrides,
                ),
            ),
        ],
    )
    preview_summary = pd.DataFrame(
        [
            {"sheet_name": "leap_combined_esto", "proposed_row_count": len(proposed_lcesto)},
            {"sheet_name": "leap_combined_ninth", "proposed_row_count": len(proposed_lcninth)},
            {"sheet_name": "ninth_pairs_to_esto_pairs", "proposed_row_count": len(proposed_nesto)},
        ]
    )
    QA_DIR.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(SUBTOTAL_CHANGE_PREVIEW_PATH, engine="openpyxl") as writer:
        preview_summary.to_excel(writer, sheet_name="summary", index=False)
        proposed_lcesto.to_excel(writer, sheet_name="leap_combined_esto_proposed", index=False)
        proposed_lcninth.to_excel(writer, sheet_name="leap_combined_ninth_proposed", index=False)
        proposed_nesto.to_excel(writer, sheet_name="ninth_pairs_to_esto_pairs_proposed", index=False)
    print(f"  subtotal proposed rows preview -> {SUBTOTAL_CHANGE_PREVIEW_PATH}")
    if not apply_subtotal_changes_to_workbook:
        print("  subtotal preview only; workbook update skipped (explicit apply flag not set)")
        return

    _archive_workbook(WORKBOOK_PATH)

    # ── leap_combined_esto ───────────────────────────────────────────────────
    sheet_name = "leap_combined_esto"
    ws = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    u, sk = _update_sheet_leap_subtotals(ws, subtotal_paths)
    print(f"  leap_is_subtotal     -> updated={u}  skipped_blank_path={sk}")
    u, nf, sk = _update_sheet_column(
        ws,
        col_name="esto_pair_is_subtotal",
        key_col_names=["esto_flow", "esto_product"],
        lookup=esto_lookup,
    )
    print(f"  esto_pair_is_subtotal -> updated={u}  not_found={nf}  skipped_blank_key={sk}")
    print(f"  reviewed subtotal overrides -> applied={_apply_subtotal_overrides_to_sheet(ws, subtotal_overrides)}")

    # ── leap_combined_ninth ──────────────────────────────────────────────────
    sheet_name = "leap_combined_ninth"
    ws = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    u, sk = _update_sheet_leap_subtotals(ws, subtotal_paths)
    print(f"  leap_is_subtotal      -> updated={u}  skipped_blank_path={sk}")
    u, nf, sk = _update_sheet_column(
        ws,
        col_name="ninth_pair_is_subtotal",
        key_col_names=["ninth_sector", "ninth_fuel"],
        lookup=ninth_lookup,
    )
    print(f"  ninth_pair_is_subtotal -> updated={u}  not_found={nf}  skipped_blank_key={sk}")
    print(f"  reviewed subtotal overrides -> applied={_apply_subtotal_overrides_to_sheet(ws, subtotal_overrides)}")

    # ── ninth_pairs_to_esto_pairs ────────────────────────────────────────────
    sheet_name = "ninth_pairs_to_esto_pairs"
    ws = wb[sheet_name]
    print(f"\nSheet: {sheet_name}")
    u, nf, sk = _update_sheet_column(
        ws,
        col_name="ninth_pair_is_subtotal",
        key_col_names=["9th_sector", "9th_fuel"],
        lookup=ninth_lookup,
    )
    print(f"  ninth_pair_is_subtotal -> updated={u}  not_found={nf}  skipped_blank_key={sk}")
    u, nf, sk = _update_sheet_column(
        ws,
        col_name="esto_pair_is_subtotal",
        key_col_names=["esto_flow", "esto_product"],
        lookup=esto_lookup,
    )
    print(f"  esto_pair_is_subtotal  -> updated={u}  not_found={nf}  skipped_blank_key={sk}")
    print(f"  reviewed subtotal overrides -> applied={_apply_subtotal_overrides_to_sheet(ws, subtotal_overrides)}")

    wb.save(WORKBOOK_PATH)
    print(f"\nSaved -> {WORKBOOK_PATH}")

    # ── Re-read updated sheets for QA ────────────────────────────────────────
    print("\nBuilding QA outputs …")
    df_lcesto = pd.read_excel(WORKBOOK_PATH, sheet_name="leap_combined_esto", dtype=object).fillna("")
    df_lcninth = pd.read_excel(WORKBOOK_PATH, sheet_name="leap_combined_ninth", dtype=object).fillna("")
    df_nesto = pd.read_excel(WORKBOOK_PATH, sheet_name="ninth_pairs_to_esto_pairs", dtype=object).fillna("")
    for df in [df_lcesto, df_lcninth, df_nesto]:
        for col in df.columns:
            df[col] = df[col].astype(str).map(_norm)

    QA_DIR.mkdir(parents=True, exist_ok=True)

    # Cardinality
    card_lcesto = _compute_cardinality(
        df_lcesto, "leap_sector_name_full_path", "raw_leap_fuel_name", "esto_flow", "esto_product"
    )
    card_lcninth = _compute_cardinality(
        df_lcninth, "leap_sector_name_full_path", "raw_leap_fuel_name", "ninth_sector", "ninth_fuel"
    )
    card_nesto = _compute_cardinality(
        df_nesto, "9th_sector", "9th_fuel", "esto_flow", "esto_product"
    )
    card_lcesto.to_csv(QA_DIR / "cardinality_leap_esto.csv", index=False)
    card_lcninth.to_csv(QA_DIR / "cardinality_leap_ninth.csv", index=False)
    card_nesto.to_csv(QA_DIR / "cardinality_ninth_esto.csv", index=False)
    print(f"  cardinality_leap_esto:  {len(card_lcesto):,} pairs  "
          f"({(card_lcesto['cardinality'] == 'one_to_one').sum():,} one-to-one)")
    print(f"  cardinality_leap_ninth: {len(card_lcninth):,} pairs  "
          f"({(card_lcninth['cardinality'] == 'one_to_one').sum():,} one-to-one)")
    print(f"  cardinality_ninth_esto: {len(card_nesto):,} pairs  "
          f"({(card_nesto['cardinality'] == 'one_to_one').sum():,} one-to-one)")

    # Migrated legacy conflict checks
    all_many_to_many = _many_to_many_conflicts([
        ("leap_combined_esto", card_lcesto),
        ("leap_combined_ninth", card_lcninth),
        ("ninth_pairs_to_esto_pairs", card_nesto),
    ])
    many_to_many, allowed_many_to_many = _split_allowed_many_to_many(all_many_to_many)
    leap_source_presence = _leap_source_presence_conflicts(df_lcesto, df_lcninth)
    all_crosswalk_conflicts = _crosswalk_target_conflicts(df_lcesto, df_lcninth, df_nesto)
    crosswalk_conflicts, allowed_crosswalk_conflicts = _split_allowed_crosswalk_conflicts(
        all_crosswalk_conflicts
    )
    leap_source_presence, allowed_leap_source_presence = _split_allowed_leap_source_presence(
        leap_source_presence
    )
    allowed_many_to_many.to_csv(QA_DIR / "many_to_many_allowed_matched.csv", index=False)
    many_to_many.to_csv(QA_DIR / "many_to_many_conflicts.csv", index=False)
    allowed_leap_source_presence.to_csv(QA_DIR / "leap_source_presence_allowed_matched.csv", index=False)
    leap_source_presence.to_csv(QA_DIR / "leap_source_presence_conflicts.csv", index=False)
    allowed_crosswalk_conflicts.to_csv(QA_DIR / "crosswalk_target_conflicts_allowed_matched.csv", index=False)
    crosswalk_conflicts.to_csv(QA_DIR / "crosswalk_target_conflicts.csv", index=False)
    print(f"  many_to_many_allowed_matched:      {len(allowed_many_to_many):,}")
    print(f"  many_to_many_conflicts:            {len(many_to_many):,}")
    print(f"  leap_source_presence_allowed_matched:{len(allowed_leap_source_presence):,}")
    print(f"  leap_source_presence_conflicts:    {len(leap_source_presence):,}")
    print(f"  crosswalk_target_conflicts_allowed_matched:{len(allowed_crosswalk_conflicts):,}")
    print(f"  crosswalk_target_conflicts:        {len(crosswalk_conflicts):,}")
    if not crosswalk_conflicts.empty:
        crosswalk_class_counts = crosswalk_conflicts["conflict_classification"].value_counts().to_dict()
        print(f"    crosswalk classifications:       {crosswalk_class_counts}")

    # Unmapped pairs
    unmapped_esto = _unmapped_esto_pairs([df_lcesto, df_nesto], esto_lookup)
    unmapped_ninth = _unmapped_ninth_pairs([df_lcninth, df_nesto], ninth_lookup)
    unmapped_esto, allowed_unmapped_esto = split_allowed_rows(
        unmapped_esto,
        sheet_name=UNMAPPED_NONZERO_ESTO_ALLOWED_SHEET,
        status_column="unmapped_nonzero_review_status",
        reason_column="unmapped_nonzero_review_reason",
        workbook_path=EXCEPTION_WORKBOOK_PATH,
    )
    unmapped_ninth, allowed_unmapped_ninth = split_allowed_rows(
        unmapped_ninth,
        sheet_name=UNMAPPED_NONZERO_NINTH_ALLOWED_SHEET,
        status_column="unmapped_nonzero_review_status",
        reason_column="unmapped_nonzero_review_reason",
        workbook_path=EXCEPTION_WORKBOOK_PATH,
    )
    unmapped_esto.to_csv(QA_DIR / "unmapped_nonzero_esto_pairs.csv", index=False)
    allowed_unmapped_esto.to_csv(QA_DIR / "unmapped_nonzero_esto_pairs_allowed_matched.csv", index=False)
    unmapped_ninth.to_csv(QA_DIR / "unmapped_nonzero_ninth_pairs.csv", index=False)
    allowed_unmapped_ninth.to_csv(QA_DIR / "unmapped_nonzero_ninth_pairs_allowed_matched.csv", index=False)
    print(f"  unmapped_nonzero_esto_pairs:  {len(unmapped_esto):,}  allowed={len(allowed_unmapped_esto):,}")
    print(f"  unmapped_nonzero_ninth_pairs: {len(unmapped_ninth):,}  allowed={len(allowed_unmapped_ninth):,}")

    # Subtotal mismatches (M6)
    mm_esto = _subtotal_mismatches(
        df_lcesto,
        "leap_sector_name_full_path", "raw_leap_fuel_name",
        "esto_flow", "esto_product",
        "leap_is_subtotal", "esto_pair_is_subtotal",
    )
    mm_ninth = _subtotal_mismatches(
        df_lcninth,
        "leap_sector_name_full_path", "raw_leap_fuel_name",
        "ninth_sector", "ninth_fuel",
        "leap_is_subtotal", "ninth_pair_is_subtotal",
    )
    mm_nesto = _subtotal_mismatches(
        df_nesto,
        "9th_sector", "9th_fuel",
        "esto_flow", "esto_product",
        "ninth_pair_is_subtotal", "esto_pair_is_subtotal",
    )
    all_mm = pd.concat([
        mm_esto.assign(sheet="leap_combined_esto"),
        mm_ninth.assign(sheet="leap_combined_ninth"),
        mm_nesto.assign(sheet="ninth_pairs_to_esto_pairs"),
    ], ignore_index=True)
    subtotal_mismatches, allowed_subtotal_mismatches = _split_allowed_subtotal_mismatches(all_mm)
    subtotal_mismatches.to_csv(QA_DIR / "subtotal_mismatches.csv", index=False)
    allowed_subtotal_mismatches.to_csv(QA_DIR / "subtotal_mismatches_allowed_matched.csv", index=False)
    _remove_stale_generated_exception_outputs()
    print(
        "  subtotal_mismatches_allowed_matched: "
        f"{len(allowed_subtotal_mismatches):,}  (matched manual allowlist)"
    )
    print(
        "  subtotal_mismatches:  "
        f"{len(subtotal_mismatches):,}  (not in manual allowlist)"
    )

    print(f"\nQA outputs written to: {QA_DIR}")

    # --- Display name sync ---------------------------------------------------
    print("\nUpdating leap_display_names …")
    from codebase.mapping_tools.update_leap_display_names import (
        EXCEPTION_WORKBOOK_PATH as _DN_EXCEPTION_PATH,
        run_display_name_update,
    )
    dn_qa = run_display_name_update(
        workbook_path=WORKBOOK_PATH,
        qa_dir=QA_DIR,
        exception_workbook_path=_DN_EXCEPTION_PATH,
    )
    _dn_orphans = int(
        dn_qa["status"].isin(["potential_issue_orphan"]).sum()
    ) if not dn_qa.empty else 0
    _dn_dups = int(dn_qa.get("duplicate_display_name", False).sum()) if not dn_qa.empty else 0
    if _dn_orphans or _dn_dups:
        print(
            f"  DISPLAY NAME ISSUES: {_dn_orphans} orphan(s), "
            f"{_dn_dups} duplicate display name(s) — see display_names_qa.csv"
        )

    # --- M3: tree structure ---------------------------------------------------
    print("\nBuilding dataset tree structures …")
    from codebase.mapping_tools.build_dataset_tree_structure import run_tree_structure_workflow
    run_tree_structure_workflow()

    summary = _write_maintenance_summary(QA_DIR / "maintenance_summary.csv")
    print(f"\nMaintenance summary: {len(summary):,} rows -> {(QA_DIR / 'maintenance_summary.csv').relative_to(REPO_ROOT)}")


if __name__ == "__main__":
    run()


