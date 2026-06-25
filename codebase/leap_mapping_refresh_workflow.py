#%%

"""
Legacy refresh workflow for config/leap_mappings.xlsx.

The current Stage 0 path for config/outlook_mappings_master.xlsx is
codebase/outlook_mapping_maintenance_workflow.py. Keep this file as legacy
reference for old workbook checks and migration ideas; do not use it for the
new canonical mapping pipeline unless a task explicitly asks for legacy
workbook maintenance.
"""

from __future__ import annotations

import os
import re
import shutil
import sys
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Sequence

import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

REPO_ROOT = Path(__file__).resolve().parents[1]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

from codebase.utilities.leap_balance_export_resolver import resolve_balance_export_workbook  # noqa: E402
from codebase.utilities.energy_balance_template_extractor import (  # noqa: E402
    TemplateBalanceExtractor,
    _parse_unit_factor_to_petajoule,
)
from codebase.utilities.leap_results_dashboard_balance import _list_balance_sheets, _pick_template_sheet  # noqa: E402


#%%
def _resolve(path: Path | str) -> Path:
    raw = str(path).replace("\\", "/")
    drive_match = re.match(r"^([a-zA-Z]):/(.*)$", raw)
    if drive_match:
        drive = drive_match.group(1).lower()
        rest = drive_match.group(2)
        if os.name == "nt":
            return Path(f"{drive.upper()}:/{rest}")
        return Path(f"/mnt/{drive}/{rest}")
    candidate = Path(raw)
    return candidate if candidate.is_absolute() else (REPO_ROOT / candidate)


MAPPING_WORKBOOK_PATH = _resolve("config/leap_mappings.xlsx")
MASTER_CONFIG_PATH = _resolve("config/master_config.xlsx")
CODEBOOK_PATH = _resolve("config/sector_fuel_codes_to_names.xlsx")
ESTO_TABLE_PATH = _resolve("data/00APEC_2025_low_with_subtotals.csv")
NINTH_TABLE_PATH = _resolve("data/merged_file_energy_ALL_20251106.csv")
OUTPUT_DIR = _resolve("outputs/mappings/mapping_checks")
RESEARCHER_MAPPINGS_PATH = _resolve("outputs/mappings/researcher_mappings.xlsx")
MISSING_PAIRS_CSV_PATH = OUTPUT_DIR / "leap_mapping_missing_pairs.csv"
DUPLICATE_MAPPINGS_CSV_PATH = OUTPUT_DIR / "leap_mapping_duplicate_mappings.csv"
TRIO_PRESENCE_CSV_PATH = OUTPUT_DIR / "leap_mapping_trio_presence_check.csv"
MAPPING_CONFLICTS_WORKBOOK_PATH = OUTPUT_DIR / "leap_mapping_conflicts.xlsx"
MAPPING_REFRESH_REPORT_PATH = OUTPUT_DIR / "leap_mapping_refresh_report.xlsx"

TRIO_PRESENCE_OUTPUT_NOTE = (
    "Filter trio_presence_csv by presence_status first. Focus most on "
    "ninth_active_esto_removed, esto_active_ninth_removed, "
    "esto_removed_ninth_active, and ninth_active_esto_missing; these are mapped "
    "rows that can change expected dashboard results. Treat fuel=Total rows, "
    "same-target parent/child mappings, old incorrect fuel rows, expected "
    "losses-sector removals like 10.01.02/10.01.03, and detailed transport "
    "rows kept remove_row=True as low priority. Sort previous_runs ascending "
    "to find rows that are new or least frequently repeated across refreshes."
)

ESTO_SHEET = "leap_combined_esto"
NINTH_SHEET = "leap_combined_ninth"
SECTOR_FUEL_CODE_TO_NAME_SHEET = "sector_fuel_code_to_name"
NINTH_PAIRS_TO_ESTO_PAIRS_SHEET = "ninth_pairs_to_esto_pairs"

BASE_YEAR = 2022
PROJECTION_YEARS: Sequence[int] = tuple(range(2023, 2061))
PROJECTION_SCENARIOS: Sequence[str] = ("reference", "target")
BALANCE_EXPORT_ECONOMY = "20_USA"
REF_BALANCE_EXPORT_DATE_ID: str | None = None
TGT_BALANCE_EXPORT_DATE_ID: str | None = None
BALANCE_TEMPLATE_SHEET = "EBal|2060"


#%%
def _clean(value: object) -> str:
    text = str(value or "").strip()
    return "" if text.lower() in {"", "nan", "none"} else text


def _norm_text(value: object) -> str:
    return " ".join(_clean(value).lower().split())


def _truthy(value: object) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "t", "yes", "y", "on"}


def _path_key(path: object) -> str:
    parts = [part.strip() for part in str(path or "").split("/") if part.strip()]
    return "/".join(_norm_text(part) for part in parts)


def _mapping_cardinality(source_target_count: int, target_source_count: int) -> str:
    if source_target_count <= 0 or target_source_count <= 0:
        return ""
    if source_target_count == 1 and target_source_count == 1:
        return "one_to_one"
    if source_target_count > 1 and target_source_count == 1:
        return "one_to_many"
    if source_target_count == 1 and target_source_count > 1:
        return "many_to_one"
    return "many_to_many"


def _subtotal_alignment(leap_is_subtotal: bool, target_is_subtotal: bool) -> str:
    if leap_is_subtotal and target_is_subtotal:
        return "aligned_subtotal"
    if (not leap_is_subtotal) and (not target_is_subtotal):
        return "aligned_non_subtotal"
    return "mismatch"


def _active_mask(frame: pd.DataFrame) -> pd.Series:
    remove_mask = frame.get("remove_row", False)
    duplicate_mask = frame.get("duplicate_to_remove", False)
    remove_mask = pd.Series(remove_mask, index=frame.index).map(_truthy)
    duplicate_mask = pd.Series(duplicate_mask, index=frame.index).map(_truthy)
    return ~(remove_mask | duplicate_mask)


def _drop_unnamed_columns(frame: pd.DataFrame) -> pd.DataFrame:
    keep_cols = [col for col in frame.columns if not str(col).startswith("Unnamed:")]
    return frame.loc[:, keep_cols].copy()


def _drop_columns_if_present(frame: pd.DataFrame, columns: Sequence[str]) -> pd.DataFrame:
    drop_cols = [col for col in columns if col in frame.columns]
    if not drop_cols:
        return frame.copy()
    return frame.drop(columns=drop_cols).copy()


def _reorder_columns(frame: pd.DataFrame, preferred_columns: Sequence[str]) -> pd.DataFrame:
    ordered = [col for col in preferred_columns if col in frame.columns]
    trailing = [col for col in frame.columns if col not in ordered]
    return frame.loc[:, ordered + trailing].copy()


def _compute_leap_subtotals(frame: pd.DataFrame) -> pd.DataFrame:
    out = frame.copy()
    for col in ["leap_sector_name_full_path", "raw_leap_fuel_name"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()

    active = _active_mask(out)
    active_paths = {
        _clean(value)
        for value in out.loc[active, "leap_sector_name_full_path"].tolist()
        if _clean(value)
    }

    def leap_sector_is_subtotal(path: object) -> bool:
        text = _clean(path)
        key = _path_key(text)
        if not key:
            return False
        if key.startswith("total "):
            return True
        prefix = f"{text}/"
        return any(other != text and other.startswith(prefix) for other in active_paths)

    def leap_fuel_is_subtotal(fuel: object) -> bool:
        key = _norm_text(fuel)
        return key == "total" or key.startswith("total ")

    leap_sector_is_subtotal = out["leap_sector_name_full_path"].map(leap_sector_is_subtotal)
    leap_fuel_is_subtotal = out["raw_leap_fuel_name"].map(leap_fuel_is_subtotal)
    out["leap_is_subtotal"] = leap_sector_is_subtotal.fillna(False).astype(bool) | leap_fuel_is_subtotal.fillna(False).astype(bool)
    return out


def _compute_pair_cardinality(frame: pd.DataFrame, target_sector_col: str, target_fuel_col: str) -> pd.DataFrame:
    """Compute cardinality of (leap_sector, leap_fuel) <-> (target_sector, target_fuel) pairs."""
    out = frame.copy()
    source_cols = ["leap_sector_name_full_path", "raw_leap_fuel_name"]
    target_cols = [target_sector_col, target_fuel_col]
    all_cols = source_cols + [c for c in target_cols if c not in source_cols]
    for col in all_cols:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()
    active = _active_mask(out)
    valid = (
        active
        & out["leap_sector_name_full_path"].ne("")
        & out["raw_leap_fuel_name"].ne("")
        & out[target_sector_col].ne("")
        & out[target_fuel_col].ne("")
    )
    pair_frame = out.loc[valid, source_cols + [target_sector_col, target_fuel_col]].copy()
    pair_frame["_source_key"] = pair_frame["leap_sector_name_full_path"] + "|||" + pair_frame["raw_leap_fuel_name"]
    pair_frame["_target_key"] = pair_frame[target_sector_col] + "|||" + pair_frame[target_fuel_col]
    pairs = pair_frame[["_source_key", "_target_key"]].drop_duplicates()
    source_count = pairs.groupby("_source_key")["_target_key"].nunique()
    target_count = pairs.groupby("_target_key")["_source_key"].nunique()
    out["_source_key"] = out["leap_sector_name_full_path"].fillna("").astype(str).str.strip() + "|||" + out["raw_leap_fuel_name"].fillna("").astype(str).str.strip()
    out["_target_key"] = out[target_sector_col].fillna("").astype(str).str.strip() + "|||" + out[target_fuel_col].fillna("").astype(str).str.strip()
    out["pair_mapping_cardinality"] = ""
    valid_rows = out["leap_sector_name_full_path"].ne("") & out["raw_leap_fuel_name"].ne("") & out[target_sector_col].ne("") & out[target_fuel_col].ne("")
    out.loc[valid_rows, "pair_mapping_cardinality"] = out.loc[valid_rows].apply(
        lambda row: _mapping_cardinality(
            int(source_count.get(row["_source_key"], 0)),
            int(target_count.get(row["_target_key"], 0)),
        ),
        axis=1,
    )
    out = out.drop(columns=["_source_key", "_target_key"])
    return out


def _apply_auto_remove_rules(frame: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, int]]:
    """Mark obvious rows as remove_row=True and annotate the reason."""
    out = frame.copy()
    for col in [
        "leap_sector_name_full_path",
        "raw_leap_fuel_name",
        "remove_row",
        "remove_row_reason",
    ]:
        if col not in out.columns:
            out[col] = ""
    out["leap_sector_name_full_path"] = out["leap_sector_name_full_path"].fillna("").astype(str).str.strip()
    out["raw_leap_fuel_name"] = out["raw_leap_fuel_name"].fillna("").astype(str).str.strip()
    out["remove_row_reason"] = out["remove_row_reason"].fillna("").astype(str).str.strip()

    fuel_total_mask = out["raw_leap_fuel_name"].map(_norm_text).eq("total")
    def _sector_ends_with_fuel(path_value: object, fuel_value: object) -> bool:
        path_text = _clean(path_value)
        fuel_text = _clean(fuel_value)
        if not path_text or not fuel_text:
            return False
        parts = [part.strip() for part in path_text.split("/") if part.strip()]
        return len(parts) > 1 and parts[-1] == fuel_text

    suffix_mask = out.apply(
        lambda row: _sector_ends_with_fuel(row["leap_sector_name_full_path"], row["raw_leap_fuel_name"])
        and _norm_text(row["raw_leap_fuel_name"]) != "total"
        and not _clean(row["leap_sector_name_full_path"]).startswith("Electricity Generation/"),
        axis=1,
    )

    existing_remove_mask = out["remove_row"].map(_truthy)
    auto_mask = fuel_total_mask | suffix_mask
    newly_removed_mask = auto_mask & ~existing_remove_mask

    out["remove_row"] = existing_remove_mask | auto_mask

    def _append_reason(existing: str, reason: str) -> str:
        if not reason:
            return existing
        if not existing:
            return reason
        if reason in existing.split(" | "):
            return existing
        return f"{existing} | {reason}"

    def _strip_auto_reasons(existing: str) -> str:
        reasons = [part.strip() for part in existing.split(" | ") if part.strip()]
        reasons = [reason for reason in reasons if reason not in {"auto_remove_total_fuel", "auto_remove_sector_fuel_suffix"}]
        return " | ".join(reasons)

    out["remove_row_reason"] = out["remove_row_reason"].map(_strip_auto_reasons)
    out.loc[fuel_total_mask, "remove_row_reason"] = out.loc[fuel_total_mask, "remove_row_reason"].map(
        lambda reason: _append_reason(reason, "auto_remove_total_fuel")
    )
    out.loc[suffix_mask, "remove_row_reason"] = out.loc[suffix_mask, "remove_row_reason"].map(
        lambda reason: _append_reason(reason, "auto_remove_sector_fuel_suffix")
    )

    diagnostics = {
        "auto_remove_total_fuel_rows": int(fuel_total_mask.sum()),
        "auto_remove_sector_fuel_suffix_rows": int(suffix_mask.sum()),
        "auto_removed_new_rows": int(newly_removed_mask.sum()),
    }
    return out, diagnostics


def _load_esto_lookup() -> pd.DataFrame:
    base_df = pd.read_csv(ESTO_TABLE_PATH)
    work = base_df.copy()
    if "is_subtotal" not in work.columns:
        work["is_subtotal"] = False
    for col in ["economy", "flows", "products", str(BASE_YEAR), "is_subtotal"]:
        if col not in work.columns:
            work[col] = ""
    work["esto_flow"] = work["flows"].fillna("").astype(str).str.strip()
    work["esto_product"] = work["products"].fillna("").astype(str).str.strip()
    work["value"] = pd.to_numeric(work[str(BASE_YEAR)], errors="coerce").fillna(0.0)
    work["is_subtotal"] = work["is_subtotal"].fillna(False).map(_truthy)
    work = work[work["esto_flow"].ne("") & work["esto_product"].ne("")].copy()
    grouped = (
        work.groupby(["esto_flow", "esto_product"], as_index=False)
        .agg(
            pair_value_sum=("value", "sum"),
            esto_pair_is_subtotal=("is_subtotal", "max"),
        )
        .reset_index(drop=True)
    )
    grouped["esto_pair_abs_sum"] = grouped["pair_value_sum"].abs()
    return grouped


def _load_ninth_lookup() -> pd.DataFrame:
    ninth_df = pd.read_csv(NINTH_TABLE_PATH)
    work = ninth_df.copy()
    for col in [
        "economy",
        "scenarios",
        "sectors",
        "sub1sectors",
        "sub2sectors",
        "sub3sectors",
        "sub4sectors",
        "fuels",
        "subfuels",
        "subtotal_layout",
        "subtotal_results",
    ]:
        if col not in work.columns:
            work[col] = ""
    for col in ["subtotal_layout", "subtotal_results"]:
        work[col] = work[col].fillna(False).map(_truthy)
    scenario_set = {str(value).strip().lower() for value in PROJECTION_SCENARIOS}
    work = work[work["scenarios"].fillna("").astype(str).str.strip().str.lower().isin(scenario_set)].copy()
    year_cols = [str(year) for year in PROJECTION_YEARS if str(year) in work.columns]
    if not year_cols or work.empty:
        return pd.DataFrame(
            columns=[
                "ninth_sector",
                "ninth_fuel",
                "ninth_pair_is_subtotal",
                "ninth_pair_abs_sum",
            ]
        )
    values = work[year_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    work["ninth_sector"] = work.apply(
        lambda row: next(
            (
                _clean(row.get(col, ""))
                for col in ["sub4sectors", "sub3sectors", "sub2sectors", "sub1sectors", "sectors"]
                if _clean(row.get(col, ""))
            ),
            "",
        ),
        axis=1,
    )
    work["ninth_fuel"] = work.apply(
        lambda row: next(
            (
                _clean(row.get(col, ""))
                for col in ["subfuels", "fuels"]
                if _clean(row.get(col, ""))
            ),
            "",
        ),
        axis=1,
    )
    work["value_abs_sum_row"] = values.abs().sum(axis=1)
    work = work[work["ninth_sector"].ne("") & work["ninth_fuel"].ne("")].copy()
    grouped = (
        work.groupby(["ninth_sector", "ninth_fuel"], as_index=False)
        .agg(
            subtotal_layout=("subtotal_layout", "max"),
            subtotal_results=("subtotal_results", "max"),
            ninth_pair_abs_sum=("value_abs_sum_row", "sum"),
        )
        .reset_index(drop=True)
    )
    grouped["ninth_pair_is_subtotal"] = (
        grouped["subtotal_layout"].fillna(False).astype(bool)
        | grouped["subtotal_results"].fillna(False).astype(bool)
    )
    return grouped


def _refresh_esto_sheet(frame: pd.DataFrame, esto_lookup: pd.DataFrame) -> pd.DataFrame:
    out = _drop_unnamed_columns(frame)
    out = _drop_columns_if_present(
        out,
        [
            "many_to_many_is_ok",
            "esto_pair_is_subtotal",
            "esto_pair_is_subtotal_x",
            "esto_pair_is_subtotal_y",
            "esto_pair_abs_sum",
            "esto_pair_abs_sum_x",
            "esto_pair_abs_sum_y",
            "leap_sector_is_subtotal_computed",
            "leap_fuel_is_subtotal_computed",
        ],
    )
    out = _compute_leap_subtotals(out)
    out = _compute_pair_cardinality(out, "esto_flow", "esto_product")
    lookup = esto_lookup.copy()
    for col in ["esto_flow", "esto_product"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()
    out = out.merge(
        lookup[["esto_flow", "esto_product", "esto_pair_is_subtotal", "esto_pair_abs_sum"]],
        on=["esto_flow", "esto_product"],
        how="left",
    )
    if "esto_pair_is_subtotal" not in out.columns:
        out["esto_pair_is_subtotal"] = False
    out["esto_pair_is_subtotal"] = out["esto_pair_is_subtotal"].fillna(False).astype(bool)
    if "esto_pair_abs_sum" not in out.columns:
        out["esto_pair_abs_sum"] = 0.0
    out["esto_pair_abs_sum"] = pd.to_numeric(out["esto_pair_abs_sum"], errors="coerce").fillna(0.0)
    total_mask = out["esto_product"].fillna("").astype(str).str.strip().str.lower().eq("19 total")
    out.loc[total_mask, "esto_pair_is_subtotal"] = True
    out["subtotal_alignment"] = out.apply(
        lambda row: _subtotal_alignment(bool(row.get("leap_is_subtotal", False)), bool(row.get("esto_pair_is_subtotal", False))),
        axis=1,
    )
    return _reorder_columns(
        out,
        [
            "leap_sector_name_original",
            "leap_sector_name_full_path",
            "raw_leap_fuel_name",
            "value",
            "esto_flow",
            "esto_product",
            "pair_mapping_cardinality",
            "leap_is_subtotal",
            "esto_pair_is_subtotal",
            "subtotal_mismatch_is_ok",
            "subtotal_alignment",
            "esto_pair_abs_sum",
            "remove_row",
            "remove_row_reason",
        ],
    )


def _refresh_ninth_sheet(frame: pd.DataFrame, ninth_lookup: pd.DataFrame) -> pd.DataFrame:
    out = _drop_unnamed_columns(frame)
    out = _drop_columns_if_present(
        out,
        [
            "many_to_many_is_ok",
            "ninth_pair_is_subtotal",
            "ninth_pair_is_subtotal_x",
            "ninth_pair_is_subtotal_y",
            "ninth_pair_abs_sum",
            "ninth_pair_abs_sum_x",
            "ninth_pair_abs_sum_y",
            "leap_sector_is_subtotal_computed",
            "leap_fuel_is_subtotal_computed",
        ],
    )
    out = _compute_leap_subtotals(out)
    out = _compute_pair_cardinality(out, "ninth_sector", "ninth_fuel")
    lookup = ninth_lookup.copy()
    for col in ["ninth_sector", "ninth_fuel"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()
    out = out.merge(
        lookup[["ninth_sector", "ninth_fuel", "ninth_pair_is_subtotal", "ninth_pair_abs_sum"]],
        on=["ninth_sector", "ninth_fuel"],
        how="left",
    )
    if "ninth_pair_is_subtotal" not in out.columns:
        out["ninth_pair_is_subtotal"] = False
    out["ninth_pair_is_subtotal"] = out["ninth_pair_is_subtotal"].fillna(False).astype(bool)
    if "ninth_pair_abs_sum" not in out.columns:
        out["ninth_pair_abs_sum"] = 0.0
    out["ninth_pair_abs_sum"] = pd.to_numeric(out["ninth_pair_abs_sum"], errors="coerce").fillna(0.0)
    total_mask = out["ninth_fuel"].fillna("").astype(str).str.strip().str.lower().eq("19_total")
    out.loc[total_mask, "ninth_pair_is_subtotal"] = True
    out["subtotal_alignment"] = out.apply(
        lambda row: _subtotal_alignment(bool(row.get("leap_is_subtotal", False)), bool(row.get("ninth_pair_is_subtotal", False))),
        axis=1,
    )
    return _reorder_columns(
        out,
        [
            "leap_sector_name_original",
            "leap_sector_name_full_path",
            "raw_leap_fuel_name",
            "value",
            "ninth_sector",
            "ninth_fuel",
            "pair_mapping_cardinality",
            "leap_is_subtotal",
            "ninth_pair_is_subtotal",
            "subtotal_mismatch_is_ok",
            "subtotal_alignment",
            "ninth_pair_abs_sum",
            "remove_row",
            "remove_row_reason",
        ],
    )


def _active_pairs(frame: pd.DataFrame, col_a: str, col_b: str) -> set[tuple[str, str]]:
    """Return the set of (col_a, col_b) pairs in active (non-removed) rows."""
    active = frame[_active_mask(frame)].copy()
    a = active[col_a].fillna("").astype(str).str.strip() if col_a in active.columns else pd.Series("", index=active.index)
    b = active[col_b].fillna("").astype(str).str.strip() if col_b in active.columns else pd.Series("", index=active.index)
    return {(av, bv) for av, bv in zip(a, b) if av and bv}


def _active_leap_source_pairs(frame: pd.DataFrame) -> set[tuple[str, str]]:
    """Return active LEAP sector/fuel source pairs from a mapping sheet."""
    active = frame[_active_mask(frame)].copy()
    for col in ["leap_sector_name_full_path", "raw_leap_fuel_name"]:
        if col not in active.columns:
            active[col] = ""
        active[col] = active[col].fillna("").astype(str).str.strip()
    return {
        (_path_key(sector), _norm_text(fuel))
        for sector, fuel in zip(active["leap_sector_name_full_path"], active["raw_leap_fuel_name"])
        if _path_key(sector) and _norm_text(fuel)
    }


def _leap_source_pair_presence_lookup(frame: pd.DataFrame) -> dict[tuple[str, str], dict[str, object]]:
    """Return active/removed presence counts for LEAP source pairs in a mapping sheet."""
    work = frame.copy()
    for col in ["leap_sector_name_full_path", "raw_leap_fuel_name", "remove_row", "duplicate_to_remove"]:
        if col not in work.columns:
            work[col] = ""
    work["leap_sector_name_full_path"] = work["leap_sector_name_full_path"].fillna("").astype(str).str.strip()
    work["raw_leap_fuel_name"] = work["raw_leap_fuel_name"].fillna("").astype(str).str.strip()
    work["_source_key"] = list(zip(
        work["leap_sector_name_full_path"].map(_path_key),
        work["raw_leap_fuel_name"].map(_norm_text),
    ))
    work = work[work["_source_key"].map(lambda key: bool(key[0] and key[1]))].copy()
    if work.empty:
        return {}

    work["_is_removed"] = work["remove_row"].map(_truthy)
    work["_is_duplicate_removed"] = work["duplicate_to_remove"].map(_truthy)
    work["_is_active"] = ~(work["_is_removed"] | work["_is_duplicate_removed"])

    lookup: dict[tuple[str, str], dict[str, object]] = {}
    for source_key, group in work.groupby("_source_key", dropna=False):
        active_count = int(group["_is_active"].sum())
        removed_count = int(group["_is_removed"].sum())
        duplicate_removed_count = int(group["_is_duplicate_removed"].sum())
        total_count = int(len(group))
        if active_count:
            state = "active"
        elif removed_count and duplicate_removed_count:
            state = "removed_or_duplicate_removed_only"
        elif removed_count:
            state = "removed_only"
        elif duplicate_removed_count:
            state = "duplicate_removed_only"
        else:
            state = "present_but_inactive"
        lookup[source_key] = {
            "state": state,
            "detail": (
                f"active={active_count}; removed={removed_count}; "
                f"duplicate_removed={duplicate_removed_count}; total={total_count}"
            ),
        }
    return lookup


def _build_duplicate_mappings(frame: pd.DataFrame, *, sheet_name: str, target_a: str, target_b: str) -> pd.DataFrame:
    """Return exact active duplicate source/target rows for one mapping sheet."""
    work = frame.copy().fillna("")
    source_cols = ["leap_sector_name_full_path", "raw_leap_fuel_name"]
    target_cols = [target_a, target_b]
    required_cols = [*source_cols, *target_cols, "remove_row", "duplicate_to_remove"]
    for col in required_cols:
        if col not in work.columns:
            work[col] = ""
        work[col] = work[col].fillna("").astype(str).str.strip()

    active = work[_active_mask(work)].copy()
    valid = active[source_cols + target_cols].apply(lambda col: col.map(_clean).ne("")).all(axis=1)
    active = active.loc[valid].copy()
    if active.empty:
        return pd.DataFrame(
            columns=[
                "sheet_name",
                "mapping_row_number",
                "duplicate_group_size",
                *source_cols,
                *target_cols,
            ]
        )

    duplicate_mask = active.duplicated(subset=[*source_cols, *target_cols], keep=False)
    duplicates = active.loc[duplicate_mask].copy()
    if duplicates.empty:
        return pd.DataFrame(
            columns=[
                "sheet_name",
                "mapping_row_number",
                "duplicate_group_size",
                *source_cols,
                *target_cols,
            ]
        )

    duplicates.insert(0, "mapping_row_number", duplicates.index + 2)
    duplicates.insert(0, "sheet_name", sheet_name)
    duplicates["duplicate_group_size"] = duplicates.groupby([*source_cols, *target_cols])[source_cols[0]].transform("size")
    return duplicates[[
        "sheet_name",
        "mapping_row_number",
        "duplicate_group_size",
        *source_cols,
        *target_cols,
    ]].reset_index(drop=True)


def _build_trio_presence_check(esto_sheet: pd.DataFrame, ninth_sheet: pd.DataFrame) -> pd.DataFrame:
    """Return row-level presence diagnostics for the two mapping sheets."""
    source_cols = ["leap_sector_name_original", "leap_sector_name_full_path", "raw_leap_fuel_name"]

    def _sheet_row_status(frame: pd.DataFrame, sheet_name: str, target_cols: list[str]) -> pd.DataFrame:
        work = frame.copy().fillna("")
        for col in source_cols + target_cols + ["remove_row", "duplicate_to_remove"]:
            if col not in work.columns:
                work[col] = ""
            work[col] = work[col].fillna("").astype(str).str.strip()
        valid = work[source_cols + target_cols].apply(lambda col: col.map(_clean).ne("")).all(axis=1)
        work = work.loc[valid].copy()
        if work.empty:
            return pd.DataFrame(
                columns=[
                    "sheet_name",
                    "mapping_row_number",
                    *source_cols,
                    *target_cols,
                    "this_row_status",
                    "this_row_is_removed",
                    "this_row_is_duplicate_removed",
                ]
            )
        work["sheet_name"] = sheet_name
        work["mapping_row_number"] = work.index + 2
        work["this_row_is_removed"] = work["remove_row"].map(_truthy)
        work["this_row_is_duplicate_removed"] = work["duplicate_to_remove"].map(_truthy)
        work["this_row_status"] = work.apply(
            lambda row: "removed_row_true"
            if row["this_row_is_removed"]
            else "duplicate_removed_row_true"
            if row["this_row_is_duplicate_removed"]
            else "active",
            axis=1,
        )
        return work[
            [
                "sheet_name",
                "mapping_row_number",
                *source_cols,
                *target_cols,
                "this_row_status",
                "this_row_is_removed",
                "this_row_is_duplicate_removed",
            ]
        ].reset_index(drop=True)

    def _first_non_empty(series: pd.Series) -> str:
        values = [str(value).strip() for value in series.tolist() if _clean(value)]
        unique_values = list(dict.fromkeys(values))
        return " | ".join(unique_values)

    def _sheet_source_summary(frame: pd.DataFrame, sheet_name: str, target_cols: list[str]) -> pd.DataFrame:
        work = frame.copy().fillna("")
        for col in source_cols + target_cols + ["remove_row", "duplicate_to_remove"]:
            if col not in work.columns:
                work[col] = ""
            work[col] = work[col].fillna("").astype(str).str.strip()
        valid = work[source_cols + target_cols].apply(lambda col: col.map(_clean).ne("")).all(axis=1)
        work = work.loc[valid].copy()
        if work.empty:
            return pd.DataFrame(
                columns=[
                    *source_cols,
                    *target_cols,
                    f"{sheet_name}_active_row_count",
                    f"{sheet_name}_removed_row_count",
                    f"{sheet_name}_duplicate_removed_row_count",
                    f"{sheet_name}_presence_state",
                ]
            )
        work[f"{sheet_name}_is_active"] = ~work["remove_row"].map(_truthy) & ~work["duplicate_to_remove"].map(_truthy)
        work[f"{sheet_name}_is_removed"] = work["remove_row"].map(_truthy)
        work[f"{sheet_name}_is_duplicate_removed"] = work["duplicate_to_remove"].map(_truthy)
        grouped = (
            work.groupby(source_cols, as_index=False)
            .agg(
                **{
                    f"{sheet_name}_active_row_count": (f"{sheet_name}_is_active", "sum"),
                    f"{sheet_name}_removed_row_count": (f"{sheet_name}_is_removed", "sum"),
                    f"{sheet_name}_duplicate_removed_row_count": (f"{sheet_name}_is_duplicate_removed", "sum"),
                    **{col: (col, _first_non_empty) for col in target_cols},
                }
            )
            .reset_index(drop=True)
        )
        for col in [
            f"{sheet_name}_active_row_count",
            f"{sheet_name}_removed_row_count",
            f"{sheet_name}_duplicate_removed_row_count",
        ]:
            grouped[col] = pd.to_numeric(grouped[col], errors="coerce").fillna(0).astype(int)
        grouped[f"{sheet_name}_presence_state"] = grouped.apply(
            lambda row: "active"
            if row[f"{sheet_name}_active_row_count"] > 0
            else "removed_only"
            if row[f"{sheet_name}_removed_row_count"] > 0
            else "duplicate_removed_only"
            if row[f"{sheet_name}_duplicate_removed_row_count"] > 0
            else "missing",
            axis=1,
        )
        return grouped

    def _comparison_status(sheet_name: str, this_row_status: str, counterpart_presence_state: str) -> str:
        if this_row_status == "active" and counterpart_presence_state == "active":
            return "both_active"
        if sheet_name == "esto":
            if this_row_status == "active" and counterpart_presence_state in {"removed_only", "duplicate_removed_only"}:
                return "esto_active_ninth_removed"
            if this_row_status in {"removed_row_true", "duplicate_removed_row_true"} and counterpart_presence_state == "active":
                return "esto_removed_ninth_active"
            if this_row_status == "active" and counterpart_presence_state == "missing":
                return "esto_active_ninth_missing"
            if this_row_status in {"removed_row_true", "duplicate_removed_row_true"} and counterpart_presence_state == "missing":
                return "esto_removed_ninth_missing"
        if sheet_name == "ninth":
            if this_row_status == "active" and counterpart_presence_state in {"removed_only", "duplicate_removed_only"}:
                return "ninth_active_esto_removed"
            if this_row_status in {"removed_row_true", "duplicate_removed_row_true"} and counterpart_presence_state == "active":
                return "ninth_removed_esto_active"
            if this_row_status == "active" and counterpart_presence_state == "missing":
                return "ninth_active_esto_missing"
            if this_row_status in {"removed_row_true", "duplicate_removed_row_true"} and counterpart_presence_state == "missing":
                return "ninth_removed_esto_missing"
        if this_row_status in {"removed_row_true", "duplicate_removed_row_true"} and counterpart_presence_state in {"removed_only", "duplicate_removed_only"}:
            return "both_removed"
        if counterpart_presence_state == "missing":
            return "actually_missing"
        return "mixed"

    def _issue_side(comparison_status: str) -> str:
        if comparison_status == "both_active":
            return "both_active"
        if comparison_status == "both_removed":
            return "both_removed"
        if comparison_status in {"esto_removed_ninth_active", "esto_removed_ninth_missing"}:
            return "esto_removed"
        if comparison_status in {"ninth_removed_esto_active", "ninth_removed_esto_missing"}:
            return "ninth_removed"
        if comparison_status in {"esto_active_ninth_removed"}:
            return "ninth_removed"
        if comparison_status in {"ninth_active_esto_removed"}:
            return "esto_removed"
        if comparison_status in {"esto_active_ninth_missing"}:
            return "ninth_missing"
        if comparison_status in {"ninth_active_esto_missing"}:
            return "esto_missing"
        if comparison_status == "actually_missing":
            return "missing"
        return comparison_status

    esto_rows = _sheet_row_status(esto_sheet, "esto", ["esto_flow", "esto_product"])
    ninth_rows = _sheet_row_status(ninth_sheet, "ninth", ["ninth_sector", "ninth_fuel"])
    esto_summary = _sheet_source_summary(esto_sheet, "esto", ["esto_flow", "esto_product"])
    ninth_summary = _sheet_source_summary(ninth_sheet, "ninth", ["ninth_sector", "ninth_fuel"])

    esto_rows = esto_rows.merge(
        ninth_summary[source_cols + ["ninth_sector", "ninth_fuel", "ninth_presence_state"]],
        on=source_cols,
        how="left",
    )
    ninth_rows = ninth_rows.merge(
        esto_summary[source_cols + ["esto_flow", "esto_product", "esto_presence_state"]],
        on=source_cols,
        how="left",
    )

    esto_rows["counterpart_presence_state"] = esto_rows["ninth_presence_state"].fillna("missing")
    ninth_rows["counterpart_presence_state"] = ninth_rows["esto_presence_state"].fillna("missing")

    esto_rows["presence_status"] = esto_rows.apply(
        lambda row: _comparison_status("esto", row["this_row_status"], row["counterpart_presence_state"]),
        axis=1,
    )
    ninth_rows["presence_status"] = ninth_rows.apply(
        lambda row: _comparison_status("ninth", row["this_row_status"], row["counterpart_presence_state"]),
        axis=1,
    )

    for work in [esto_rows, ninth_rows]:
        work["comparison_status"] = work["presence_status"]
        work["row_status"] = work["this_row_status"]
        work["issue_side"] = work["comparison_status"].map(_issue_side)
        work["missing_reason"] = work["comparison_status"].map(lambda value: "" if value == "both_active" else value)
        work["has_removed_row"] = work["this_row_is_removed"]
        work["has_duplicate_removed_row"] = work["this_row_is_duplicate_removed"]

    combined = pd.concat([esto_rows, ninth_rows], ignore_index=True)
    combined["issue_side"] = combined.apply(
        lambda row: _issue_side(str(row.get("comparison_status", ""))),
        axis=1,
    )
    combined["is_issue_row"] = combined["comparison_status"].ne("both_active")

    return combined.sort_values(
        ["is_issue_row", "sheet_name", *source_cols, "mapping_row_number"],
        ascending=[False, True, True, True, True, True],
    ).reset_index(drop=True)


def _active_mapping_rows(frame: pd.DataFrame, target_cols: Sequence[str]) -> pd.DataFrame:
    """Return active rows with nonblank LEAP source and target columns."""
    source_cols = ["leap_sector_name_original", "leap_sector_name_full_path", "raw_leap_fuel_name"]
    work = frame.copy().fillna("")
    for col in [*source_cols, *target_cols, "remove_row", "duplicate_to_remove", "pair_mapping_cardinality"]:
        if col not in work.columns:
            work[col] = ""
        work[col] = work[col].fillna("").astype(str).str.strip()
    work["mapping_row_number"] = work.index + 2
    active = work[_active_mask(work)].copy()
    valid = active[["leap_sector_name_full_path", "raw_leap_fuel_name", *target_cols]].apply(
        lambda col: col.map(_clean).ne("")
    ).all(axis=1)
    return active.loc[valid].copy()


def _build_many_to_many_conflicts(esto_sheet: pd.DataFrame, ninth_sheet: pd.DataFrame) -> pd.DataFrame:
    """Return active rows whose pair cardinality is many_to_many."""
    records: list[pd.DataFrame] = []
    sheet_specs = [
        (ESTO_SHEET, esto_sheet, ["esto_flow", "esto_product"]),
        (NINTH_SHEET, ninth_sheet, ["ninth_sector", "ninth_fuel"]),
    ]
    source_cols = ["leap_sector_name_full_path", "raw_leap_fuel_name"]
    for sheet_name, frame, target_cols in sheet_specs:
        work = _active_mapping_rows(frame, target_cols)
        if "pair_mapping_cardinality" not in work.columns:
            work = _compute_pair_cardinality(work, target_cols[0], target_cols[1])
        work = work[work["pair_mapping_cardinality"].fillna("").astype(str).str.strip().eq("many_to_many")].copy()
        if work.empty:
            continue
        work.insert(0, "conflict_type", "many_to_many_mapping")
        work.insert(1, "sheet_name", sheet_name)
        keep_cols = [
            "conflict_type",
            "sheet_name",
            "mapping_row_number",
            *source_cols,
            *target_cols,
            "pair_mapping_cardinality",
        ]
        records.append(work.loc[:, keep_cols])
    if not records:
        return pd.DataFrame(
            columns=[
                "conflict_type",
                "sheet_name",
                "mapping_row_number",
                "leap_sector_name_full_path",
                "raw_leap_fuel_name",
                "esto_flow",
                "esto_product",
                "ninth_sector",
                "ninth_fuel",
                "pair_mapping_cardinality",
            ]
        )
    return pd.concat(records, ignore_index=True).fillna("")


def _build_missing_between_sheet_conflicts(trio_presence: pd.DataFrame) -> pd.DataFrame:
    """Return active rows that are missing or removed in the counterpart sheet."""
    if trio_presence.empty:
        return pd.DataFrame()
    conflict_statuses = {
        "esto_active_ninth_missing",
        "ninth_active_esto_missing",
    }
    out = trio_presence[trio_presence["presence_status"].isin(conflict_statuses)].copy()
    if out.empty:
        return pd.DataFrame(
            columns=[
                "conflict_type",
                "sheet_name",
                "mapping_row_number",
                "presence_status",
                "leap_sector_name_full_path",
                "raw_leap_fuel_name",
                "esto_flow",
                "esto_product",
                "ninth_sector",
                "ninth_fuel",
                "counterpart_presence_state",
            ]
        )
    out.insert(0, "conflict_type", "active_mapping_missing_from_counterpart")
    keep_cols = [
        "conflict_type",
        "sheet_name",
        "mapping_row_number",
        "presence_status",
        "leap_sector_name_full_path",
        "raw_leap_fuel_name",
        "esto_flow",
        "esto_product",
        "ninth_sector",
        "ninth_fuel",
        "counterpart_presence_state",
    ]
    return out.loc[:, keep_cols].fillna("").reset_index(drop=True)


def _build_crosswalk_target_conflicts(
    esto_sheet: pd.DataFrame,
    ninth_sheet: pd.DataFrame,
    ninth_to_esto_pairs: pd.DataFrame,
) -> pd.DataFrame:
    """
    Compare active LEAP mapping targets against master 9th -> ESTO pair mappings.

    A row is reported when an active 9th mapping for a LEAP source has no matching
    active ESTO target implied by ninth_pairs_to_esto_pairs. The conflict type is
    split by cardinality so strict one-to-one mismatches are separated from
    one-to-many / many-to-one rows that need review.
    """
    source_cols = ["leap_sector_name_full_path", "raw_leap_fuel_name"]
    esto_sheet = _compute_pair_cardinality(esto_sheet, "esto_flow", "esto_product")
    ninth_sheet = _compute_pair_cardinality(ninth_sheet, "ninth_sector", "ninth_fuel")
    esto_active = _active_mapping_rows(esto_sheet, ["esto_flow", "esto_product"])
    ninth_active = _active_mapping_rows(ninth_sheet, ["ninth_sector", "ninth_fuel"])

    if esto_active.empty or ninth_active.empty:
        return pd.DataFrame(
            columns=[
                "conflict_type",
                "leap_sector_name_full_path",
                "raw_leap_fuel_name",
                "ninth_sector",
                "ninth_fuel",
                "implied_esto_targets",
                "active_esto_targets",
                "esto_cardinalities",
                "ninth_cardinality",
                "ninth_mapping_row_number",
            ]
        )

    pairs = _filter_researcher_rows(ninth_to_esto_pairs).copy().fillna("")
    for col in ["9th_sector", "9th_fuel", "esto_flow", "esto_product"]:
        if col not in pairs.columns:
            pairs[col] = ""
        pairs[col] = pairs[col].fillna("").astype(str).str.strip()
    pairs = pairs[pairs[["9th_sector", "9th_fuel", "esto_flow", "esto_product"]].apply(lambda col: col.map(_clean).ne("")).all(axis=1)].copy()
    pairs = pairs.drop_duplicates(subset=["9th_sector", "9th_fuel", "esto_flow", "esto_product"])

    esto_targets = (
        esto_active.groupby(source_cols, as_index=False)
        .agg(
            active_esto_targets=(
                "esto_flow",
                lambda series: " | ".join(
                    sorted(
                        {
                            f"{flow} || {product}"
                            for flow, product in zip(
                                series.astype(str),
                                esto_active.loc[series.index, "esto_product"].astype(str),
                            )
                            if _clean(flow) and _clean(product)
                        }
                    )
                ),
            ),
            esto_cardinalities=(
                "pair_mapping_cardinality",
                lambda series: " | ".join(
                    sorted({str(value).strip() for value in series.tolist() if _clean(value)})
                ),
            ),
        )
    )
    active_esto_pairs = set(
        zip(
            esto_active["leap_sector_name_full_path"].astype(str),
            esto_active["raw_leap_fuel_name"].astype(str),
            esto_active["esto_flow"].astype(str),
            esto_active["esto_product"].astype(str),
        )
    )

    merged = ninth_active.merge(
        pairs,
        left_on=["ninth_sector", "ninth_fuel"],
        right_on=["9th_sector", "9th_fuel"],
        how="left",
    )
    merged["_ninth_row_key"] = merged["mapping_row_number"].astype(str)
    merged["has_master_pair"] = (
        merged["esto_flow"].fillna("").astype(str).str.strip().ne("")
        & merged["esto_product"].fillna("").astype(str).str.strip().ne("")
    )
    merged["has_matching_esto_target"] = merged.apply(
        lambda row: (
            row["leap_sector_name_full_path"],
            row["raw_leap_fuel_name"],
            row["esto_flow"],
            row["esto_product"],
        )
        in active_esto_pairs,
        axis=1,
    )

    def _joined_targets(group: pd.DataFrame) -> str:
        values = sorted(
            {
                f"{flow} || {product}"
                for flow, product in zip(group["esto_flow"].astype(str), group["esto_product"].astype(str))
                if _clean(flow) and _clean(product)
            }
        )
        return " | ".join(values)

    grouped = (
        merged.groupby("_ninth_row_key", as_index=False)
        .agg(
            leap_sector_name_full_path=("leap_sector_name_full_path", "first"),
            raw_leap_fuel_name=("raw_leap_fuel_name", "first"),
            ninth_sector=("ninth_sector", "first"),
            ninth_fuel=("ninth_fuel", "first"),
            ninth_cardinality=("pair_mapping_cardinality", "first"),
            ninth_mapping_row_number=("mapping_row_number", "first"),
            implied_esto_targets=("esto_flow", lambda series: _joined_targets(merged.loc[series.index])),
            has_master_pair=("has_master_pair", "max"),
            has_matching_esto_target=("has_matching_esto_target", "max"),
        )
        .reset_index(drop=True)
    )
    grouped = grouped.merge(esto_targets, on=source_cols, how="left")
    has_active_esto_target = grouped["active_esto_targets"].fillna("").astype(str).str.strip().ne("")
    conflicts = grouped[
        has_active_esto_target
        & (~grouped["has_master_pair"] | ~grouped["has_matching_esto_target"])
    ].copy()
    if conflicts.empty:
        return pd.DataFrame(
            columns=[
                "conflict_type",
                "leap_sector_name_full_path",
                "raw_leap_fuel_name",
                "ninth_sector",
                "ninth_fuel",
                "implied_esto_targets",
                "active_esto_targets",
                "esto_cardinalities",
                "ninth_cardinality",
                "ninth_mapping_row_number",
            ]
        )

    def _split_cardinalities(value: object) -> set[str]:
        return {part.strip() for part in str(value or "").split("|") if part.strip()}

    def _target_conflict_type(row: pd.Series) -> str:
        if not bool(row["has_master_pair"]):
            return "ninth_pair_missing_from_master_crosswalk"
        esto_cardinalities = _split_cardinalities(row.get("esto_cardinalities", ""))
        ninth_cardinality = _clean(row.get("ninth_cardinality", ""))
        if esto_cardinalities == {"one_to_one"} and ninth_cardinality == "one_to_one":
            return "strict_one_to_one_target_mismatch"
        if "many_to_many" in esto_cardinalities or ninth_cardinality == "many_to_many":
            return "many_to_many_target_review"
        return "non_strict_cardinality_target_review"

    conflicts["conflict_type"] = conflicts.apply(
        _target_conflict_type,
        axis=1,
    )
    keep_cols = [
        "conflict_type",
        "leap_sector_name_full_path",
        "raw_leap_fuel_name",
        "ninth_sector",
        "ninth_fuel",
        "implied_esto_targets",
        "active_esto_targets",
        "esto_cardinalities",
        "ninth_cardinality",
        "ninth_mapping_row_number",
    ]
    return conflicts.loc[:, keep_cols].fillna("").drop_duplicates().reset_index(drop=True)


def _active_ninth_to_esto_pairs(ninth_to_esto_pairs: pd.DataFrame) -> pd.DataFrame:
    """Return active/non-faulty rows from master 9th -> ESTO pair mapping."""
    pairs = _filter_researcher_rows(ninth_to_esto_pairs).copy().fillna("")
    for col in ["9th_sector", "9th_fuel", "esto_flow", "esto_product"]:
        if col not in pairs.columns:
            pairs[col] = ""
        pairs[col] = pairs[col].fillna("").astype(str).str.strip()
    pairs = pairs[
        pairs[["9th_sector", "9th_fuel", "esto_flow", "esto_product"]]
        .apply(lambda col: col.map(_clean).ne(""))
        .all(axis=1)
    ].copy()
    return pairs.drop_duplicates(subset=["9th_sector", "9th_fuel", "esto_flow", "esto_product"])


def _build_implied_missing_crosswalk_pairs(
    esto_sheet: pd.DataFrame,
    ninth_sheet: pd.DataFrame,
    ninth_to_esto_pairs: pd.DataFrame,
) -> pd.DataFrame:
    """
    Build candidate ninth_pairs_to_esto_pairs rows implied by active LEAP mappings.

    The candidates are inferred by joining active leap_combined_ninth and
    leap_combined_esto rows on the same LEAP sector/fuel source pair, then
    removing exact pairs already present in ninth_pairs_to_esto_pairs.
    """
    columns = [
        "candidate_status",
        "would_create_many_to_many",
        "candidate_crosswalk_cardinality",
        "9th_sector",
        "9th_fuel",
        "esto_flow",
        "esto_product",
        "leap_sector_name_full_path",
        "raw_leap_fuel_name",
        "ninth_cardinality",
        "esto_cardinality",
        "ninth_mapping_row_number",
        "esto_mapping_row_number",
    ]
    source_cols = ["leap_sector_name_full_path", "raw_leap_fuel_name"]
    esto_sheet = _compute_pair_cardinality(esto_sheet, "esto_flow", "esto_product")
    ninth_sheet = _compute_pair_cardinality(ninth_sheet, "ninth_sector", "ninth_fuel")
    esto_active = _active_mapping_rows(esto_sheet, ["esto_flow", "esto_product"])
    ninth_active = _active_mapping_rows(ninth_sheet, ["ninth_sector", "ninth_fuel"])
    if esto_active.empty or ninth_active.empty:
        return pd.DataFrame(columns=columns)

    esto_keep = [
        *source_cols,
        "esto_flow",
        "esto_product",
        "pair_mapping_cardinality",
        "mapping_row_number",
    ]
    ninth_keep = [
        *source_cols,
        "ninth_sector",
        "ninth_fuel",
        "pair_mapping_cardinality",
        "mapping_row_number",
    ]
    implied = ninth_active[ninth_keep].merge(
        esto_active[esto_keep],
        on=source_cols,
        how="inner",
        suffixes=("_ninth", "_esto"),
    )
    if implied.empty:
        return pd.DataFrame(columns=columns)

    implied = implied.rename(
        columns={
            "pair_mapping_cardinality_ninth": "ninth_cardinality",
            "pair_mapping_cardinality_esto": "esto_cardinality",
            "mapping_row_number_ninth": "ninth_mapping_row_number",
            "mapping_row_number_esto": "esto_mapping_row_number",
        }
    )
    master_pairs = _active_ninth_to_esto_pairs(ninth_to_esto_pairs)
    existing_keys = set(
        zip(
            master_pairs["9th_sector"].astype(str),
            master_pairs["9th_fuel"].astype(str),
            master_pairs["esto_flow"].astype(str),
            master_pairs["esto_product"].astype(str),
        )
    )
    implied["_pair_key"] = list(
        zip(
            implied["ninth_sector"].astype(str),
            implied["ninth_fuel"].astype(str),
            implied["esto_flow"].astype(str),
            implied["esto_product"].astype(str),
        )
    )
    implied = implied[~implied["_pair_key"].isin(existing_keys)].copy()
    if implied.empty:
        return pd.DataFrame(columns=columns)

    combined_pairs = pd.concat(
        [
            master_pairs.rename(columns={"9th_sector": "ninth_sector", "9th_fuel": "ninth_fuel"})[
                ["ninth_sector", "ninth_fuel", "esto_flow", "esto_product"]
            ],
            implied[["ninth_sector", "ninth_fuel", "esto_flow", "esto_product"]],
        ],
        ignore_index=True,
    ).drop_duplicates()
    combined_pairs["_source_key"] = combined_pairs["ninth_sector"].astype(str) + "|||" + combined_pairs["ninth_fuel"].astype(str)
    combined_pairs["_target_key"] = combined_pairs["esto_flow"].astype(str) + "|||" + combined_pairs["esto_product"].astype(str)
    source_count = combined_pairs.groupby("_source_key")["_target_key"].nunique()
    target_count = combined_pairs.groupby("_target_key")["_source_key"].nunique()

    implied["_source_key"] = implied["ninth_sector"].astype(str) + "|||" + implied["ninth_fuel"].astype(str)
    implied["_target_key"] = implied["esto_flow"].astype(str) + "|||" + implied["esto_product"].astype(str)
    implied["candidate_crosswalk_cardinality"] = implied.apply(
        lambda row: _mapping_cardinality(
            int(source_count.get(row["_source_key"], 0)),
            int(target_count.get(row["_target_key"], 0)),
        ),
        axis=1,
    )
    implied["would_create_many_to_many"] = implied["candidate_crosswalk_cardinality"].eq("many_to_many")
    implied["candidate_status"] = implied["would_create_many_to_many"].map(
        lambda value: "review_many_to_many_before_adding" if bool(value) else "candidate_to_add"
    )
    implied = implied.rename(columns={"ninth_sector": "9th_sector", "ninth_fuel": "9th_fuel"})
    return (
        implied.loc[:, columns]
        .fillna("")
        .drop_duplicates()
        .sort_values(["would_create_many_to_many", "9th_sector", "9th_fuel", "esto_flow", "esto_product"], ascending=[False, True, True, True, True])
        .reset_index(drop=True)
    )


def build_mapping_conflict_report(
    esto_sheet: pd.DataFrame,
    ninth_sheet: pd.DataFrame,
    ninth_to_esto_pairs: pd.DataFrame,
) -> dict[str, pd.DataFrame]:
    """Build conflict-check sheets for mapping maintenance."""
    trio_presence = _build_trio_presence_check(esto_sheet, ninth_sheet)
    many_to_many = _build_many_to_many_conflicts(esto_sheet, ninth_sheet)
    missing_between_sheets = _build_missing_between_sheet_conflicts(trio_presence)
    crosswalk_target_conflicts = _build_crosswalk_target_conflicts(
        esto_sheet,
        ninth_sheet,
        ninth_to_esto_pairs,
    )
    implied_missing_crosswalk = _build_implied_missing_crosswalk_pairs(
        esto_sheet,
        ninth_sheet,
        ninth_to_esto_pairs,
    )
    summary_records = [
        {"check_name": "many_to_many", "row_count": int(len(many_to_many))},
        {"check_name": "missing_between_sheets", "row_count": int(len(missing_between_sheets))},
        {"check_name": "crosswalk_target_conflicts", "row_count": int(len(crosswalk_target_conflicts))},
        {"check_name": "implied_missing_crosswalk", "row_count": int(len(implied_missing_crosswalk))},
    ]
    return {
        "summary": pd.DataFrame(summary_records),
        "many_to_many": many_to_many,
        "missing_between_sheets": missing_between_sheets,
        "crosswalk_target_conflicts": crosswalk_target_conflicts,
        "implied_missing_crosswalk": implied_missing_crosswalk,
    }


def _write_mapping_conflict_report(
    report_sheets: dict[str, pd.DataFrame],
    output_path: Path = MAPPING_CONFLICTS_WORKBOOK_PATH,
) -> None:
    """Write the mapping conflict report workbook."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        _assert_not_open(output_path)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, sheet_df in report_sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)


def _write_excel_report(output_path: Path, sheets: dict[str, pd.DataFrame]) -> None:
    """Write an Excel workbook with basic filter/freeze formatting."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        _assert_not_open(output_path)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, sheet_df in sheets.items():
            safe_name = sheet_name[:31]
            sheet_df.to_excel(writer, sheet_name=safe_name, index=False)
            worksheet = writer.book[safe_name]
            worksheet.freeze_panes = "A2"
            if worksheet.max_row >= 1 and worksheet.max_column >= 1:
                worksheet.auto_filter.ref = worksheet.dimensions
            for column_cells in worksheet.columns:
                header = str(column_cells[0].value or "")
                width = min(max(len(header) + 2, 12), 42)
                worksheet.column_dimensions[column_cells[0].column_letter].width = width


def _build_mapping_refresh_report_sheets(
    *,
    gaps: pd.DataFrame,
    duplicate_mappings: pd.DataFrame,
    trio_presence_issues: pd.DataFrame,
    mapping_conflict_report: dict[str, pd.DataFrame],
) -> dict[str, pd.DataFrame]:
    """Build the one-workbook mapping refresh report sheets."""
    conflict_summary = mapping_conflict_report.get("summary", pd.DataFrame())
    report_rows = [
        {
            "section": "coverage_gaps",
            "row_count": int(len(gaps)),
            "sheet_name": "coverage_gaps",
            "description": "Source or target pairs missing from active mapping rows.",
        },
        {
            "section": "duplicate_mappings",
            "row_count": int(len(duplicate_mappings)),
            "sheet_name": "duplicate_mappings",
            "description": "Exact duplicate active mapping rows.",
        },
        {
            "section": "trio_presence_issues",
            "row_count": int(len(trio_presence_issues)),
            "sheet_name": "trio_presence_issues",
            "description": "Active/removed/missing mismatches between leap_combined_esto and leap_combined_ninth.",
        },
    ]
    for row in conflict_summary.to_dict("records"):
        report_rows.append(
            {
                "section": str(row.get("check_name", "")),
                "row_count": int(row.get("row_count", 0) or 0),
                "sheet_name": str(row.get("check_name", ""))[:31],
                "description": "Mapping conflict check.",
            }
        )

    readme = pd.DataFrame(
        [
            {
                "item": "start_here",
                "detail": "Open the summary sheet first, then inspect only sheets with nonzero row_count.",
            },
            {
                "item": "most_actionable_sheets",
                "detail": "coverage_gaps, trio_presence_issues, strict one-to-one rows in crosswalk_target_conflicts.",
            },
            {
                "item": "many_to_many",
                "detail": "These rows are not always wrong, but they need explicit review because one LEAP pair and one target pair both map multiple ways.",
            },
            {
                "item": "non_strict_cardinality_target_review",
                "detail": "These target differences involve one-to-many or many-to-one cardinality, so they are review items rather than strict conflicts.",
            },
            {
                "item": "implied_missing_crosswalk",
                "detail": "Candidate ninth_pairs_to_esto_pairs rows inferred from active LEAP combined mappings. Review rows marked would_create_many_to_many before adding.",
            },
            {
                "item": "trio_presence_note",
                "detail": TRIO_PRESENCE_OUTPUT_NOTE,
            },
        ]
    )
    summary = pd.DataFrame(report_rows)
    output_inventory = pd.DataFrame(
        [
            {
                "output_type": "primary_excel_report",
                "path": str(MAPPING_REFRESH_REPORT_PATH),
                "note": "One workbook with all current mapping refresh checks.",
            },
            {
                "output_type": "coverage_gaps_csv",
                "path": str(COVERAGE_GAPS_PATH),
                "note": "Kept for compatibility with existing workflows.",
            },
            {
                "output_type": "duplicate_mappings_csv",
                "path": str(DUPLICATE_MAPPINGS_CSV_PATH),
                "note": "Kept for compatibility with existing workflows.",
            },
            {
                "output_type": "trio_presence_csv",
                "path": str(TRIO_PRESENCE_CSV_PATH),
                "note": "Kept for compatibility with existing workflows.",
            },
        ]
    )

    sheets: dict[str, pd.DataFrame] = {
        "README": readme,
        "summary": summary,
        "output_inventory": output_inventory,
        "coverage_gaps": gaps,
        "duplicate_mappings": duplicate_mappings,
        "trio_presence_issues": trio_presence_issues,
    }
    for sheet_name, sheet_df in mapping_conflict_report.items():
        if sheet_name == "summary":
            continue
        sheets[sheet_name] = sheet_df
    return sheets


def _write_mapping_refresh_report(
    *,
    gaps: pd.DataFrame,
    duplicate_mappings: pd.DataFrame,
    trio_presence_issues: pd.DataFrame,
    mapping_conflict_report: dict[str, pd.DataFrame],
    output_path: Path = MAPPING_REFRESH_REPORT_PATH,
) -> None:
    """Write the primary one-workbook report for mapping refresh checks."""
    sheets = _build_mapping_refresh_report_sheets(
        gaps=gaps,
        duplicate_mappings=duplicate_mappings,
        trio_presence_issues=trio_presence_issues,
        mapping_conflict_report=mapping_conflict_report,
    )
    _write_excel_report(output_path, sheets)


def _resolve_balance_workbook_for_mapping_check(*, scenario: str, date_id: str | None) -> Path:
    return resolve_balance_export_workbook(
        economy=BALANCE_EXPORT_ECONOMY,
        scenario=scenario,
        date_id=date_id,
    )


def _extract_raw_balance_workbook(workbook_path: Path) -> pd.DataFrame:
    """
    Extract raw nonzero LEAP balance rows without loading or applying mappings.

    The mapping workbook may be incomplete or temporarily invalid while this
    maintenance workflow is being used to find gaps, so this raw extraction must
    not depend on mapping rows being valid.
    """
    chosen_template = _pick_template_sheet(workbook_path, BALANCE_TEMPLATE_SHEET)
    extractor = TemplateBalanceExtractor(
        template_sheet=chosen_template,
        mapping_pairs_path=MAPPING_WORKBOOK_PATH,
        codebook_path=CODEBOOK_PATH,
        reinterpret_fuel_rows_as_parent_sector=False,
        explicit_pair_mappings_only=True,
    )
    workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=False)
    if chosen_template not in workbook.sheetnames:
        raise ValueError(f"Template sheet {chosen_template!r} not found in workbook: {workbook_path}")

    template_layout = extractor._extract_layout(workbook[chosen_template])
    selected_sheets = _list_balance_sheets(workbook_path)
    if not selected_sheets:
        raise ValueError(f"No balance sheets found in workbook: {workbook_path}")

    frames: list[pd.DataFrame] = []
    for sheet_name in selected_sheets:
        worksheet = workbook[sheet_name]
        meta = extractor._extract_metadata(worksheet)
        try:
            sheet_layout = extractor._extract_layout(worksheet)
        except ValueError:
            sheet_layout = template_layout
        extracted = extractor._extract_sheet_matrix(worksheet, template=sheet_layout)
        if extracted.empty:
            continue
        extracted.insert(0, "source_sheet", sheet_name)
        extracted.insert(1, "source_workbook", str(workbook_path))
        extracted["area"] = str(meta.get("area", ""))
        extracted["scenario"] = str(meta.get("scenario", ""))
        extracted["year"] = meta.get("year")
        extracted["units"] = str(meta.get("units", ""))
        factor, parse_status, prefix_label, base_label = _parse_unit_factor_to_petajoule(
            str(meta.get("units", ""))
        )
        extracted["value_original"] = extracted["value"]
        extracted["units_original"] = extracted["units"]
        extracted["unit_to_petajoule_factor"] = factor
        extracted["unit_parse_status"] = parse_status
        extracted["unit_prefix"] = prefix_label
        extracted["unit_base"] = base_label
        if factor is not None:
            extracted["value_petajoule"] = pd.to_numeric(extracted["value"], errors="coerce") * float(factor)
        else:
            extracted["value_petajoule"] = pd.NA
        extracted["units_petajoule"] = "Petajoule"
        extracted = extracted[
            pd.to_numeric(extracted.get("value_petajoule", pd.Series(index=extracted.index)), errors="coerce")
            .fillna(0.0)
            .ne(0.0)
        ].copy()
        if not extracted.empty:
            frames.append(extracted)

    if not frames:
        return pd.DataFrame()
    raw_long = pd.concat(frames, ignore_index=True, sort=False)
    dedupe_cols = [
        "source_sheet",
        "leap_sector_name",
        "leap_sector_name_original",
        "leap_sector_name_full_path",
        "leap_fuel_name",
        "value",
    ]
    dedupe_cols_present = [col for col in dedupe_cols if col in raw_long.columns]
    if dedupe_cols_present:
        raw_long = raw_long.drop_duplicates(subset=dedupe_cols_present).reset_index(drop=True)
    return raw_long


def _load_raw_leap_balance_lookup() -> pd.DataFrame:
    """
    Return nonzero raw LEAP balance sector/fuel pairs from REF and TGT exports.

    This intentionally uses raw extractor output rather than mapped rows, so it
    can find LEAP source pairs that are absent from both mapping sheets.
    """
    workbooks = [
        ("Reference", _resolve_balance_workbook_for_mapping_check(scenario="REF", date_id=REF_BALANCE_EXPORT_DATE_ID)),
        ("Target", _resolve_balance_workbook_for_mapping_check(scenario="TGT", date_id=TGT_BALANCE_EXPORT_DATE_ID)),
    ]
    frames: list[pd.DataFrame] = []
    for scenario_label, workbook_path in workbooks:
        raw = _extract_raw_balance_workbook(workbook_path)
        if raw.empty:
            continue
        raw["scenario"] = raw.get("scenario", "").fillna("").astype(str).str.strip()
        raw.loc[raw["scenario"].eq(""), "scenario"] = scenario_label
        if "raw_leap_fuel_name" not in raw.columns:
            raw["raw_leap_fuel_name"] = raw.get("leap_fuel_name", "")
        for col in ["source_sheet", "leap_sector_name_full_path", "raw_leap_fuel_name"]:
            if col not in raw.columns:
                raw[col] = ""
            raw[col] = raw[col].fillna("").astype(str).str.strip()
        if "year" not in raw.columns:
            raw["year"] = pd.NA
        raw["year"] = pd.to_numeric(raw["year"], errors="coerce")
        raw["value_petajoule"] = pd.to_numeric(
            raw.get("value_petajoule", raw.get("value", pd.NA)),
            errors="coerce",
        ).fillna(0.0)
        raw = raw[
            raw["leap_sector_name_full_path"].ne("")
            & raw["raw_leap_fuel_name"].ne("")
            & raw["value_petajoule"].ne(0)
        ].copy()
        if raw.empty:
            continue
        frames.append(
            raw[
                [
                    "source_sheet",
                    "scenario",
                    "year",
                    "leap_sector_name_full_path",
                    "raw_leap_fuel_name",
                    "value_petajoule",
                ]
            ]
        )

    columns = [
        "leap_sector_name_full_path",
        "raw_leap_fuel_name",
        "leap_pair_is_subtotal",
        "leap_pair_abs_sum",
        "raw_leap_row_count",
        "raw_leap_source_sheet_count",
        "raw_leap_scenarios",
        "raw_leap_year_min",
        "raw_leap_year_max",
    ]
    if not frames:
        return pd.DataFrame(columns=columns)

    combined = pd.concat(frames, ignore_index=True, sort=False)
    combined["_value_abs"] = pd.to_numeric(combined["value_petajoule"], errors="coerce").fillna(0.0).abs()
    combined = combined[combined["_value_abs"].gt(0)].copy()
    if combined.empty:
        return pd.DataFrame(columns=columns)
    subtotal_frame = _compute_leap_subtotals(
        combined[
            ["leap_sector_name_full_path", "raw_leap_fuel_name"]
        ].assign(remove_row=False, duplicate_to_remove=False)
    )
    combined["leap_pair_is_subtotal"] = subtotal_frame["leap_is_subtotal"].fillna(False).astype(bool)

    grouped = (
        combined.groupby(["leap_sector_name_full_path", "raw_leap_fuel_name"], as_index=False)
        .agg(
            leap_pair_is_subtotal=("leap_pair_is_subtotal", "max"),
            leap_pair_abs_sum=("_value_abs", "sum"),
            raw_leap_row_count=("_value_abs", "size"),
            raw_leap_source_sheet_count=("source_sheet", lambda values: int(values.astype(str).str.strip().nunique())),
            raw_leap_scenarios=("scenario", lambda values: "|".join(sorted({str(value).strip() for value in values if str(value).strip()}))),
            raw_leap_year_min=("year", "min"),
            raw_leap_year_max=("year", "max"),
        )
        .reset_index(drop=True)
    )
    return grouped[columns]


def _build_coverage_gaps(
    esto_sheet: pd.DataFrame,
    ninth_sheet: pd.DataFrame,
    esto_lookup: pd.DataFrame,
    ninth_lookup: pd.DataFrame,
    raw_leap_lookup: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """
    Return a DataFrame of all coverage gaps: pairs with abs values > 0 that are
    missing from the active mapping rows.

    Columns: gap_type, sheet_name, original_dataset, original_pair_is_subtotal,
    key_col_1, key_col_2, pair_1, pair_2, abs_sum, mapping_presence_state,
    mapping_presence_detail
      gap_type values:
        "esto_missing"   - esto data pair not in any active esto mapping row
        "ninth_missing"  - 9th data pair not in any active ninth mapping row
        "leap_unmapped_esto"  - LEAP pair with value > 0 but no esto target in esto sheet
        "leap_unmapped_ninth" - LEAP pair with value > 0 but no ninth target in ninth sheet
        "raw_leap_missing_esto_mapping" - raw LEAP export pair not present in active esto mapping rows
        "raw_leap_missing_ninth_mapping" - raw LEAP export pair not present in active ninth mapping rows
    """
    records: list[dict] = []

    def _lookup_subtotal_flag(frame: pd.DataFrame, key_a: str, key_b: str, value_a: str, value_b: str, subtotal_col: str) -> bool:
        if subtotal_col not in frame.columns:
            return False
        mask = frame[key_a].astype(str).str.strip().eq(value_a) & frame[key_b].astype(str).str.strip().eq(value_b)
        if not bool(mask.any()):
            return False
        return bool(frame.loc[mask, subtotal_col].fillna(False).astype(bool).any())

    # --- 1. ESTO data pairs missing from the esto mapping ---
    esto_data_pairs = set(
        zip(
            esto_lookup.loc[esto_lookup["esto_pair_abs_sum"] > 0, "esto_flow"].astype(str).str.strip(),
            esto_lookup.loc[esto_lookup["esto_pair_abs_sum"] > 0, "esto_product"].astype(str).str.strip(),
        )
    )
    esto_mapped_pairs = _active_pairs(esto_sheet, "esto_flow", "esto_product")
    for flow, product in sorted(esto_data_pairs - esto_mapped_pairs):
        abs_sum = float(
            esto_lookup.loc[
                esto_lookup["esto_flow"].astype(str).str.strip().eq(flow)
                & esto_lookup["esto_product"].astype(str).str.strip().eq(product),
                "esto_pair_abs_sum",
            ].sum()
        )
        records.append(
            {
                "gap_type": "esto_missing",
                "sheet_name": ESTO_SHEET,
                "original_dataset": "esto",
                "original_pair_is_subtotal": _lookup_subtotal_flag(
                    esto_lookup,
                    "esto_flow",
                    "esto_product",
                    flow,
                    product,
                    "esto_pair_is_subtotal",
                ),
                "key_col_1": "esto_flow",
                "key_col_2": "esto_product",
                "pair_1": flow,
                "pair_2": product,
                "abs_sum": abs_sum,
                "mapping_presence_state": "target_pair_missing_from_active_mappings",
                "mapping_presence_detail": "",
            }
        )

    # --- 2. 9th data pairs missing from the ninth mapping ---
    ninth_data_pairs = set(
        zip(
            ninth_lookup.loc[ninth_lookup["ninth_pair_abs_sum"] > 0, "ninth_sector"].astype(str).str.strip(),
            ninth_lookup.loc[ninth_lookup["ninth_pair_abs_sum"] > 0, "ninth_fuel"].astype(str).str.strip(),
        )
    )
    ninth_mapped_pairs = _active_pairs(ninth_sheet, "ninth_sector", "ninth_fuel")
    for sector, fuel in sorted(ninth_data_pairs - ninth_mapped_pairs):
        abs_sum = float(
            ninth_lookup.loc[
                ninth_lookup["ninth_sector"].astype(str).str.strip().eq(sector)
                & ninth_lookup["ninth_fuel"].astype(str).str.strip().eq(fuel),
                "ninth_pair_abs_sum",
            ].sum()
        )
        records.append(
            {
                "gap_type": "ninth_missing",
                "sheet_name": NINTH_SHEET,
                "original_dataset": "ninth",
                "original_pair_is_subtotal": _lookup_subtotal_flag(
                    ninth_lookup,
                    "ninth_sector",
                    "ninth_fuel",
                    sector,
                    fuel,
                    "ninth_pair_is_subtotal",
                ),
                "key_col_1": "ninth_sector",
                "key_col_2": "ninth_fuel",
                "pair_1": sector,
                "pair_2": fuel,
                "abs_sum": abs_sum,
                "mapping_presence_state": "target_pair_missing_from_active_mappings",
                "mapping_presence_detail": "",
            }
        )

    # --- 3. LEAP pairs with abs(value) > 0 that are unmapped ---
    for sheet, gap_type, sheet_name, target_a, target_b in [
        (esto_sheet, "leap_unmapped_esto", ESTO_SHEET, "esto_flow", "esto_product"),
        (ninth_sheet, "leap_unmapped_ninth", NINTH_SHEET, "ninth_sector", "ninth_fuel"),
    ]:
        active = _compute_leap_subtotals(sheet)[_active_mask(sheet)].copy()
        if "value" not in active.columns:
            continue
        values = pd.to_numeric(active["value"], errors="coerce").fillna(0.0).abs()
        leap_sector = active["leap_sector_name_full_path"].fillna("").astype(str).str.strip() if "leap_sector_name_full_path" in active.columns else pd.Series("", index=active.index)
        leap_fuel = active["raw_leap_fuel_name"].fillna("").astype(str).str.strip() if "raw_leap_fuel_name" in active.columns else pd.Series("", index=active.index)
        leap_is_subtotal = active["leap_is_subtotal"].fillna(False).astype(bool) if "leap_is_subtotal" in active.columns else pd.Series(False, index=active.index)
        ta = active[target_a].fillna("").astype(str).str.strip() if target_a in active.columns else pd.Series("", index=active.index)
        tb = active[target_b].fillna("").astype(str).str.strip() if target_b in active.columns else pd.Series("", index=active.index)
        unmapped_mask = (values > 0) & (ta.eq("") | tb.eq(""))
        for sector, fuel in sorted(set(zip(leap_sector[unmapped_mask], leap_fuel[unmapped_mask]))):
            abs_sum = float(values[unmapped_mask & leap_sector.eq(sector) & leap_fuel.eq(fuel)].sum())
            is_subtotal = bool(leap_is_subtotal[unmapped_mask & leap_sector.eq(sector) & leap_fuel.eq(fuel)].any())
            records.append(
                {
                    "gap_type": gap_type,
                    "sheet_name": sheet_name,
                    "original_dataset": "leap",
                    "original_pair_is_subtotal": is_subtotal,
                    "key_col_1": "leap_sector_name_full_path",
                    "key_col_2": "raw_leap_fuel_name",
                    "pair_1": sector,
                    "pair_2": fuel,
                    "abs_sum": abs_sum,
                    "mapping_presence_state": "active_source_row_missing_target",
                    "mapping_presence_detail": "",
                }
            )

    # --- 4. Raw LEAP export pairs absent from the mapping sheets entirely ---
    if raw_leap_lookup is not None and not raw_leap_lookup.empty:
        raw = raw_leap_lookup.copy()
        for col in ["leap_sector_name_full_path", "raw_leap_fuel_name"]:
            if col not in raw.columns:
                raw[col] = ""
            raw[col] = raw[col].fillna("").astype(str).str.strip()
        if "leap_pair_abs_sum" not in raw.columns:
            raw["leap_pair_abs_sum"] = 0.0
        raw["leap_pair_abs_sum"] = pd.to_numeric(raw["leap_pair_abs_sum"], errors="coerce").fillna(0.0)
        raw = raw[
            raw["leap_sector_name_full_path"].ne("")
            & raw["raw_leap_fuel_name"].ne("")
            & raw["leap_pair_abs_sum"].gt(0)
        ].copy()

        active_source_pairs = {
            "raw_leap_missing_esto_mapping": (
                ESTO_SHEET,
                _active_leap_source_pairs(esto_sheet),
                _leap_source_pair_presence_lookup(esto_sheet),
            ),
            "raw_leap_missing_ninth_mapping": (
                NINTH_SHEET,
                _active_leap_source_pairs(ninth_sheet),
                _leap_source_pair_presence_lookup(ninth_sheet),
            ),
        }
        for gap_type, (sheet_name, mapped_source_pairs, presence_lookup) in active_source_pairs.items():
            for row in raw.itertuples(index=False):
                sector = str(getattr(row, "leap_sector_name_full_path", "")).strip()
                fuel = str(getattr(row, "raw_leap_fuel_name", "")).strip()
                if not sector or not fuel:
                    continue
                source_key = (_path_key(sector), _norm_text(fuel))
                if source_key in mapped_source_pairs:
                    continue
                presence = presence_lookup.get(source_key, {})
                records.append(
                    {
                        "gap_type": gap_type,
                        "sheet_name": sheet_name,
                        "original_dataset": "leap_balance_export",
                        "original_pair_is_subtotal": bool(getattr(row, "leap_pair_is_subtotal", False)),
                        "key_col_1": "leap_sector_name_full_path",
                        "key_col_2": "raw_leap_fuel_name",
                        "pair_1": sector,
                        "pair_2": fuel,
                        "abs_sum": float(getattr(row, "leap_pair_abs_sum", 0.0) or 0.0),
                        "mapping_presence_state": str(presence.get("state", "actually_missing")),
                        "mapping_presence_detail": str(presence.get("detail", "")),
                    }
                )

    return pd.DataFrame(
        records,
        columns=[
            "gap_type",
            "sheet_name",
            "original_dataset",
            "original_pair_is_subtotal",
            "key_col_1",
            "key_col_2",
            "pair_1",
            "pair_2",
            "abs_sum",
            "mapping_presence_state",
            "mapping_presence_detail",
        ],
    )


COVERAGE_GAPS_PATH = MISSING_PAIRS_CSV_PATH


def _report_coverage_gaps(gaps: pd.DataFrame, *, error_on_gaps: bool) -> None:
    """Write gaps CSV and either raise or warn depending on *error_on_gaps*."""
    COVERAGE_GAPS_PATH.parent.mkdir(parents=True, exist_ok=True)
    gaps.to_csv(COVERAGE_GAPS_PATH, index=False)

    if gaps.empty:
        return

    summary_lines: list[str] = []
    for gap_type, group in gaps.groupby("gap_type"):
        summary_lines.append(f"  {gap_type}: {len(group)} pair(s)")
    try:
        report_path = COVERAGE_GAPS_PATH.relative_to(REPO_ROOT)
    except ValueError:
        report_path = COVERAGE_GAPS_PATH
    summary = (
        f"{len(gaps)} coverage gap(s) found in leap_mappings.xlsx "
        f"(written to {report_path}):\n" + "\n".join(summary_lines)
    )

    if error_on_gaps:
        raise ValueError(summary)
    else:
        import warnings
        warnings.warn(summary, stacklevel=3)


def _report_duplicate_mappings(duplicates: pd.DataFrame) -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    duplicates.to_csv(DUPLICATE_MAPPINGS_CSV_PATH, index=False)
    if duplicates.empty:
        return
    summary = (
        f"{len(duplicates)} exact duplicate mapping row(s) found in leap_mappings.xlsx "
        f"(written to {DUPLICATE_MAPPINGS_CSV_PATH.relative_to(REPO_ROOT)})."
    )
    import warnings

    warnings.warn(summary, stacklevel=3)


TRIO_PRESENCE_HISTORY_KEY_COLUMNS = [
    "sheet_name",
    "presence_status",
    "leap_sector_name_original",
    "leap_sector_name_full_path",
    "raw_leap_fuel_name",
    "esto_flow",
    "esto_product",
    "ninth_sector",
    "ninth_fuel",
]


def _trio_presence_history_key(frame: pd.DataFrame) -> pd.Series:
    """Build a stable row key for carrying trio-presence history between runs."""
    work = frame.copy()
    for col in TRIO_PRESENCE_HISTORY_KEY_COLUMNS:
        if col not in work.columns:
            work[col] = ""
        work[col] = work[col].fillna("").astype(str).map(_norm_text)
    return work[TRIO_PRESENCE_HISTORY_KEY_COLUMNS].agg("|||".join, axis=1)


def _add_trio_presence_previous_runs(row_exclusive: pd.DataFrame) -> pd.DataFrame:
    """Add the count of prior refresh outputs where each issue row appeared."""
    out = row_exclusive.copy()
    if "previous_runs" in out.columns:
        out = out.drop(columns=["previous_runs"])
    out.insert(2, "previous_runs", 0)
    if out.empty or not TRIO_PRESENCE_CSV_PATH.exists():
        return out

    previous = pd.read_csv(TRIO_PRESENCE_CSV_PATH).fillna("")
    if previous.empty:
        return out

    previous_key = _trio_presence_history_key(previous)
    # If the previous file had no history column, appearing there still means
    # the row was present in one previous refresh output.
    if "previous_runs" in previous.columns:
        previous_count = pd.to_numeric(previous["previous_runs"], errors="coerce").fillna(0).astype(int) + 1
    else:
        previous_count = pd.Series(1, index=previous.index)

    prior_counts = previous_count.groupby(previous_key).max().to_dict()
    out["previous_runs"] = _trio_presence_history_key(out).map(prior_counts).fillna(0).astype(int)
    return out


def _write_trio_presence_csv(trio_presence: pd.DataFrame) -> pd.DataFrame:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    row_exclusive = trio_presence[~trio_presence["presence_status"].isin({"both_active"})].copy()
    row_exclusive = _add_trio_presence_previous_runs(row_exclusive)
    row_exclusive.to_csv(TRIO_PRESENCE_CSV_PATH, index=False)
    if row_exclusive.empty:
        return row_exclusive
    summary = row_exclusive.groupby(["sheet_name", "presence_status"], as_index=False).size().rename(columns={"size": "row_count"}).sort_values(["sheet_name", "presence_status"])
    summary_parts = [
        f"{row.sheet_name}:{row.presence_status}: {int(row.row_count)}"
        for row in summary.itertuples(index=False)
    ]
    import warnings

    warnings.warn(
        "Row presence mismatches found in leap_combined_esto and leap_combined_ninth "
        f"(written to {TRIO_PRESENCE_CSV_PATH.relative_to(REPO_ROOT)}): "
        + ", ".join(summary_parts),
        stacklevel=3,
    )
    return row_exclusive


ARCHIVE_DIR = MAPPING_WORKBOOK_PATH.parent / "archive"


def _backup_workbook(path: Path) -> Path:
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    backup_path = ARCHIVE_DIR / f"{path.stem}.before_refresh_mapping_maintenance_columns_{pd.Timestamp.now():%Y%m%d_%H%M%S}{path.suffix}"
    shutil.copy2(path, backup_path)
    return backup_path


def _assert_not_open(path: Path) -> None:
    """Raise a clear error if the file is locked (open in Excel or another process)."""
    try:
        with open(path, "r+b"):
            pass
    except PermissionError:
        raise PermissionError(
            f"{path.name} is open in another application (e.g. Excel). "
            "Close it and re-run the workflow."
        ) from None


def _replace_sheet_with_dataframe(workbook_path: Path, sheet_name: str, frame: pd.DataFrame) -> None:
    """Replace one sheet in-place while preserving every other sheet in the workbook."""
    try:
        workbook = load_workbook(workbook_path)
    except IndexError as exc:
        raise RuntimeError(
            f"{workbook_path.name} can be read in streaming mode but openpyxl cannot "
            "load it in editable mode. Close and re-save the workbook in Excel, or "
            "restore a recent copy from config/archive, then re-run this workflow."
        ) from exc
    if sheet_name in workbook.sheetnames:
        sheet_index = workbook.sheetnames.index(sheet_name)
        del workbook[sheet_name]
        worksheet = workbook.create_sheet(title=sheet_name, index=sheet_index)
    else:
        worksheet = workbook.create_sheet(title=sheet_name)
    for row in dataframe_to_rows(frame, index=False, header=True):
        worksheet.append(row)
    workbook.save(workbook_path)


def _read_mapping_sheet(sheet_name: str) -> pd.DataFrame:
    """Read a mapping workbook sheet, falling back to openpyxl read-only mode."""
    try:
        return pd.read_excel(MAPPING_WORKBOOK_PATH, sheet_name=sheet_name, dtype=object).fillna("")
    except IndexError:
        try:
            workbook = load_workbook(MAPPING_WORKBOOK_PATH, read_only=True, data_only=False)
            if sheet_name not in workbook.sheetnames:
                raise ValueError(f"Sheet {sheet_name!r} not found in {MAPPING_WORKBOOK_PATH}")
            rows = list(workbook[sheet_name].iter_rows(values_only=True))
        except IndexError:
            rows = _read_xlsx_sheet_values_xml(MAPPING_WORKBOOK_PATH, sheet_name)
        if not rows:
            return pd.DataFrame()
        headers = [str(value or "").strip() for value in rows[0]]
        data = list(rows[1:])
        width = len(headers)
        padded = [tuple(list(row[:width]) + [""] * max(0, width - len(row))) for row in data]
        return pd.DataFrame(padded, columns=headers).fillna("")


def _read_xlsx_sheet_values_xml(workbook_path: Path, sheet_name: str) -> list[tuple[object, ...]]:
    """Read one XLSX sheet directly from XML, ignoring styles that can break openpyxl."""
    ns = {
        "main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
        "rel": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
        "pkgrel": "http://schemas.openxmlformats.org/package/2006/relationships",
    }

    def _col_index(cell_ref: str) -> int:
        letters = "".join(ch for ch in str(cell_ref) if ch.isalpha()).upper()
        value = 0
        for letter in letters:
            value = value * 26 + (ord(letter) - ord("A") + 1)
        return max(value - 1, 0)

    with zipfile.ZipFile(workbook_path) as archive:
        shared_strings: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            shared_root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            for si in shared_root.findall("main:si", ns):
                shared_strings.append("".join(t.text or "" for t in si.findall(".//main:t", ns)))

        workbook_root = ET.fromstring(archive.read("xl/workbook.xml"))
        rel_root = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        rel_targets = {
            rel.attrib.get("Id", ""): rel.attrib.get("Target", "")
            for rel in rel_root.findall("pkgrel:Relationship", ns)
        }
        target = ""
        for sheet in workbook_root.findall(".//main:sheet", ns):
            if sheet.attrib.get("name", "") == sheet_name:
                target = rel_targets.get(sheet.attrib.get(f"{{{ns['rel']}}}id", ""), "")
                break
        if not target:
            raise ValueError(f"Sheet {sheet_name!r} not found in {workbook_path}")
        sheet_path = target.lstrip("/")
        if not sheet_path.startswith("xl/"):
            sheet_path = f"xl/{sheet_path}"

        sheet_root = ET.fromstring(archive.read(sheet_path))
        rows: list[tuple[object, ...]] = []
        for row in sheet_root.findall(".//main:sheetData/main:row", ns):
            values: list[object] = []
            for cell in row.findall("main:c", ns):
                col_idx = _col_index(cell.attrib.get("r", ""))
                while len(values) <= col_idx:
                    values.append("")
                cell_type = cell.attrib.get("t", "")
                raw_value = cell.findtext("main:v", default="", namespaces=ns)
                if cell_type == "s" and str(raw_value).strip().isdigit():
                    shared_idx = int(raw_value)
                    value = shared_strings[shared_idx] if shared_idx < len(shared_strings) else ""
                elif cell_type == "inlineStr":
                    value = "".join(t.text or "" for t in cell.findall(".//main:t", ns))
                else:
                    value = raw_value
                values[col_idx] = value
            rows.append(tuple(values))
        return rows


def _read_master_config_sheet(sheet_name: str) -> pd.DataFrame:
    """Read one sheet from config/master_config.xlsx."""
    if not MASTER_CONFIG_PATH.exists():
        raise FileNotFoundError(f"Missing master config workbook: {MASTER_CONFIG_PATH}")
    return pd.read_excel(MASTER_CONFIG_PATH, sheet_name=sheet_name, dtype=object).fillna("")


def _filter_researcher_rows(frame: pd.DataFrame) -> pd.DataFrame:
    """Keep rows that are not marked as removed, duplicate-removed, or faulty."""
    out = frame.copy()
    active = pd.Series(True, index=out.index)
    for col in ["remove_row", "duplicate_to_remove", "removed", "is_removed", "faulty mapping", "faulty_mapping"]:
        if col in out.columns:
            active &= ~out[col].map(_truthy)
    return out.loc[active].copy()


def _pair_cardinality_for_columns(
    frame: pd.DataFrame,
    *,
    source_cols: Sequence[str],
    target_cols: Sequence[str],
    cardinality_col: str = "cardinality",
) -> pd.DataFrame:
    """Add cardinality between source column pairs and target column pairs."""
    out = frame.copy()
    for col in [*source_cols, *target_cols]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()

    valid = pd.Series(True, index=out.index)
    for col in [*source_cols, *target_cols]:
        valid &= out[col].ne("")

    pairs = out.loc[valid, [*source_cols, *target_cols]].drop_duplicates().copy()
    if pairs.empty:
        out[cardinality_col] = ""
        return out

    pairs["_source_key"] = pairs[list(source_cols)].agg("|||".join, axis=1)
    pairs["_target_key"] = pairs[list(target_cols)].agg("|||".join, axis=1)
    source_count = pairs.groupby("_source_key")["_target_key"].nunique()
    target_count = pairs.groupby("_target_key")["_source_key"].nunique()

    out["_source_key"] = out[list(source_cols)].agg("|||".join, axis=1)
    out["_target_key"] = out[list(target_cols)].agg("|||".join, axis=1)
    out[cardinality_col] = ""
    out.loc[valid, cardinality_col] = out.loc[valid].apply(
        lambda row: _mapping_cardinality(
            int(source_count.get(row["_source_key"], 0)),
            int(target_count.get(row["_target_key"], 0)),
        ),
        axis=1,
    )
    return out.drop(columns=["_source_key", "_target_key"])


def _researcher_export_frame(
    frame: pd.DataFrame,
    *,
    column_rename: dict[str, str],
    include_name: bool = False,
) -> pd.DataFrame:
    """Return the narrow researcher-facing mapping columns."""
    out = _filter_researcher_rows(_drop_unnamed_columns(frame))
    out = out.rename(columns={old: new for old, new in column_rename.items() if old in out.columns})
    if "pair_mapping_cardinality" in out.columns and "cardinality" not in out.columns:
        out = out.rename(columns={"pair_mapping_cardinality": "cardinality"})
    if "sector_mapping_cardinality" in out.columns and "cardinality" not in out.columns:
        out = out.rename(columns={"sector_mapping_cardinality": "cardinality"})
    if "cardinality" not in out.columns:
        out["cardinality"] = ""

    requested_cols = ["9th_sector", "9th_fuel", "esto_flow", "esto_product", "leap_flow", "leap_product"]
    for col in requested_cols:
        if col not in out.columns:
            out[col] = ""

    final_cols = [*requested_cols, "cardinality"]
    if include_name and "name" in out.columns:
        final_cols.insert(-1, "name")
    out = out.loc[:, final_cols].copy()
    for col in out.columns:
        out[col] = out[col].fillna("").astype(str).str.strip()
    return out.drop_duplicates().reset_index(drop=True)


def build_researcher_mappings_workbook(
    output_path: Path | str = RESEARCHER_MAPPINGS_PATH,
) -> dict[str, object]:
    """
    Write a narrow researcher-facing mapping workbook.

    The output keeps active rows only and writes four sheets:
    leap_combined_esto, leap_combined_ninth, sector_fuel_code_to_name, and
    ninth_pairs_to_esto_pairs.
    """
    output_path = _resolve(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    if output_path.exists():
        _assert_not_open(output_path)

    esto = _compute_pair_cardinality(_read_mapping_sheet(ESTO_SHEET), "esto_flow", "esto_product")
    ninth = _compute_pair_cardinality(_read_mapping_sheet(NINTH_SHEET), "ninth_sector", "ninth_fuel")
    code_to_name = _read_master_config_sheet(SECTOR_FUEL_CODE_TO_NAME_SHEET)
    ninth_to_esto = _pair_cardinality_for_columns(
        _read_master_config_sheet(NINTH_PAIRS_TO_ESTO_PAIRS_SHEET),
        source_cols=["9th_sector", "9th_fuel"],
        target_cols=["esto_flow", "esto_product"],
    )

    sheets = {
        ESTO_SHEET: _researcher_export_frame(
            esto,
            column_rename={
                "leap_sector_name_full_path": "leap_flow",
                "raw_leap_fuel_name": "leap_product",
            },
        ),
        NINTH_SHEET: _researcher_export_frame(
            ninth,
            column_rename={
                "leap_sector_name_full_path": "leap_flow",
                "raw_leap_fuel_name": "leap_product",
            },
        ),
        SECTOR_FUEL_CODE_TO_NAME_SHEET: _researcher_export_frame(
            code_to_name,
            column_rename={
                "9th_label": "9th_fuel",
                "esto_label": "esto_product",
            },
            include_name=True,
        ),
        NINTH_PAIRS_TO_ESTO_PAIRS_SHEET: _researcher_export_frame(
            ninth_to_esto,
            column_rename={},
        ),
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, sheet_df in sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)

    return {
        "researcher_mappings_workbook": str(output_path),
        "sheet_rows": {sheet_name: int(len(sheet_df)) for sheet_name, sheet_df in sheets.items()},
    }


def build_mapping_conflicts_workbook(
    output_path: Path | str = MAPPING_CONFLICTS_WORKBOOK_PATH,
) -> dict[str, object]:
    """Build the mapping conflict report from the current workbook state."""
    output_path = _resolve(output_path)
    esto_sheet = _compute_pair_cardinality(_read_mapping_sheet(ESTO_SHEET), "esto_flow", "esto_product")
    ninth_sheet = _compute_pair_cardinality(_read_mapping_sheet(NINTH_SHEET), "ninth_sector", "ninth_fuel")
    ninth_to_esto_pairs = _read_master_config_sheet(NINTH_PAIRS_TO_ESTO_PAIRS_SHEET)
    report_sheets = build_mapping_conflict_report(esto_sheet, ninth_sheet, ninth_to_esto_pairs)
    _write_mapping_conflict_report(report_sheets, output_path)
    return {
        "mapping_conflicts_workbook": str(output_path),
        "sheet_rows": {sheet_name: int(len(sheet_df)) for sheet_name, sheet_df in report_sheets.items()},
    }


def run_workflow(*, error_on_gaps: bool = True, check_raw_leap_coverage: bool = True) -> dict[str, object]:
    """
    Refresh mapping maintenance columns.

    Parameters
    ----------
    error_on_gaps:
        If True (default), raise a ValueError when coverage gaps are found.
        If False, emit a warning and continue; gaps are still written to
        outputs/mappings/mapping_checks/leap_mapping_missing_pairs.csv.
    check_raw_leap_coverage:
        If True (default), read the latest REF/TGT LEAP balance exports and
        report nonzero raw LEAP sector/fuel pairs missing from the mapping
        workbook.
    """
    if not MAPPING_WORKBOOK_PATH.exists():
        raise FileNotFoundError(f"Missing mapping workbook: {MAPPING_WORKBOOK_PATH}")
    _assert_not_open(MAPPING_WORKBOOK_PATH)
    backup_path = _backup_workbook(MAPPING_WORKBOOK_PATH)

    esto_sheet = _read_mapping_sheet(ESTO_SHEET)
    ninth_sheet = _read_mapping_sheet(NINTH_SHEET)

    esto_lookup = _load_esto_lookup()
    ninth_lookup = _load_ninth_lookup()

    refreshed_esto = _refresh_esto_sheet(esto_sheet, esto_lookup)
    refreshed_ninth = _refresh_ninth_sheet(ninth_sheet, ninth_lookup)

    refreshed_esto, esto_auto_remove = _apply_auto_remove_rules(refreshed_esto)
    refreshed_ninth, ninth_auto_remove = _apply_auto_remove_rules(refreshed_ninth)
    refreshed_esto = _compute_pair_cardinality(refreshed_esto, "esto_flow", "esto_product")
    refreshed_ninth = _compute_pair_cardinality(refreshed_ninth, "ninth_sector", "ninth_fuel")

    auto_remove_summary = (
        "Auto-remove rules applied: "
        f"ESTO total fuels={esto_auto_remove['auto_remove_total_fuel_rows']}, "
        f"ESTO suffix matches={esto_auto_remove['auto_remove_sector_fuel_suffix_rows']}, "
        f"9th total fuels={ninth_auto_remove['auto_remove_total_fuel_rows']}, "
        f"9th suffix matches={ninth_auto_remove['auto_remove_sector_fuel_suffix_rows']}, "
        f"newly marked rows={esto_auto_remove['auto_removed_new_rows'] + ninth_auto_remove['auto_removed_new_rows']}."
    )
    import warnings

    warnings.warn(auto_remove_summary, stacklevel=3)

    raw_leap_lookup = _load_raw_leap_balance_lookup() if check_raw_leap_coverage else pd.DataFrame()
    gaps = _build_coverage_gaps(refreshed_esto, refreshed_ninth, esto_lookup, ninth_lookup, raw_leap_lookup)

    refreshed_esto = refreshed_esto.sort_values(
        ["leap_sector_name_full_path", "raw_leap_fuel_name", "esto_flow", "esto_product"],
        key=lambda col: col.fillna("").astype(str).str.lower(),
        na_position="last",
    ).reset_index(drop=True)
    refreshed_ninth = refreshed_ninth.sort_values(
        ["leap_sector_name_full_path", "raw_leap_fuel_name", "ninth_sector", "ninth_fuel"],
        key=lambda col: col.fillna("").astype(str).str.lower(),
        na_position="last",
    ).reset_index(drop=True)

    duplicate_esto = _build_duplicate_mappings(
        refreshed_esto,
        sheet_name=ESTO_SHEET,
        target_a="esto_flow",
        target_b="esto_product",
    )
    duplicate_ninth = _build_duplicate_mappings(
        refreshed_ninth,
        sheet_name=NINTH_SHEET,
        target_a="ninth_sector",
        target_b="ninth_fuel",
    )
    duplicate_mappings = pd.concat([duplicate_esto, duplicate_ninth], ignore_index=True)
    _report_duplicate_mappings(duplicate_mappings)

    trio_presence = _build_trio_presence_check(refreshed_esto, refreshed_ninth)
    trio_presence_issues = _write_trio_presence_csv(trio_presence)
    mapping_conflict_report = build_mapping_conflict_report(
        refreshed_esto,
        refreshed_ninth,
        _read_master_config_sheet(NINTH_PAIRS_TO_ESTO_PAIRS_SHEET),
    )
    _write_mapping_refresh_report(
        gaps=gaps,
        duplicate_mappings=duplicate_mappings,
        trio_presence_issues=trio_presence_issues,
        mapping_conflict_report=mapping_conflict_report,
    )
    _report_coverage_gaps(gaps, error_on_gaps=error_on_gaps)

    _replace_sheet_with_dataframe(MAPPING_WORKBOOK_PATH, ESTO_SHEET, refreshed_esto)
    _replace_sheet_with_dataframe(MAPPING_WORKBOOK_PATH, NINTH_SHEET, refreshed_ninth)

    return {
        "mapping_workbook": str(MAPPING_WORKBOOK_PATH),
        "backup_workbook": str(backup_path),
        "coverage_gaps_csv": str(COVERAGE_GAPS_PATH),
        "coverage_gaps_count": int(len(gaps)),
        "raw_leap_source_pairs_checked": int(len(raw_leap_lookup)),
        "raw_leap_missing_mapping_count": int(
            gaps["gap_type"].isin(
                {"raw_leap_missing_esto_mapping", "raw_leap_missing_ninth_mapping"}
            ).sum()
        ) if "gap_type" in gaps.columns else 0,
        "duplicate_mappings_csv": str(DUPLICATE_MAPPINGS_CSV_PATH),
        "duplicate_mappings_count": int(len(duplicate_mappings)),
        "trio_presence_csv": str(TRIO_PRESENCE_CSV_PATH),
        "trio_presence_count": int(len(trio_presence)),
        "mapping_refresh_report": str(MAPPING_REFRESH_REPORT_PATH),
        "mapping_conflicts_count": int(
            sum(
                len(sheet_df)
                for sheet_name, sheet_df in mapping_conflict_report.items()
                if sheet_name != "summary"
            )
        ),
        "auto_remove_total_fuel_rows_esto": int(esto_auto_remove["auto_remove_total_fuel_rows"]),
        "auto_remove_sector_fuel_suffix_rows_esto": int(esto_auto_remove["auto_remove_sector_fuel_suffix_rows"]),
        "auto_removed_new_rows_esto": int(esto_auto_remove["auto_removed_new_rows"]),
        "auto_remove_total_fuel_rows_ninth": int(ninth_auto_remove["auto_remove_total_fuel_rows"]),
        "auto_remove_sector_fuel_suffix_rows_ninth": int(ninth_auto_remove["auto_remove_sector_fuel_suffix_rows"]),
        "auto_removed_new_rows_ninth": int(ninth_auto_remove["auto_removed_new_rows"]),
        "leap_combined_esto_rows": int(len(refreshed_esto)),
        "leap_combined_ninth_rows": int(len(refreshed_ninth)),
    }


#%%
RUN_WORKFLOW = True
# Set to False to emit a warning instead of raising an error when coverage gaps are found.
ERROR_ON_GAPS = False
CREATE_RESEARCHER_MAPPINGS = False
CREATE_MAPPING_CONFLICTS = False

WORKFLOW_RESULT: dict[str, object] | None = None
if __name__ == "__main__" and RUN_WORKFLOW:
    WORKFLOW_RESULT = run_workflow(error_on_gaps=ERROR_ON_GAPS)
    print("[OK] Mapping maintenance columns refreshed.")
    for key, value in WORKFLOW_RESULT.items():
        print(f"- {key}: {value}")
        if key == "trio_presence_csv":
            print(f"- trio_presence_note: {TRIO_PRESENCE_OUTPUT_NOTE}")

RESEARCHER_MAPPINGS_RESULT: dict[str, object] | None = None
if __name__ == "__main__" and CREATE_RESEARCHER_MAPPINGS:
    RESEARCHER_MAPPINGS_RESULT = build_researcher_mappings_workbook()
    print("[OK] Researcher mappings workbook written.")
    for key, value in RESEARCHER_MAPPINGS_RESULT.items():
        print(f"- {key}: {value}")

MAPPING_CONFLICTS_RESULT: dict[str, object] | None = None
if __name__ == "__main__" and CREATE_MAPPING_CONFLICTS:
    MAPPING_CONFLICTS_RESULT = build_mapping_conflicts_workbook()
    print("[OK] Mapping conflicts workbook written.")
    for key, value in MAPPING_CONFLICTS_RESULT.items():
        print(f"- {key}: {value}")
#%%
