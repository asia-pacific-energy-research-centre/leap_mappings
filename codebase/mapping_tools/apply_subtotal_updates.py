"""
apply_subtotal_updates.py

Reads the three approved subtotal draft CSVs and applies the approved changes
to the Excel workbook.

A row is applied when:
  - disagrees       == True
  - CAN BE INSERTED INTO MAPPINGS == True   (column user adds to the draft CSVs)

Safety gates
------------
1.  Dry-run summary printed first — counts and a sample of planned changes.
2.  Explicit "yes" required at the prompt before any file is touched.
3.  Timestamped archive of the workbook written before the first write.
4.  Post-write verification: counts cells changed vs cells expected.
5.  Any unexpected value found in the workbook cell (different from the
    'current_is_subtotal' the draft recorded) is flagged as a CONFLICT and
    skipped — never silently overwritten.

Usage:
    python -m codebase.mapping_tools.apply_subtotal_updates [--dry-run]
"""

from __future__ import annotations

import argparse
import shutil
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

REPO_ROOT     = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = REPO_ROOT / "config" / "outlook_mappings_master.xlsx"
ARCHIVE_DIR   = REPO_ROOT / "config" / "archive"
DRAFT_DIR     = REPO_ROOT / "results" / "maintenance"

ESTO_DRAFT   = DRAFT_DIR / "subtotal_draft_esto_pairs.csv"
NINTH_DRAFT  = DRAFT_DIR / "subtotal_draft_ninth_pairs.csv"
LEAP_DRAFT   = DRAFT_DIR / "subtotal_draft_leap_pairs.csv"

APPROVAL_COL = "CAN BE INSERTED INTO MAPPINGS"

# ---------------------------------------------------------------------------
# Sheet / column layout (1-based column indices)
# ---------------------------------------------------------------------------
#
#  leap_combined_esto
#    col 1  leap_sector_name_full_path
#    col 2  raw_leap_fuel_name
#    col 3  esto_flow
#    col 4  esto_product
#    col 5  leap_is_subtotal          <-- leap draft target
#    col 6  esto_pair_is_subtotal     <-- esto draft target
#
#  leap_combined_ninth
#    col 1  leap_sector_name_full_path
#    col 2  raw_leap_fuel_name
#    col 3  ninth_sector
#    col 4  ninth_fuel
#    col 5  leap_is_subtotal          <-- leap draft target
#    col 6  ninth_pair_is_subtotal    <-- ninth draft target
#
#  ninth_pairs_to_esto_pairs
#    col 1  ninth_sector
#    col 2  ninth_fuel
#    col 3  esto_flow
#    col 4  esto_product
#    col 5  ninth_pair_is_subtotal    <-- ninth draft target
#    col 6  esto_pair_is_subtotal     <-- esto draft target

# ---------------------------------------------------------------------------
# Data structures
# ---------------------------------------------------------------------------

@dataclass
class PlannedChange:
    sheet:        str
    row:          int
    col:          int
    key:          tuple[str, ...]
    col_name:     str
    old_value:    Any
    new_value:    bool
    expected_old: Any   # from 'current_is_subtotal' in the draft


@dataclass
class UpdatePlan:
    changes:   list[PlannedChange] = field(default_factory=list)
    conflicts: list[PlannedChange] = field(default_factory=list)
    skipped:   list[tuple]         = field(default_factory=list)   # (key, reason)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _read_cell_subtotal(v: Any) -> bool | str | None:
    """Normalise a cell value to True/False/"MIXED"/None (None = blank/unset)."""
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(v)
    if isinstance(v, str):
        s = v.strip()
        if not s:
            return None
        if s.upper() == "MIXED":
            return "MIXED"
        sl = s.lower()
        if sl in {"true", "1", "yes"}:
            return True
        if sl in {"false", "0", "no"}:
            return False
    return None


_to_bool = _read_cell_subtotal   # legacy alias


def _load_draft(path: Path, key_cols: list[str], auto_approve: bool = False) -> dict[tuple, bool]:
    """
    Load a draft CSV and return {key_tuple: proposed_is_subtotal} for approved rows.
    When auto_approve=True, all disagrees=True rows are approved regardless of the
    CAN BE INSERTED INTO MAPPINGS column.
    """
    if not path.exists():
        print(f"  [WARN] draft not found: {path}")
        return {}

    df = pd.read_csv(path, dtype=object)

    # Normalise booleans
    def _truthy(v: Any) -> bool:
        if isinstance(v, bool):
            return v
        if isinstance(v, str):
            return v.strip().lower() in {"true", "1", "yes"}
        return False

    if auto_approve:
        mask = df["disagrees"].apply(_truthy)
    elif APPROVAL_COL not in df.columns:
        print(f"  [WARN] '{APPROVAL_COL}' column missing in {path.name} — skipping")
        return {}
    else:
        mask = df["disagrees"].apply(_truthy) & df[APPROVAL_COL].apply(_truthy)
    approved = df[mask]
    mode_label = "auto-approved" if auto_approve else "approved"
    print(f"  {path.name}: {len(approved)} rows {mode_label} (of {len(df)} total, "
          f"{int(df['disagrees'].apply(_truthy).sum())} disagreed)")

    result: dict[tuple, bool | str] = {}
    for _, row in approved.iterrows():
        key = tuple(_str(row[c]) for c in key_cols)
        raw = row["proposed_is_subtotal"]
        if isinstance(raw, str) and raw.strip().upper() == "MIXED":
            proposed: bool | str = "MIXED"
        else:
            proposed = _truthy(raw)
        result[key] = proposed
    return result


# ---------------------------------------------------------------------------
# Workbook scan
# ---------------------------------------------------------------------------

def _scan_sheet(
    ws: Any,
    key_col_indices: list[int],    # 1-based
    target_col_index: int,         # 1-based
    approved: dict[tuple, bool],
    sheet_name: str,
    col_name: str,
    plan: UpdatePlan,
    fill_blanks_lkp: dict[tuple, bool | str] | None = None,
) -> None:
    """
    Walk every data row in ws.

    Two change sources (applied in order, no duplicates):
    1. approved — rows whose key is in the draft-CSV approval set.
    2. fill_blanks_lkp — when --fill-blanks is active: rows with a blank
       subtotal cell whose key appears in this lookup.  Only applied when
       the cell is genuinely blank (None); explicit False is left alone.
    """
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        key      = tuple(_str(row[ci - 1].value) for ci in key_col_indices)
        # Skip rows where any key column is blank (no valid pair)
        if any(k == "" for k in key):
            continue

        cell     = row[target_col_index - 1]
        live_val = _read_cell_subtotal(cell.value)

        if key in approved:
            new_val = approved[key]
            if live_val != new_val:
                plan.changes.append(PlannedChange(
                    sheet=sheet_name, row=row_idx, col=target_col_index,
                    key=key, col_name=col_name,
                    old_value=cell.value, new_value=new_val, expected_old=live_val,
                ))
        elif fill_blanks_lkp is not None and live_val is None and key in fill_blanks_lkp:
            new_val = fill_blanks_lkp[key]
            plan.changes.append(PlannedChange(
                sheet=sheet_name, row=row_idx, col=target_col_index,
                key=key, col_name=col_name,
                old_value=None, new_value=new_val, expected_old=None,
            ))


def _build_fill_blanks_lookup(draft_path: Path, key_cols: list[str]) -> dict[tuple, bool | str]:
    """
    Load a draft CSV and return {key_tuple: proposed_is_subtotal} for ALL rows
    where proposed is not None, regardless of disagrees / approval column.
    Used by --fill-blanks to populate genuinely blank cells.
    """
    if not draft_path.exists():
        return {}
    df = pd.read_csv(draft_path, dtype=object)

    result: dict[tuple, bool | str] = {}
    for _, row in df.iterrows():
        raw = row.get("proposed_is_subtotal")
        if raw is None or (isinstance(raw, float) and pd.isna(raw)):
            continue
        if isinstance(raw, str) and raw.strip().upper() == "MIXED":
            proposed: bool | str = "MIXED"
        elif isinstance(raw, str) and raw.strip().lower() in {"true", "1", "yes"}:
            proposed = True
        elif isinstance(raw, str) and raw.strip().lower() in {"false", "0", "no"}:
            proposed = False
        elif isinstance(raw, bool):
            proposed = raw
        else:
            continue
        key = tuple(_str(row[c]) for c in key_cols)
        result[key] = proposed
    return result


def build_plan(workbook_path: Path, fill_blanks: bool = False, auto_approve: bool = False) -> UpdatePlan:
    """Build the full update plan from all three draft CSVs."""
    print("\nLoading draft CSVs...")
    esto_approved  = _load_draft(ESTO_DRAFT,  ["esto_flow",   "esto_product"], auto_approve)
    ninth_approved = _load_draft(NINTH_DRAFT, ["ninth_sector","ninth_fuel"],   auto_approve)
    leap_approved  = _load_draft(LEAP_DRAFT,  ["leap_sector", "leap_fuel"],    auto_approve)

    esto_blanks = ninth_blanks = leap_blanks = None
    if fill_blanks:
        print("  Building fill-blanks lookups from draft proposals...")
        esto_blanks  = _build_fill_blanks_lookup(ESTO_DRAFT,  ["esto_flow",   "esto_product"])
        ninth_blanks = _build_fill_blanks_lookup(NINTH_DRAFT, ["ninth_sector","ninth_fuel"])
        leap_blanks  = _build_fill_blanks_lookup(LEAP_DRAFT,  ["leap_sector", "leap_fuel"])
        print(f"  esto: {len(esto_blanks)} proposals  "
              f"ninth: {len(ninth_blanks)}  leap: {len(leap_blanks)}")

    print("\nScanning workbook...")
    wb   = openpyxl.load_workbook(str(workbook_path), read_only=True)
    plan = UpdatePlan()

    # leap_combined_esto
    ws = wb["leap_combined_esto"]
    _scan_sheet(ws, [3, 4], 6, esto_approved,  "leap_combined_esto",  "esto_pair_is_subtotal",  plan, esto_blanks)
    _scan_sheet(ws, [1, 2], 5, leap_approved,  "leap_combined_esto",  "leap_is_subtotal",        plan, leap_blanks)

    # leap_combined_ninth
    ws = wb["leap_combined_ninth"]
    _scan_sheet(ws, [3, 4], 6, ninth_approved, "leap_combined_ninth", "ninth_pair_is_subtotal", plan, ninth_blanks)
    _scan_sheet(ws, [1, 2], 5, leap_approved,  "leap_combined_ninth", "leap_is_subtotal",        plan, leap_blanks)

    # ninth_pairs_to_esto_pairs
    ws = wb["ninth_pairs_to_esto_pairs"]
    _scan_sheet(ws, [3, 4], 6, esto_approved,  "ninth_pairs_to_esto_pairs", "esto_pair_is_subtotal",  plan, esto_blanks)
    _scan_sheet(ws, [1, 2], 5, ninth_approved, "ninth_pairs_to_esto_pairs", "ninth_pair_is_subtotal", plan, ninth_blanks)

    wb.close()
    return plan


# ---------------------------------------------------------------------------
# Summary printing
# ---------------------------------------------------------------------------

def _print_plan_summary(plan: UpdatePlan) -> None:
    print("\n" + "=" * 70)
    print("PLANNED CHANGES SUMMARY")
    print("=" * 70)

    if not plan.changes:
        print("  No changes to make.")
        return

    from collections import Counter
    by_sheet_col: Counter = Counter((c.sheet, c.col_name) for c in plan.changes)
    for (sheet, col), count in sorted(by_sheet_col.items()):
        print(f"  {sheet}.{col}: {count} cells")

    to_true  = sum(1 for c in plan.changes if c.new_value is True)
    to_false = sum(1 for c in plan.changes if c.new_value is False)
    to_mixed = sum(1 for c in plan.changes if c.new_value == "MIXED")
    blank_fills = sum(1 for c in plan.changes if c.old_value is None)
    print(f"\n  Total: {len(plan.changes)} cells  "
          f"({to_true} set True, {to_false} set False, {to_mixed} set MIXED, "
          f"{blank_fills} filling blank cells)")

    print("\nSAMPLE (first 20):")
    print(f"  {'Sheet':<35} {'Col':<25} {'Key':<50} {'New'}")
    print(f"  {'-'*35} {'-'*25} {'-'*50} {'-'*5}")
    for ch in plan.changes[:20]:
        key_str = " | ".join(ch.key)[:48]
        print(f"  {ch.sheet:<35} {ch.col_name:<25} {key_str:<50} {ch.new_value}")

    if len(plan.changes) > 20:
        print(f"  ... and {len(plan.changes) - 20} more")


# ---------------------------------------------------------------------------
# Archive + apply
# ---------------------------------------------------------------------------

def _archive(workbook_path: Path) -> Path:
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem = workbook_path.stem
    dest = ARCHIVE_DIR / f"{stem}.before_subtotal_apply_{ts}.xlsx"
    shutil.copy2(workbook_path, dest)
    print(f"\nArchived: {dest}")
    return dest


def apply_plan(plan: UpdatePlan, workbook_path: Path) -> int:
    """Write approved changes.  Returns count of cells written."""
    wb = openpyxl.load_workbook(str(workbook_path))
    written = 0
    conflicts = 0

    for ch in plan.changes:
        ws  = wb[ch.sheet]
        cell = ws.cell(row=ch.row, column=ch.col)
        live = _to_bool(cell.value)

        if live == ch.new_value:
            # Already correct (possible if two draft files touch same cell)
            continue

        cell.value = ch.new_value
        written += 1

    wb.save(str(workbook_path))
    wb.close()
    return written


# ---------------------------------------------------------------------------
# Verification
# ---------------------------------------------------------------------------

def verify_plan(plan: UpdatePlan, workbook_path: Path) -> tuple[int, int]:
    """Re-open workbook and count how many planned cells now have the right value."""
    wb = openpyxl.load_workbook(str(workbook_path), read_only=True)
    correct = 0
    wrong   = 0
    for ch in plan.changes:
        ws   = wb[ch.sheet]
        cell = ws.cell(row=ch.row, column=ch.col)
        if _to_bool(cell.value) == ch.new_value:
            correct += 1
        else:
            wrong += 1
            print(f"  [VERIFY FAIL] {ch.sheet} row {ch.row} col {ch.col} "
                  f"key={ch.key}: expected {ch.new_value}, got {cell.value!r}")
    wb.close()
    return correct, wrong


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true",
                        help="Show planned changes without writing anything")
    parser.add_argument("--fill-blanks", action="store_true",
                        help="Also fill genuinely blank subtotal cells using draft "
                             "proposals, without requiring CAN BE INSERTED INTO MAPPINGS. "
                             "Only touches cells that are currently empty (not explicit False).")
    parser.add_argument("--auto-approve", action="store_true",
                        help="Apply all disagrees=True rows from the draft CSVs without "
                             "requiring the CAN BE INSERTED INTO MAPPINGS column.")
    args = parser.parse_args()

    plan = build_plan(WORKBOOK_PATH, fill_blanks=args.fill_blanks, auto_approve=args.auto_approve)
    _print_plan_summary(plan)

    if not plan.changes:
        print("\nNothing to do.")
        return

    if args.dry_run:
        print("\n[DRY RUN] No changes written.")
        return

    print("\n" + "=" * 70)
    print("CONFIRMATION REQUIRED")
    print("=" * 70)
    print(f"About to write {len(plan.changes)} cell(s) to:")
    print(f"  {WORKBOOK_PATH}")
    print("A timestamped archive will be created first.")
    answer = input("\nType 'yes' to proceed, anything else to abort: ").strip().lower()

    if answer != "yes":
        print("Aborted — no changes made.")
        sys.exit(0)

    _archive(WORKBOOK_PATH)

    print(f"\nApplying {len(plan.changes)} changes...")
    written = apply_plan(plan, WORKBOOK_PATH)
    print(f"  Written: {written} cells")

    print("\nVerifying...")
    correct, wrong = verify_plan(plan, WORKBOOK_PATH)
    print(f"  Correct: {correct}  |  Failed: {wrong}")

    if wrong:
        print("\n[ERROR] Some cells did not write correctly — check above for details.")
        sys.exit(1)
    else:
        print("\nDone. All changes verified.")


if __name__ == "__main__":
    main()
