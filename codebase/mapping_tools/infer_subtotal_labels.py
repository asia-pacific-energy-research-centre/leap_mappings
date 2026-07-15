"""
infer_subtotal_labels.py

Derives subtotal labels from structural tree hierarchy and rollup sheets,
and checks rollup-rule input consistency.  Produces review CSVs so the user
can audit the proposed labels before updating the workbook.

Outputs (results/maintenance/):
    subtotal_draft_esto_pairs.csv    -- current vs proposed esto_pair_is_subtotal
                                       for every pair in leap_combined_esto and
                                       ninth_pairs_to_esto_pairs
    subtotal_draft_ninth_pairs.csv   -- current vs proposed ninth_pair_is_subtotal
                                       for every pair in leap_combined_ninth and
                                       ninth_pairs_to_esto_pairs
    subtotal_draft_leap_pairs.csv    -- current vs proposed leap_is_subtotal
                                       for every pair in leap_combined_esto and
                                       leap_combined_ninth
    rollup_consistency.csv           -- per-output-group input consistency check;
                                       MIXED_WARNING rows need manual review

Design
------
Subtotal status for a mapping row is derived from two authoritative sources,
tried in order:

1. Structural trees (esto_tree.csv / ninth_tree.csv / leap_tree.csv): a node is
   is_subtotal when it has children in the dataset hierarchy.

2. Rollup sheets (esto_rollup_rules / ninth_rollup_rules / leap_rollup_rules):
   the Subtotal column on each rollup-output row is the explicit ground truth for
   that combined label, including the value "MIXED" for outputs that aggregate
   across both subtotal and leaf inputs.

A small manual override list covers "semantic totals" that have no explicit
children in the ESTO tree but are accounting roll-ups of the whole dataset
(e.g. 07, 12, 13).
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Any

import pandas as pd

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

def _find_repo_root() -> Path:
    here = Path(__file__).resolve()
    for parent in [here, *here.parents]:
        if (parent / "config" / "outlook_mappings_master.xlsx").exists():
            return parent
    raise RuntimeError("Could not locate repo root (no config/outlook_mappings_master.xlsx found).")


REPO_ROOT = _find_repo_root()
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

TREE_DIR          = REPO_ROOT / "results" / "tree_structure"
WORKBOOK_PATH     = REPO_ROOT / "config" / "outlook_mappings_master.xlsx"
EXCEPTION_PATH    = REPO_ROOT / "config" / "mapping_issue_exception_sets.xlsx"
OUTPUT_DIR        = REPO_ROOT / "results" / "maintenance"

# ---------------------------------------------------------------------------
# Manual overrides
# Flows that are accounting totals with no explicit children in the ESTO tree.
# ---------------------------------------------------------------------------

ESTO_FLOW_MANUAL_SUBTOTAL_OVERRIDES: frozenset[str] = frozenset({
    "07 Total primary energy supply",
    "12 Total final consumption",
    "13 Total final energy consumption",
})

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _truthy(val: Any) -> bool:
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.strip().lower() in {"true", "1", "yes"}
    return False


def _read_subtotal(val: Any) -> bool | str | None:
    """
    Read a subtotal cell value preserving the MIXED class and blank state.
    Returns True, False, "MIXED", or None (cell is blank/unset).
    Blank is kept as None so that missing labels are distinguishable from
    explicit False and are correctly flagged as needing to be filled.
    """
    if val is None:
        return None
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        s = val.strip()
        if not s:
            return None
        if s.upper() == "MIXED":
            return "MIXED"
        if s.lower() in {"true", "1", "yes"}:
            return True
        if s.lower() in {"false", "0", "no"}:
            return False
    return None


def _subtotal_disagrees(proposed: bool | str | None, current: bool | str | None) -> bool:
    """True when the proposal is a concrete value that differs from current.
    current=None means the cell is blank — any concrete proposal disagrees."""
    if proposed is None:
        return False
    return proposed != current


def load_subtotal_exceptions() -> dict[str, set[tuple[str, str]]]:
    """
    Read the subtotal_label_exceptions sheet from mapping_issue_exception_sets.xlsx.

    Returns a dict keyed by draft_type ('esto_pairs', 'ninth_pairs', 'leap_pairs'),
    each value a set of (key_1, key_2) tuples that are accepted as-is and should
    not be flagged as disagreements in the draft CSVs.

    Only rows with enabled=True are included.
    """
    result: dict[str, set[tuple[str, str]]] = {
        "esto_pairs": set(),
        "ninth_pairs": set(),
        "leap_pairs": set(),
    }
    if not EXCEPTION_PATH.exists():
        return result

    import openpyxl as _xl
    wb = _xl.load_workbook(str(EXCEPTION_PATH), read_only=True)
    if "subtotal_label_exceptions" not in wb.sheetnames:
        wb.close()
        return result

    ws = wb["subtotal_label_exceptions"]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    for row in ws.iter_rows(min_row=2, values_only=True):
        rec = dict(zip(headers, row))
        if not _truthy(rec.get("enabled", True)):
            continue
        dtype = str(rec.get("draft_type") or "").strip()
        k1    = str(rec.get("key_1") or "").strip()
        k2    = str(rec.get("key_2") or "").strip()
        if dtype in result and k1:
            result[dtype].add((k1, k2))

    wb.close()
    return result


def _load_tree(dataset: str) -> pd.DataFrame:
    path = TREE_DIR / f"{dataset}_tree.csv"
    if not path.exists():
        print(f"  [WARN] Tree file not found: {path}  -- run build_dataset_tree_structure.py first")
        return pd.DataFrame(columns=["dataset", "axis", "code", "label", "level",
                                     "parent_code", "is_subtotal"])
    return pd.read_csv(path)


def build_axis_lookup(tree: pd.DataFrame, axis: str) -> dict[str, bool]:
    """
    Return {code: is_subtotal} for one axis.  Missing code -> key absent.

    For trees that store hierarchical paths as slash-joined codes
    (e.g. ninth_tree: "09_total_transformation_sector/09_01_electricity_plants"),
    the lookup is indexed by BOTH the full path AND the final segment so that
    mapping sheets that store only the leaf code can still match.
    When both keys exist the full-path entry wins (it is more specific).
    """
    sub = tree[tree["axis"] == axis][["code", "is_subtotal"]].dropna(subset=["code"])
    result: dict[str, bool] = {}
    for _, row in sub.iterrows():
        full_code = str(row["code"]).strip()
        is_sub    = bool(row["is_subtotal"])
        result[full_code] = is_sub
        # Also index by last segment for slash-joined hierarchical codes
        if "/" in full_code:
            leaf_segment = full_code.rsplit("/", 1)[-1].strip()
            if leaf_segment not in result:   # full path takes priority
                result[leaf_segment] = is_sub
    return result


def build_rollup_subtotal_lookup(wb: Any) -> dict[str, dict[str, bool | str]]:
    """
    Read the Subtotal column from all three rollup sheets and return per-sheet
    lookups keyed by the rolled (output) label.

    Returns:
        {
            "esto":  {rolled_esto_flow: True/False/"MIXED", ...},
            "ninth": {rolled_ninth_sector: True/False/"MIXED", ...},
            "leap":  {rolled_leap_sector_name_full_path: True/False/"MIXED", ...},
        }

    When multiple input rows share the same output label with different Subtotal
    values (should not happen if the sheet is filled consistently), the first
    non-None value wins and a warning is printed.
    """
    configs = [
        ("esto_rollup_rules",  "rolled_esto_flow",                  "esto"),
        ("ninth_rollup_rules", "rolled_ninth_sector",                  "ninth"),
        ("leap_rollup_rules",  "rolled_leap_sector_name_full_path",  "leap"),
    ]
    result: dict[str, dict[str, bool | str]] = {}

    for sheet_name, key_col, axis in configs:
        lkp: dict[str, bool | str] = {}
        _, rows = _read_sheet_as_dicts(wb, sheet_name)
        for row in rows:
            key = str(row.get(key_col) or "").strip()
            if not key:
                continue
            sub_raw = row.get("Subtotal")
            if sub_raw is None:
                continue
            sub_val: bool | str = _read_subtotal(sub_raw)
            if key in lkp and lkp[key] != sub_val:
                print(f"  [WARN] {sheet_name}: inconsistent Subtotal for '{key}': "
                      f"{lkp[key]!r} vs {sub_val!r} — keeping first")
            else:
                lkp[key] = sub_val
        result[axis] = lkp

    return result


def _classify_input_set(flags: list[bool | None]) -> str:
    """
    Given a list of is_subtotal values for a group of rollup inputs
    (None = not found in tree), return a classification string.
    """
    known    = [v for v in flags if v is not None]
    n_unknown = sum(1 for v in flags if v is None)
    unknown_suffix = "_some_unknown" if n_unknown else ""

    if not known:
        return "all_unknown"

    n_sub  = sum(1 for v in known if v)
    n_leaf = sum(1 for v in known if not v)

    if n_sub > 0 and n_leaf > 0:
        return "MIXED_WARNING"           # always flag, even with unknowns

    return ("all_subtotal" if n_sub else "all_leaf") + unknown_suffix


# ---------------------------------------------------------------------------
# Rollup consistency analysis
# ---------------------------------------------------------------------------

def _read_sheet_as_dicts(wb: Any, sheet_name: str) -> tuple[list[str], list[dict]]:
    ws = wb[sheet_name]
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    rows = [
        {h: v for h, v in zip(headers, row)}
        for row in ws.iter_rows(min_row=2, values_only=True)
    ]
    return headers, rows


def analyze_rollup_consistency(
    wb: Any,
    esto_tree: pd.DataFrame,
    ninth_tree: pd.DataFrame,
    leap_tree: pd.DataFrame,
) -> pd.DataFrame:
    """
    For each rollup sheet, group rules by their output label and classify the
    input set's subtotal consistency.  Flags MIXED_WARNING where inputs span
    leaf and subtotal levels.

    Also looks up the output label in the relevant tree to get a
    tree-derived subtotal status for the output (None if not in tree).
    """
    esto_flow_lkp  = build_axis_lookup(esto_tree,  "flow")
    esto_prod_lkp  = build_axis_lookup(esto_tree,  "product")
    ninth_sec_lkp  = build_axis_lookup(ninth_tree, "sector")
    ninth_fuel_lkp = build_axis_lookup(ninth_tree, "fuel")
    leap_sec_lkp   = build_axis_lookup(leap_tree,  "sector")
    leap_fuel_lkp  = build_axis_lookup(leap_tree,  "fuel")

    sheet_configs = [
        (
            "leap_rollup_rules",
            "input_leap_sector_name_full_path", leap_sec_lkp,
            "input_raw_leap_fuel_name",         leap_fuel_lkp,
            "rolled_leap_sector_name_full_path", leap_sec_lkp,
            "rolled_raw_leap_fuel_name",         leap_fuel_lkp,
        ),
        (
            "esto_rollup_rules",
            "input_esto_flow",    esto_flow_lkp,
            "input_esto_product", esto_prod_lkp,
            "rolled_esto_flow",    esto_flow_lkp,
            "rolled_esto_product", esto_prod_lkp,
        ),
        (
            "ninth_rollup_rules",
            "input_ninth_sector", ninth_sec_lkp,
            "input_ninth_fuel",   ninth_fuel_lkp,
            "rolled_ninth_sector", ninth_sec_lkp,
            "rolled_ninth_fuel",   ninth_fuel_lkp,
        ),
    ]

    records: list[dict] = []

    for (sheet_name,
         in_primary_col, in_primary_lkp,
         in_secondary_col, in_secondary_lkp,
         out_primary_col, out_primary_lkp,
         out_secondary_col, out_secondary_lkp) in sheet_configs:

        _, rows = _read_sheet_as_dicts(wb, sheet_name)

        # Group active rows by (rolled_primary, rolled_secondary)
        groups: dict[tuple[str, str], list[tuple[str, str]]] = {}
        for row in rows:
            in_p = str(row.get(in_primary_col) or "").strip()
            if not in_p:
                continue
            if not _truthy(row.get("include", True)):
                continue
            in_s  = str(row.get(in_secondary_col) or "").strip()
            out_p = str(row.get(out_primary_col) or "").strip()
            out_s = str(row.get(out_secondary_col) or "").strip()
            key = (out_p, out_s)
            groups.setdefault(key, []).append((in_p, in_s))

        for (out_p, out_s), inputs in groups.items():
            # Input classification on primary axis (the most semantically important)
            in_primary_flags: list[bool | None] = [
                in_primary_lkp.get(inp) for inp, _ in inputs
            ]
            consistency = _classify_input_set(in_primary_flags)

            # Secondary axis consistency (informational)
            in_sec_flags: list[bool | None] = [
                in_secondary_lkp.get(sec) if sec else None for _, sec in inputs
            ]
            consistency_secondary = _classify_input_set(in_sec_flags) if any(
                s for _, s in inputs
            ) else "not_applicable"

            # Tree lookup for the output
            out_tree_primary   = out_primary_lkp.get(out_p)         # None = not in tree
            out_tree_secondary = out_secondary_lkp.get(out_s) if out_s else None

            # Manual override for ESTO flow outputs
            manual_override_note = ""
            if sheet_name == "esto_rollup_rules" and out_p in ESTO_FLOW_MANUAL_SUBTOTAL_OVERRIDES:
                out_tree_primary = True
                manual_override_note = "ESTO_FLOW_MANUAL_OVERRIDE"

            in_tree_labels  = [k for k in in_primary_lkp if k in {i for i, _ in inputs}]
            unknown_inputs  = [i for i, _ in inputs if i not in in_primary_lkp]

            records.append({
                "sheet":                      sheet_name,
                "rolled_primary":             out_p,
                "rolled_secondary":           out_s,
                "rolled_primary_in_tree":     out_p in out_primary_lkp,
                "rolled_primary_tree_is_subtotal": out_tree_primary,
                "manual_override":            manual_override_note,
                "input_count":                len(inputs),
                "input_primary_labels":       " | ".join(i for i, _ in inputs),
                "unknown_input_primaries":    " | ".join(unknown_inputs) if unknown_inputs else "",
                "input_primary_consistency":  consistency,
                "input_secondary_consistency": consistency_secondary,
                "has_warning":               "MIXED_WARNING" in consistency or "MIXED_WARNING" in consistency_secondary,
            })

    return pd.DataFrame(records).sort_values(
        ["has_warning", "sheet", "rolled_primary"], ascending=[False, True, True]
    ).reset_index(drop=True)


# ---------------------------------------------------------------------------
# Mapping sheet draft: esto_pair_is_subtotal
# ---------------------------------------------------------------------------

def _propose_esto_pair(
    esto_flow: str,
    esto_product: str,
    flow_lkp: dict[str, bool],
    prod_lkp: dict[str, bool],
    esto_rollup_lkp: dict[str, bool | str],
) -> tuple[bool | str | None, str]:
    """
    Return (proposed_is_subtotal, source).
    proposed may be True, False, "MIXED", or None (needs manual review).
    Priority: manual_override > tree > rollup_sheet.
    """
    if esto_flow in ESTO_FLOW_MANUAL_SUBTOTAL_OVERRIDES:
        return True, "manual_override"

    # Rollup sheet takes priority for rollup output labels.
    if esto_flow in esto_rollup_lkp:
        return esto_rollup_lkp[esto_flow], "rollup_sheet"

    flow_sub = flow_lkp.get(esto_flow)
    prod_sub = prod_lkp.get(esto_product)

    if flow_sub is True:
        return True, "tree_flow"
    if flow_sub is False:
        if prod_sub is True:
            return True, "tree_product"
        return False, "tree_flow"

    if prod_sub is True:
        return True, "tree_product_only"
    return None, "not_in_tree"


def generate_esto_pair_draft(
    wb: Any,
    esto_tree: pd.DataFrame,
    rollup_lkps: dict[str, dict[str, bool | str]],
    exceptions: dict[str, set[tuple[str, str]]] | None = None,
) -> pd.DataFrame:
    """
    For each unique (esto_flow, esto_product) pair across leap_combined_esto
    and ninth_pairs_to_esto_pairs, return a review table comparing the current
    esto_pair_is_subtotal value to the tree/rollup-derived proposal.
    """
    flow_lkp = build_axis_lookup(esto_tree, "flow")
    prod_lkp = build_axis_lookup(esto_tree, "product")
    esto_rollup_lkp = rollup_lkps.get("esto", {})

    seen: dict[tuple[str, str], dict] = {}

    for sheet_name, flow_col, prod_col, sub_col in [
        ("leap_combined_esto",      "esto_flow",  "esto_product", "esto_pair_is_subtotal"),
        ("ninth_pairs_to_esto_pairs", "esto_flow", "esto_product", "esto_pair_is_subtotal"),
    ]:
        _, rows = _read_sheet_as_dicts(wb, sheet_name)
        for row in rows:
            f = str(row.get(flow_col) or "").strip()
            p = str(row.get(prod_col) or "").strip()
            if not f or not p:
                continue
            key = (f, p)
            if key in seen:
                seen[key]["sheets"].add(sheet_name)
                continue
            current = _read_subtotal(row.get(sub_col))
            proposed, source = _propose_esto_pair(f, p, flow_lkp, prod_lkp, esto_rollup_lkp)
            excepted = exceptions is not None and key in exceptions.get("esto_pairs", set())
            if excepted:
                proposed, source = current, "accepted_exception"
            seen[key] = {
                "esto_flow":              f,
                "esto_product":           p,
                "sheets":                 {sheet_name},
                "current_is_subtotal":    current,
                "proposed_is_subtotal":   proposed,
                "proposal_source":        source,
                "disagrees":             _subtotal_disagrees(proposed, current),
                "needs_manual_review":   (proposed is None),
            }

    df = pd.DataFrame([
        {**rec, "sheets": " | ".join(sorted(rec["sheets"]))}
        for rec in seen.values()
    ]).sort_values(
        ["disagrees", "needs_manual_review", "esto_flow", "esto_product"],
        ascending=[False, False, True, True],
    ).reset_index(drop=True)
    return df


# ---------------------------------------------------------------------------
# Mapping sheet draft: ninth_pair_is_subtotal
# ---------------------------------------------------------------------------

def generate_ninth_pair_draft(
    wb: Any,
    ninth_tree: pd.DataFrame,
    rollup_lkps: dict[str, dict[str, bool | str]],
    exceptions: dict[str, set[tuple[str, str]]] | None = None,
) -> pd.DataFrame:
    """
    For each unique (ninth_sector, ninth_fuel) pair across leap_combined_ninth
    and ninth_pairs_to_esto_pairs, return a review table comparing the current
    ninth_pair_is_subtotal value to the tree/rollup-derived proposal.
    """
    sec_lkp  = build_axis_lookup(ninth_tree, "sector")
    fuel_lkp = build_axis_lookup(ninth_tree, "fuel")
    ninth_rollup_lkp = rollup_lkps.get("ninth", {})

    seen: dict[tuple[str, str], dict] = {}

    for sheet_name, sec_col, fuel_col, sub_col in [
        ("leap_combined_ninth",       "ninth_sector", "ninth_fuel", "ninth_pair_is_subtotal"),
        ("ninth_pairs_to_esto_pairs", "ninth_sector",   "ninth_fuel",  "ninth_pair_is_subtotal"),
    ]:
        _, rows = _read_sheet_as_dicts(wb, sheet_name)
        for row in rows:
            sec  = str(row.get(sec_col)  or "").strip()
            fuel = str(row.get(fuel_col) or "").strip()
            if not sec or not fuel:
                continue
            key = (sec, fuel)
            if key in seen:
                seen[key]["sheets"].add(sheet_name)
                continue

            current = _read_subtotal(row.get(sub_col))

            # Rollup sheet takes priority for rollup output labels.
            if sec in ninth_rollup_lkp:
                proposed, source = ninth_rollup_lkp[sec], "rollup_sheet"
            else:
                sec_sub  = sec_lkp.get(sec)
                fuel_sub = fuel_lkp.get(fuel)

                if sec_sub is True:
                    proposed, source = True, "tree_sector"
                elif sec_sub is False:
                    if fuel_sub is True:
                        proposed, source = True, "tree_fuel"
                    else:
                        proposed, source = False, "tree_sector"
                else:
                    if fuel_sub is True:
                        proposed, source = True, "tree_fuel_only"
                    else:
                        proposed, source = None, "not_in_tree"

            excepted = exceptions is not None and key in exceptions.get("ninth_pairs", set())
            if excepted:
                proposed, source = current, "accepted_exception"
            seen[key] = {
                "ninth_sector":          sec,
                "ninth_fuel":            fuel,
                "sheets":                {sheet_name},
                "current_is_subtotal":   current,
                "proposed_is_subtotal":  proposed,
                "proposal_source":       source,
                "disagrees":            _subtotal_disagrees(proposed, current),
                "needs_manual_review":  (proposed is None),
            }

    df = pd.DataFrame([
        {**rec, "sheets": " | ".join(sorted(rec["sheets"]))}
        for rec in seen.values()
    ]).sort_values(
        ["disagrees", "needs_manual_review", "ninth_sector", "ninth_fuel"],
        ascending=[False, False, True, True],
    ).reset_index(drop=True)
    return df


# ---------------------------------------------------------------------------
# Mapping sheet draft: leap_is_subtotal
# ---------------------------------------------------------------------------

def generate_leap_pair_draft(
    wb: Any,
    leap_tree: pd.DataFrame,
    rollup_lkps: dict[str, dict[str, bool | str]],
    exceptions: dict[str, set[tuple[str, str]]] | None = None,
) -> pd.DataFrame:
    """
    For each unique (leap_sector_name_full_path, raw_leap_fuel_name) pair across
    leap_combined_esto and leap_combined_ninth, return a review table comparing
    the current leap_is_subtotal value to the tree/rollup-derived proposal.

    Priority: tree_sector > tree_fuel > rollup_sheet > not_in_tree.
    Rollup sheet covers custom combined labels (e.g. "Power", "Transport") whose
    Subtotal column the user has explicitly set, including "MIXED".
    """
    sec_lkp  = build_axis_lookup(leap_tree, "sector")
    fuel_lkp = build_axis_lookup(leap_tree, "fuel")
    leap_rollup_lkp = rollup_lkps.get("leap", {})

    seen: dict[tuple[str, str], dict] = {}

    for sheet_name in ["leap_combined_esto", "leap_combined_ninth"]:
        _, rows = _read_sheet_as_dicts(wb, sheet_name)
        for row in rows:
            sec  = str(row.get("leap_sector_name_full_path") or "").strip()
            fuel = str(row.get("raw_leap_fuel_name")         or "").strip()
            if not sec:
                continue
            key = (sec, fuel)
            if key in seen:
                seen[key]["sheets"].add(sheet_name)
                continue

            current  = _read_subtotal(row.get("leap_is_subtotal"))

            # Rollup sheet takes priority for rollup output labels — the user
            # explicitly sets Subtotal there, overriding the structural tree.
            if sec in leap_rollup_lkp:
                proposed, source = leap_rollup_lkp[sec], "rollup_sheet"
            else:
                sec_sub  = sec_lkp.get(sec)
                fuel_sub = fuel_lkp.get(fuel) if fuel else None

                if sec_sub is True:
                    proposed, source = True, "tree_sector"
                elif sec_sub is False:
                    if fuel_sub is True:
                        proposed, source = True, "tree_fuel"
                    else:
                        proposed, source = False, "tree_sector"
                else:
                    if fuel_sub is True:
                        proposed, source = True, "tree_fuel_only"
                    else:
                        proposed, source = None, "not_in_tree"

            excepted = exceptions is not None and key in exceptions.get("leap_pairs", set())
            if excepted:
                proposed, source = current, "accepted_exception"
            seen[key] = {
                "leap_sector":           sec,
                "leap_fuel":             fuel,
                "sheets":                {sheet_name},
                "current_is_subtotal":   current,
                "proposed_is_subtotal":  proposed,
                "proposal_source":       source,
                "disagrees":            _subtotal_disagrees(proposed, current),
                "needs_manual_review":  (proposed is None),
            }

    df = pd.DataFrame([
        {**rec, "sheets": " | ".join(sorted(rec["sheets"]))}
        for rec in seen.values()
    ]).sort_values(
        ["disagrees", "needs_manual_review", "leap_sector", "leap_fuel"],
        ascending=[False, False, True, True],
    ).reset_index(drop=True)
    return df


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    import openpyxl

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    print("Loading trees...")
    esto_tree  = _load_tree("esto")
    ninth_tree = _load_tree("ninth")
    leap_tree  = _load_tree("leap")

    print("Loading exceptions...")
    exceptions = load_subtotal_exceptions()
    n_exc = sum(len(v) for v in exceptions.values())
    print(f"  {n_exc} accepted exceptions  "
          f"(esto:{len(exceptions['esto_pairs'])}  "
          f"ninth:{len(exceptions['ninth_pairs'])}  "
          f"leap:{len(exceptions['leap_pairs'])})")

    print("Loading workbook...")
    wb = openpyxl.load_workbook(str(WORKBOOK_PATH), read_only=True)

    # --- Rollup subtotal lookup (from explicit Subtotal columns) -----------
    print("Reading rollup sheet Subtotal columns...")
    rollup_lkps = build_rollup_subtotal_lookup(wb)
    for axis, lkp in rollup_lkps.items():
        n_mixed = sum(1 for v in lkp.values() if v == "MIXED")
        print(f"  {axis}: {len(lkp)} rollup outputs  ({n_mixed} MIXED)")

    # --- Rollup consistency ------------------------------------------------
    print("Analysing rollup rule consistency...")
    rollup_df = analyze_rollup_consistency(wb, esto_tree, ninth_tree, leap_tree)
    n_warn = int(rollup_df["has_warning"].sum())
    out_path = OUTPUT_DIR / "rollup_consistency.csv"
    rollup_df.to_csv(out_path, index=False)
    print(f"  Written: {out_path}  ({len(rollup_df)} rollup groups, {n_warn} with MIXED_WARNING)")

    # --- ESTO pair draft ---------------------------------------------------
    print("Generating esto_pair_is_subtotal draft...")
    esto_df = generate_esto_pair_draft(wb, esto_tree, rollup_lkps, exceptions)
    n_dis  = int(esto_df["disagrees"].sum())
    n_rev  = int(esto_df["needs_manual_review"].sum())
    out_path = OUTPUT_DIR / "subtotal_draft_esto_pairs.csv"
    esto_df.to_csv(out_path, index=False)
    print(f"  Written: {out_path}  ({len(esto_df)} pairs, {n_dis} disagree with current, {n_rev} need manual review)")

    # --- Ninth pair draft --------------------------------------------------
    print("Generating ninth_pair_is_subtotal draft...")
    ninth_df = generate_ninth_pair_draft(wb, ninth_tree, rollup_lkps, exceptions)
    n_dis  = int(ninth_df["disagrees"].sum())
    n_rev  = int(ninth_df["needs_manual_review"].sum())
    out_path = OUTPUT_DIR / "subtotal_draft_ninth_pairs.csv"
    ninth_df.to_csv(out_path, index=False)
    print(f"  Written: {out_path}  ({len(ninth_df)} pairs, {n_dis} disagree with current, {n_rev} need manual review)")

    # --- LEAP pair draft ---------------------------------------------------
    print("Generating leap_is_subtotal draft...")
    leap_df = generate_leap_pair_draft(wb, leap_tree, rollup_lkps, exceptions)
    n_dis  = int(leap_df["disagrees"].sum())
    n_rev  = int(leap_df["needs_manual_review"].sum())
    out_path = OUTPUT_DIR / "subtotal_draft_leap_pairs.csv"
    leap_df.to_csv(out_path, index=False)
    print(f"  Written: {out_path}  ({len(leap_df)} pairs, {n_dis} disagree with current, {n_rev} need manual review)")

    # Summary
    warnings = rollup_df[rollup_df["has_warning"]]
    if not warnings.empty:
        print("\n=== MIXED_WARNING rollup groups (investigate before assigning subtotal) ===")
        for _, r in warnings.iterrows():
            print(f"  [{r['sheet']}] {r['rolled_primary']} <- {r['input_primary_labels']}")

    print("\nDone.")


if __name__ == "__main__":
    main()
