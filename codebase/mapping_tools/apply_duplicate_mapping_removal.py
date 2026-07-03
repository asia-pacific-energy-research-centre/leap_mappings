"""
apply_duplicate_mapping_removal.py

Physically removes exact-duplicate mapping rows from
config/outlook_mappings_master.xlsx, keeping one copy of each duplicate group.

A duplicate is an *active* row (remove_row / duplicate_to_remove both falsey)
whose full key-pair -> key-pair tuple is identical to another active row on the
same sheet:

  leap_combined_esto          (leap_sector_name_full_path, raw_leap_fuel_name)
                              -> (esto_flow, esto_product)
  leap_combined_ninth         (leap_sector_name_full_path, raw_leap_fuel_name)
                              -> (ninth_sector, ninth_fuel)
  ninth_pairs_to_esto_pairs   (9th_sector, 9th_fuel) -> (esto_flow, esto_product)

Counterpart to the duplicate_mappings.csv report produced by
codebase/outlook_mapping_maintenance_workflow.py. Because the removed rows are
exact duplicates of a retained row, deleting them is lossless: the mapping the
row expressed is still present via the copy that is kept (the lowest workbook
row of each group).

Safety gates
------------
1.  Requires results/maintenance/duplicate_mappings.csv to exist — run the
    maintenance workflow first so the duplicates have been reported/reviewed.
2.  Rows to delete are re-derived from the live workbook (not trusted from the
    CSV row numbers), so the plan is correct even if the workbook was edited
    after the report was written.
3.  Dry-run summary printed first; nothing is written without an explicit "yes".
4.  Timestamped archive of the workbook is written before any deletion.
5.  Post-write verification re-scans the workbook and fails if any active
    duplicate remains.

Usage:
    python -m codebase.mapping_tools.apply_duplicate_mapping_removal [--dry-run]
"""

from __future__ import annotations

import argparse
import shutil
import sys
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[2]
WORKBOOK_PATH = REPO_ROOT / "config" / "outlook_mappings_master.xlsx"
ARCHIVE_DIR = REPO_ROOT / "config" / "archive"
DUPLICATE_CSV_PATH = REPO_ROOT / "results" / "maintenance" / "duplicate_mappings.csv"

# ---------------------------------------------------------------------------
# Sheet key layout: source pair + target pair used to define a duplicate.
# ---------------------------------------------------------------------------

SHEET_KEY_CONFIG: dict[str, list[str]] = {
    "leap_combined_esto": [
        "leap_sector_name_full_path", "raw_leap_fuel_name", "esto_flow", "esto_product",
    ],
    "leap_combined_ninth": [
        "leap_sector_name_full_path", "raw_leap_fuel_name", "ninth_sector", "ninth_fuel",
    ],
    "ninth_pairs_to_esto_pairs": [
        "9th_sector", "9th_fuel", "esto_flow", "esto_product",
    ],
}

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _norm(value: Any) -> str:
    """Strip and collapse whitespace; treat nan/none as blank (matches workflow)."""
    text = " ".join(str(value or "").split())
    return "" if text.lower() in {"nan", "none"} else text


def _truthy(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "t", "yes", "y", "on"}


def _header_index(ws: Any, header_name: str) -> int | None:
    """Return the 1-based column index of header_name in row 1, or None."""
    for cell in ws[1]:
        if _norm(cell.value) == _norm(header_name):
            return cell.column
    return None


@dataclass
class PlannedDeletion:
    sheet: str
    row: int                 # 1-based Excel row
    key: tuple[str, ...]
    kept_row: int            # the row of the copy being retained


@dataclass
class RemovalPlan:
    deletions: list[PlannedDeletion] = field(default_factory=list)
    missing_sheets: list[str] = field(default_factory=list)
    missing_columns: list[tuple[str, str]] = field(default_factory=list)  # (sheet, column)


# ---------------------------------------------------------------------------
# Plan build (re-derived from the live workbook)
# ---------------------------------------------------------------------------


def _scan_sheet_duplicates(ws: Any, sheet_name: str, key_cols: list[str], plan: RemovalPlan) -> None:
    """Mark all-but-the-first active copy of each duplicate key group for deletion."""
    key_idx = [_header_index(ws, col) for col in key_cols]
    remove_idx = _header_index(ws, "remove_row")
    dup_idx = _header_index(ws, "duplicate_to_remove")

    for col, idx in zip(key_cols, key_idx):
        if idx is None:
            plan.missing_columns.append((sheet_name, col))
    if any(idx is None for idx in key_idx):
        return

    seen: dict[tuple[str, ...], int] = {}   # key -> first retained row
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        key = tuple(_norm(row[i - 1].value) for i in key_idx)
        if any(part == "" for part in key):
            continue
        is_removed = remove_idx is not None and _truthy(row[remove_idx - 1].value)
        is_dup_removed = dup_idx is not None and _truthy(row[dup_idx - 1].value)
        if is_removed or is_dup_removed:
            continue  # already inactive — leave it alone
        if key in seen:
            plan.deletions.append(
                PlannedDeletion(sheet=sheet_name, row=row_idx, key=key, kept_row=seen[key])
            )
        else:
            seen[key] = row_idx


def build_plan(workbook_path: Path) -> RemovalPlan:
    plan = RemovalPlan()
    wb = openpyxl.load_workbook(str(workbook_path), read_only=True)
    try:
        for sheet_name, key_cols in SHEET_KEY_CONFIG.items():
            if sheet_name not in wb.sheetnames:
                plan.missing_sheets.append(sheet_name)
                continue
            _scan_sheet_duplicates(wb[sheet_name], sheet_name, key_cols, plan)
    finally:
        wb.close()
    return plan


# ---------------------------------------------------------------------------
# Summary
# ---------------------------------------------------------------------------


def _print_plan_summary(plan: RemovalPlan) -> None:
    print("\n" + "=" * 70)
    print("PLANNED DUPLICATE ROW DELETIONS")
    print("=" * 70)

    for sheet in plan.missing_sheets:
        print(f"  [WARN] sheet not found, skipped: {sheet}")
    for sheet, col in plan.missing_columns:
        print(f"  [WARN] {sheet}: key column not found, sheet skipped: {col}")

    if not plan.deletions:
        print("  No duplicate rows to delete.")
        return

    from collections import Counter
    by_sheet: Counter = Counter(d.sheet for d in plan.deletions)
    for sheet, count in sorted(by_sheet.items()):
        print(f"  {sheet}: {count} redundant row(s) to delete")
    print(f"\n  Total: {len(plan.deletions)} row(s) to delete "
          f"(one copy of each group is retained).")

    print("\nSAMPLE (first 20):")
    print(f"  {'Sheet':<28} {'DelRow':>6} {'KeepRow':>7}  Key")
    print(f"  {'-'*28} {'-'*6} {'-'*7}  {'-'*40}")
    for d in plan.deletions[:20]:
        key_str = " | ".join(d.key)[:60]
        print(f"  {d.sheet:<28} {d.row:>6} {d.kept_row:>7}  {key_str}")
    if len(plan.deletions) > 20:
        print(f"  ... and {len(plan.deletions) - 20} more")


# ---------------------------------------------------------------------------
# Archive + apply
# ---------------------------------------------------------------------------


def _archive(workbook_path: Path) -> Path:
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = ARCHIVE_DIR / f"{workbook_path.stem}.before_duplicate_removal_{ts}{workbook_path.suffix}"
    shutil.copy2(workbook_path, dest)
    print(f"\nArchived: {dest}")
    return dest


def apply_plan(plan: RemovalPlan, workbook_path: Path) -> int:
    """Delete the planned rows.  Returns count of rows deleted."""
    wb = openpyxl.load_workbook(str(workbook_path))
    deleted = 0
    # Delete per sheet, highest row first, so earlier row indices stay valid.
    from collections import defaultdict
    by_sheet: dict[str, list[int]] = defaultdict(list)
    for d in plan.deletions:
        by_sheet[d.sheet].append(d.row)
    for sheet_name, rows in by_sheet.items():
        ws = wb[sheet_name]
        for row_idx in sorted(rows, reverse=True):
            ws.delete_rows(row_idx, 1)
            deleted += 1
    wb.save(str(workbook_path))
    wb.close()
    return deleted


def verify(workbook_path: Path) -> int:
    """Re-scan and return the number of active duplicate rows still present."""
    plan = build_plan(workbook_path)
    return len(plan.deletions)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--dry-run", action="store_true",
                        help="Show planned deletions without writing anything")
    args = parser.parse_args()

    if not WORKBOOK_PATH.exists():
        print(f"[ERROR] Workbook not found: {WORKBOOK_PATH}")
        sys.exit(1)
    if not DUPLICATE_CSV_PATH.exists():
        print(f"[ERROR] Duplicate report not found: {DUPLICATE_CSV_PATH}")
        print("        Run codebase/outlook_mapping_maintenance_workflow.py first "
              "(apply mode) to generate and review it.")
        sys.exit(1)

    reported = pd.read_csv(DUPLICATE_CSV_PATH)
    print(f"Duplicate report: {len(reported):,} row(s) across "
          f"{reported['sheet_name'].nunique() if not reported.empty else 0} sheet(s) "
          f"-> {DUPLICATE_CSV_PATH}")

    plan = build_plan(WORKBOOK_PATH)
    _print_plan_summary(plan)

    if not plan.deletions:
        print("\nNothing to do.")
        return

    if args.dry_run:
        print("\n[DRY RUN] No changes written.")
        return

    print("\n" + "=" * 70)
    print("CONFIRMATION REQUIRED")
    print("=" * 70)
    print(f"About to delete {len(plan.deletions)} row(s) from:")
    print(f"  {WORKBOOK_PATH}")
    print("A timestamped archive will be created first.")
    answer = input("\nType 'yes' to proceed, anything else to abort: ").strip().lower()
    if answer != "yes":
        print("Aborted — no changes made.")
        sys.exit(0)

    _archive(WORKBOOK_PATH)

    print(f"\nDeleting {len(plan.deletions)} row(s)...")
    deleted = apply_plan(plan, WORKBOOK_PATH)
    print(f"  Deleted: {deleted} row(s)")

    print("\nVerifying...")
    remaining = verify(WORKBOOK_PATH)
    if remaining:
        print(f"  [ERROR] {remaining} active duplicate row(s) still remain — check above.")
        sys.exit(1)
    print("  No active duplicates remain. Done.")


if __name__ == "__main__":
    main()
