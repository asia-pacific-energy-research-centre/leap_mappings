#%%
"""Flip leap_is_subtotal False -> True for rows flagged by the M6 subtotal
mismatch check (results/maintenance/subtotal_mismatches.csv).

Each flagged row has a leaf-level LEAP source (leap_is_subtotal=False)
mapped to an aggregate target (esto_pair_is_subtotal/ninth_pair_is_subtotal
=True) where a more specific, non-subtotal target also exists. The fix is to
align the LEAP source side with the target: set leap_is_subtotal=True so the
pair is treated consistently as a subtotal-to-subtotal mapping.

Edits are made directly to config/outlook_mappings_master.xlsx (no exception
or override sheet routing) since Stage 0's lookup-based recompute-and-write
of these columns has been disabled -- direct edits now persist across
maintenance reruns.
"""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[2]
MAPPING_WORKBOOK_PATH = REPO_ROOT / "config" / "outlook_mappings_master.xlsx"
MISMATCHES_CSV_PATH = REPO_ROOT / "results" / "maintenance" / "subtotal_mismatches.csv"
ARCHIVE_DIR = REPO_ROOT / "config" / "archive"


def _archive(path: Path) -> Path:
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_path = ARCHIVE_DIR / f"{path.stem}.before_subtotal_source_flip_{timestamp}{path.suffix}"
    shutil.copy2(path, archive_path)
    return archive_path

SHEET_KEYS = {
    "leap_combined_esto": ("leap_sector_name_full_path", "raw_leap_fuel_name", "esto_flow", "esto_product"),
    "leap_combined_ninth": ("leap_sector_name_full_path", "raw_leap_fuel_name", "ninth_sector", "ninth_fuel"),
    "ninth_pairs_to_esto_pairs": ("9th_sector", "9th_fuel", "esto_flow", "esto_product"),
}
SOURCE_COLUMN = {
    "leap_combined_esto": "leap_is_subtotal",
    "leap_combined_ninth": "leap_is_subtotal",
    "ninth_pairs_to_esto_pairs": "ninth_pair_is_subtotal",
}


def _norm(value: object) -> str:
    if pd.isna(value):
        return ""
    return " ".join(str(value).split())


def _header_map(ws) -> dict[str, int]:
    return {_norm(cell.value): cell.column for cell in ws[1] if _norm(cell.value)}


def run(
    mapping_workbook_path: Path = MAPPING_WORKBOOK_PATH,
    mismatches_csv_path: Path = MISMATCHES_CSV_PATH,
) -> dict[str, object]:
    mismatches = pd.read_csv(mismatches_csv_path, dtype=object)
    if mismatches.empty:
        raise ValueError("subtotal_mismatches.csv is empty; nothing to apply.")

    wb = openpyxl.load_workbook(mapping_workbook_path)
    archive_path = _archive(mapping_workbook_path)

    sheet_header_maps = {
        sheet_name: _header_map(wb[sheet_name]) for sheet_name in SHEET_KEYS
    }
    sheet_row_indexes: dict[str, dict[tuple[str, ...], list[int]]] = {}
    for sheet_name, keys in SHEET_KEYS.items():
        ws = wb[sheet_name]
        headers = sheet_header_maps[sheet_name]
        key_cols = [headers[k] for k in keys]
        index: dict[tuple[str, ...], list[int]] = {}
        for row in range(2, ws.max_row + 1):
            key = tuple(_norm(ws.cell(row, col).value) for col in key_cols)
            index.setdefault(key, []).append(row)
        sheet_row_indexes[sheet_name] = index

    updated_rows = 0
    unmatched: list[str] = []
    already_true: list[str] = []

    for _, mismatch_row in mismatches.iterrows():
        sheet_name = _norm(mismatch_row["sheet"])
        keys = SHEET_KEYS[sheet_name]
        key = tuple(_norm(mismatch_row[k]) for k in keys)
        rows = sheet_row_indexes[sheet_name].get(key, [])
        if not rows:
            unmatched.append(f"{sheet_name}: {key}")
            continue

        ws = wb[sheet_name]
        headers = sheet_header_maps[sheet_name]
        source_col = headers[SOURCE_COLUMN[sheet_name]]
        for row in rows:
            cell = ws.cell(row, source_col)
            current = _norm(cell.value)
            if current.lower() == "true":
                already_true.append(f"{sheet_name} row {row}")
                continue
            cell.value = True
            updated_rows += 1

    if unmatched:
        preview = "\n".join(unmatched[:20])
        raise ValueError(f"{len(unmatched)} mismatch row(s) had no matching workbook row:\n{preview}")

    wb.save(mapping_workbook_path)

    result = {
        "mismatch_rows": len(mismatches),
        "updated_cells": updated_rows,
        "already_true_skipped": len(already_true),
        "archive": str(archive_path),
    }
    print(result)
    return result


#%%
if __name__ == "__main__":
    run()
