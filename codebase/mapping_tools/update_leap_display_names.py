"""
update_leap_display_names.py

Keeps the leap_display_names sheet of outlook_mappings_master.xlsx in sync
with the LEAP name information embedded in leap_combined_esto and
leap_combined_ninth.

For each external code the "implied LEAP name" is derived as:
    esto_product  / ninth_fuel   -> raw_leap_fuel_name
    esto_flow     / ninth_sector -> last segment of leap_sector_name_full_path

Rules applied
-------------
1.  Code has a single unique LEAP name that differs from the auto-stripped
    code  -> propose an entry in leap_display_names with
             matches_original_product_flow_name = False.
2.  Code has a single unique LEAP name that matches the auto-stripped code
    -> if a display-names entry exists, propose matches_original = True
       (no override needed); otherwise no action.
3.  Code maps to multiple different LEAP names (many-to-one)  ->  flagged as
    a conflict; no proposal.
4.  Code is in leap_display_names but absent from both combined sheets  ->
    flagged as a potential issue in the QA output; entry is not removed.

This module NEVER writes to the master workbook.  Proposed additions/updates
are reported to a review CSV; the user reviews them, sets the approval column,
and applies the approved rows with the standalone helper
``codebase.mapping_tools.apply_display_name_updates``.

QA / review output
------------------
results/maintenance/display_names_qa.csv               — full QA of every code
results/maintenance/display_names_proposed_updates.csv — proposed writes to review
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

REPO_ROOT = Path(__file__).resolve().parents[2]
if str(REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(REPO_ROOT))

WORKBOOK_PATH = REPO_ROOT / "config" / "outlook_mappings_master.xlsx"
EXCEPTION_WORKBOOK_PATH = REPO_ROOT / "config" / "mapping_issue_exception_sets.xlsx"
QA_DIR = REPO_ROOT / "results" / "maintenance"

# Review CSV listing proposed leap_display_names writes. run_display_name_update
# only reports here; the standalone apply_display_name_updates helper writes the
# approved rows into the master.
PROPOSED_UPDATES_FILENAME = "display_names_proposed_updates.csv"
PROPOSED_UPDATES_APPROVAL_COL = "CAN BE INSERTED INTO MAPPINGS"
PROPOSED_UPDATES_COLUMNS = [
    "code_type",
    "code",
    "auto_name",
    "current_leap_display_name",
    "proposed_leap_display_name",
    "proposed_matches_original_product_flow_name",
    "status",
    "note",
    PROPOSED_UPDATES_APPROVAL_COL,
]

_NUMERIC_PREFIX_RE = re.compile(r"^[\d.]+\s+")


# ---------------------------------------------------------------------------
# Helpers (mirrored from maintenance workflow for self-contained use)
# ---------------------------------------------------------------------------

def _norm(value: object) -> str:
    text = " ".join(str(value or "").split())
    return "" if text.lower() in {"nan", "none"} else text


def _truthy(value: object) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "t", "yes", "y", "on"}


def _strip_numeric_prefix(code: str) -> str:
    return _NUMERIC_PREFIX_RE.sub("", code).strip()


def _branch_end(full_path: str) -> str:
    """Return the last slash-separated segment of a LEAP branch path."""
    parts = [p.strip() for p in full_path.split("/") if p.strip()]
    return parts[-1] if parts else ""


def _active_rows(df: pd.DataFrame) -> pd.DataFrame:
    remove = df.get("remove_row", pd.Series(False, index=df.index)).map(_truthy)
    duplicate = df.get("duplicate_to_remove", pd.Series(False, index=df.index)).map(
        _truthy
    )
    return df[~(remove | duplicate)].copy()


# ---------------------------------------------------------------------------
# Core: derive implied LEAP names from combined sheets
# ---------------------------------------------------------------------------

def _build_implied_leap_names(
    leap_esto_df: pd.DataFrame,
    leap_ninth_df: pd.DataFrame,
) -> dict[tuple[str, str], set[str]]:
    """
    Build (code_type, code) -> set[implied LEAP name] from active rows of
    both combined sheets.

    esto_product  : raw_leap_fuel_name
    esto_flow     : last segment of leap_sector_name_full_path
    ninth_fuel    : raw_leap_fuel_name
    ninth_sector  : last segment of leap_sector_name_full_path
    """
    implied: dict[tuple[str, str], set[str]] = {}

    for _, row in _active_rows(leap_esto_df).iterrows():
        product = _norm(row.get("esto_product", ""))
        fuel = _norm(row.get("raw_leap_fuel_name", ""))
        if product and fuel:
            implied.setdefault(("esto_product", product), set()).add(fuel)

        flow = _norm(row.get("esto_flow", ""))
        path = _norm(row.get("leap_sector_name_full_path", ""))
        end = _branch_end(path)
        if flow and end:
            implied.setdefault(("esto_flow", flow), set()).add(end)

    for _, row in _active_rows(leap_ninth_df).iterrows():
        fuel_ninth = _norm(row.get("ninth_fuel", ""))
        fuel_leap = _norm(row.get("raw_leap_fuel_name", ""))
        if fuel_ninth and fuel_leap:
            implied.setdefault(("ninth_fuel", fuel_ninth), set()).add(fuel_leap)

        sector = _norm(row.get("ninth_sector", ""))
        path = _norm(row.get("leap_sector_name_full_path", ""))
        end = _branch_end(path)
        if sector and end:
            implied.setdefault(("ninth_sector", sector), set()).add(end)

    return implied


# ---------------------------------------------------------------------------
# Workbook writer
# ---------------------------------------------------------------------------

def _read_display_names_sheet(wb) -> tuple[list[str], list[list]]:
    """Return (headers, rows_data) for the leap_display_names sheet.

    rows_data is a list-of-lists; each inner list matches headers by position.
    Working at the list level avoids issues with duplicate column names.
    """
    ws = wb["leap_display_names"]
    headers = [_norm(cell.value) for cell in ws[1]]
    rows_data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows_data.append([_norm(v) for v in row])
    return headers, rows_data


def _apply_updates_to_workbook(
    wb,
    headers: list[str],
    rows_data: list[list],
    updates: list[tuple[tuple[str, str], str, str, bool]],
) -> None:
    """
    Apply (key, auto_name, leap_display_name, matches_original) updates to
    rows_data in-place, then rewrite the leap_display_names sheet in wb.

    rows_data is a list-of-lists (one list per data row, same order as headers).
    """

    def _hcol(name: str) -> int | None:
        return headers.index(name) if name in headers else None

    auto_name_col = _hcol("auto_name")
    display_col = _hcol("leap_display_name")
    matches_col = _hcol("matches_original_product_flow_name")
    code_type_col = _hcol("code_type")
    code_col = _hcol("code")

    # Build lookup: (code_type, code) -> row index into rows_data
    key_to_idx: dict[tuple[str, str], int] = {}
    for i, row in enumerate(rows_data):
        ct = row[code_type_col] if code_type_col is not None else ""
        code = row[code_col] if code_col is not None else ""
        if ct and code:
            key_to_idx[(ct, code)] = i

    for key, auto_name, leap_display_name, matches_original in updates:
        cur_idx = key_to_idx.get(key)
        if cur_idx is not None:
            row = rows_data[cur_idx]
            if auto_name_col is not None:
                row[auto_name_col] = auto_name
            if display_col is not None:
                row[display_col] = leap_display_name
            if matches_col is not None:
                row[matches_col] = str(matches_original)
        else:
            new_row = [""] * len(headers)
            if code_type_col is not None:
                new_row[code_type_col] = key[0]
            if code_col is not None:
                new_row[code_col] = key[1]
            if auto_name_col is not None:
                new_row[auto_name_col] = auto_name
            if display_col is not None:
                new_row[display_col] = leap_display_name
            if matches_col is not None:
                new_row[matches_col] = str(matches_original)
            rows_data.append(new_row)

    # Rewrite the sheet
    ws = wb["leap_display_names"]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for row in rows_data:
        ws.append(row)


# ---------------------------------------------------------------------------
# Exception list loader
# ---------------------------------------------------------------------------

def _load_exceptions(
    exception_workbook_path: Path = EXCEPTION_WORKBOOK_PATH,
) -> set[tuple[str, str]]:
    """
    Return a set of (code_type, code) pairs from the display_names_exceptions
    sheet that have enabled=True.  These are skipped on future update passes.
    """
    if not exception_workbook_path.exists():
        return set()
    try:
        df = pd.read_excel(
            exception_workbook_path,
            sheet_name="display_names_exceptions",
            dtype=object,
        ).fillna("")
    except Exception:
        return set()
    excluded: set[tuple[str, str]] = set()
    for _, row in df.iterrows():
        if not _truthy(row.get("enabled", "")):
            continue
        ct = _norm(row.get("code_type", ""))
        code = _norm(row.get("code", ""))
        if ct and code:
            excluded.add((ct, code))
    return excluded


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def run_display_name_update(
    workbook_path: Path = WORKBOOK_PATH,
    qa_dir: Path = QA_DIR,
    exception_workbook_path: Path = EXCEPTION_WORKBOOK_PATH,
) -> pd.DataFrame:
    """
    Check and update leap_display_names against the combined mapping sheets.

    Returns a QA DataFrame summarising every code examined.
    """
    exceptions = _load_exceptions(exception_workbook_path)

    # ── Load combined sheets ────────────────────────────────────────────────
    leap_esto_df = pd.read_excel(
        workbook_path, sheet_name="leap_combined_esto", dtype=object
    ).fillna("")
    leap_ninth_df = pd.read_excel(
        workbook_path, sheet_name="leap_combined_ninth", dtype=object
    ).fillna("")
    for df in [leap_esto_df, leap_ninth_df]:
        for col in df.columns:
            df[col] = df[col].astype(str).map(_norm)

    implied = _build_implied_leap_names(leap_esto_df, leap_ninth_df)

    # ── Load current leap_display_names ─────────────────────────────────────
    wb = openpyxl.load_workbook(workbook_path)
    headers, rows_data = _read_display_names_sheet(wb)

    def _hcol(name: str) -> int | None:
        return headers.index(name) if name in headers else None

    ct_col = _hcol("code_type")
    code_col = _hcol("code")
    display_name_col = _hcol("leap_display_name")
    matches_flag_col = _hcol("matches_original_product_flow_name")

    current_lookup: dict[tuple[str, str], int] = {}
    for i, row in enumerate(rows_data):
        ct = row[ct_col] if ct_col is not None else ""
        code = row[code_col] if code_col is not None else ""
        if ct and code:
            current_lookup[(ct, code)] = i

    # ── Evaluate every code ─────────────────────────────────────────────────
    all_codes = set(implied.keys()) | set(current_lookup.keys())
    qa_rows: list[dict] = []
    updates: list[tuple[tuple[str, str], str, str, bool]] = []

    for key in sorted(all_codes, key=lambda k: (k[0], k[1].lower())):
        code_type, code = key

        if key in exceptions:
            qa_rows.append({
                "code_type": code_type,
                "code": code,
                "auto_name": _strip_numeric_prefix(code),
                "current_leap_display_name": "",
                "implied_leap_name": " | ".join(sorted(implied.get(key, set()))),
                "status": "skipped_exception",
                "note": "In display_names_exceptions — skipped",
            })
            continue

        auto_name = _strip_numeric_prefix(code)

        cur_idx = current_lookup.get(key)
        cur_row = rows_data[cur_idx] if cur_idx is not None else None
        cur_display = (
            cur_row[display_name_col]
            if cur_row is not None and display_name_col is not None
            else ""
        )
        cur_matches_flag = (
            cur_row[matches_flag_col]
            if cur_row is not None and matches_flag_col is not None
            else ""
        )

        leap_names = implied.get(key, set())

        if not leap_names:
            # In leap_display_names but not referenced in either combined sheet
            qa_rows.append({
                "code_type": code_type,
                "code": code,
                "auto_name": auto_name,
                "current_leap_display_name": cur_display,
                "implied_leap_name": "",
                "status": "potential_issue_orphan",
                "note": (
                    "Code is in leap_display_names but not found in "
                    "leap_combined_esto or leap_combined_ninth"
                ),
            })
            continue

        implied_names_str = " | ".join(sorted(leap_names))

        if len(leap_names) > 1:
            qa_rows.append({
                "code_type": code_type,
                "code": code,
                "auto_name": auto_name,
                "current_leap_display_name": cur_display,
                "implied_leap_name": implied_names_str,
                "status": "conflict_multiple_leap_names",
                "note": "Maps to multiple different LEAP names — manual review required",
            })
            continue

        unique_leap_name = next(iter(leap_names))
        matches_auto = unique_leap_name == auto_name

        if matches_auto:
            if cur_idx is not None and not _truthy(cur_matches_flag):
                # Entry exists but not flagged as matching — update the flag
                updates.append((key, auto_name, unique_leap_name, True))
                qa_rows.append({
                    "code_type": code_type,
                    "code": code,
                    "auto_name": auto_name,
                    "current_leap_display_name": cur_display,
                    "implied_leap_name": implied_names_str,
                    "status": "updated_matches_auto_flag",
                    "note": "LEAP name matches auto-name; set matches_original_product_flow_name=True",
                })
            else:
                qa_rows.append({
                    "code_type": code_type,
                    "code": code,
                    "auto_name": auto_name,
                    "current_leap_display_name": cur_display,
                    "implied_leap_name": implied_names_str,
                    "status": "ok_auto_name",
                    "note": "",
                })
        else:
            # LEAP name differs from auto-name — needs an explicit entry
            if cur_idx is None:
                updates.append((key, auto_name, unique_leap_name, False))
                qa_rows.append({
                    "code_type": code_type,
                    "code": code,
                    "auto_name": auto_name,
                    "current_leap_display_name": "",
                    "implied_leap_name": implied_names_str,
                    "status": "added",
                    "note": f"New entry: leap_display_name set to {unique_leap_name!r}",
                })
            elif cur_display == unique_leap_name:
                qa_rows.append({
                    "code_type": code_type,
                    "code": code,
                    "auto_name": auto_name,
                    "current_leap_display_name": cur_display,
                    "implied_leap_name": implied_names_str,
                    "status": "consistent",
                    "note": "",
                })
            else:
                updates.append((key, auto_name, unique_leap_name, False))
                qa_rows.append({
                    "code_type": code_type,
                    "code": code,
                    "auto_name": auto_name,
                    "current_leap_display_name": cur_display,
                    "implied_leap_name": implied_names_str,
                    "status": "updated",
                    "note": (
                        f"leap_display_name changed from {cur_display!r} "
                        f"to {unique_leap_name!r}"
                    ),
                })

    # ── Report proposed updates (no workbook write) ─────────────────────────
    # This module never edits the master.  Proposed additions/updates are
    # written to a review CSV; approved rows are applied later with the
    # standalone apply_display_name_updates helper.
    wb.close()
    if updates:
        print(
            f"  leap_display_names: {len(updates)} proposed update(s) reported "
            "for review (workbook NOT modified)"
        )
    else:
        print("  leap_display_names: no updates proposed")

    # ── Duplicate display-name check ────────────────────────────────────────
    # For each (code_type, code), determine the effective display name that
    # will be used after this run, then flag any that are shared across
    # different codes within the same code_type.
    _EFFECTIVE_STATUSES = {
        "ok_auto_name", "consistent", "added", "updated",
        "updated_matches_auto_flag",
    }

    # Build (code_type, effective_display_name) -> list[code]
    from collections import defaultdict
    effective: dict[tuple[str, str], list[str]] = defaultdict(list)
    for row in qa_rows:
        if row["status"] not in _EFFECTIVE_STATUSES:
            continue
        ct = row["code_type"]
        # implied_leap_name is the resolved single name for these statuses;
        # fall back to auto_name when implied is blank (ok_auto_name case)
        display = row["implied_leap_name"] or row["auto_name"]
        if ct and display:
            effective[(ct, display)].append(row["code"])

    duplicate_keys: set[tuple[str, str]] = {
        (ct, name)
        for (ct, name), codes in effective.items()
        if len(codes) > 1
    }

    # Annotate qa_rows with duplicate flag
    for row in qa_rows:
        ct = row["code_type"]
        display = row["implied_leap_name"] or row["auto_name"]
        is_dup = (ct, display) in duplicate_keys
        row["duplicate_display_name"] = is_dup
        if is_dup:
            sharing = [c for c in effective[(ct, display)] if c != row["code"]]
            row["note"] = (
                (row["note"] + " | " if row["note"] else "")
                + f"duplicate_display_name: also used by {sharing}"
            )

    n_dups = sum(1 for r in qa_rows if r.get("duplicate_display_name"))
    if n_dups:
        print(f"  WARNING: {n_dups} code(s) share a display name within their code_type")

    # ── Proposed-updates review CSV ─────────────────────────────────────────
    # One row per proposed write, with a blank approval column for the user.
    qa_dir.mkdir(parents=True, exist_ok=True)
    qa_by_key = {(r["code_type"], r["code"]): r for r in qa_rows}
    proposed_rows = []
    for key, auto_name, leap_display_name, matches_original in updates:
        qa_row = qa_by_key.get(key, {})
        proposed_rows.append({
            "code_type": key[0],
            "code": key[1],
            "auto_name": auto_name,
            "current_leap_display_name": qa_row.get("current_leap_display_name", ""),
            "proposed_leap_display_name": leap_display_name,
            "proposed_matches_original_product_flow_name": matches_original,
            "status": qa_row.get("status", ""),
            "note": qa_row.get("note", ""),
            PROPOSED_UPDATES_APPROVAL_COL: "",
        })
    proposed_df = pd.DataFrame(proposed_rows, columns=PROPOSED_UPDATES_COLUMNS)
    proposed_path = qa_dir / PROPOSED_UPDATES_FILENAME
    try:
        proposed_df.to_csv(proposed_path, index=False)
        if updates:
            print(
                f"  proposed display-name updates -> {proposed_path.relative_to(REPO_ROOT)} "
                f"(set '{PROPOSED_UPDATES_APPROVAL_COL}'=TRUE, then run "
                "apply_display_name_updates)"
            )
    except PermissionError:
        fallback = proposed_path.with_stem(proposed_path.stem + "_new")
        proposed_df.to_csv(fallback, index=False)
        print(f"  proposed display-name updates (locked, wrote to fallback) -> {fallback.relative_to(REPO_ROOT)}")

    # ── QA output ───────────────────────────────────────────────────────────
    qa_df = pd.DataFrame(qa_rows)
    qa_path = qa_dir / "display_names_qa.csv"
    _CSV_EXCLUDED_STATUSES = {"skipped_exception", "consistent", "ok_auto_name"}
    qa_csv_df = (
        qa_df[~qa_df["status"].isin(_CSV_EXCLUDED_STATUSES)]
        if not qa_df.empty
        else qa_df
    )
    try:
        qa_csv_df.to_csv(qa_path, index=False)
        print(f"  leap_display_names QA -> {qa_path.relative_to(REPO_ROOT)}")
    except PermissionError:
        fallback = qa_path.with_stem(qa_path.stem + "_new")
        qa_csv_df.to_csv(fallback, index=False)
        print(f"  leap_display_names QA (locked, wrote to fallback) -> {fallback.relative_to(REPO_ROOT)}")

    status_counts = qa_df["status"].value_counts().to_dict() if not qa_df.empty else {}
    for status, count in sorted(status_counts.items()):
        print(f"    {status}: {count}")

    return qa_df


def check_display_name_issues(
    workbook_path: Path = WORKBOOK_PATH,
    exception_workbook_path: Path = EXCEPTION_WORKBOOK_PATH,
) -> dict[str, int]:
    """
    Read-only check of leap_display_names consistency.

    Derives implied LEAP names from the combined sheets and returns a summary
    dict of issue counts.  Does NOT write to the workbook or produce a QA CSV.

    Keys in the returned dict:
        orphan_count       — codes in leap_display_names not in any combined sheet
        duplicate_count    — codes whose display name is shared within a code_type
    """
    leap_esto_df = pd.read_excel(
        workbook_path, sheet_name="leap_combined_esto", dtype=object
    ).fillna("")
    leap_ninth_df = pd.read_excel(
        workbook_path, sheet_name="leap_combined_ninth", dtype=object
    ).fillna("")
    for df in [leap_esto_df, leap_ninth_df]:
        for col in df.columns:
            df[col] = df[col].astype(str).map(_norm)

    implied = _build_implied_leap_names(leap_esto_df, leap_ninth_df)
    exceptions = _load_exceptions(exception_workbook_path)

    wb = openpyxl.load_workbook(workbook_path, read_only=True)
    headers, rows_data = _read_display_names_sheet(wb)
    wb.close()

    def _hcol(name: str) -> int | None:
        return headers.index(name) if name in headers else None

    ct_col = _hcol("code_type")
    code_col = _hcol("code")
    display_col = _hcol("leap_display_name")

    # Build current entries
    current_entries: list[tuple[str, str, str]] = []
    for row in rows_data:
        ct = row[ct_col] if ct_col is not None else ""
        code = row[code_col] if code_col is not None else ""
        display = row[display_col] if display_col is not None else ""
        if ct and code:
            current_entries.append((ct, code, display))

    current_keys = {(ct, code) for ct, code, _ in current_entries}

    # Orphans: in leap_display_names but not in combined sheets (excluding exceptions)
    orphan_count = sum(
        1 for (ct, code, _) in current_entries
        if (ct, code) not in implied and (ct, code) not in exceptions
    )

    # Duplicate display names within a code_type (excluding exceptions)
    from collections import defaultdict
    by_name: dict[tuple[str, str], list[str]] = defaultdict(list)
    for key, names in implied.items():
        if key in exceptions or len(names) != 1:
            continue
        ct, code = key
        leap_name = next(iter(names))
        by_name[(ct, leap_name)].append(code)

    duplicate_count = sum(
        len(codes)
        for codes in by_name.values()
        if len(codes) > 1
    )

    return {"orphan_count": orphan_count, "duplicate_count": duplicate_count}


# ---------------------------------------------------------------------------
# Script entry-point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    qa = run_display_name_update()
    n_issues = (
        qa["status"]
        .isin(["potential_issue_orphan"])
        .sum()
        if not qa.empty
        else 0
    )
    n_dups = int(qa["duplicate_display_name"].sum()) if "duplicate_display_name" in qa.columns else 0
    print(f"\nDone. {n_issues} orphan(s), {n_dups} duplicate display name(s).")
