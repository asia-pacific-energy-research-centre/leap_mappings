#%%
"""Apply reviewed subtotal suggestions and synchronize subtotal exceptions.

Rows marked INSERT=True receive the suggested subtotal value in both mapped
systems. Rows marked INSERT=False retain their workbook values and become the
complete subtotal_mismatch_allowed exception set. Existing subtotal exceptions
not represented by a False review decision are removed.
"""

from __future__ import annotations

import shutil
from datetime import datetime
from pathlib import Path

import openpyxl
import pandas as pd

#%%
# --- Stable paths and sheet definitions ---

REPO_ROOT = Path(__file__).resolve().parents[2]
MAPPING_WORKBOOK_PATH = REPO_ROOT / "config" / "outlook_mappings_master.xlsx"
EXCEPTION_WORKBOOK_PATH = REPO_ROOT / "config" / "mapping_issue_exception_sets.xlsx"
REVIEW_CSV_PATH = REPO_ROOT / "results" / "maintenance" / "subtotal_mismatch_suggested_improvements.csv"
ARCHIVE_DIR = REPO_ROOT / "config" / "archive"
EXCEPTION_SHEET = "subtotal_mismatch_allowed"
OVERRIDE_SHEET = "subtotal_label_overrides"

OVERRIDE_HEADERS = [
    "enabled",
    "sheet",
    "leap_sector_name_full_path",
    "raw_leap_fuel_name",
    "ninth_sector",
    "ninth_fuel",
    "9th_sector",
    "9th_fuel",
    "esto_flow",
    "esto_product",
    "leap_is_subtotal",
    "ninth_pair_is_subtotal",
    "esto_pair_is_subtotal",
    "notes",
]

SHEET_CONFIGS = {
    "leap_combined_ninth": {
        "keys": ("leap_sector_name_full_path", "raw_leap_fuel_name", "ninth_sector", "ninth_fuel"),
        "subtotal_columns": ("leap_is_subtotal", "ninth_pair_is_subtotal"),
    },
    "leap_combined_esto": {
        "keys": ("leap_sector_name_full_path", "raw_leap_fuel_name", "esto_flow", "esto_product"),
        "subtotal_columns": ("leap_is_subtotal", "esto_pair_is_subtotal"),
    },
    "ninth_pairs_to_esto_pairs": {
        "keys": ("9th_sector", "9th_fuel", "esto_flow", "esto_product"),
        "subtotal_columns": ("ninth_pair_is_subtotal", "esto_pair_is_subtotal"),
    },
}


#%%
# --- Helpers ---

def _normalise_text(value: object) -> str:
    if pd.isna(value):
        return ""
    return " ".join(str(value).split())


def _parse_decision(value: object) -> bool:
    text = _normalise_text(value).upper()
    if text == "TRUE":
        return True
    if text == "FALSE":
        return False
    raise ValueError(f"INSERT must be TRUE or FALSE, received {value!r}")


def _parse_bool(value: object) -> bool | None:
    text = _normalise_text(value).lower()
    if not text:
        return None
    if text in {"true", "1", "yes", "y"}:
        return True
    if text in {"false", "0", "no", "n"}:
        return False
    raise ValueError(f"Expected a boolean or blank, received {value!r}")


def _header_map(ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    return {
        _normalise_text(cell.value): cell.column
        for cell in ws[1]
        if _normalise_text(cell.value)
    }


def _archive(path: Path) -> Path:
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    archive_path = ARCHIVE_DIR / f"{path.stem}.before_subtotal_review_{timestamp}{path.suffix}"
    shutil.copy2(path, archive_path)
    return archive_path


def _validate_review(review: pd.DataFrame) -> None:
    if review.empty:
        raise ValueError("The subtotal review CSV is empty.")
    if review.duplicated(["sheet_name", "workbook_row"]).any():
        raise ValueError("Duplicate (sheet_name, workbook_row) rows found in review CSV.")
    review["INSERT"].map(_parse_decision)


def _validate_workbook_rows(
    workbook: openpyxl.Workbook,
    review: pd.DataFrame,
) -> None:
    errors: list[str] = []
    for _, review_row in review.iterrows():
        sheet_name = _normalise_text(review_row["sheet_name"])
        config = SHEET_CONFIGS.get(sheet_name)
        if config is None or sheet_name not in workbook.sheetnames:
            errors.append(f"Unknown sheet {sheet_name!r}")
            continue
        ws = workbook[sheet_name]
        headers = _header_map(ws)
        workbook_row = int(review_row["workbook_row"])
        for key_column in config["keys"]:
            expected = _normalise_text(review_row.get(key_column, ""))
            actual = _normalise_text(ws.cell(workbook_row, headers[key_column]).value)
            if expected != actual:
                errors.append(
                    f"{sheet_name} row {workbook_row} {key_column}: "
                    f"review={expected!r}, workbook={actual!r}"
                )
    if errors:
        preview = "\n".join(errors[:20])
        raise ValueError(f"Review rows no longer match the workbook:\n{preview}")


def _apply_true_decisions(
    workbook: openpyxl.Workbook,
    review: pd.DataFrame,
) -> int:
    updated_cells = 0
    approved = review[review["INSERT"].map(_parse_decision)].copy()
    for _, review_row in approved.iterrows():
        sheet_name = _normalise_text(review_row["sheet_name"])
        ws = workbook[sheet_name]
        headers = _header_map(ws)
        workbook_row = int(review_row["workbook_row"])
        for subtotal_column in SHEET_CONFIGS[sheet_name]["subtotal_columns"]:
            suggested_column = f"suggested_{subtotal_column}"
            suggested_value = _parse_bool(review_row.get(suggested_column, ""))
            if suggested_value is None:
                raise ValueError(
                    f"Missing {suggested_column} for {sheet_name} row {workbook_row}"
                )
            ws.cell(workbook_row, headers[subtotal_column]).value = suggested_value
            updated_cells += 1
    return updated_cells


def _build_exception_rows(
    mapping_workbook: openpyxl.Workbook,
    review: pd.DataFrame,
    exception_headers: list[str],
) -> list[list[object]]:
    rejected = review[~review["INSERT"].map(_parse_decision)].copy()
    output_rows: list[list[object]] = []
    seen_rows: set[tuple[str, ...]] = set()
    for _, review_row in rejected.iterrows():
        sheet_name = _normalise_text(review_row["sheet_name"])
        ws = mapping_workbook[sheet_name]
        headers = _header_map(ws)
        workbook_row = int(review_row["workbook_row"])
        values: dict[str, object] = {
            "enabled": True,
            "sheet": sheet_name,
            "notes": "Reviewed subtotal suggestion rejected; retain current mapping subtotal values.",
        }
        for key_column in SHEET_CONFIGS[sheet_name]["keys"]:
            values[key_column] = ws.cell(workbook_row, headers[key_column]).value
        for subtotal_column in SHEET_CONFIGS[sheet_name]["subtotal_columns"]:
            values[subtotal_column] = ws.cell(workbook_row, headers[subtotal_column]).value
        output_row = [values.get(header, "") for header in exception_headers]
        comparison_key = tuple(_normalise_text(value) for value in output_row)
        if comparison_key not in seen_rows:
            output_rows.append(output_row)
            seen_rows.add(comparison_key)
    return output_rows


def _build_override_rows(
    mapping_workbook: openpyxl.Workbook,
    review: pd.DataFrame,
) -> list[list[object]]:
    """Return one unique reviewed override row per mapping key combination."""
    output_rows: list[list[object]] = []
    seen_rows: set[tuple[str, ...]] = set()
    for _, review_row in review.iterrows():
        sheet_name = _normalise_text(review_row["sheet_name"])
        ws = mapping_workbook[sheet_name]
        headers = _header_map(ws)
        workbook_row = int(review_row["workbook_row"])
        values: dict[str, object] = {
            "enabled": True,
            "sheet": sheet_name,
            "notes": "Reviewed subtotal decision; overrides Stage 0 computed subtotal values.",
        }
        for key_column in SHEET_CONFIGS[sheet_name]["keys"]:
            values[key_column] = ws.cell(workbook_row, headers[key_column]).value
        for subtotal_column in SHEET_CONFIGS[sheet_name]["subtotal_columns"]:
            values[subtotal_column] = ws.cell(workbook_row, headers[subtotal_column]).value
        output_row = [values.get(header, "") for header in OVERRIDE_HEADERS]
        comparison_key = tuple(_normalise_text(value) for value in output_row[:-1])
        if comparison_key not in seen_rows:
            output_rows.append(output_row)
            seen_rows.add(comparison_key)
    return output_rows


def run(
    mapping_workbook_path: Path = MAPPING_WORKBOOK_PATH,
    exception_workbook_path: Path = EXCEPTION_WORKBOOK_PATH,
    review_csv_path: Path = REVIEW_CSV_PATH,
) -> dict[str, object]:
    review = pd.read_csv(review_csv_path, dtype=object)
    _validate_review(review)

    mapping_wb = openpyxl.load_workbook(mapping_workbook_path)
    exception_wb = openpyxl.load_workbook(exception_workbook_path)
    _validate_workbook_rows(mapping_wb, review)

    mapping_archive = _archive(mapping_workbook_path)
    exception_archive = _archive(exception_workbook_path)

    updated_cells = _apply_true_decisions(mapping_wb, review)

    exception_ws = exception_wb[EXCEPTION_SHEET]
    exception_headers = [_normalise_text(cell.value) for cell in exception_ws[1]]
    exception_rows = _build_exception_rows(mapping_wb, review, exception_headers)
    if exception_ws.max_row > 1:
        exception_ws.delete_rows(2, exception_ws.max_row - 1)
    for row_values in exception_rows:
        exception_ws.append(row_values)

    if OVERRIDE_SHEET in exception_wb.sheetnames:
        override_ws = exception_wb[OVERRIDE_SHEET]
        exception_wb.remove(override_ws)
    override_ws = exception_wb.create_sheet(OVERRIDE_SHEET)
    override_ws.append(OVERRIDE_HEADERS)
    override_rows = _build_override_rows(mapping_wb, review)
    for row_values in override_rows:
        override_ws.append(row_values)

    mapping_wb.save(mapping_workbook_path)
    exception_wb.save(exception_workbook_path)

    result = {
        "approved_rows": int(review["INSERT"].map(_parse_decision).sum()),
        "rejected_rows": len(exception_rows),
        "override_rows": len(override_rows),
        "updated_cells": updated_cells,
        "mapping_archive": mapping_archive,
        "exception_archive": exception_archive,
    }
    print(result)
    return result


#%%
# --- Notebook run block ---

APPLY_SUBTOTAL_REVIEW = True

if APPLY_SUBTOTAL_REVIEW:
    APPLY_RESULT = run()

#%%
