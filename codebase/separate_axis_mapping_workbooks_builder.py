#%%
"""Build, validate, and promote separate-axis mapping workbooks in Python."""

#%%
from __future__ import annotations

import hashlib
import json
import shutil
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.drawing.image import Image as WorksheetImage
from openpyxl.worksheet.datavalidation import DataValidation

from codebase.workbook_python_helpers import (
    COLORS,
    new_workbook,
    read_csv_rows,
    render_sheet_preview,
    reopen_workbook,
    scan_formula_errors,
    style_data_sheet,
    style_readme_table,
    style_title,
    validate_literal_booleans,
    write_rows,
)


# --- Stable paths and workbook contract ------------------------------------

REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_ROOT = REPO_ROOT / "outputs" / "separate_axis_mapping_refresh" / "workbooks"
PREVIEW_ROOT = OUTPUT_ROOT / "previews"
SPLIT_MANIFEST_PATH = OUTPUT_ROOT / "split_workbook_manifest.json"
COMPILER_MANIFEST_PATH = (
    REPO_ROOT
    / "outputs"
    / "separate_axis_mapping_refresh"
    / "compiler"
    / "workbook_manifest.json"
)
GENERATION_MANIFEST_PATH = REPO_ROOT / "config" / "outlook_mappings_generation_manifest.json"
EDITABLE_DUPLICATE_AUDIT_PATH = OUTPUT_ROOT / "editable_duplicate_cleanup.json"
REPLACEMENT_SUMMARY_PATH = OUTPUT_ROOT / "generated_master_replacement_summary.json"
MAPPING_WORKFLOW_DIAGRAM_PATH = REPO_ROOT / "docs" / "diagrams" / "mapping_production_workflow.png"

EXCEPTION_SHEET_NAME = "exceptions"
EXCEPTION_HEADERS = [
    "exception_type",
    "enabled",
    "mapping_name",
    "comparison_scope",
    "axis_name",
    "source_keys",
    "target_keys",
    "notes",
]
INITIAL_EXCEPTION_ROW = [
    "allowed_many_to_many_component",
    False,
    "leap_to_esto",
    "BOTH",
    "flow",
    "Demand\\All demand aggregated\\Road|Road|Transport non road/Freight non road/Rail|Transport non road/Nonspecified transport|Transport non road/Passenger non road/Rail",
    "15.02 Road|15.03 Rail|15.06 Non-specified transport",
    "Approved aggregate road and rail hierarchy bridge.",
]
EDITABLE_MANUAL_SHEET_NAMES = [
    "leap_rollup_rules",
    "esto_rollup_rules",
    "ninth_rollup_rules",
    "rollup_label_overrides",
]


# --- General helpers --------------------------------------------------------

def _sha256(path: str | Path) -> str:
    """Return a stable SHA-256 hash for a workbook or manifest."""
    digest = hashlib.sha256()
    with Path(path).open("rb") as file:
        for chunk in iter(lambda: file.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _load_manifest(path: str | Path = SPLIT_MANIFEST_PATH) -> dict[str, object]:
    """Read the source-table manifest and validate its required keys."""
    manifest_path = Path(path)
    if not manifest_path.exists():
        raise FileNotFoundError(f"Split workbook manifest is missing: {manifest_path}")
    manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    required = {
        "editable_axis_workbook_path",
        "generated_pair_workbook_path",
        "generated_master_workbook_path",
        "canonical_master_path",
        "editable_sources",
        "pair_sources",
        "compiled_sources",
        "compiled_columns",
    }
    missing = sorted(required.difference(manifest))
    if missing:
        raise ValueError(f"Split workbook manifest is missing keys: {missing}")
    return manifest


def _rows_for_source(manifest: dict[str, object], relative_path: str) -> list[list[object]]:
    """Load one manifest-relative CSV with literal Boolean values."""
    return read_csv_rows(OUTPUT_ROOT / relative_path)


def _copy_sheet_values(source_sheet: object, target_workbook: object, title: str) -> object:
    """Copy worksheet values into a newly created, consistently styled sheet."""
    target = target_workbook.create_sheet(title)
    for row in source_sheet.iter_rows(values_only=True):
        target.append(list(row))
    style_data_sheet(target, editable=False)
    return target


def _replace_sheet_at_index(workbook: object, sheet_name: str, rows: list[list[object]]) -> object:
    """Replace a sheet body without changing the workbook's sheet order."""
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Canonical workbook is missing sheet: {sheet_name}")
    index = workbook.sheetnames.index(sheet_name)
    workbook.remove(workbook[sheet_name])
    sheet = workbook.create_sheet(sheet_name, index)
    write_rows(sheet, rows)
    style_data_sheet(sheet, editable=False)
    return sheet


def _copy_sheet_at_index(
    source_workbook: object,
    target_workbook: object,
    sheet_name: str,
    target_index: int,
) -> object:
    """Replace one target sheet with values from another workbook."""
    if sheet_name not in source_workbook.sheetnames:
        raise ValueError(f"Editable workbook is missing manual sheet: {sheet_name}")
    if sheet_name in target_workbook.sheetnames:
        target_workbook.remove(target_workbook[sheet_name])
    sheet = target_workbook.create_sheet(sheet_name, target_index)
    for row in source_workbook[sheet_name].iter_rows(values_only=True):
        sheet.append(list(row))
    style_data_sheet(sheet, editable=False)
    return sheet


def _add_workflow_diagram(sheet: object, heading_row: int, image_top_row: int) -> None:
    """Embed the maintained production workflow image when available."""
    sheet.merge_cells(start_row=heading_row, start_column=1, end_row=heading_row, end_column=8)
    sheet.cell(heading_row, 1, "Production workflow (rendered from docs/diagrams/mapping_production_workflow.mmd)")
    if not MAPPING_WORKFLOW_DIAGRAM_PATH.exists():
        return
    image = WorksheetImage(MAPPING_WORKFLOW_DIAGRAM_PATH)
    image.width = 1180
    image.height = 235
    sheet.add_image(image, f"A{image_top_row}")


def _refresh_master_guide(workbook: object) -> None:
    """Replace the canonical Guide with the generated-workbook instructions."""
    index = workbook.sheetnames.index("Guide")
    workbook.remove(workbook["Guide"])
    guide = workbook.create_sheet("Guide", index)
    style_title(
        guide,
        "Generated mapping compatibility workbook",
        "GENERATED COMPATIBILITY INTERFACE — edit outlook_mappings_single_axis.xlsx, not generated pair or rollup sheets here.",
        COLORS["warning"],
        COLORS["warning_text"],
    )
    style_readme_table(
        guide,
        6,
        [
            ["Where people edit", "config/outlook_mappings_single_axis.xlsx is the human-maintained authority for the six axes, accepted exact pairs, exceptions, and rollups."],
            ["Preliminary production gate", "After an editable-contract change, save and close Excel, then run codebase/separate_axis_mapping_refresh_workflow.py before Stage 1."],
            ["Generated pair sheets", "leap_combined_esto, leap_combined_ninth, and ninth_pairs_to_esto_pairs are compiled compatibility outputs. Do not edit their bodies."],
            ["Generated rollup copies", "leap_rollup_rules, esto_rollup_rules, and ninth_rollup_rules are copied from the editable workbook."],
            ["Preserved live sheets", "Display-name, unique-code, and compatibility sheets remain active reference inputs."],
            ["Reserved sheet", "rollup_label_overrides is retained for schema compatibility."],
            ["Deletion candidates", "Keep unused compatibility sheets empty until a coordinated workbook-contract migration."],
            ["Then run", "Run the affected mapping pipeline stages and review relationship, cardinality, hierarchy, lineage, and value-preservation QA."],
            ["QA exceptions", "Reviewed diagnostic exceptions belong in config/mapping_issue_exception_sets.xlsx. They never create or repair a mapping relationship."],
            ["Detailed documentation", "See docs/guide_outlook_mappings_master.md, docs/mappings_system.md, config/README.md, and docs/separate_axis_mapping_pipeline.md."],
        ],
    )
    _add_workflow_diagram(guide, 18, 19)
    guide.freeze_panes = "A5"


# --- Editable-workbook cleanup ---------------------------------------------

def _mapping_key_headers(sheet_name: str, headers: list[str]) -> list[str]:
    scoped = {
        "leap_sector_to_esto": ["leap_sector", "esto_flow"],
        "leap_fuel_to_esto": ["leap_fuel", "esto_product"],
        "ninth_sector_to_esto": ["ninth_sector", "esto_flow"],
        "ninth_fuel_to_esto": ["ninth_fuel", "esto_product"],
    }
    return scoped.get(sheet_name, headers)


def _consolidated_scope(scopes: set[str]) -> str:
    cleaned = {scope for scope in scopes if scope}
    if "BOTH" in cleaned or {"ESTO", "ESTO_EXTENDED"}.issubset(cleaned):
        return "BOTH"
    return next(iter(cleaned), "")


def _ensure_exception_sheet(workbook: object) -> tuple[bool, bool]:
    """Ensure the exceptions contract and literal enabled column exist."""
    added = False
    enabled_added = False
    if EXCEPTION_SHEET_NAME not in workbook.sheetnames:
        sheet = workbook.create_sheet(EXCEPTION_SHEET_NAME)
        write_rows(sheet, [EXCEPTION_HEADERS, INITIAL_EXCEPTION_ROW])
        style_data_sheet(sheet, editable=True)
        added = True
    sheet = workbook[EXCEPTION_SHEET_NAME]
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    if "enabled" not in headers:
        column = sheet.max_column + 1
        sheet.cell(1, column, "enabled")
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row, column, False)
        style_data_sheet(sheet, editable=True)
        enabled_added = True
    return added, enabled_added


def _ensure_manual_sheets(
    workbook: object,
    canonical_master_path: Path,
) -> list[str]:
    """Restore missing human-maintained rollup sheets from the canonical master."""
    missing = [name for name in EDITABLE_MANUAL_SHEET_NAMES if name not in workbook.sheetnames]
    if not missing:
        return []
    canonical = load_workbook(canonical_master_path, data_only=False)
    for name in missing:
        sheet = workbook.create_sheet(name)
        for row in canonical[name].iter_rows(values_only=True):
            sheet.append(list(row))
        style_data_sheet(sheet, editable=True)
    return missing


def _deduplicate_editable_sheet(sheet: object, sheet_name: str) -> dict[str, object]:
    """Keep the first exact mapping key and consolidate duplicate ESTO scopes."""
    headers = [str(cell.value or "").strip() for cell in sheet[1]]
    indexes = {header: index for index, header in enumerate(headers)}
    key_headers = _mapping_key_headers(sheet_name, headers)
    missing = [header for header in key_headers if header not in indexes]
    if missing:
        raise ValueError(f"{sheet_name} is missing editable mapping columns: {missing}")
    scope_index = indexes.get("esto_dataset_scope")
    retained: list[list[object]] = []
    seen: dict[tuple[str, ...], dict[str, object]] = {}
    duplicates: list[dict[str, object]] = []
    for row_number, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        row = list(values[: len(headers)])
        key = tuple(str(row[index] or "").strip() for index in (indexes[name] for name in key_headers))
        if not any(key):
            retained.append(row)
            continue
        if key in seen:
            record = seen[key]
            if scope_index is not None:
                record["scopes"].add(str(row[scope_index] or "").strip().upper())
            duplicates.append(
                {
                    "workbook_row_number": row_number,
                    "retained_workbook_row_number": record["workbook_row_number"],
                    "mapping_key": list(key),
                    "scope_consolidated_into": None,
                }
            )
            continue
        scopes = set()
        if scope_index is not None:
            scopes.add(str(row[scope_index] or "").strip().upper())
        seen[key] = {
            "workbook_row_number": row_number,
            "retained_index": len(retained),
            "scopes": scopes,
        }
        retained.append(row)
    if scope_index is not None:
        by_row = {record["workbook_row_number"]: record for record in seen.values()}
        for record in seen.values():
            retained[record["retained_index"]][scope_index] = _consolidated_scope(record["scopes"])
        for duplicate in duplicates:
            duplicate["scope_consolidated_into"] = _consolidated_scope(
                by_row[duplicate["retained_workbook_row_number"]]["scopes"]
            )
    if duplicates:
        sheet.delete_rows(2, max(1, sheet.max_row - 1))
        for row in retained:
            sheet.append(row)
        style_data_sheet(sheet, editable=True)
    return {
        "input_row_count": max(0, sheet.max_row - 1 + len(duplicates)),
        "retained_row_count": len(retained),
        "duplicate_rows_removed": len(duplicates),
        "duplicates": duplicates,
    }


def clean_editable_workbook_duplicates(
    editable_workbook_path: str | Path,
    canonical_master_path: str | Path,
    editable_sheet_names: list[str],
) -> dict[str, object]:
    """Normalise the editable workbook, write an audit, and validate the rewrite."""
    editable_path = Path(editable_workbook_path)
    workbook = load_workbook(editable_path, data_only=False)
    exception_added, enabled_added = _ensure_exception_sheet(workbook)
    manual_added = _ensure_manual_sheets(workbook, Path(canonical_master_path))
    sheet_audits = {
        name: _deduplicate_editable_sheet(workbook[name], name)
        for name in editable_sheet_names
    }
    changed = [name for name, audit in sheet_audits.items() if audit["duplicate_rows_removed"]]
    rewritten = bool(changed or exception_added or enabled_added or manual_added)
    audit = {
        "checked_at_utc": datetime.now(timezone.utc).isoformat(),
        "workbook_path": str(editable_path),
        "checked_sheets": editable_sheet_names,
        "duplicate_rows_removed": sum(audit["duplicate_rows_removed"] for audit in sheet_audits.values()),
        "workbook_rewritten": rewritten,
        "changed_sheets": changed,
        "exception_sheet_added": exception_added,
        "exception_sheet_enabled_column_added": enabled_added,
        "manual_sheets_added": manual_added,
        "sheets": sheet_audits,
    }
    if rewritten:
        errors = scan_formula_errors(workbook)
        if errors:
            raise ValueError("Editable workbook formula errors:\n- " + "\n- ".join(errors))
        temporary = editable_path.with_suffix(".deduplicate.tmp.xlsx")
        workbook.save(temporary)
        reopened = reopen_workbook(temporary)
        remaining = sum(
            _deduplicate_editable_sheet(reopened[name], name)["duplicate_rows_removed"]
            for name in editable_sheet_names
        )
        if remaining:
            temporary.unlink(missing_ok=True)
            raise ValueError(f"Editable workbook duplicate cleanup did not converge: {remaining}")
        shutil.copyfile(temporary, editable_path)
        temporary.unlink()
        for name in sorted(set(changed + manual_added + ([EXCEPTION_SHEET_NAME] if exception_added or enabled_added else []))):
            render_sheet_preview(reopened, name, PREVIEW_ROOT / "editable_duplicate_cleanup" / f"{name}.png")
    EDITABLE_DUPLICATE_AUDIT_PATH.parent.mkdir(parents=True, exist_ok=True)
    EDITABLE_DUPLICATE_AUDIT_PATH.write_text(json.dumps(audit, indent=2), encoding="utf-8")
    return audit


# --- Workbook builders ------------------------------------------------------

def build_editable_workbook(manifest: dict[str, object]) -> Path:
    """Intentionally rebuild the human-edited workbook from narrow CSV sources."""
    workbook = new_workbook()
    readme = workbook.create_sheet("README")
    style_title(
        readme,
        "Single-axis mapping contract",
        "EDIT THIS WORKBOOK — axis mappings and accepted extra key pairs are the human-maintained source of truth.",
        COLORS["green"],
        COLORS["green_text"],
    )
    style_readme_table(
        readme,
        6,
        [
            ["Start here", "This is the human-maintained mapping contract. Yellow cells are editable; generated pair sheets live in other workbooks."],
            ["What people edit", "Maintain the six single-axis mapping sheets, accepted-extra-pair sheets, exceptions, and rollup sheets in this workbook."],
            ["Why axes are separate", "Sector/flow and fuel/product meanings are maintained once, then combined only for accepted exact source and target pairs."],
            ["ESTO dataset scope", "BOTH applies to ESTO and ESTO Extended; the individual labels restrict a relationship to one registry."],
            ["Generated outputs", "The refresh writes exact-pair evidence and the compatibility master. Do not edit generated pair sheets."],
            ["Extra key pairs", "Each row accepts one exact dataset pair that would otherwise be excluded. Delete the row to withdraw it."],
            ["Duplicate rows", "Each refresh keeps the first occurrence of an exact mapping key and removes later duplicates."],
            ["Subtotals and rollups", "Maintain rollup rules here. The refresh copies them into the generated master."],
            ["Run order", "Save and close Excel, run separate_axis_mapping_refresh_workflow.py, then run the affected mapping stages."],
            ["Review after generation", "Inspect mapping, cardinality, hierarchy, lineage, and value-preservation QA."],
            ["Current status", str(manifest.get("prototype_status", "Production contract."))],
        ],
    )
    _add_workflow_diagram(readme, 19, 20)
    readme.freeze_panes = "A5"
    for sheet_name, relative_path in manifest["editable_sources"].items():
        sheet = workbook.create_sheet(sheet_name)
        write_rows(sheet, _rows_for_source(manifest, relative_path))
        style_data_sheet(sheet, editable=True)
        headers = [str(cell.value or "") for cell in sheet[1]]
        if "esto_dataset_scope" in headers and sheet.max_row > 1:
            column = headers.index("esto_dataset_scope") + 1
            validation = DataValidation(type="list", formula1='"BOTH,ESTO,ESTO_EXTENDED"', allow_blank=False)
            sheet.add_data_validation(validation)
            validation.add(f"{sheet.cell(2, column).coordinate}:{sheet.cell(sheet.max_row, column).coordinate}")
    _ensure_exception_sheet(workbook)
    _ensure_manual_sheets(workbook, Path(manifest["canonical_master_path"]))
    expected = ["README", *manifest["editable_sources"].keys(), EXCEPTION_SHEET_NAME, *EDITABLE_MANUAL_SHEET_NAMES]
    if workbook.sheetnames != expected:
        raise ValueError(f"Editable workbook sheet contract changed: {workbook.sheetnames} vs {expected}")
    output = Path(manifest["editable_axis_workbook_path"])
    workbook.save(output)
    return output


def build_pair_workbook(manifest: dict[str, object]) -> tuple[Path, list[str]]:
    """Build and reopen the generated exact-pair registry workbook."""
    workbook = new_workbook()
    readme = workbook.create_sheet("README")
    style_title(
        readme,
        "Generated key-pair registries",
        "GENERATED — DO NOT EDIT. Rebuild from dataset evidence and the reviewed single-axis contract.",
        COLORS["warning"],
        COLORS["warning_text"],
    )
    style_readme_table(
        readme,
        6,
        [
            ["Purpose", "Make dataset-specific sector/fuel and flow/product key pairs explicit and reusable by the mapping compiler."],
            ["Possible combinations", "ESTO, ESTO Extended, and Ninth sheets contain discovered axis combinations and reviewed extra pairs."],
            ["Historical rule", f"Ordinary ESTO pairs are eligible at the {manifest['historical_boundary_year']} boundary or when reviewed."],
            ["Projection rule", f"Ninth pairs are eligible after {manifest['historical_boundary_year']} or when reviewed."],
            ["LEAP authority", str(manifest.get("leap_registry_authority", ""))],
            ["LEAP boundary", "Exact model pairs, the balance-report grid, and canonical rollup-derived pairs define the registry."],
            ["Subtotal labels", "pair_is_subtotal is generated from the source registry."],
            ["Compilation", "Dataset-specific eligibility and graph safety gates apply before promotion."],
        ],
    )
    for sheet_name, relative_path in manifest["pair_sources"].items():
        sheet = workbook.create_sheet(sheet_name)
        write_rows(sheet, _rows_for_source(manifest, relative_path))
        style_data_sheet(sheet, editable=False)
    errors = scan_formula_errors(workbook)
    if errors:
        raise ValueError("Generated pair workbook formula errors:\n- " + "\n- ".join(errors))
    output = Path(manifest["generated_pair_workbook_path"])
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    reopened = reopen_workbook(output)
    boolean_columns = validate_literal_booleans(reopened, manifest["pair_sources"].keys())
    for name in reopened.sheetnames:
        render_sheet_preview(reopened, name, PREVIEW_ROOT / "generated_pairs" / f"{name}.png")
    return output, boolean_columns


def validate_master_contract(
    workbook: object,
    manifest: dict[str, object],
    expected_sheet_names: list[str],
) -> dict[str, object]:
    """Validate sheet order, compiled headers, Boolean storage, and formulas."""
    if workbook.sheetnames != expected_sheet_names:
        raise ValueError(
            f"Generated master sheet order changed: {workbook.sheetnames} vs {expected_sheet_names}"
        )
    row_counts: dict[str, int] = {}
    for sheet_name, expected_headers in manifest["compiled_columns"].items():
        sheet = workbook[sheet_name]
        actual_headers = [cell.value for cell in sheet[1]][: len(expected_headers)]
        if actual_headers != expected_headers:
            raise ValueError(
                f"Generated master header mismatch for {sheet_name}: {actual_headers} vs {expected_headers}"
            )
        row_counts[sheet_name] = max(0, sheet.max_row - 1)
    formula_errors = scan_formula_errors(workbook)
    if formula_errors:
        raise ValueError("Generated master formula errors:\n- " + "\n- ".join(formula_errors))
    return {
        "sheetCount": len(workbook.sheetnames),
        "sheetNames": workbook.sheetnames,
        "rowCounts": row_counts,
        "booleanColumns": validate_literal_booleans(workbook, manifest["compiled_sources"].keys()),
        "formulaErrors": formula_errors,
    }


def build_generated_master(manifest: dict[str, object]) -> tuple[Path, dict[str, object], dict[str, object]]:
    """Replace generated bodies and rollups while preserving the master contract."""
    canonical = Path(manifest["canonical_master_path"])
    editable = Path(manifest["editable_axis_workbook_path"])
    workbook = load_workbook(canonical, data_only=False)
    editable_workbook = load_workbook(editable, data_only=False)
    expected_sheet_names = list(workbook.sheetnames)
    replacement_summary: dict[str, object] = {}
    for sheet_name, relative_path in manifest["compiled_sources"].items():
        source_rows = _rows_for_source(manifest, relative_path)
        canonical_headers = [cell.value for cell in workbook[sheet_name][1]]
        generated_headers = source_rows[0]
        if canonical_headers[: len(generated_headers)] != generated_headers:
            raise ValueError(
                f"Header mismatch for {sheet_name}: {canonical_headers} vs {generated_headers}"
            )
        canonical_capacity = max(0, workbook[sheet_name].max_row - 1)
        _replace_sheet_at_index(workbook, sheet_name, source_rows)
        replacement_summary[sheet_name] = {
            "canonicalCapacity": canonical_capacity,
            "generatedDataRows": max(0, len(source_rows) - 1),
            "columns": generated_headers,
            "presentation": "Recreated at the canonical sheet index using only contract columns.",
        }
    for name in EDITABLE_MANUAL_SHEET_NAMES:
        index = workbook.sheetnames.index(name)
        _copy_sheet_at_index(editable_workbook, workbook, name, index)
        replacement_summary[name] = {"presentation": "Copied from the editable single-axis workbook."}
    _refresh_master_guide(workbook)
    output = Path(manifest["generated_master_workbook_path"])
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)
    reopened = reopen_workbook(output)
    validation = validate_master_contract(reopened, manifest, expected_sheet_names)
    for name in ["Guide", *manifest["compiled_sources"].keys()]:
        render_sheet_preview(reopened, name, PREVIEW_ROOT / "generated_master" / f"{name}.png")
    replacement_document = {
        "replacementSummary": replacement_summary,
        "contractValidation": validation,
        "inspection": {
            name: [list(row) for row in reopened[name].iter_rows(min_row=1, max_row=min(6, reopened[name].max_row), values_only=True)]
            for name in manifest["compiled_sources"].keys()
        },
    }
    REPLACEMENT_SUMMARY_PATH.write_text(json.dumps(replacement_document, indent=2, default=str), encoding="utf-8")
    return output, replacement_summary, validation


# --- Orchestration and promotion -------------------------------------------

def build_separate_axis_mapping_workbooks(
    promote_master: bool = True,
    rebuild_editable_workbook: bool = False,
    clean_editable_duplicates: bool = True,
    split_manifest_path: str | Path = SPLIT_MANIFEST_PATH,
    original_canonical_master_sha256: str | None = None,
) -> dict[str, object]:
    """Build all generated workbooks and optionally promote the master."""
    manifest = _load_manifest(split_manifest_path)
    editable_path = Path(manifest["editable_axis_workbook_path"])
    canonical_path = Path(manifest["canonical_master_path"])
    if rebuild_editable_workbook:
        build_editable_workbook(manifest)
    if not editable_path.exists():
        raise FileNotFoundError(f"Editable single-axis workbook is missing: {editable_path}")
    duplicate_cleanup = None
    if clean_editable_duplicates:
        duplicate_cleanup = clean_editable_workbook_duplicates(
            editable_path,
            canonical_path,
            list(manifest["editable_sources"].keys()),
        )
    pair_path, pair_boolean_columns = build_pair_workbook(manifest)
    candidate_path, replacement_summary, candidate_validation = build_generated_master(manifest)
    result: dict[str, object] = {
        "status": "candidate_generated_not_promoted",
        "editableWorkbookPath": str(editable_path),
        "pairWorkbookPath": str(pair_path),
        "generatedMasterPath": str(candidate_path),
        "editableDuplicateAuditPath": str(EDITABLE_DUPLICATE_AUDIT_PATH),
        "editableDuplicateCleanup": duplicate_cleanup,
        "pairBooleanColumns": pair_boolean_columns,
        "replacementSummary": replacement_summary,
        "candidateValidation": candidate_validation,
        "promoted": False,
    }
    if not promote_master:
        return result
    if not COMPILER_MANIFEST_PATH.exists():
        raise FileNotFoundError(f"Compiler manifest is missing: {COMPILER_MANIFEST_PATH}")
    compiler_manifest = json.loads(COMPILER_MANIFEST_PATH.read_text(encoding="utf-8"))
    prior_hash = _sha256(canonical_path)
    candidate_hash = _sha256(candidate_path)
    existing = None
    if GENERATION_MANIFEST_PATH.exists():
        try:
            existing = json.loads(GENERATION_MANIFEST_PATH.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            existing = None
    original_hash = (
        (existing or {}).get("hashes", {}).get("original_canonical_master_sha256")
        or original_canonical_master_sha256
        or (existing or {}).get("hashes", {}).get("prior_canonical_master_sha256")
        or prior_hash
    )
    backup_path = OUTPUT_ROOT / f"prior_canonical_master_{prior_hash[:12]}.xlsx"
    shutil.copyfile(canonical_path, backup_path)
    original_backup = (existing or {}).get("original_canonical_backup_path")
    if not original_backup:
        possible = OUTPUT_ROOT / f"prior_canonical_master_{original_hash[:12]}.xlsx"
        original_backup = str(possible) if possible.exists() else None
    temporary = canonical_path.with_suffix(".separate_axis_refresh.tmp.xlsx")
    shutil.copyfile(candidate_path, temporary)
    shutil.copyfile(temporary, canonical_path)
    temporary.unlink()
    promoted = reopen_workbook(canonical_path)
    promoted_validation = validate_master_contract(
        promoted, manifest, candidate_validation["sheetNames"]
    )
    promoted_hash = _sha256(canonical_path)
    if promoted_hash != candidate_hash:
        raise ValueError(
            f"Promoted workbook hash {promoted_hash} does not match candidate hash {candidate_hash}."
        )
    generation_manifest = {
        "status": "promoted_and_reopened",
        "writer": "python_openpyxl",
        "generated_at_utc": datetime.now(timezone.utc).isoformat(),
        "contract_version": "separate_axis_mapping_contract_v1",
        "historical_boundary_year": manifest["historical_boundary_year"],
        "editable_axis_workbook_path": str(editable_path),
        "generated_pair_workbook_path": str(pair_path),
        "canonical_master_path": str(canonical_path),
        "candidate_path": str(candidate_path),
        "prior_canonical_backup_path": str(backup_path),
        "original_canonical_backup_path": original_backup,
        "hashes": {
            "original_canonical_master_sha256": original_hash,
            "prior_canonical_master_sha256": prior_hash,
            "promoted_master_sha256": promoted_hash,
            "generated_pair_workbook_sha256": _sha256(pair_path),
            "editable_axis_workbook_sha256": _sha256(editable_path),
            "compiler_manifest_sha256": _sha256(COMPILER_MANIFEST_PATH),
            "split_manifest_sha256": _sha256(split_manifest_path),
        },
        "compiled_counts": manifest.get("compiled_counts", {}),
        "pair_counts": manifest.get("pair_counts", {}),
        "compiler_summary": compiler_manifest.get("summary", {}),
        "leap_pair_registry_manifest": compiler_manifest.get("leap_pair_registry_manifest"),
        "provisional_relationship_policy": "provisionally_accepted",
        "semantic_review_debt": {
            "within_axis_many_to_many_components": compiler_manifest.get("summary", {}).get("blocking_within_axis_many_to_many_components"),
            "additional_compiled_relationships": compiler_manifest.get("summary", {}).get("generated_relationship_governance_rows"),
        },
        "validation": promoted_validation,
        "editable_duplicate_cleanup": duplicate_cleanup,
        "boolean_storage": "literal_boolean_no_checkbox_controls",
        "rollup_boundary": "Manual rollups remain workbook rules; graph-generated Common rows are not written back as manual rollups.",
        "rollback": "Restore config/outlook_mappings_master.xlsx from Git, then rerun the separate-axis refresh when ready.",
    }
    GENERATION_MANIFEST_PATH.write_text(json.dumps(generation_manifest, indent=2), encoding="utf-8")
    result.update(
        {
            "status": "promoted_and_reopened",
            "generationManifest": generation_manifest,
            "promotedValidation": promoted_validation,
            "promoted": True,
        }
    )
    return result


# --- Frequently changed run flags ------------------------------------------

BUILD_SEPARATE_AXIS_MAPPING_WORKBOOKS = False
PROMOTE_MASTER = True
REBUILD_EDITABLE_WORKBOOK = False


#%%
if __name__ == "__main__" and BUILD_SEPARATE_AXIS_MAPPING_WORKBOOKS:
    BUILD_RESULT = build_separate_axis_mapping_workbooks(
        promote_master=PROMOTE_MASTER,
        rebuild_editable_workbook=REBUILD_EDITABLE_WORKBOOK,
    )
    print(json.dumps(BUILD_RESULT, indent=2, default=str))


#%%
