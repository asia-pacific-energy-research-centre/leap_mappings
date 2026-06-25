from __future__ import annotations

import re
import os
from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pandas as pd

from codebase.utilities.master_config import MASTER_CONFIG_PATH, config_table_exists, read_config_table
from codebase.utilities.workflow_outputs import build_workflow_output_layout, write_output_manifest


EXPECTED_FLOW_COLUMNS = {"flows", "flow"}
EXPECTED_PRODUCT_COLUMNS = {"products", "product"}
SECTOR_COLUMNS = {"sectors", "sub1sectors", "sub2sectors", "sub3sectors", "sub4sectors"}
FUEL_COLUMNS = {"fuels", "subfuels"}
MAX_SCAN_COLS = 200
MAX_SCAN_ROWS = 2000

UNIT_PREFIX_TO_SCALE = {
    "none": 1.0,
    "thousand": 1e3,
    "million": 1e6,
    "billion": 1e9,
    "trillion": 1e12,
}

BASE_UNIT_TO_PETAJOULE = {
    # Joule-family
    "joule": 1e-15,
    "kilojoule": 1e-12,
    "megajoule": 1e-9,
    "gigajoule": 1e-6,
    "terajoule": 1e-3,
    "petajoule": 1.0,
    "exajoule": 1e3,
    # Power-time units
    "kilowatt hour": 3.6e-9,
    "megawatt hour": 3.6e-6,
    "gigawatt hour": 3.6e-3,
    "terawatt hour": 3.6,
    "megawatt year": 0.031536,  # 1 MW * 8760 h
    # Thermal units
    "british thermal unit": 1.05505585262e-12,
    "million btu": 1.05505585262e-6,
    "quad": 1055.05585262,  # quadrillion BTU
    "therm eu": 1.055e-7,
    "therm us": 1.05505585262e-7,
    "therm": 1.05505585262e-7,
    # Other energy units
    "kilocalorie": 4.184e-12,
    "barrel of oil equivalent": 6.1178632e-6,
    "tonnes of oil equivalent": 4.1868e-5,
    "tonnes of coal equivalent": 2.93076e-5,
    "foot pound": 1.3558179483314e-15,
    "erg": 1e-22,
    "electron volt": 1.602176634e-34,
    "electon volt": 1.602176634e-34,  # keep typo variant seen in UI
    "megaelectron volt": 1.602176634e-28,
    # Transport-equivalent units (common default approximations)
    "gallons gasoline equiv": 1.2132e-7,
    "gallons gasoline equivalent": 1.2132e-7,
    "gallons diesel equivalent": 1.358e-7,
}


def _normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text or text.lower() in {"nan", "none"}:
        return ""
    text = text.replace("&", "and")
    text = re.sub(r"[^a-zA-Z0-9]+", " ", text).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def _to_float(value: object) -> float | None:
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    if pd.isna(number):
        return None
    return float(number)


def _resolve_path(path: str | Path, repo_root: Path) -> Path:
    if isinstance(path, Path):
        candidate = path
    else:
        normalized = str(path).replace("\\", "/")
        drive_match = re.match(r"^([a-zA-Z]):/(.*)$", normalized)
        if drive_match:
            drive = drive_match.group(1).lower()
            rest = drive_match.group(2)
            if os.name == "nt":
                return Path(f"{drive.upper()}:/{rest}")
            return Path(f"/mnt/{drive}/{rest}")
        candidate = Path(normalized)
    if candidate.is_absolute():
        return candidate
    return repo_root / candidate


def _sheet_is_balance_like(name: str) -> bool:
    key = _normalize_text(name)
    return (
        key.startswith("energy balance")
        or key.startswith("targt energy balance")
        or bool(re.match(r"^eb\s*\d+", key))
    )


def _write_remove_row_conflict_preview(sheet_label: str, conflict_rows: pd.DataFrame) -> Path:
    repo_root = Path(__file__).resolve().parents[2]
    output_dir = repo_root / "outputs" / "mappings" / "mapping_checks"
    output_dir.mkdir(parents=True, exist_ok=True)

    safe_label = re.sub(r"[^A-Za-z0-9._-]+", "_", str(sheet_label).strip()) or "mapping"
    output_path = output_dir / f"{safe_label}_remove_row_conflicts.csv"
    conflict_rows.to_csv(output_path, index=False)

    try:
        link = output_path.resolve().as_uri()
    except ValueError:
        link = str(output_path.resolve())

    print(f"[WRITE] remove-row conflict preview written to: {output_path}")
    print(f"[LINK] {link}")
    return output_path


def _normalize_unit_key(value: object) -> str:
    key = _normalize_text(value)
    key = key.replace("  ", " ")
    key = re.sub(r"\s+", " ", key).strip()
    return key


def _parse_unit_factor_to_petajoule(unit_text: str) -> tuple[float | None, str, str, str]:
    """
    Return (factor_to_petajoule, parse_status, prefix_label, base_label).

    Examples:
    - "Billion Gigajoule" -> (1000.0, "parsed", "billion", "gigajoule")
    - "Petajoule" -> (1.0, "parsed", "none", "petajoule")
    """
    normalized = _normalize_unit_key(unit_text)
    if not normalized:
        return None, "missing_units", "", ""

    # First try direct mapping (e.g., "million btu" is a base unit label).
    direct_key = normalized
    if direct_key.endswith("s") and direct_key[:-1] in BASE_UNIT_TO_PETAJOULE:
        direct_key = direct_key[:-1]
    if direct_key in BASE_UNIT_TO_PETAJOULE:
        return float(BASE_UNIT_TO_PETAJOULE[direct_key]), "parsed", "none", direct_key

    prefix = "none"
    base_key = normalized

    for token in ["trillion", "billion", "million", "thousand", "none"]:
        candidate = f"{token} "
        if normalized.startswith(candidate):
            prefix = token
            base_key = normalized[len(candidate) :].strip()
            break

    base_key = base_key.strip()
    if base_key.endswith("s") and base_key[:-1] in BASE_UNIT_TO_PETAJOULE:
        base_key = base_key[:-1]
    base_factor = BASE_UNIT_TO_PETAJOULE.get(base_key)
    if base_factor is None:
        return None, "unknown_unit", prefix, base_key

    scale = UNIT_PREFIX_TO_SCALE.get(prefix, 1.0)
    factor = float(scale) * float(base_factor)
    return factor, "parsed", prefix, base_key


def _infer_subtotal_flag(row: pd.Series) -> bool:
    sector_name = _normalize_text(row.get("leap_sector_name", ""))
    fuel_name = _normalize_text(row.get("leap_fuel_name", ""))
    leap_sector = str(row.get("leap_sector", "") or "").strip().lower()
    leap_fuel = str(row.get("leap_fuel", "") or "").strip().lower()
    esto_flow = str(row.get("esto_flow", "") or "").strip().lower()
    esto_product = str(row.get("esto_product", "") or "").strip().lower()

    if sector_name.startswith("total ") or fuel_name.startswith("total "):
        return True
    if any(token in leap_sector for token in ["_total_", "_subtotal_"]) or any(
        token in leap_fuel for token in ["_total_", "_subtotal_"]
    ):
        return True
    if leap_sector.endswith("_total") or leap_fuel.endswith("_total"):
        return True
    if any(
        text.startswith(prefix)
        for text in [esto_flow, esto_product]
        for prefix in ["07 total ", "09 total ", "12 total ", "13 total ", "19 total "]
    ):
        return True
    return False


def _build_simple_output(mapped_long: pd.DataFrame) -> pd.DataFrame:
    out = mapped_long.copy()
    if "value_petajoule" in out.columns:
        value_col = pd.to_numeric(out["value_petajoule"], errors="coerce")
    else:
        value_col = pd.to_numeric(out.get("value", pd.Series(index=out.index)), errors="coerce")

    simple = pd.DataFrame(
        {
            "Year": pd.to_numeric(out.get("year", pd.Series(index=out.index)), errors="coerce").astype("Int64"),
            "values": value_col,
            "esto product": out.get("esto_product", "").astype(str),
            "leap product": out.get("leap_fuel", "").astype(str),
            "esto flow": out.get("esto_flow", "").astype(str),
            "leap flow": out.get("leap_sector", "").astype(str),
            "subtotal": out.apply(_infer_subtotal_flag, axis=1),
        }
    )
    simple["values"] = simple["values"].astype(float)
    full_mapping = (
        simple["esto product"].fillna("").astype(str).str.strip().ne("")
        & simple["leap product"].fillna("").astype(str).str.strip().ne("")
        & simple["esto flow"].fillna("").astype(str).str.strip().ne("")
        & simple["leap flow"].fillna("").astype(str).str.strip().ne("")
    )
    simple = simple[full_mapping].copy()
    simple = simple.sort_values(
        ["Year", "leap flow", "leap product", "esto flow", "esto product"],
        na_position="last",
    ).reset_index(drop=True)
    return simple


@dataclass
class TemplateLayout:
    flows: list[str]
    fuels: list[str]


class TemplateBalanceExtractor:
    """Extract LEAP energy-balance sheets and map to LEAP/ESTO keys."""

    def __init__(
        self,
        *,
        template_sheet: str,
        mapping_pairs_path: Path,
        codebook_path: Path,
        reinterpret_fuel_rows_as_parent_sector: bool = False,
        explicit_pair_mappings_only: bool = False,
        allow_descendant_mapping_expansion: bool = True,
    ) -> None:
        self.template_sheet = template_sheet
        self.mapping_pairs_path = mapping_pairs_path
        self.codebook_path = codebook_path
        self.reinterpret_fuel_rows_as_parent_sector = bool(reinterpret_fuel_rows_as_parent_sector)
        self.explicit_pair_mappings_only = bool(explicit_pair_mappings_only)
        self.allow_descendant_mapping_expansion = bool(allow_descendant_mapping_expansion)
        self._flow_name_to_codes: dict[str, list[str]] = {}
        self._fuel_name_to_codes: dict[str, list[str]] = {}
        self._flow_name_to_esto: dict[str, list[str]] = {}
        self._fuel_name_to_esto: dict[str, list[str]] = {}
        self._canonical_pair_to_esto: dict[tuple[str, str], list[tuple[str, str]]] = {}
        self._balance_name_pair_to_esto: dict[tuple[str, str], list[tuple[str, str]]] = {}
        self._balance_code_pair_to_esto: dict[tuple[str, str], list[tuple[str, str]]] = {}
        self._balance_name_pair_to_ninth: dict[tuple[str, str], tuple[str, str]] = {}
        self._balance_full_path_pair_to_esto: dict[tuple[str, str], list[dict[str, object]]] = {}
        self._balance_full_path_pair_to_ninth: dict[tuple[str, str], list[dict[str, object]]] = {}
        self._balance_full_path_pairs_to_remove: set[tuple[str, str]] = set()
        self._balance_full_path_pairs_with_removed_rows: set[tuple[str, str]] = set()
        self._balance_present_source_keys_by_sheet: dict[str, set[tuple[str, str]]] = {}
        self._balance_detail_mode: str = "detailed"
        self._balance_sheet_detail_modes: dict[str, str] = {}
        self.many_to_many_is_ok_diagnostics = pd.DataFrame()
        self._flow_code_cache: dict[str, list[str]] = {}
        self._fuel_code_cache: dict[str, list[str]] = {}
        self._flow_esto_cache: dict[str, list[str]] = {}
        self._fuel_esto_cache: dict[str, list[str]] = {}
        self._alias_map = {
            "total primary supply": "total primary energy supply",
            "total final energy demand": "total final energy consumption",
            "total transformation": "total transformation sector",
            "transmission and distribution": "transmission and distribution losses",
            "non specified transformation": "Transfers nonspecified",
            "ng liquefaction": "liquefaction regasification plants",
            "lng regasification": "liquefaction regasification plants",
            "oil refining": "oil refineries",
            "heat production": "heat plants",
            "electricity generation": "electricity plants",
            "industry": "industry sector",
            "transport non road": "transport sector",
            "freight road": "transport sector",
            "passenger road": "transport sector",
            "coal sub bituminous": "sub bituminous coal",
        }

    def _read_mapping_pairs_table(self, sheet_name: str, **kwargs: object) -> pd.DataFrame:
        """Read explicitly supplied LEAP mapping workbook sheets before master-config aliases."""
        path = Path(self.mapping_pairs_path)
        if path.exists() and path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
            try:
                workbook = pd.ExcelFile(path)
                if sheet_name in workbook.sheet_names:
                    return pd.read_excel(path, sheet_name=sheet_name, **kwargs)
            except Exception:
                pass
        return read_config_table(self.mapping_pairs_path, sheet_name=sheet_name, **kwargs)

    def load_mappings(self) -> None:
        self.many_to_many_is_ok_diagnostics = pd.DataFrame()

        def add_unique(lookup: dict[str, list[str]], key: str, value: str) -> None:
            if not key or not value:
                return
            lookup.setdefault(key, [])
            if value not in lookup[key]:
                lookup[key].append(value)

        codebook = read_config_table(self.codebook_path, sheet_name="code_to_name", dtype=str).fillna("")
        for _, row in codebook.iterrows():
            name = _normalize_text(row.get("name", ""))
            ninth_label = str(row.get("9th_label", "")).strip()
            ninth_column = _normalize_text(row.get("9th_column", ""))
            esto_label = str(row.get("esto_label", "")).strip()
            esto_column = _normalize_text(row.get("esto_column", ""))
            if not name:
                continue
            if ninth_label and ninth_column in SECTOR_COLUMNS:
                self._flow_name_to_codes.setdefault(name, []).append(ninth_label)
            if ninth_label and ninth_column in FUEL_COLUMNS:
                self._fuel_name_to_codes.setdefault(name, []).append(ninth_label)
            if esto_label and esto_column in EXPECTED_FLOW_COLUMNS:
                self._flow_name_to_esto.setdefault(name, []).append(esto_label)
            if esto_label and esto_column in EXPECTED_PRODUCT_COLUMNS:
                self._fuel_name_to_esto.setdefault(name, []).append(esto_label)

        esto_leap = read_config_table(self.codebook_path, sheet_name="ESTO_LEAP_names", dtype=str).fillna("")
        for _, row in esto_leap.iterrows():
            category = _normalize_text(row.get("category", ""))
            leap_name = _normalize_text(row.get("leap_name", ""))
            esto_label = str(row.get("original_label", "")).strip()
            if category == "products" and leap_name and esto_label:
                self._fuel_name_to_esto.setdefault(leap_name, []).append(esto_label)

        if not self.explicit_pair_mappings_only:
            pairs = read_config_table(self.mapping_pairs_path, dtype=str).fillna("")
            for _, row in pairs.iterrows():
                sector = str(row.get("9th_sector", "")).strip()
                fuel = str(row.get("9th_fuel", "")).strip()
                flow = str(row.get("esto_flow", "")).strip()
                product = str(row.get("esto_product", "")).strip()
                if not (sector and fuel and flow and product):
                    continue
                key = (sector, fuel)
                self._canonical_pair_to_esto.setdefault(key, [])
                pair = (flow, product)
                if pair not in self._canonical_pair_to_esto[key]:
                    self._canonical_pair_to_esto[key].append(pair)

        # Legacy optional mapping sheets are disabled in explicit-pair mode.
        # Balance workflows should resolve only through leap_combined_esto and
        # leap_combined_ninth from the explicitly supplied mapping workbook.
        if not self.explicit_pair_mappings_only:
            try:
                name_pairs = self._read_mapping_pairs_table("leap_name_to_esto_pair", dtype=str).fillna("")
            except Exception:
                name_pairs = pd.DataFrame()
            if not name_pairs.empty:
                for _, row in name_pairs.iterrows():
                    sector_name = self._canonicalize_label(str(row.get("leap_sector_name", "")).strip())
                    fuel_name = self._canonicalize_label(str(row.get("leap_fuel_name", "")).strip())
                    flow = str(row.get("esto_flow", "")).strip()
                    product = str(row.get("esto_product", "")).strip()
                    if not (sector_name and fuel_name and flow and product):
                        continue
                    key = (sector_name, fuel_name)
                    pair = (flow, product)
                    self._balance_name_pair_to_esto.setdefault(key, [])
                    if pair not in self._balance_name_pair_to_esto[key]:
                        self._balance_name_pair_to_esto[key].append(pair)

            try:
                code_pairs = self._read_mapping_pairs_table("leap_code_to_esto_pair", dtype=str).fillna("")
            except Exception:
                code_pairs = pd.DataFrame()
            if not code_pairs.empty:
                for _, row in code_pairs.iterrows():
                    sector = str(row.get("leap_sector", "")).strip()
                    fuel = str(row.get("leap_fuel", "")).strip()
                    flow = str(row.get("esto_flow", "")).strip()
                    product = str(row.get("esto_product", "")).strip()
                    if not (sector and fuel and flow and product):
                        continue
                    key = (sector, fuel)
                    pair = (flow, product)
                    self._balance_code_pair_to_esto.setdefault(key, [])
                    if pair not in self._balance_code_pair_to_esto[key]:
                        self._balance_code_pair_to_esto[key].append(pair)

        # New LEAP balance mapping workbook shape:
        # config/leap_mappings.xlsx contains proposed one-dimensional and
        # pair-level LEAP label mappings rather than the older optional sheet
        # names above. Use exact pair mappings first, then feed the one-
        # dimensional sheets into the existing name->code/name->ESTO lookups.
        if not self.explicit_pair_mappings_only:
            try:
                sector_ninth = self._read_mapping_pairs_table("sector_ninth_final_proposed", dtype=str).fillna("")
            except Exception:
                sector_ninth = pd.DataFrame()
            for _, row in sector_ninth.iterrows():
                key = self._canonicalize_label(str(row.get("leap_sector_name", "")).strip())
                add_unique(self._flow_name_to_codes, key, str(row.get("ninth_sector", "")).strip())

            try:
                fuel_ninth = self._read_mapping_pairs_table("fuel_ninth_final_proposed", dtype=str).fillna("")
            except Exception:
                fuel_ninth = pd.DataFrame()
            for _, row in fuel_ninth.iterrows():
                key = self._canonicalize_label(str(row.get("leap_fuel_name", "")).strip())
                add_unique(self._fuel_name_to_codes, key, str(row.get("ninth_fuel", "")).strip())

            try:
                sector_esto = self._read_mapping_pairs_table("sector_flow_final_proposed", dtype=str).fillna("")
            except Exception:
                sector_esto = pd.DataFrame()
            for _, row in sector_esto.iterrows():
                key = self._canonicalize_label(str(row.get("leap_sector_name", "")).strip())
                add_unique(self._flow_name_to_esto, key, str(row.get("esto_flow", "")).strip())

            try:
                fuel_esto = self._read_mapping_pairs_table("fuel_product_final_proposed", dtype=str).fillna("")
            except Exception:
                fuel_esto = pd.DataFrame()
            for _, row in fuel_esto.iterrows():
                key = self._canonicalize_label(str(row.get("leap_fuel_name", "")).strip())
                add_unique(self._fuel_name_to_esto, key, str(row.get("esto_product", "")).strip())

        def truthy(value: object) -> bool:
            return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on", "t"}

        def clean(value: object) -> str:
            return str(value or "").strip()

        def path_key(value: object) -> str:
            parts = [part.strip() for part in str(value or "").split("/") if part.strip()]
            return "/".join(self._canonicalize_label(part) for part in parts)

        def active_mask(frame: pd.DataFrame) -> pd.Series:
            duplicate = (
                frame["duplicate_to_remove"].map(truthy)
                if "duplicate_to_remove" in frame.columns
                else pd.Series(False, index=frame.index)
            )
            remove_row = (
                frame["remove_row"].map(truthy)
                if "remove_row" in frame.columns
                else (
                    frame["remove_duplicate_row"].map(truthy)
                    if "remove_duplicate_row" in frame.columns
                    else pd.Series(False, index=frame.index)
                )
            )
            return ~(duplicate | remove_row)

        def load_esto_subtotal_lookup() -> dict[tuple[str, str], bool]:
            try:
                subtotals = read_config_table(
                    self.mapping_pairs_path.parent / "ESTO_subtotal_mapping.xlsx",
                    dtype=str,
                ).fillna("")
            except Exception:
                return {}
            required = {"flows", "products", "is_subtotal"}
            if not required.issubset(set(subtotals.columns)):
                return {}
            return {
                (clean(row.get("flows", "")), clean(row.get("products", ""))): truthy(row.get("is_subtotal", False))
                for _, row in subtotals.iterrows()
                if clean(row.get("flows", "")) and clean(row.get("products", ""))
            }

        def load_ninth_subtotal_lookup() -> dict[tuple[str, str], bool]:
            for sheet_name in ["ninth_pair_subtotal_mapping", "ninth_end_node_sector_fuel_subtotal_mapping"]:
                try:
                    subtotals = read_config_table(MASTER_CONFIG_PATH, sheet_name=sheet_name, dtype=str).fillna("")
                except Exception:
                    continue
                sector_col = next(
                    (col for col in ["ninth_sector", "9th_sector", "sector", "sectors"] if col in subtotals.columns),
                    "",
                )
                fuel_col = next(
                    (col for col in ["ninth_fuel", "9th_fuel", "fuel", "fuels", "subfuels"] if col in subtotals.columns),
                    "",
                )
                subtotal_col = next(
                    (col for col in ["is_subtotal", "ninth_is_subtotal", "subtotal_results", "subtotal"] if col in subtotals.columns),
                    "",
                )
                if not (sector_col and fuel_col and subtotal_col):
                    continue
                return {
                    (clean(row.get(sector_col, "")), clean(row.get(fuel_col, ""))): truthy(row.get(subtotal_col, False))
                    for _, row in subtotals.iterrows()
                    if clean(row.get(sector_col, "")) and clean(row.get(fuel_col, ""))
                }
            return {}

        esto_subtotal_lookup = load_esto_subtotal_lookup()
        ninth_subtotal_lookup = load_ninth_subtotal_lookup()

        def annotate_mapping_frame(
            frame: pd.DataFrame,
            *,
            sheet_label: str,
            target_sector_col: str,
            target_fuel_col: str,
            target_subtotal_col: str,
            target_subtotal_lookup: dict[tuple[str, str], bool],
        ) -> pd.DataFrame:
            out = frame.copy()
            if target_subtotal_col == "ninth_pair_is_subtotal" and target_subtotal_col not in out.columns:
                if "ninth_pair_subtotal_status" in out.columns:
                    out[target_subtotal_col] = out["ninth_pair_subtotal_status"]
            for col in ["leap_sector_name_full_path", "raw_leap_fuel_name", target_sector_col, target_fuel_col]:
                if col not in out.columns:
                    out[col] = ""
                out[col] = out[col].fillna("").astype(str).str.strip()
            if "subtotal_mismatch_is_ok" not in out.columns:
                out["subtotal_mismatch_is_ok"] = False
            out["subtotal_mismatch_is_ok"] = out["subtotal_mismatch_is_ok"].map(truthy)
            if "many_to_many_is_ok" in out.columns:
                out["legacy_many_to_many_is_ok"] = out["many_to_many_is_ok"].map(truthy)
            else:
                out["legacy_many_to_many_is_ok"] = False

            active = active_mask(out)
            active_paths = {
                clean(value)
                for value in out.loc[active, "leap_sector_name_full_path"].tolist()
                if clean(value)
            }

            def leap_sector_is_subtotal(path: object) -> bool:
                text = clean(path)
                key = self._canonicalize_path_key(text)
                if not key:
                    return False
                if key.startswith("total "):
                    return True
                prefix = f"{text}/"
                return any(other != text and other.startswith(prefix) for other in active_paths)

            def leap_fuel_is_subtotal(fuel: object) -> bool:
                key = self._canonicalize_label(clean(fuel))
                return key == "total" or key.startswith("total ")

            def target_fuel_requires_subtotal(fuel: object) -> bool:
                text = clean(fuel)
                key = self._canonicalize_label(text)
                return key in {"19 total", "19_total"} or text.lower() in {"19 total", "19_total"}

            out["leap_sector_is_subtotal_computed"] = out["leap_sector_name_full_path"].map(leap_sector_is_subtotal)
            out["leap_fuel_is_subtotal_computed"] = out["raw_leap_fuel_name"].map(leap_fuel_is_subtotal)
            out["leap_is_subtotal_computed"] = (
                out["leap_sector_is_subtotal_computed"] | out["leap_fuel_is_subtotal_computed"]
            )
            if "leap_is_subtotal" in out.columns:
                leap_total_flag_mismatch = (
                    active
                    & out["leap_fuel_is_subtotal_computed"]
                    & ~out["leap_is_subtotal"].map(truthy)
                )
                leap_total_mismatch_rows = out[leap_total_flag_mismatch].copy()
                if not leap_total_mismatch_rows.empty:
                    preview = (
                        leap_total_mismatch_rows[
                            ["leap_sector_name_full_path", "raw_leap_fuel_name", "leap_is_subtotal"]
                        ]
                        .drop_duplicates()
                        .head(30)
                        .to_dict("records")
                    )
                    raise ValueError(
                        f"{sheet_label} has LEAP Total fuel rows that are not flagged as subtotal. "
                        f"Set leap_is_subtotal=True for these rows. Preview: {preview}"
                    )

            valid = (
                active
                & out["leap_sector_name_full_path"].map(clean).ne("")
                & out["raw_leap_fuel_name"].map(clean).ne("")
                & out[target_sector_col].map(clean).ne("")
                & out[target_fuel_col].map(clean).ne("")
            )
            exact_duplicate_cols = [
                "leap_sector_name_full_path",
                "raw_leap_fuel_name",
                target_sector_col,
                target_fuel_col,
            ]
            exact_duplicate_active = out.loc[valid].duplicated(subset=exact_duplicate_cols, keep=False)
            exact_duplicate_rows = out.loc[valid].loc[exact_duplicate_active].copy()
            if not exact_duplicate_rows.empty:
                exact_duplicate_rows["_mapping_row_number"] = exact_duplicate_rows.index + 2
                preview_cols = ["_mapping_row_number", *exact_duplicate_cols]
                preview = exact_duplicate_rows[preview_cols].head(30).to_dict("records")
                raise ValueError(
                    f"{sheet_label} contains exact duplicate active mappings. "
                    f"Each active LEAP source/target pair should appear once. Delete the duplicate rows, "
                    f"or mark them with duplicate_to_remove=True if you intentionally keep a non-suppressing "
                    f"audit copy. Preview: {preview}"
                )

            fuel_pairs = out.loc[valid, ["raw_leap_fuel_name", target_fuel_col]].drop_duplicates()
            fuel_source_count = fuel_pairs.groupby("raw_leap_fuel_name")[target_fuel_col].nunique()
            fuel_target_count = fuel_pairs.groupby(target_fuel_col)["raw_leap_fuel_name"].nunique()

            # Pair-level cardinality: (leap_sector, leap_fuel) <-> (target_sector, target_fuel)
            _pair_frame = out.loc[valid, ["leap_sector_name_full_path", "raw_leap_fuel_name", target_sector_col, target_fuel_col]].copy()
            _pair_frame["_src"] = _pair_frame["leap_sector_name_full_path"].str.strip() + "|||" + _pair_frame["raw_leap_fuel_name"].str.strip()
            _pair_frame["_tgt"] = _pair_frame[target_sector_col].str.strip() + "|||" + _pair_frame[target_fuel_col].str.strip()
            _pairs = _pair_frame[["_src", "_tgt"]].drop_duplicates()
            _pair_src_count = _pairs.groupby("_src")["_tgt"].nunique()
            _pair_tgt_count = _pairs.groupby("_tgt")["_src"].nunique()

            out["pair_mapping_cardinality_computed"] = ""
            out["fuel_mapping_cardinality_computed"] = ""
            out.loc[valid, "_src"] = out.loc[valid, "leap_sector_name_full_path"].fillna("").astype(str).str.strip() + "|||" + out.loc[valid, "raw_leap_fuel_name"].fillna("").astype(str).str.strip()
            out.loc[valid, "_tgt"] = out.loc[valid, target_sector_col].fillna("").astype(str).str.strip() + "|||" + out.loc[valid, target_fuel_col].fillna("").astype(str).str.strip()
            out.loc[valid, "pair_mapping_cardinality_computed"] = out.loc[valid].apply(
                lambda row: _mapping_cardinality(
                    int(_pair_src_count.get(row["_src"], 0)),
                    int(_pair_tgt_count.get(row["_tgt"], 0)),
                ),
                axis=1,
            )
            out.loc[valid, "fuel_mapping_cardinality_computed"] = out.loc[valid].apply(
                lambda row: _mapping_cardinality(
                    int(fuel_source_count.get(row["raw_leap_fuel_name"], 0)),
                    int(fuel_target_count.get(row[target_fuel_col], 0)),
                ),
                axis=1,
            )
            out = out.drop(columns=["_src", "_tgt"], errors="ignore")

            def _norm_cardinality(value: object) -> str:
                return str(value or "").strip().lower().replace("-", "_").replace(" ", "_")

            for authored_col, computed_col, label in [
                ("pair_mapping_cardinality", "pair_mapping_cardinality_computed", "sector+fuel pair"),
                ("fuel_mapping_cardinality", "fuel_mapping_cardinality_computed", "fuel"),
            ]:
                if authored_col in out.columns and out[authored_col].fillna("").astype(str).str.strip().ne("").any():
                    authored = out[authored_col].fillna("").astype(str).map(_norm_cardinality)
                    computed = out[computed_col].fillna("").astype(str).map(_norm_cardinality)
                    mismatch_mask = valid & out[authored_col].fillna("").astype(str).str.strip().ne("") & authored.ne(computed)
                    mismatch_rows = out[mismatch_mask].copy()
                    if not mismatch_rows.empty:
                        preview_cols = [
                            "leap_sector_name_full_path",
                            "raw_leap_fuel_name",
                            target_sector_col,
                            target_fuel_col,
                            authored_col,
                            computed_col,
                        ]
                        preview = mismatch_rows[preview_cols].drop_duplicates().head(30).to_dict("records")
                        raise ValueError(
                            f"{sheet_label} contains {label} mapping cardinality values that do not match the computed "
                            f"cardinality. Update {authored_col!r} to match the resolved mapping or leave it blank if "
                            f"you do not want authored cardinality validation. Preview: {preview}"
                        )

            if target_subtotal_lookup:
                out[target_subtotal_col] = out.apply(
                    lambda row: target_subtotal_lookup.get(
                        (clean(row.get(target_sector_col, "")), clean(row.get(target_fuel_col, ""))),
                        truthy(row.get(target_subtotal_col, False)),
                    ),
                    axis=1,
                )
            else:
                out[target_subtotal_col] = out.get(target_subtotal_col, False)
                out[target_subtotal_col] = out[target_subtotal_col].map(truthy)

            target_total_flag_mismatch = (
                valid
                & out[target_fuel_col].map(target_fuel_requires_subtotal)
                & ~out[target_subtotal_col].fillna(False).map(truthy)
            )
            target_total_mismatch_rows = out[target_total_flag_mismatch].copy()
            if not target_total_mismatch_rows.empty:
                preview_cols = [
                    "leap_sector_name_full_path",
                    "raw_leap_fuel_name",
                    target_sector_col,
                    target_fuel_col,
                    target_subtotal_col,
                ]
                preview = target_total_mismatch_rows[preview_cols].drop_duplicates().head(30).to_dict("records")
                raise ValueError(
                    f"{sheet_label} has target Total fuel rows that are not flagged as subtotal. "
                    f"Ninth 19_total and ESTO 19 Total must be subtotal. Preview: {preview}"
                )

            subtotal_mismatch = (
                valid
                & out["leap_is_subtotal_computed"].ne(out[target_subtotal_col].fillna(False).map(truthy))
                & ~out["subtotal_mismatch_is_ok"]
            )
            mismatch_rows = out[subtotal_mismatch].copy()
            if not mismatch_rows.empty:
                preview_cols = [
                    "leap_sector_name_full_path",
                    "raw_leap_fuel_name",
                    target_sector_col,
                    target_fuel_col,
                    "leap_is_subtotal_computed",
                    target_subtotal_col,
                    "subtotal_mismatch_is_ok",
                ]
                preview = mismatch_rows[preview_cols].drop_duplicates().head(30).to_dict("records")
                raise ValueError(
                    f"{sheet_label} contains subtotal mismatches. Set subtotal_mismatch_is_ok=True "
                    f"only for intentional subtotal-to-non-subtotal mappings. Preview: {preview}"
                )

            pair_many_to_many = out["pair_mapping_cardinality_computed"].eq("many_to_many")
            many_to_many_rows = out[
                valid
                & pair_many_to_many
                & ~out["leap_is_subtotal_computed"]
            ].copy()
            if not many_to_many_rows.empty:
                many_to_many_rows["_diagnostic_sheet"] = sheet_label
                many_to_many_rows["_diagnostic_issue"] = "non_subtotal_many_to_many_mapping"
                many_to_many_rows["_diagnostic_explanation"] = (
                    "Active non-subtotal LEAP sector/fuel pair has computed many-to-many "
                    "cardinality. The extractor records this for audit; concrete safety "
                    "checks are handled by duplicate, subtotal, and dashboard exposure diagnostics."
                )
                diagnostic_cols = [
                    "_diagnostic_sheet",
                    "_diagnostic_issue",
                    "_diagnostic_explanation",
                    "leap_sector_name_full_path",
                    "raw_leap_fuel_name",
                    target_sector_col,
                    target_fuel_col,
                    "pair_mapping_cardinality",
                    "pair_mapping_cardinality_computed",
                    "fuel_mapping_cardinality_computed",
                    "leap_is_subtotal_computed",
                    target_subtotal_col,
                    "legacy_many_to_many_is_ok",
                ]
                for col in diagnostic_cols:
                    if col not in many_to_many_rows.columns:
                        many_to_many_rows[col] = ""
                diagnostic = many_to_many_rows[diagnostic_cols].drop_duplicates().rename(
                    columns={
                        target_sector_col: "target_sector",
                        target_fuel_col: "target_fuel",
                        target_subtotal_col: "target_pair_is_subtotal",
                    }
                )
                self.many_to_many_is_ok_diagnostics = pd.concat(
                    [self.many_to_many_is_ok_diagnostics, diagnostic],
                    ignore_index=True,
                )

            return out

        def _mapping_cardinality(source_target_count: int, target_source_count: int) -> str:
            if source_target_count <= 0 or target_source_count <= 0:
                return ""
            if source_target_count == 1 and target_source_count == 1:
                return "one_to_one"
            if source_target_count > 1 and target_source_count == 1:
                return "one_to_many"
            if source_target_count == 1 and target_source_count > 1:
                return "many_to_one"
            return "many_to_many"

        def add_full_path_esto(row: pd.Series) -> None:
            source_path = str(row.get("leap_sector_name_full_path", "")).strip()
            source_fuel = str(row.get("raw_leap_fuel_name", "")).strip()
            remove_row = truthy(row.get("remove_row", row.get("remove_duplicate_row", False)))
            if source_path and source_fuel and remove_row:
                self._balance_full_path_pairs_with_removed_rows.add(
                    (path_key(source_path), self._canonicalize_label(source_fuel))
                )
                return
            if truthy(row.get("duplicate_to_remove", False)):
                return
            flow = str(row.get("esto_flow", "")).strip()
            product = str(row.get("esto_product", "")).strip()
            if not (source_path and source_fuel and flow and product):
                return
            key = (path_key(source_path), self._canonicalize_label(source_fuel))
            record = {
                "esto_flow": flow,
                "esto_product": product,
                "esto_pair_abs_sum": pd.to_numeric(row.get("esto_pair_abs_sum", pd.NA), errors="coerce"),
                "candidate_leap_sector_name_full_path": str(
                    row.get("candidate_leap_sector_name_full_path", "")
                ).strip()
                or source_path,
                "candidate_leap_fuel_name": str(row.get("candidate_leap_fuel_name", "")).strip()
                or source_fuel,
                "candidate_rule": str(row.get("candidate_rule", "")).strip(),
                "pair_mapping_cardinality": str(row.get("pair_mapping_cardinality_computed", "")).strip(),
                "fuel_mapping_cardinality": str(row.get("fuel_mapping_cardinality_computed", "")).strip(),
                "leap_is_subtotal": truthy(row.get("leap_is_subtotal_computed", False)),
                "subtotal_mismatch_is_ok": truthy(row.get("subtotal_mismatch_is_ok", False)),
                "esto_is_subtotal": truthy(row.get("esto_pair_is_subtotal", False)),
            }
            bucket = self._balance_full_path_pair_to_esto.setdefault(key, [])
            if record not in bucket:
                bucket.append(record)

        def add_full_path_ninth(row: pd.Series) -> None:
            source_path = str(row.get("leap_sector_name_full_path", "")).strip()
            source_fuel = str(row.get("raw_leap_fuel_name", "")).strip()
            remove_row = truthy(row.get("remove_row", row.get("remove_duplicate_row", False)))
            if source_path and source_fuel and remove_row:
                self._balance_full_path_pairs_with_removed_rows.add(
                    (path_key(source_path), self._canonicalize_label(source_fuel))
                )
                return
            if truthy(row.get("duplicate_to_remove", False)):
                return
            ninth_sector = str(row.get("ninth_sector", "")).strip()
            ninth_fuel = str(row.get("ninth_fuel", "")).strip()
            if not (source_path and source_fuel and ninth_sector and ninth_fuel):
                return
            key = (path_key(source_path), self._canonicalize_label(source_fuel))
            record = {
                "ninth_sector": ninth_sector,
                "ninth_fuel": ninth_fuel,
                "candidate_leap_sector_name_full_path": str(
                    row.get("candidate_leap_sector_name_full_path", "")
                ).strip()
                or source_path,
                "candidate_leap_fuel_name": str(row.get("candidate_leap_fuel_name", "")).strip()
                or source_fuel,
                "candidate_rule": str(row.get("candidate_rule", "")).strip(),
                "pair_mapping_cardinality": str(row.get("pair_mapping_cardinality_computed", "")).strip(),
                "fuel_mapping_cardinality": str(row.get("fuel_mapping_cardinality_computed", "")).strip(),
                "leap_is_subtotal": truthy(row.get("leap_is_subtotal_computed", False)),
                "subtotal_mismatch_is_ok": truthy(row.get("subtotal_mismatch_is_ok", False)),
                "ninth_is_subtotal": truthy(
                    row.get("ninth_pair_is_subtotal", row.get("ninth_pair_subtotal_status", False))
                ),
            }
            bucket = self._balance_full_path_pair_to_ninth.setdefault(key, [])
            if record not in bucket:
                bucket.append(record)

        combined_sheet = "leap_combined_esto"
        try:
            combined = self._read_mapping_pairs_table(combined_sheet, dtype=str).fillna("")
        except Exception as exc:
            raise ValueError(
                f"Explicit balance mapping sheet {combined_sheet!r} not found in {self.mapping_pairs_path}."
            ) from exc
        if not combined.empty:
            combined = annotate_mapping_frame(
                combined,
                sheet_label="leap_combined_esto",
                target_sector_col="esto_flow",
                target_fuel_col="esto_product",
                target_subtotal_col="esto_pair_is_subtotal",
                target_subtotal_lookup=esto_subtotal_lookup,
            )
            for _, row in combined.iterrows():
                add_full_path_esto(row)
                sector_name = self._canonicalize_label(str(row.get("leap_sector_name", "")).strip())
                fuel_name = self._canonicalize_label(str(row.get("leap_fuel_name", "")).strip())
                flow = str(row.get("esto_flow", "")).strip()
                product = str(row.get("esto_product", "")).strip()
                if sector_name and fuel_name and flow and product:
                    key = (sector_name, fuel_name)
                    pair = (flow, product)
                    self._balance_name_pair_to_esto.setdefault(key, [])
                    if pair not in self._balance_name_pair_to_esto[key]:
                        self._balance_name_pair_to_esto[key].append(pair)

                ninth_sector = str(row.get("ninth_sector", "")).strip()
                ninth_fuel = str(row.get("ninth_fuel", "")).strip()
                if sector_name and fuel_name and ninth_sector and ninth_fuel:
                    self._balance_name_pair_to_ninth[(sector_name, fuel_name)] = (ninth_sector, ninth_fuel)

        try:
            combined_ninth = self._read_mapping_pairs_table("leap_combined_ninth", dtype=str).fillna("")
        except Exception as exc:
            raise ValueError(
                f"Explicit balance mapping sheet 'leap_combined_ninth' not found in {self.mapping_pairs_path}."
            ) from exc
        if not combined_ninth.empty:
            combined_ninth = annotate_mapping_frame(
                combined_ninth,
                sheet_label="leap_combined_ninth",
                target_sector_col="ninth_sector",
                target_fuel_col="ninth_fuel",
                target_subtotal_col="ninth_pair_is_subtotal",
                target_subtotal_lookup=ninth_subtotal_lookup,
            )
        for _, row in combined_ninth.iterrows():
            add_full_path_ninth(row)
            sector_name = self._canonicalize_label(str(row.get("leap_sector_name", "")).strip())
            fuel_name = self._canonicalize_label(str(row.get("leap_fuel_name", "")).strip())
            ninth_sector = str(row.get("ninth_sector", "")).strip()
            ninth_fuel = str(row.get("ninth_fuel", "")).strip()
            if sector_name and fuel_name and ninth_sector and ninth_fuel:
                self._balance_name_pair_to_ninth[(sector_name, fuel_name)] = (ninth_sector, ninth_fuel)

        active_full_path_keys = set(self._balance_full_path_pair_to_esto) | set(self._balance_full_path_pair_to_ninth)
        self._balance_full_path_pairs_to_remove = (
            self._balance_full_path_pairs_with_removed_rows - active_full_path_keys
        )

        for lookup in [
            self._flow_name_to_codes,
            self._fuel_name_to_codes,
            self._flow_name_to_esto,
            self._fuel_name_to_esto,
        ]:
            for key, values in list(lookup.items()):
                deduped: list[str] = []
                for value in values:
                    if value and value not in deduped:
                        deduped.append(value)
                lookup[key] = deduped

    def _canonicalize_label(self, label: str) -> str:
        key = _normalize_text(label)
        return self._alias_map.get(key, key)

    def _canonicalize_path_key(self, path: object) -> str:
        parts = [part.strip() for part in str(path or "").split("/") if part.strip()]
        return "/".join(self._canonicalize_label(part) for part in parts)

    def _extract_metadata(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, str | int | None]:
        title_text = str(ws.cell(1, 1).value or "").strip()
        meta_text = str(ws.cell(2, 1).value or "").strip()

        area_match = re.search(r'Area\s+"([^"]+)"', title_text)
        scenario_match = re.search(r"Scenario:\s*([^,]+)", meta_text, flags=re.IGNORECASE)
        year_match = re.search(r"Year:\s*(\d{4})", meta_text, flags=re.IGNORECASE)
        units_match = re.search(r"Units:\s*(.+)$", meta_text, flags=re.IGNORECASE)

        return {
            "area": area_match.group(1).strip() if area_match else "",
            "scenario": scenario_match.group(1).strip() if scenario_match else "",
            "year": int(year_match.group(1)) if year_match else None,
            "units": units_match.group(1).strip() if units_match else "",
        }

    def _extract_layout(self, ws: openpyxl.worksheet.worksheet.Worksheet) -> TemplateLayout:
        fuels: list[str] = []
        col = 2
        while col <= MAX_SCAN_COLS:
            value = ws.cell(3, col).value
            text = str(value).strip() if value not in (None, "") else ""
            if not text:
                break
            fuels.append(text)
            col += 1

        flows: list[str] = []
        row = 4
        while row <= MAX_SCAN_ROWS:
            value = ws.cell(row, 1).value
            text = str(value).strip() if value not in (None, "") else ""
            if not text:
                break
            flows.append(text)
            row += 1

        if not fuels or not flows:
            raise ValueError("Template layout extraction failed: no fuel columns or no flow rows found.")

        return TemplateLayout(flows=flows, fuels=fuels)

    def _extract_sheet_matrix(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        *,
        template: TemplateLayout,
    ) -> pd.DataFrame:
        def _row_indent(raw_label: str, excel_indent: object) -> float:
            if excel_indent is not None:
                try:
                    parsed_indent = float(excel_indent)
                    if parsed_indent > 0:
                        return parsed_indent
                except (TypeError, ValueError):
                    pass
            leading_spaces = len(raw_label) - len(raw_label.lstrip(" "))
            return float(leading_spaces)

        def _row_kind(label: str) -> str:
            key = self._canonicalize_label(label)
            has_flow = bool(self._flow_name_to_codes.get(key) or self._flow_name_to_esto.get(key))
            has_fuel = bool(self._fuel_name_to_codes.get(key) or self._fuel_name_to_esto.get(key))
            if has_fuel and not has_flow:
                return "fuel_like"
            if has_flow and not has_fuel:
                return "sector_like"
            if has_flow and has_fuel:
                return "both"
            return "unknown"

        def _is_electricity_generation(label: str) -> bool:
            return self._canonicalize_label(label) == "electricity generation"

        def _resolve_effective_sector_name(flow_rows: list[dict[str, object]], idx: int) -> str:
            current = flow_rows[idx]
            current_name = str(current["name"])
            current_kind = str(current["kind"])
            current_indent = float(current["indent"])
            if current_kind != "fuel_like":
                return current_name

            parent_name = ""
            # Some balance layouts place parent sectors below children, so search down first.
            for j in range(idx + 1, len(flow_rows)):
                candidate = flow_rows[j]
                if float(candidate["indent"]) < current_indent and str(candidate["kind"]) != "fuel_like":
                    parent_name = str(candidate["name"])
                    break
            if not parent_name:
                for j in range(idx - 1, -1, -1):
                    candidate = flow_rows[j]
                    if float(candidate["indent"]) < current_indent and str(candidate["kind"]) != "fuel_like":
                        parent_name = str(candidate["name"])
                        break
            if not parent_name:
                return current_name

            # Under Electricity Generation, child labels are plant types, not fuels to collapse.
            if _is_electricity_generation(parent_name):
                return current_name
            return parent_name

        sheet_fuels: list[str] = []
        col = 2
        while col <= MAX_SCAN_COLS:
            value = ws.cell(3, col).value
            text = str(value).strip() if value not in (None, "") else ""
            if not text:
                break
            sheet_fuels.append(text)
            col += 1

        sheet_flows: list[dict[str, object]] = []
        row = 4
        while row <= MAX_SCAN_ROWS:
            cell = ws.cell(row, 1)
            value = cell.value
            raw_text = str(value) if value not in (None, "") else ""
            text = raw_text.strip()
            if not text:
                break
            sheet_flows.append(
                {
                    "name": text,
                    "raw_name": raw_text,
                    "row_idx": row,
                    "indent": _row_indent(raw_text, getattr(cell.alignment, "indent", None)),
                }
            )
            row += 1

        for flow_row in sheet_flows:
            flow_row["kind"] = _row_kind(str(flow_row["name"]))

        def _section_for_rows(rows: list[dict[str, object]]) -> None:
            total_transformation_idx: int | None = None
            total_final_idx: int | None = None
            for idx, row_info in enumerate(rows):
                key = self._canonicalize_label(str(row_info["name"]))
                if key == "total transformation sector" and total_transformation_idx is None:
                    total_transformation_idx = idx
                if key == "total final energy consumption" and total_final_idx is None:
                    total_final_idx = idx
            for idx, row_info in enumerate(rows):
                if total_transformation_idx is None or idx <= total_transformation_idx:
                    section = "supply_transformation"
                elif total_final_idx is None or idx <= total_final_idx:
                    section = "demand"
                else:
                    section = "final"
                row_info["leap_balance_section"] = section

        def _clean_path_parts(parts: list[str]) -> list[str]:
            cleaned: list[str] = []
            for part in parts:
                clean = str(part or "").strip()
                if not clean:
                    continue
                cleaned.append(clean)
            return cleaned

        def _assign_normal_paths(rows: list[dict[str, object]]) -> None:
            stack: list[dict[str, object]] = []
            for row_info in rows:
                indent = float(row_info["indent"])
                while stack and float(stack[-1]["indent"]) >= indent:
                    stack.pop()
                path_parts = _clean_path_parts([str(parent["name"]) for parent in stack] + [str(row_info["name"])])
                row_info["leap_sector_name_full_path"] = "/".join(path_parts)
                stack.append(row_info)

        def _assign_reverse_paths(rows: list[dict[str, object]]) -> None:
            stack: list[dict[str, object]] = []
            for row_info in reversed(rows):
                indent = float(row_info["indent"])
                while stack and float(stack[-1]["indent"]) >= indent:
                    stack.pop()
                path_parts = _clean_path_parts([str(parent["name"]) for parent in reversed(stack)] + [str(row_info["name"])])
                row_info["leap_sector_name_full_path"] = "/".join(path_parts)
                stack.append(row_info)

        _section_for_rows(sheet_flows)
        section_indices: dict[str, list[int]] = {"supply_transformation": [], "demand": [], "final": []}
        for idx, flow_row in enumerate(sheet_flows):
            section_indices.setdefault(str(flow_row.get("leap_balance_section", "")), []).append(idx)
        for section, indices in section_indices.items():
            if not indices:
                continue
            rows = [sheet_flows[idx] for idx in indices]
            if section == "supply_transformation":
                _assign_reverse_paths(rows)
            elif section == "demand":
                _assign_normal_paths(rows)
            else:
                for row_info in rows:
                    row_info["leap_sector_name_full_path"] = str(row_info["name"])

        for i, flow_row in enumerate(sheet_flows):
            current_indent = float(flow_row["indent"])
            has_fuel_children = False
            has_sector_children = False
            j = i + 1
            while j < len(sheet_flows):
                child = sheet_flows[j]
                child_indent = float(child["indent"])
                if child_indent <= current_indent:
                    break
                child_kind = str(child.get("kind", ""))
                if child_kind == "fuel_like":
                    has_fuel_children = True
                elif child_kind in {"sector_like", "both"}:
                    has_sector_children = True
                j += 1
            flow_row["has_fuel_children"] = has_fuel_children
            flow_row["has_sector_children"] = has_sector_children

        for i, flow_row in enumerate(sheet_flows):
            if self.reinterpret_fuel_rows_as_parent_sector:
                flow_row["effective_sector_name"] = _resolve_effective_sector_name(sheet_flows, i)
            else:
                flow_row["effective_sector_name"] = str(flow_row["name"])

        flow_occurrences: dict[str, list[dict[str, object]]] = {}
        for flow_info in sheet_flows:
            flow_occurrences.setdefault(str(flow_info["name"]), []).append(flow_info)
        flow_occurrence_index: dict[str, int] = {}
        fuel_to_col = {fuel: idx + 2 for idx, fuel in enumerate(sheet_fuels)}

        records: list[dict[str, object]] = []
        for flow in template.flows:
            flow_text = str(flow)
            occurrence_idx = flow_occurrence_index.get(flow_text, 0)
            candidates = flow_occurrences.get(flow_text, [])
            flow_info = candidates[occurrence_idx] if occurrence_idx < len(candidates) else None
            flow_occurrence_index[flow_text] = occurrence_idx + 1
            for fuel in template.fuels:
                row_idx = int(flow_info["row_idx"]) if flow_info is not None else None
                col_idx = fuel_to_col.get(fuel)
                value = None
                if row_idx is not None and col_idx is not None:
                    value = _to_float(ws.cell(row_idx, col_idx).value)
                effective_sector = str(flow_info.get("effective_sector_name", flow_text)) if flow_info is not None else flow_text
                records.append(
                    {
                        "leap_sector_name_raw": flow_text,
                        "leap_sector_name": effective_sector,
                        "leap_sector_name_original": flow_text,
                        "leap_sector_name_full_path": (
                            str(flow_info.get("leap_sector_name_full_path", flow_text)) if flow_info is not None else flow_text
                        ),
                        "leap_fuel_name": fuel,
                        "leap_fuel_name_raw": fuel,
                        "value": value,
                        "leap_balance_section": (
                            str(flow_info.get("leap_balance_section", "")) if flow_info is not None else ""
                        ),
                        "sector_name_reassigned": bool(effective_sector != flow),
                        "leap_sector_row_kind": str(flow_info.get("kind", "")) if flow_info is not None else "",
                        "leap_sector_row_has_fuel_children": (
                            bool(flow_info.get("has_fuel_children", False)) if flow_info is not None else False
                        ),
                        "leap_sector_row_has_sector_children": (
                            bool(flow_info.get("has_sector_children", False)) if flow_info is not None else False
                        ),
                        "flow_present_in_sheet": row_idx is not None,
                        "fuel_present_in_sheet": col_idx is not None,
                    }
                )

        return pd.DataFrame(records)

    def _lookup_with_fallback(self, key: str, lookup: dict[str, list[str]]) -> list[str]:
        normalized = self._canonicalize_label(key)
        values = lookup.get(normalized)
        if values:
            return values

        best = ""
        best_score = 0.0
        for candidate in lookup.keys():
            overlap = len(set(normalized.split()) & set(candidate.split()))
            denom = max(len(set(normalized.split()) | set(candidate.split())), 1)
            score = overlap / denom
            if score > best_score:
                best = candidate
                best_score = score
        if best and best_score >= 0.8:
            return lookup.get(best, [])
        return []

    def _map_row_records(self, row: pd.Series) -> list[dict[str, str]]:
        flow_label = str(row["leap_sector_name"])
        fuel_label = str(row["leap_fuel_name"])
        full_path_label = str(row.get("leap_sector_name_full_path", flow_label))
        original_flow_label = str(row.get("leap_sector_name_original", row.get("leap_sector_name_raw", flow_label)))
        flow_label_key = self._canonicalize_label(flow_label)
        fuel_label_key = self._canonicalize_label(fuel_label)
        full_path_key = (self._canonicalize_path_key(full_path_label), fuel_label_key)
        source_sheet_key = str(row.get("source_sheet", "") or "").strip()

        if self.explicit_pair_mappings_only:
            flow_codes: list[str] = []
            fuel_codes: list[str] = []
            fallback_flows: list[str] = []
            fallback_products: list[str] = []
        else:
            if flow_label not in self._flow_code_cache:
                self._flow_code_cache[flow_label] = self._lookup_with_fallback(flow_label, self._flow_name_to_codes)
            if fuel_label not in self._fuel_code_cache:
                self._fuel_code_cache[fuel_label] = self._lookup_with_fallback(fuel_label, self._fuel_name_to_codes)
            if flow_label not in self._flow_esto_cache:
                self._flow_esto_cache[flow_label] = self._lookup_with_fallback(flow_label, self._flow_name_to_esto)
            if fuel_label not in self._fuel_esto_cache:
                self._fuel_esto_cache[fuel_label] = self._lookup_with_fallback(fuel_label, self._fuel_name_to_esto)

            flow_codes = self._flow_code_cache[flow_label]
            fuel_codes = self._fuel_code_cache[fuel_label]
            fallback_flows = self._flow_esto_cache[flow_label]
            fallback_products = self._fuel_esto_cache[fuel_label]

        sector_code = flow_codes[0] if flow_codes else ""
        fuel_code = fuel_codes[0] if fuel_codes else ""
        ninth_pair = self._balance_name_pair_to_ninth.get((flow_label_key, fuel_label_key))
        if ninth_pair:
            sector_code, fuel_code = ninth_pair

        full_path_esto = self._balance_full_path_pair_to_esto.get(full_path_key, [])
        full_path_ninth = self._balance_full_path_pair_to_ninth.get(full_path_key, [])

        def _descendant_records(
            lookup: dict[tuple[str, str], list[dict[str, object]]],
        ) -> list[dict[str, object]]:
            descendants: list[dict[str, object]] = []
            prefix = f"{full_path_key[0]}/"
            for (path_key, descendant_fuel_key), records in lookup.items():
                if descendant_fuel_key != full_path_key[1]:
                    continue
                if path_key == full_path_key[0] or not path_key.startswith(prefix):
                    continue
                descendants.extend(records)
            return descendants

        descendant_esto = _descendant_records(self._balance_full_path_pair_to_esto)
        descendant_ninth = _descendant_records(self._balance_full_path_pair_to_ninth)
        present_source_keys = self._balance_present_source_keys_by_sheet.get(source_sheet_key, set())
        descendant_source_present = any(
            descendant_fuel_key == full_path_key[1]
            and descendant_path_key.startswith(f"{full_path_key[0]}/")
            for descendant_path_key, descendant_fuel_key in present_source_keys
        )
        # Some LEAP balance exports expose a mapped child branch only as its parent
        # row. Use descendant mappings only when that child row is absent from
        # the current sheet, otherwise the parent would double-count the child.
        use_descendant_records = (
            self.allow_descendant_mapping_expansion
            and not self.explicit_pair_mappings_only
            and bool(descendant_esto or descendant_ninth)
            and (self._balance_detail_mode == "less_detail" or not full_path_esto)
            and not descendant_source_present
        )
        if use_descendant_records:
            if descendant_esto:
                full_path_esto = descendant_esto
            if descendant_ninth:
                full_path_ninth = descendant_ninth

        # Detailed workbooks use remove_row markers to suppress duplicate parent rows.
        # Less-detail workbooks keep removable parent rows only when they can be
        # expanded across descendant child mappings.
        remove_row = (
            full_path_key in self._balance_full_path_pairs_to_remove and not use_descendant_records
        ) or (descendant_source_present and not full_path_esto)
        full_path_meta = (full_path_esto[0] if full_path_esto else full_path_ninth[0]) if (full_path_esto or full_path_ninth) else {}
        mapped_sector_path = str(full_path_meta.get("candidate_leap_sector_name_full_path", "") or full_path_label)
        mapped_fuel_label = str(full_path_meta.get("candidate_leap_fuel_name", "") or fuel_label)
        candidate_rule = str(full_path_meta.get("candidate_rule", "") or "")
        leap_is_subtotal = bool(full_path_meta.get("leap_is_subtotal", False))
        pair_mapping_cardinality = str(full_path_meta.get("pair_mapping_cardinality", "") or "")
        fuel_mapping_cardinality = str(full_path_meta.get("fuel_mapping_cardinality", "") or "")
        subtotal_mismatch_is_ok = bool(full_path_meta.get("subtotal_mismatch_is_ok", False))

        if full_path_ninth:
            sector_codes: list[str] = []
            fuel_codes_from_full_path: list[str] = []
            for ninth_record in full_path_ninth:
                ninth_sector = str(ninth_record.get("ninth_sector", "")).strip()
                ninth_fuel = str(ninth_record.get("ninth_fuel", "")).strip()
                if ninth_sector and ninth_sector not in sector_codes:
                    sector_codes.append(ninth_sector)
                if ninth_fuel and ninth_fuel not in fuel_codes_from_full_path:
                    fuel_codes_from_full_path.append(ninth_fuel)
            sector_code = "|".join(sector_codes)
            fuel_code = "|".join(fuel_codes_from_full_path)

        canonical = (
            self._canonical_pair_to_esto.get((sector_code, fuel_code), [])
            if sector_code and fuel_code
            else []
        )
        balance_code = (
            self._balance_code_pair_to_esto.get((sector_code, fuel_code), [])
            if sector_code and fuel_code
            else []
        )
        balance_name = self._balance_name_pair_to_esto.get((flow_label_key, fuel_label_key), [])

        target_pairs: list[tuple[str, str]] = []
        mapping_status = "unmapped"
        mapping_method = ""
        esto_mapping_found = False
        ninth_mapping_found = bool(full_path_ninth or ninth_pair)
        if full_path_esto:
            target_pairs = [
                (str(record.get("esto_flow", "")).strip(), str(record.get("esto_product", "")).strip())
                for record in full_path_esto
            ]
            target_pairs = [pair for idx, pair in enumerate(target_pairs) if pair != ("", "") and pair not in target_pairs[:idx]]
            esto_mapping_found = bool(target_pairs)
            mapping_status = "mapped" if full_path_ninth else "partial_full_path_pair"
            if use_descendant_records:
                mapping_method = "module_full_path_pair" if len(target_pairs) == 1 else "module_full_path_pair_multiple"
            else:
                mapping_method = "full_path_pair" if len(target_pairs) == 1 else "full_path_pair_multiple"
        elif not self.explicit_pair_mappings_only and balance_code:
            target_pairs = balance_code
            esto_mapping_found = True
            mapping_status = "mapped"
            mapping_method = "balance_code_pair" if len(balance_code) == 1 else "balance_code_pair_multiple"
        elif not self.explicit_pair_mappings_only and balance_name:
            target_pairs = balance_name
            esto_mapping_found = True
            mapping_status = "mapped"
            mapping_method = "balance_name_pair" if len(balance_name) == 1 else "balance_name_pair_multiple"
        elif not self.explicit_pair_mappings_only and canonical:
            target_pairs = canonical
            esto_mapping_found = True
            mapping_status = "mapped"
            mapping_method = "canonical_pair" if len(canonical) == 1 else "canonical_pair_multiple"
        elif not self.explicit_pair_mappings_only:
            fallback_flow = fallback_flows[0] if fallback_flows else ""
            fallback_product = fallback_products[0] if fallback_products else ""
            if fallback_flow and fallback_product:
                target_pairs = [(fallback_flow, fallback_product)]
                esto_mapping_found = True
                mapping_status = "mapped_fallback"
                mapping_method = "codebook_name"
            elif fallback_flow or fallback_product:
                mapping_status = "partial_fallback"
                mapping_method = "codebook_name_partial"

        if not target_pairs:
            target_pairs = [("", "")]

        weights: list[float] = []
        for esto_flow, esto_product in target_pairs:
            matching_esto = next(
                (
                    record
                    for record in full_path_esto
                    if str(record.get("esto_flow", "")).strip() == esto_flow
                    and str(record.get("esto_product", "")).strip() == esto_product
                ),
                {},
            )
            weight = pd.to_numeric(matching_esto.get("esto_pair_abs_sum", pd.NA), errors="coerce")
            weights.append(float(weight) if pd.notna(weight) and float(weight) > 0 else 0.0)

        allocation_method = "direct"
        allocation_shares = [1.0] * len(target_pairs)
        if len(target_pairs) > 1:
            weight_total = float(sum(weights))
            if weight_total > 0:
                allocation_shares = [float(weight) / weight_total for weight in weights]
                allocation_method = "proportional_esto_pair_abs_sum"
            else:
                share = 1.0 / float(len(target_pairs))
                allocation_shares = [share for _ in target_pairs]
                allocation_method = "equal_split"

        match_resolution = "module_only" if use_descendant_records or self._balance_detail_mode == "less_detail" else "detailed"

        records: list[dict[str, str]] = []
        for (esto_flow, esto_product), allocation_share in zip(target_pairs, allocation_shares):
            matching_esto = next(
                (
                    record
                    for record in full_path_esto
                    if str(record.get("esto_flow", "")).strip() == esto_flow
                    and str(record.get("esto_product", "")).strip() == esto_product
                ),
                {},
            )
            records.append(
                {
                    "source_sheet": source_sheet_key,
                    "leap_sector_name": flow_label,
                    "mapped_leap_sector_name": mapped_sector_path,
                    "leap_sector_name_original": original_flow_label,
                    "leap_sector_name_full_path": full_path_label,
                    "leap_fuel_name": fuel_label,
                    "mapped_leap_fuel_name": mapped_fuel_label,
                    "leap_sector": sector_code,
                    "leap_fuel": fuel_code,
                    "esto_flow": esto_flow,
                    "esto_product": esto_product,
                    "allocation_share": allocation_share,
                    "allocation_method": allocation_method,
                    "match_resolution": match_resolution,
                    "source_value_petajoule": pd.NA,
                    "mapping_status": mapping_status,
                    "mapping_method": mapping_method,
                    "mapping_key_sector": full_path_label,
                    "mapping_key_fuel": fuel_label,
                    "mapping_candidate_rule": candidate_rule,
                    "pair_mapping_cardinality": pair_mapping_cardinality,
                    "fuel_mapping_cardinality": fuel_mapping_cardinality,
                    "remove_row": remove_row,
                    "subtotal_mismatch_is_ok": subtotal_mismatch_is_ok,
                    "esto_mapping_found": esto_mapping_found,
                    "ninth_mapping_found": ninth_mapping_found,
                    "leap_is_subtotal": leap_is_subtotal,
                    "esto_is_subtotal": bool(matching_esto.get("esto_is_subtotal", False)),
                    "ninth_is_subtotal": any(
                        bool(record.get("ninth_is_subtotal", False))
                        for record in full_path_ninth
                    ) if full_path_ninth else False,
                    "leap_sector_candidates": "|".join(flow_codes),
                    "leap_fuel_candidates": "|".join(fuel_codes),
                    "esto_flow_candidates": (
                        "|".join([p[0] for p in target_pairs])
                        if target_pairs and target_pairs != [("", "")]
                        else "|".join(fallback_flows)
                    ),
                    "esto_product_candidates": (
                        "|".join([p[1] for p in target_pairs])
                        if target_pairs and target_pairs != [("", "")]
                            else "|".join(fallback_products)
                    ),
                }
            )
        return records

    def _map_row(self, row: pd.Series) -> pd.Series:
        return pd.Series(self._map_row_records(row)[0])

    def extract_workbook(
        self,
        workbook_path: Path,
        *,
        include_zero_values: bool = True,
        sheet_name_filter: list[str] | None = None,
        convert_units_to_petajoule: bool = True,
    ) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, object]]:
        wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=False)
        if self.template_sheet not in wb.sheetnames:
            raise ValueError(f"Template sheet {self.template_sheet!r} not found in workbook.")

        template_layout = self._extract_layout(wb[self.template_sheet])

        selected_sheets: list[str] = []
        for name in wb.sheetnames:
            if sheet_name_filter and name not in sheet_name_filter:
                continue
            if not sheet_name_filter and not _sheet_is_balance_like(name):
                continue
            selected_sheets.append(name)

        all_rows: list[pd.DataFrame] = []
        coverage_rows: list[dict[str, object]] = []
        unit_rows: list[dict[str, object]] = []

        for sheet_name in selected_sheets:
            ws = wb[sheet_name]
            meta = self._extract_metadata(ws)
            try:
                sheet_layout = self._extract_layout(ws)
            except ValueError:
                sheet_layout = template_layout
            extracted = self._extract_sheet_matrix(ws, template=sheet_layout)
            extracted.insert(0, "source_sheet", sheet_name)
            extracted.insert(1, "source_workbook", str(workbook_path))
            extracted["area"] = str(meta.get("area", ""))
            extracted["scenario"] = str(meta.get("scenario", ""))
            extracted["year"] = meta.get("year")
            extracted["units"] = str(meta.get("units", ""))

            if convert_units_to_petajoule:
                factor, parse_status, prefix_label, base_label = _parse_unit_factor_to_petajoule(
                    str(meta.get("units", ""))
                )
                extracted["value_original"] = extracted["value"]
                extracted["units_original"] = extracted["units"]
                extracted["unit_to_petajoule_factor"] = factor
                extracted["unit_parse_status"] = parse_status
                extracted["unit_prefix"] = prefix_label
                extracted["unit_base"] = base_label
                if factor is not None:
                    extracted["value_petajoule"] = pd.to_numeric(extracted["value"], errors="coerce") * float(factor)
                else:
                    extracted["value_petajoule"] = pd.NA
                extracted["units_petajoule"] = "Petajoule"

                unit_rows.append(
                    {
                        "source_sheet": sheet_name,
                        "year": meta.get("year"),
                        "scenario": str(meta.get("scenario", "")),
                        "units_original": str(meta.get("units", "")),
                        "unit_to_petajoule_factor": factor,
                        "unit_parse_status": parse_status,
                        "unit_prefix": prefix_label,
                        "unit_base": base_label,
                    }
                )

            if not include_zero_values:
                extracted = extracted[(extracted["value"].notna()) & (extracted["value"] != 0.0)].copy()
            all_rows.append(extracted)

            missing_flows = extracted.loc[~extracted["flow_present_in_sheet"], "leap_sector_name"].drop_duplicates()
            missing_fuels = extracted.loc[~extracted["fuel_present_in_sheet"], "leap_fuel_name"].drop_duplicates()
            coverage_rows.append(
                {
                    "source_sheet": sheet_name,
                    "template_flow_count": len(sheet_layout.flows),
                    "template_fuel_count": len(sheet_layout.fuels),
                    "matched_flow_count": int(
                        extracted["flow_present_in_sheet"].sum() / max(len(sheet_layout.fuels), 1)
                    ),
                    "matched_fuel_count": int(
                        extracted["fuel_present_in_sheet"].sum() / max(len(sheet_layout.flows), 1)
                    ),
                    "missing_flow_names": "|".join(sorted(set(missing_flows))),
                    "missing_fuel_names": "|".join(sorted(set(missing_fuels))),
                }
            )

        if not all_rows:
            raise ValueError("No balance-like sheets found to extract.")

        extracted_long = pd.concat(all_rows, ignore_index=True)
        # The extraction loop iterates over template.flows which may contain duplicate
        # flow names (the same sector appearing in multiple ESTO flow sections). Since
        # flow_to_row maps by name only, all occurrences read the same balance cell and
        # produce identical rows. Drop those duplicates now so they don't sum multiple
        # times when aggregated by ESTO flow/product pair.
        _dedup_cols = [
            "source_sheet",
            "leap_sector_name",
            "leap_sector_name_original",
            "leap_sector_name_full_path",
            "leap_fuel_name",
            "value",
        ]
        _dedup_cols_present = [c for c in _dedup_cols if c in extracted_long.columns]
        if _dedup_cols_present:
            extracted_long = extracted_long.drop_duplicates(subset=_dedup_cols_present).reset_index(drop=True)
        sheet_detail_rows: list[dict[str, object]] = []
        sheet_detail_modes: dict[str, str] = {}
        for sheet_name, sheet_frame in extracted_long.groupby("source_sheet", dropna=False):
            sheet_name_text = str(sheet_name or "").strip()
            if sheet_frame.empty:
                continue
            path_series = sheet_frame.get("leap_sector_name_full_path", pd.Series(dtype=str)).fillna("").astype(str)
            slash_mask = path_series.str.contains("/", regex=False)
            slash_count = int(slash_mask.sum())
            row_count = int(len(sheet_frame))
            slash_ratio = float(slash_count / row_count) if row_count else 0.0
            detail_mode = "less_detail" if slash_ratio < 0.05 else "detailed"
            sheet_detail_modes[sheet_name_text] = detail_mode
            sheet_detail_rows.append(
                {
                    "source_sheet": sheet_name_text,
                    "row_count": row_count,
                    "slash_row_count": slash_count,
                    "slash_row_ratio": slash_ratio,
                    "detail_mode": detail_mode,
                }
            )
        detail_profile = pd.DataFrame(
            sheet_detail_rows,
            columns=["source_sheet", "row_count", "slash_row_count", "slash_row_ratio", "detail_mode"],
        )
        workbook_detail_mode = (
            "less_detail"
            if int((detail_profile["detail_mode"].eq("less_detail")).sum()) >= int((detail_profile["detail_mode"].eq("detailed")).sum())
            else "detailed"
        )
        self._balance_detail_mode = workbook_detail_mode
        self._balance_sheet_detail_modes = sheet_detail_modes
        present_source_keys_by_sheet: dict[str, set[tuple[str, str]]] = {}
        for sheet_name, sheet_frame in extracted_long.groupby("source_sheet", dropna=False):
            sheet_key = str(sheet_name or "").strip()
            keys: set[tuple[str, str]] = set()
            for _, source_row in sheet_frame.iterrows():
                source_path = self._canonicalize_path_key(source_row.get("leap_sector_name_full_path", ""))
                source_fuel = self._canonicalize_label(source_row.get("leap_fuel_name", ""))
                if source_path and source_fuel:
                    keys.add((source_path, source_fuel))
            present_source_keys_by_sheet[sheet_key] = keys
        self._balance_present_source_keys_by_sheet = present_source_keys_by_sheet

        key_cols = ["source_sheet", "leap_sector_name", "leap_sector_name_original", "leap_sector_name_full_path", "leap_fuel_name"]
        unique_keys = extracted_long[key_cols].drop_duplicates().reset_index(drop=True)
        mapped_records: list[dict[str, str]] = []
        for _, key_row in unique_keys.iterrows():
            mapped_records.extend(self._map_row_records(key_row))
        unique_mapped = pd.DataFrame(mapped_records)
        mapped = extracted_long.merge(unique_mapped, on=key_cols, how="left")
        removed_rows = 0
        if "remove_row" in mapped.columns:
            remove_mask = mapped["remove_row"].fillna(False).astype(bool)
            removed_rows = int(remove_mask.sum())
            mapped = mapped.loc[~remove_mask].copy()
        if "allocation_share" not in mapped.columns:
            mapped["allocation_share"] = 1.0
        mapped["allocation_share"] = pd.to_numeric(mapped["allocation_share"], errors="coerce").fillna(1.0)
        mapped["source_value_petajoule"] = pd.to_numeric(mapped.get("value_petajoule", pd.NA), errors="coerce")
        if "match_resolution" not in mapped.columns:
            mapped["match_resolution"] = ""
        if "allocation_method" not in mapped.columns:
            mapped["allocation_method"] = ""
        mapped["allocated_value_petajoule"] = mapped["source_value_petajoule"] * mapped["allocation_share"]
        mapped["value_petajoule"] = mapped["allocated_value_petajoule"]
        mapped["leap_sector_name_raw_effective"] = mapped["leap_sector_name"]
        mapped["leap_fuel_name_raw"] = mapped["leap_fuel_name"]
        for source_col, target_col in [
            ("mapped_leap_sector_name", "leap_sector_name"),
            ("mapped_leap_fuel_name", "leap_fuel_name"),
        ]:
            if source_col in mapped.columns:
                candidate = mapped[source_col].fillna("").astype(str).str.strip()
                mapped[target_col] = mapped[target_col].where(candidate.eq(""), candidate)
        diagnostics = (
            mapped.groupby(["source_sheet", "mapping_status"], dropna=False)
            .size()
            .reset_index(name="row_count")
            .sort_values(["source_sheet", "row_count"], ascending=[True, False])
                .reset_index(drop=True)
        )
        matching_diagnostics = mapped[
            [
                "source_workbook",
                "source_sheet",
                "scenario",
                "year",
                "leap_sector_name_original",
                "leap_sector_name_full_path",
                "leap_fuel_name_raw",
                "leap_sector_name",
                "leap_fuel_name",
                "esto_flow",
                "esto_product",
                "source_value_petajoule",
                "allocation_share",
                "allocated_value_petajoule",
                "match_resolution",
                "allocation_method",
                "mapping_status",
                "mapping_method",
            ]
        ].copy()
        matching_diagnostics["source_sheet_detail_mode"] = matching_diagnostics["source_sheet"].map(
            lambda value: sheet_detail_modes.get(str(value or "").strip(), "detailed")
        )
        matching_diagnostics["detail_mode"] = workbook_detail_mode
        summary = {
            "template_sheet": self.template_sheet,
            "selected_sheet_count": len(selected_sheets),
            "row_count": int(len(mapped)),
            "mapped_rows": int((mapped["mapping_status"] == "mapped").sum()),
            "mapped_fallback_rows": int((mapped["mapping_status"] == "mapped_fallback").sum()),
            "ambiguous_rows": int((mapped["mapping_status"] == "ambiguous_canonical_pair").sum()),
            "unmapped_rows": int((mapped["mapping_status"] == "unmapped").sum()),
            "partial_rows": int((mapped["mapping_status"] == "partial_fallback").sum()),
            "removed_rows": removed_rows,
            "detail_mode": workbook_detail_mode,
            "detail_sheet_count": int(len(detail_profile)),
            "detail_sheet_less_detail_count": int((detail_profile["detail_mode"] == "less_detail").sum()) if not detail_profile.empty else 0,
            "detail_sheet_detailed_count": int((detail_profile["detail_mode"] == "detailed").sum()) if not detail_profile.empty else 0,
            "detail_path_row_ratio": float(detail_profile["slash_row_ratio"].mean()) if not detail_profile.empty else 0.0,
            "detailed_match_rows": int((matching_diagnostics["match_resolution"] == "detailed").sum()),
            "module_only_match_rows": int((matching_diagnostics["match_resolution"] == "module_only").sum()),
            "proportional_allocation_rows": int(
                matching_diagnostics["allocation_method"].astype(str).str.contains("proportional", na=False).sum()
            ),
        }
        if unit_rows:
            unit_df = pd.DataFrame(unit_rows)
            summary["unit_parse_parsed_sheet_count"] = int((unit_df["unit_parse_status"] == "parsed").sum())
            summary["unit_parse_unknown_sheet_count"] = int((unit_df["unit_parse_status"] == "unknown_unit").sum())
            summary["unit_parse_missing_sheet_count"] = int((unit_df["unit_parse_status"] == "missing_units").sum())
        else:
            unit_df = pd.DataFrame(
                columns=[
                    "source_sheet",
                    "year",
                    "scenario",
                    "units_original",
                    "unit_to_petajoule_factor",
                    "unit_parse_status",
                    "unit_prefix",
                    "unit_base",
                ]
            )

        return extracted_long, mapped, pd.DataFrame(coverage_rows), unit_df, {
            "summary": summary,
            "diagnostics": diagnostics,
            "matching_diagnostics": matching_diagnostics,
            "many_to_many_is_ok_diagnostics": self.many_to_many_is_ok_diagnostics.copy(),
            "detail_profile": detail_profile,
            "detail_mode": workbook_detail_mode,
        }


def run_template_balance_extraction(
    *,
    workbook_path: str | Path,
    output_dir: str | Path,
    template_sheet: str = "Targt Energy Balance 18",
    mapping_pairs_path: str | Path = "config/ninth_pairs_to_esto_pairs.xlsx",
    codebook_path: str | Path = "config/sector_fuel_codes_to_names.xlsx",
    include_zero_values: bool = True,
    sheet_name_filter: list[str] | None = None,
    convert_units_to_petajoule: bool = True,
) -> dict[str, object]:
    repo_root = Path(__file__).resolve().parents[2]
    workbook = _resolve_path(workbook_path, repo_root)
    output = _resolve_path(output_dir, repo_root)
    mapping_pairs = _resolve_path(mapping_pairs_path, repo_root)
    codebook = _resolve_path(codebook_path, repo_root)

    if not workbook.exists():
        raise FileNotFoundError(f"Missing required input: {workbook}")
    for path in [mapping_pairs, codebook]:
        if not config_table_exists(path):
            raise FileNotFoundError(f"Missing required input: {path}")

    layout = build_workflow_output_layout(output)

    extractor = TemplateBalanceExtractor(
        template_sheet=template_sheet,
        mapping_pairs_path=mapping_pairs,
        codebook_path=codebook,
    )
    extractor.load_mappings()
    raw_long, mapped_long, coverage, unit_diag, report = extractor.extract_workbook(
        workbook,
        include_zero_values=include_zero_values,
        sheet_name_filter=sheet_name_filter,
        convert_units_to_petajoule=convert_units_to_petajoule,
    )
    simple_main = _build_simple_output(mapped_long)

    raw_path = layout.analysis / "balance_template_extracted_long.csv"
    mapped_path = layout.analysis / "balance_template_extracted_mapped.csv"
    main_simple_path = layout.root / "balance_template_main_output_pj.csv"
    coverage_path = layout.checks / "balance_template_coverage.csv"
    unit_diag_path = layout.checks / "balance_template_units_to_petajoule.csv"
    diag_path = layout.mapping / "balance_template_mapping_diagnostics.csv"
    many_to_many_diag_path = layout.mapping / "many_to_many_is_ok_diagnostics.csv"
    summary_path = layout.runtime / "balance_template_summary.csv"

    raw_long.to_csv(raw_path, index=False)
    mapped_long.to_csv(mapped_path, index=False)
    simple_main.to_csv(main_simple_path, index=False)
    coverage.to_csv(coverage_path, index=False)
    unit_diag.to_csv(unit_diag_path, index=False)
    report["diagnostics"].to_csv(diag_path, index=False)
    report["many_to_many_is_ok_diagnostics"].to_csv(many_to_many_diag_path, index=False)
    pd.DataFrame([report["summary"]]).to_csv(summary_path, index=False)
    manifest_path = write_output_manifest(
        out_dir=layout.root,
        primary_outputs={"main_output_csv": str(main_simple_path)},
        supporting_outputs={
            "raw_csv": str(raw_path),
            "mapped_csv": str(mapped_path),
            "coverage_csv": str(coverage_path),
            "unit_diagnostics_csv": str(unit_diag_path),
            "diagnostics_csv": str(diag_path),
            "many_to_many_is_ok_diagnostics_csv": str(many_to_many_diag_path),
            "summary_csv": str(summary_path),
        },
        primary_output_descriptions={
            "main_output_csv": "Primary simple balance table with ESTO and LEAP flow/product columns in petajoules.",
        },
        supporting_output_descriptions={
            "raw_csv": "Raw long extraction from the workbook before mapping.",
            "mapped_csv": "Mapped long extraction with LEAP and ESTO codes attached.",
            "coverage_csv": "Coverage summary for mapped and unmapped template rows.",
            "unit_diagnostics_csv": "Unit parsing and conversion diagnostics.",
            "diagnostics_csv": "Row-level mapping diagnostics from template extraction.",
            "many_to_many_is_ok_diagnostics_csv": "Active non-subtotal many-to-many mappings recorded for audit.",
            "summary_csv": "One-row summary of the extraction run.",
        },
        notes=[
            "The simple output stays at the workflow root.",
            "Raw/mapped extracts and diagnostics are grouped under supporting_files/.",
        ],
    )

    return {
        "summary": report["summary"],
        "main_output_csv": str(main_simple_path),
        "raw_csv": str(raw_path),
        "mapped_csv": str(mapped_path),
        "coverage_csv": str(coverage_path),
        "unit_diagnostics_csv": str(unit_diag_path),
        "diagnostics_csv": str(diag_path),
        "many_to_many_is_ok_diagnostics_csv": str(many_to_many_diag_path),
        "summary_csv": str(summary_path),
        "output_manifest_json": str(manifest_path),
    }
