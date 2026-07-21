"""
Extract LEAP balance workbook sheets into long dataframes and comparisons.

This utility module contains the workbook extraction, mapping override,
coverage, comparison, and dashboard-rendering helpers used by the balance
workflow entry scripts. Most callers should use one of the workflow scripts,
but this file is the place to import from when you only need the dataframe
ingestion functions.
"""

from __future__ import annotations

import json
import os
import re
from collections import defaultdict
from pathlib import Path
from typing import Any, Iterable, Sequence

import pandas as pd
from codebase.mappings.canonical_mapping import ConfigTableRef, split_config_table_ref
from codebase.utilities.leap_balance_export_resolver import (
    BALANCE_EXPORT_TRUST_FILENAME_SCENARIO,
    resolve_balance_export_workbook,
)
from codebase.utilities.master_config import config_table_exists, read_config_table

from codebase.utilities.energy_balance_template_extractor import TemplateBalanceExtractor
from codebase.utilities.leap_results_dashboard_utils import (
    _aggregate_display_rows_to_total,
    _prepare_ninth_projection_frame,
    _prepare_render_long,
    _safe_token,
    build_charts,
    make_chart,
    pull_base_year_value,
    pull_projection_series,
)
from codebase.utilities.leap_results_dashboard_v2.config_loader import load_mapping_inputs
from codebase.utilities.leap_results_dashboard_v2.reference_loader import load_reference_tables


REPO_ROOT = Path(__file__).resolve().parents[2]


DEFAULT_BALANCE_EXPORT_ECONOMY = "20_USA"
DEFAULT_REF_BALANCE_EXPORT_DATE_ID: str | None = None
DEFAULT_TGT_BALANCE_EXPORT_DATE_ID: str | None = None
DEFAULT_REF_WORKBOOK_PATH = resolve_balance_export_workbook(
    economy=DEFAULT_BALANCE_EXPORT_ECONOMY,
    scenario="REF",
    date_id=DEFAULT_REF_BALANCE_EXPORT_DATE_ID,
)
DEFAULT_TGT_WORKBOOK_PATH = resolve_balance_export_workbook(
    economy=DEFAULT_BALANCE_EXPORT_ECONOMY,
    scenario="TGT",
    date_id=DEFAULT_TGT_BALANCE_EXPORT_DATE_ID,
)
DEFAULT_MAPPING_PAIRS_PATH = (REPO_ROOT / "config/master_config.xlsx", "ninth_pairs_to_esto_pairs")
DEFAULT_CODEBOOK_PATH = REPO_ROOT / "config/sector_fuel_codes_to_names.xlsx"

DEFAULT_SHEET_MAP_PATH = REPO_ROOT / "config/leap_results_sheet_map.csv"
DEFAULT_BACKUP_MAPPINGS_PATH = REPO_ROOT / "config/backup_leap_mappings.xlsx"
DEFAULT_EXPLICIT_MAPPINGS_PATH = REPO_ROOT / "config/leap_results_explicit_mappings.csv"
DEFAULT_EXPLICIT_REASSIGNMENTS_PATH = REPO_ROOT / "config/leap_results_explicit_reassignments.csv"
DEFAULT_SYNTHETIC_REFERENCE_ROWS_PATH = REPO_ROOT / "config/synthetic_reference_rows.csv"
DEFAULT_BASE_TABLE_PATH = REPO_ROOT / "data/00APEC_2025_low_with_subtotals.csv"
DEFAULT_PROJECTION_TABLE_PATH = REPO_ROOT / "data/merged_file_energy_ALL_20251106.csv"


DEFAULT_EMPTY_PAGE_NOTICE = (
    "No mapped balance-table data is available for this page in the current run. "
    "The page is kept to preserve the configured dashboard structure."
)

TRANSFORMATION_INPUT_MEASURE = "Inputs (PJ)"
TRANSFORMATION_OUTPUT_MEASURE = "Outputs (PJ)"
TRANSFORMATION_DASHBOARD_TOP_GROUPS = {"Power", "Refining", "Other transformation"}


def _product_is_total(value: object) -> bool:
    text = " ".join(str(value or "").strip().lower().split())
    return text == "total" or bool(re.match(r"^\d+(?:[._]\d+)*\s+total$", text))

BALANCE_ESTO_MAPPING_COLUMNS = [
    "leap_sector_name_original",
    "leap_sector_name_full_path",
    "raw_leap_fuel_name",
    "value",
    "esto_flow",
    "esto_product",
    "pair_mapping_cardinality",
    "leap_is_subtotal",
    "esto_pair_is_subtotal",
    "subtotal_mismatch_is_ok",
    "subtotal_alignment",
    "esto_pair_abs_sum",
    "remove_row",
    "remove_row_reason",
]

BALANCE_NINTH_MAPPING_COLUMNS = [
    "leap_sector_name_original",
    "leap_sector_name_full_path",
    "raw_leap_fuel_name",
    "value",
    "ninth_sector",
    "ninth_fuel",
    "pair_mapping_cardinality",
    "leap_is_subtotal",
    "ninth_pair_is_subtotal",
    "subtotal_mismatch_is_ok",
    "subtotal_alignment",
    "ninth_pair_abs_sum",
    "remove_row",
    "remove_row_reason",
]

ESTO_FLOW_LABEL_FALLBACKS = {
    "01": "Production",
    "02": "Imports",
    "03": "Exports",
    "04": "International marine bunkers",
    "05": "International aviation bunkers",
    "06": "Stock changes",
    "07": "Total primary energy supply",
    "08": "Transfers",
    "08.01": "Recycled products",
    "08.02": "Interproduct transfers",
    "08.03": "Products transferred",
    "08.04": "Gas separation",
    "08.99": "Transfers nonspecified",
    "09": "Total transformation sector",
    "09.01": "Main activity producer",
    "09.02": "Autoproducers",
    "09.03": "Heat pumps",
    "09.04": "Electric boilers",
    "09.05": "Chemical heat for electricity production",
    "09.06": "Gas processing plants",
    "09.07": "Oil refineries",
    "09.08": "Coal transformation",
    "09.09": "Petrochemical industry",
    "09.10": "Biofuels processing",
    "09.11": "Charcoal processing",
    "09.12": "Non-specified transformation",
    "09.13": "Hydrogen transformation",
    "10": "Losses & own use",
    "10.01": "Own use",
    "10.02": "Transmission and distribution losses",
    "11": "Statistical discrepancy",
    "12": "Total final consumption",
    "13": "Total final energy consumption",
    "14": "Industry sector",
    "14.01": "Mining and quarrying",
    "14.02": "Construction",
    "14.03": "Manufacturing",
    "15": "Transport sector",
    "16": "Other sector",
    "16.01": "Commercial and public services",
    "17": "Non-energy use",
    "18": "Electricity output in GWh",
    "19": "Heat output in PJ",
}

ESTO_DASHBOARD_GROUP_ORDER = {
    "Buildings": 0,
    "Bunkers": 1,
    "Industry sector": 2,
    "Transport sector": 3,
    "Other sector": 4,
    "Power": 5,
    "Refining": 6,
    "Other transformation": 7,
    "Losses & own use": 8,
    "Supply": 9,
    "Demand": 10,
}

BALANCE_DASHBOARD_MAJOR_SECTOR_PAGES = {
    "Buildings",
    "Bunkers",
    "Industry sector",
    "Transport sector",
    "Other sector",
    "Power",
    "Refining",
    "Other transformation",
    "Supply",
}

BALANCE_DASHBOARD_TOP_LABELS = {
    "Industry sector": "Industry",
    "Transport sector": "Transport",
    "Other sector": "Others",
}

# Dashboard hierarchy field dictionary for the ESTO-axis migration.
#
# - sheet: legacy chart namespace derived from ESTO flow. It is retained for
#   compatibility, but it is not a rendered dashboard page.
# - page_key/page_label: rendered top-level dashboard page, e.g. Buildings.
# - chart_group_key/chart_group_label: chart group within a page, e.g.
#   Commercial and public services, Residential, or Buildings Total.
# - fuel_label: fuel-level series within the chart group. Existing fuel labels
#   remain unchanged so mapping logic and rendered series names stay stable.
BALANCE_DASHBOARD_HIERARCHY_FIELDS = [
    "sheet",
    "page_key",
    "page_label",
    "chart_group_key",
    "chart_group_label",
    "fuel_label",
]

BALANCE_DASHBOARD_PAGE_FALLBACK_KEY = "unknown_page"


def _resolve(path: Path | str) -> Path:
    raw = str(path).replace("\\", "/")
    drive_match = re.match(r"^([a-zA-Z]):/(.*)$", raw)
    if drive_match:
        drive = drive_match.group(1).lower()
        rest = drive_match.group(2)
        if os.name == "nt":
            return Path(f"{drive.upper()}:/{rest}")
        return Path(f"/mnt/{drive}/{rest}")
    candidate = Path(raw)
    return candidate if candidate.is_absolute() else (REPO_ROOT / candidate)


def _resolve_config_table_ref(ref: ConfigTableRef) -> ConfigTableRef:
    path, sheet_name = split_config_table_ref(ref)
    resolved_path = _resolve(path)
    return (resolved_path, sheet_name) if sheet_name else resolved_path


def _clean_token(value: object) -> str:
    text = str(value or "").strip()
    return "" if text.lower() in {"", "nan", "none"} else text


def _hierarchy_key_token(value: object, fallback: str = "") -> str:
    token = _safe_token(str(value or "").replace("\\", "_")).strip("_").lower()
    return token or fallback


def _dashboard_hierarchy_from_path(
    path: Sequence[object],
    *,
    entry_kind: str = "direct",
    measure: object = "",
    fallback_label: object = "",
) -> dict[str, str]:
    """Return explicit page/chart-group fields from a rendered dashboard path."""
    raw_path = tuple(_clean_token(part) for part in list(path or []) if _clean_token(part))
    fallback_text = _clean_token(fallback_label) or "Unknown"
    if not raw_path:
        raw_path = (fallback_text,)
    display_path = tuple(
        _display_balance_path_part(part, is_top=idx == 0)
        for idx, part in enumerate(raw_path)
    )
    page_label = display_path[0] if display_path else fallback_text
    page_key = _hierarchy_key_token(page_label, BALANCE_DASHBOARD_PAGE_FALLBACK_KEY)
    kind = _clean_token(entry_kind) or "direct"
    leaf_label = display_path[-1] if display_path else page_label
    if kind == "aggregate":
        chart_group_label = f"{page_label} Total" if len(display_path) <= 1 else f"{leaf_label} Total"
    else:
        chart_group_label = leaf_label
    key_parts = [page_key, *raw_path, kind]
    measure_text = _clean_token(measure)
    if measure_text and measure_text != "Energy balance (PJ)":
        key_parts.append(measure_text)
    chart_group_key = "chart_group__" + "__".join(
        _hierarchy_key_token(part, "x") for part in key_parts if _clean_token(part)
    )
    return {
        "page_key": page_key,
        "page_label": page_label,
        "chart_group_key": chart_group_key,
        "chart_group_label": chart_group_label,
    }


def _sheet_catalog_hierarchy(sheet_catalog: dict[str, Any], sheet: object, *, measure: object = "") -> dict[str, str]:
    sheet_text = _clean_token(sheet)
    cfg = dict(sheet_catalog.get(sheet_text, {}) or {})
    path = list(cfg.get("path", []) or [])
    if not path:
        path = [cfg.get("display_label", "") or sheet_text]
    return _dashboard_hierarchy_from_path(path, entry_kind="direct", measure=measure, fallback_label=sheet_text)


def _backfill_dashboard_hierarchy(
    frame: pd.DataFrame,
    *,
    sheet_catalog: dict[str, Any] | None = None,
    sheet_col: str = "sheet",
) -> pd.DataFrame:
    """Ensure page/chart-group fields exist, deriving from legacy sheet when needed."""
    out = pd.DataFrame() if frame is None else frame.copy()
    if out.empty:
        for col in ["page_key", "page_label", "chart_group_key", "chart_group_label"]:
            if col not in out.columns:
                out[col] = ""
        return out
    if sheet_col not in out.columns:
        out[sheet_col] = ""
    if "measure" not in out.columns:
        out["measure"] = ""
    for col in [sheet_col, "measure", "page_key", "page_label", "chart_group_key", "chart_group_label"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()
    catalog = sheet_catalog or {}
    missing = out["page_key"].eq("") | out["page_label"].eq("") | out["chart_group_key"].eq("") | out["chart_group_label"].eq("")
    if not missing.any():
        return out
    for idx in out.index[missing]:
        hierarchy = _sheet_catalog_hierarchy(catalog, out.at[idx, sheet_col], measure=out.at[idx, "measure"])
        for col, value in hierarchy.items():
            if not out.at[idx, col]:
                out.at[idx, col] = value
    return out


def _normalize_scenario(value: object) -> str:
    text = _clean_token(value).lower()
    if text in {"reference", "ref"}:
        return "Reference"
    if text in {"target", "tgt"}:
        return "Target"
    return _clean_token(value)


def _normalize_text(value: object) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("&", "and")
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def _to_bool(value: object, default: bool = False) -> bool:
    if value is None:
        return default
    text = str(value).strip().lower()
    if text in {"1", "true", "yes", "y", "on", "t"}:
        return True
    if text in {"0", "false", "no", "n", "off", "f"}:
        return False
    return default


def _load_active_balance_mapping_crosswalk(
    mapping_workbook_path: Path | str | None,
    *,
    esto_sheet_name: str = "leap_combined_esto",
    ninth_sheet_name: str = "leap_combined_ninth",
) -> pd.DataFrame:
    """Return active ESTO->LEAP->9th mapping rows from the balance mapping workbook."""
    columns = [
        "leap_sector_name_full_path",
        "raw_leap_fuel_name",
        "esto_flow",
        "esto_product",
        "ninth_sector",
        "ninth_fuel",
        "esto_pair_mapping_cardinality",
        "ninth_pair_mapping_cardinality",
        "leap_is_subtotal",
        "esto_pair_is_subtotal",
        "ninth_pair_is_subtotal",
    ]
    if mapping_workbook_path is None:
        return pd.DataFrame(columns=columns)
    workbook = _resolve(mapping_workbook_path)
    if not workbook.exists():
        return pd.DataFrame(columns=columns)

    def _read_active(sheet_name: str, target_cols: list[str]) -> pd.DataFrame:
        try:
            frame = pd.read_excel(workbook, sheet_name=sheet_name, dtype=str).fillna("")
        except Exception:
            return pd.DataFrame()
        required_cols = [
            "leap_sector_name_full_path",
            "raw_leap_fuel_name",
            "remove_row",
            "duplicate_to_remove",
            "pair_mapping_cardinality",
            "leap_is_subtotal",
            *target_cols,
        ]
        for col in required_cols:
            if col not in frame.columns:
                frame[col] = ""
            frame[col] = frame[col].fillna("").astype(str).str.strip()
        active = ~frame["remove_row"].map(_to_bool) & ~frame["duplicate_to_remove"].map(_to_bool)
        for col in target_cols:
            active &= frame[col].ne("")
        active &= frame["leap_sector_name_full_path"].ne("") & frame["raw_leap_fuel_name"].ne("")
        frame = frame.loc[active].copy()
        if frame.empty:
            return frame
        frame["_leap_sector_key"] = frame["leap_sector_name_full_path"].str.strip().str.lower()
        frame["_leap_fuel_key"] = frame["raw_leap_fuel_name"].str.strip().str.lower()
        return frame

    esto = _read_active(esto_sheet_name, ["esto_flow", "esto_product"])
    ninth = _read_active(ninth_sheet_name, ["ninth_sector", "ninth_fuel"])
    if esto.empty or ninth.empty:
        return pd.DataFrame(columns=columns)

    for col in ["esto_pair_is_subtotal"]:
        if col not in esto.columns:
            esto[col] = ""
    for col in ["ninth_pair_is_subtotal"]:
        if col not in ninth.columns:
            ninth[col] = ""

    joined = esto.merge(
        ninth,
        on=["_leap_sector_key", "_leap_fuel_key"],
        how="inner",
        suffixes=("_esto", "_ninth"),
    )
    if joined.empty:
        return pd.DataFrame(columns=columns)

    out = pd.DataFrame(
        {
            "leap_sector_name_full_path": joined["leap_sector_name_full_path_esto"],
            "raw_leap_fuel_name": joined["raw_leap_fuel_name_esto"],
            "esto_flow": joined["esto_flow"],
            "esto_product": joined["esto_product"],
            "ninth_sector": joined["ninth_sector"],
            "ninth_fuel": joined["ninth_fuel"],
            "esto_pair_mapping_cardinality": joined["pair_mapping_cardinality_esto"],
            "ninth_pair_mapping_cardinality": joined["pair_mapping_cardinality_ninth"],
            "leap_is_subtotal": (
                joined["leap_is_subtotal_esto"].map(_to_bool)
                | joined["leap_is_subtotal_ninth"].map(_to_bool)
            ),
            "esto_pair_is_subtotal": joined["esto_pair_is_subtotal"].map(_to_bool),
            "ninth_pair_is_subtotal": joined["ninth_pair_is_subtotal"].map(_to_bool),
        }
    )
    for col in ["esto_flow", "esto_product", "ninth_sector", "ninth_fuel"]:
        out[col] = out[col].fillna("").astype(str).str.strip()
    out = out[
        out["esto_flow"].ne("")
        & out["esto_product"].ne("")
        & out["ninth_sector"].ne("")
        & out["ninth_fuel"].ne("")
    ].copy()
    return out[columns].drop_duplicates().reset_index(drop=True)


def _nonzero_ninth_projection_pairs(
    ninth_df: pd.DataFrame | None,
    *,
    projection_economy: str,
    projection_scenarios: Iterable[str],
    projection_years: Sequence[int],
) -> set[tuple[str, str]]:
    """Return deepest nonzero 9th sector/fuel pairs for the requested projection slice."""
    if ninth_df is None or ninth_df.empty or not projection_years:
        return set()
    work = ninth_df.copy()
    for col in ["economy", "scenarios", "sectors", "sub1sectors", "sub2sectors", "sub3sectors", "sub4sectors", "fuels", "subfuels"]:
        if col not in work.columns:
            work[col] = ""
    year_cols = [str(year) for year in projection_years if str(year) in work.columns]
    if not year_cols:
        year_cols = [year for year in projection_years if year in work.columns]
    if not year_cols:
        return set()
    scenario_values = {str(value).strip().lower() for value in projection_scenarios if str(value).strip()}
    work = work[
        work["economy"].fillna("").astype(str).str.strip().eq(str(projection_economy).strip())
        & work["scenarios"].fillna("").astype(str).str.strip().str.lower().isin(scenario_values)
    ].copy()
    if work.empty:
        return set()
    values = work[year_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
    work = work.loc[values.abs().gt(1e-12).any(axis=1)].copy()
    if work.empty:
        return set()

    def _deepest(row: pd.Series, cols: list[str]) -> str:
        tokens = [_clean_token(row.get(col, "")) for col in cols]
        tokens = [token for token in tokens if token and token.lower() != "x"]
        return tokens[-1] if tokens else ""

    sector_cols = ["sectors", "sub1sectors", "sub2sectors", "sub3sectors", "sub4sectors"]
    fuel_cols = ["fuels", "subfuels"]
    pairs: set[tuple[str, str]] = set()
    for _, row in work.iterrows():
        sector = _deepest(row, sector_cols)
        fuel = _deepest(row, fuel_cols)
        if sector and fuel:
            pairs.add((sector, fuel))
    return pairs


def _load_json(path: Path | str) -> dict[str, Any]:
    resolved = _resolve(path)
    if not resolved.exists():
        return {}
    return json.loads(resolved.read_text(encoding="utf-8"))


def _infer_subtotal_flag(row: pd.Series) -> bool:
    sector_name = _normalize_text(row.get("leap_sector_name", ""))
    fuel_name = _normalize_text(row.get("leap_fuel_name", ""))
    leap_sector = _clean_token(row.get("leap_sector", "")).lower()
    leap_fuel = _clean_token(row.get("leap_fuel", "")).lower()
    esto_flow = _clean_token(row.get("esto_flow", "")).lower()
    esto_product = _clean_token(row.get("esto_product", "")).lower()

    if sector_name.startswith("total ") or fuel_name.startswith("total "):
        return True
    if any(token in leap_sector for token in ["_total_", "_subtotal_"]) or any(
        token in leap_fuel for token in ["_total_", "_subtotal_"]
    ):
        return True
    if leap_sector.endswith("_total") or leap_fuel.endswith("_total"):
        return True
    if leap_fuel == "19_total":
        return True
    if any(
        text.startswith(prefix)
        for text in [esto_flow, esto_product]
        for prefix in ["07 total ", "09 total ", "12 total ", "13 total ", "19 total "]
    ):
        return True
    return False


def _apply_balance_row_filters(
    combined: pd.DataFrame,
    row_filters: dict[str, Any],
) -> pd.DataFrame:
    out = combined.copy()

    if _to_bool(row_filters.get("exclude_unmet_requirements", False), default=False):
        unmatched = pd.Series(False, index=out.index)
        for col in [
            "leap_sector_name_full_path",
            "mapping_key_sector",
            "leap_sector_name",
            "leap_sector_name_original",
        ]:
            if col not in out.columns:
                continue
            normalized = out[col].fillna("").astype(str).map(_normalize_text)
            unmatched = unmatched | normalized.eq("unmet requirements") | normalized.str.startswith(
                "unmet requirements "
            )
        out = out[~unmatched].copy()

    if _to_bool(row_filters.get("exclude_subtotals", False), default=False):
        out = out[~out["is_subtotal"].fillna(False).astype(bool)].copy()
    if _to_bool(row_filters.get("drop_zero_values", False), default=False):
        out = out[pd.to_numeric(out["value_petajoule"], errors="coerce").fillna(0).ne(0)].copy()
    min_year = row_filters.get("min_year")
    max_year = row_filters.get("max_year")
    if min_year is not None:
        out = out[out["year"].ge(int(min_year))].copy()
    if max_year is not None:
        out = out[out["year"].le(int(max_year))].copy()
    return out


def _pick_template_sheet(workbook_path: Path, preferred_sheet: str) -> str:
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        if preferred_sheet in wb.sheetnames:
            return preferred_sheet
        ebal = [name for name in wb.sheetnames if str(name).strip().lower().startswith("ebal|")]
        if ebal:
            return sorted(ebal, reverse=True)[0]
        return wb.sheetnames[0]
    finally:
        wb.close()


def _list_balance_sheets(workbook_path: Path) -> list[str]:
    import openpyxl

    wb = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    try:
        names = [str(name).strip() for name in wb.sheetnames]
    finally:
        wb.close()

    selected: list[str] = []
    for name in names:
        key = name.lower()
        if key.startswith("ebal|"):
            selected.append(name)
            continue
        if key.startswith("energy balance") or key.startswith("targt energy balance"):
            selected.append(name)
            continue
    return selected


def _load_codebook_name_maps(codebook_path: Path) -> tuple[dict[str, str], dict[str, str]]:
    codebook = read_config_table(codebook_path, sheet_name="code_to_name", dtype=str).fillna("")
    flow_map: dict[str, str] = {}
    fuel_map: dict[str, str] = {}
    for _, row in codebook.iterrows():
        ninth_label = _clean_token(row.get("ninth_label", ""))
        name = _clean_token(row.get("name", ""))
        ninth_column = _normalize_text(row.get("ninth_column", ""))
        if not ninth_label or not name:
            continue
        if ninth_column in {"sectors", "sub1sectors", "sub2sectors", "sub3sectors", "sub4sectors"}:
            flow_map[ninth_label] = name
        if ninth_column in {"fuels", "subfuels"}:
            fuel_map[ninth_label] = name
    return flow_map, fuel_map


def _apply_mapping_overrides(frame: pd.DataFrame, overrides: list[dict[str, Any]]) -> tuple[pd.DataFrame, pd.DataFrame]:
    if frame.empty or not overrides:
        return frame, pd.DataFrame(columns=["override_index", "applied_rows", "match", "set"])

    out = frame.copy()
    reports: list[dict[str, Any]] = []
    alias = {
        "scenario": "scenario",
        "year": "year",
        "leap_flow": "leap_sector",
        "leap_flow_name": "leap_sector_name",
        "leap_product": "leap_fuel",
        "leap_product_name": "leap_fuel_name",
        "esto_flow": "esto_flow",
        "esto_product": "esto_product",
        "sheet": "sheet",
    }

    for idx, override in enumerate(overrides):
        if not _to_bool(override.get("active", True), default=True):
            continue
        match = override.get("match", {}) or {}
        set_values = override.get("set", {}) or {}
        if not isinstance(match, dict) or not isinstance(set_values, dict):
            continue

        mask = pd.Series(True, index=out.index)
        for key, expected in match.items():
            col = alias.get(str(key), str(key))
            if col not in out.columns:
                mask &= False
                continue
            if col == "year":
                expected_num = pd.to_numeric(pd.Series([expected]), errors="coerce").iloc[0]
                mask &= pd.to_numeric(out[col], errors="coerce").eq(expected_num)
            else:
                mask &= out[col].fillna("").astype(str).str.strip().eq(str(expected).strip())

        hit_count = int(mask.sum())
        if hit_count:
            for key, value in set_values.items():
                col = alias.get(str(key), str(key))
                if col not in out.columns:
                    out[col] = ""
                out.loc[mask, col] = value
            note = _clean_token(set_values.get("mapping_note", ""))
            if note:
                if "mapping_note" not in out.columns:
                    out["mapping_note"] = ""
                current = out.loc[mask, "mapping_note"].fillna("").astype(str).str.strip()
                out.loc[mask, "mapping_note"] = current.where(current.ne(""), note)

        reports.append(
            {
                "override_index": idx,
                "applied_rows": hit_count,
                "match": json.dumps(match, ensure_ascii=True),
                "set": json.dumps(set_values, ensure_ascii=True),
            }
        )

    return out, pd.DataFrame(reports)


def _build_flow_and_product_labels(
    frame: pd.DataFrame,
    *,
    flow_name_map: dict[str, str],
    fuel_name_map: dict[str, str],
    label_overrides: dict[str, Any],
) -> pd.DataFrame:
    out = frame.copy()

    flow_override = {
        str(k): str(v)
        for k, v in (label_overrides.get("flow", {}) or {}).items()
        if _clean_token(k) and _clean_token(v)
    }
    product_override = {
        str(k): str(v)
        for k, v in (label_overrides.get("product", {}) or {}).items()
        if _clean_token(k) and _clean_token(v)
    }

    out["flow_label"] = out["leap_sector"].map(flow_name_map).fillna("")
    missing_flow = out["flow_label"].eq("")
    out.loc[missing_flow, "flow_label"] = out.loc[missing_flow, "leap_sector_name"].fillna("").astype(str)
    out["flow_label"] = out.apply(
        lambda row: flow_override.get(_clean_token(row.get("leap_sector", "")), row.get("flow_label", "")),
        axis=1,
    )

    out["fuel_label"] = out["leap_fuel"].map(fuel_name_map).fillna("")
    missing_fuel = out["fuel_label"].eq("")
    out.loc[missing_fuel, "fuel_label"] = out.loc[missing_fuel, "leap_fuel_name"].fillna("").astype(str)
    out["fuel_label"] = out.apply(
        lambda row: product_override.get(_clean_token(row.get("leap_fuel", "")), row.get("fuel_label", "")),
        axis=1,
    )

    return out


def _coalesce_unique(series: pd.Series) -> str:
    cleaned = [str(v).strip() for v in series.fillna("").astype(str).tolist() if str(v).strip()]
    unique = sorted(set(cleaned))
    if not unique:
        return ""
    return unique[0]


def _coalesce_pipe_tokens_unique(series: pd.Series) -> str:
    """Join unique pipe-delimited mapping tokens without dropping later rows."""
    tokens: list[str] = []
    for value in series.fillna("").astype(str).tolist():
        for token in str(value).split("|"):
            token = token.strip()
            if token:
                tokens.append(token)
    unique = sorted(set(tokens))
    return "|".join(unique)


def _series_or_default(frame: pd.DataFrame, column: str, default: object = "") -> pd.Series:
    if column in frame.columns:
        return frame[column]
    return pd.Series(default, index=frame.index)


def _simple_leap_working_frame(leap_long: pd.DataFrame) -> pd.DataFrame:
    working = leap_long.copy()
    for col in ["leap_sector_name", "leap_fuel_name", "is_subtotal"]:
        if col not in working.columns:
            working[col] = ""
    if working["leap_sector_name"].fillna("").astype(str).str.strip().eq("").all():
        working["leap_sector_name"] = _series_or_default(working, "sector_name")
    if working["leap_fuel_name"].fillna("").astype(str).str.strip().eq("").all():
        working["leap_fuel_name"] = _series_or_default(working, "fuel_label")
    return working


def build_simple_leap_balance_table(leap_long: pd.DataFrame) -> pd.DataFrame:
    """Return a compact LEAP-to-ESTO balance mapping table for direct analysis."""
    columns = [
        "scenario",
        "year",
        "leap_sector",
        "esto_flow",
        "leap_fuel",
        "esto_product",
        "value_pj",
    ]
    if leap_long.empty:
        return pd.DataFrame(columns=columns)

    working = _simple_leap_working_frame(leap_long)

    out = pd.DataFrame(
        {
            "scenario": _series_or_default(working, "scenario"),
            "year": pd.to_numeric(_series_or_default(working, "year"), errors="coerce").astype("Int64"),
            "leap_sector": _series_or_default(working, "leap_sector_name"),
            "esto_flow": _series_or_default(working, "esto_flow"),
            "leap_fuel": _series_or_default(working, "leap_fuel_name"),
            "esto_product": _series_or_default(working, "esto_product"),
            "value_pj": pd.to_numeric(_series_or_default(working, "leap_value", pd.NA), errors="coerce"),
        }
    )
    return out.sort_values(["scenario", "year", "leap_sector", "leap_fuel"], kind="mergesort").reset_index(drop=True)


def convert_leap_balances_to_esto_long_table(
    *,
    ref_workbook_path: Path | str = DEFAULT_REF_WORKBOOK_PATH,
    tgt_workbook_path: Path | str = DEFAULT_TGT_WORKBOOK_PATH,
    template_sheet: str = "EBal|2060",
    mapping_pairs_path: ConfigTableRef = DEFAULT_MAPPING_PAIRS_PATH,
    codebook_path: Path | str = DEFAULT_CODEBOOK_PATH,
    structure_config: dict[str, Any] | None = None,
    known_issues: dict[str, Any] | None = None,
    projection_economy: str = "20_USA",
    max_output_year: int | None = None,
    explicit_pair_mappings_only: bool = False,
    allow_descendant_mapping_expansion: bool = True,
) -> dict[str, Any]:
    """
    Convert REF/TGT LEAP balance exports into an ESTO-pair long balance table.

    This is the dashboard-independent entry point for the LEAP balance to ESTO
    axis conversion. It keeps the detailed mapped LEAP rows and returns the
    compact table with one value per scenario/year/ESTO flow/product pair.
    """
    ingestion = load_balance_leap_long_esto_axis(
        ref_workbook_path=ref_workbook_path,
        tgt_workbook_path=tgt_workbook_path,
        template_sheet=template_sheet,
        mapping_pairs_path=mapping_pairs_path,
        codebook_path=codebook_path,
        structure_config=structure_config,
        known_issues=known_issues,
        projection_economy=projection_economy,
        explicit_pair_mappings_only=explicit_pair_mappings_only,
        allow_descendant_mapping_expansion=allow_descendant_mapping_expansion,
    )
    leap_long = ingestion["leap_long"].copy()
    if max_output_year is not None and not leap_long.empty:
        leap_long = leap_long[pd.to_numeric(leap_long["year"], errors="coerce").le(max_output_year)].copy()
        ingestion = {**ingestion, "leap_long": leap_long}

    esto_long = build_simple_leap_balance_table(leap_long)
    return {
        "esto_long": esto_long,
        "leap_long": leap_long,
        "mapping_status": ingestion["mapping_status"],
        "issues": ingestion["issues"],
        "override_report": ingestion["override_report"],
        "pre_group_leap_mapped": ingestion.get("pre_group_leap_mapped", pd.DataFrame()),
        "pre_group_incomplete_rows": ingestion.get("pre_group_incomplete_rows", pd.DataFrame()),
        "unit_diagnostics": ingestion["unit_diagnostics"],
        "coverage": ingestion["coverage"],
        "matching_diagnostics": ingestion.get("matching_diagnostics", pd.DataFrame()),
        "resolved_structure": ingestion.get("resolved_structure", structure_config or {}),
        "auto_sheet_rows": ingestion.get("auto_sheet_rows", pd.DataFrame()),
        "extraction_summary": ingestion["extraction_summary"],
        "ingestion": ingestion,
    }


def build_simple_leap_ninth_balance_table(leap_long: pd.DataFrame) -> pd.DataFrame:
    """Return a compact LEAP-to-9th balance mapping table for direct analysis."""
    columns = [
        "scenario",
        "year",
        "leap_sector",
        "ninth_sector",
        "leap_fuel",
        "ninth_fuel",
        "value_pj",
        "subtotal",
    ]
    if leap_long.empty:
        return pd.DataFrame(columns=columns)

    working = _simple_leap_working_frame(leap_long)
    subtotal = _series_or_default(working, "is_subtotal", False).fillna(False).astype(bool)
    out = pd.DataFrame(
        {
            "scenario": _series_or_default(working, "scenario"),
            "year": pd.to_numeric(_series_or_default(working, "year"), errors="coerce").astype("Int64"),
            "leap_sector": _series_or_default(working, "leap_sector_name"),
            "ninth_sector": _series_or_default(working, "sector_code_9th"),
            "leap_fuel": _series_or_default(working, "leap_fuel_name"),
            "ninth_fuel": _series_or_default(working, "ninth_fuel_code"),
            "value_pj": pd.to_numeric(_series_or_default(working, "leap_value", pd.NA), errors="coerce"),
            "subtotal": subtotal,
        }
    )
    return out.sort_values(["scenario", "year", "leap_sector", "leap_fuel"], kind="mergesort").reset_index(drop=True)


def build_simple_ninth_balance_table(
    *,
    comparison_long: pd.DataFrame,
    mapping_status: pd.DataFrame,
) -> pd.DataFrame:
    """Return 9th projection rows mapped to ESTO flow/product pairs in PJ."""
    columns = [
        "scenario",
        "year",
        "ninth_sector",
        "esto_flow",
        "ninth_fuel",
        "esto_product",
        "value_pj",
    ]
    if comparison_long.empty or mapping_status.empty:
        return pd.DataFrame(columns=columns)

    keys = ["sheet", "measure", "fuel_label"]
    projection = comparison_long.copy()
    projection = projection[projection.get("source", "").fillna("").astype(str).str.strip().eq("projection")].copy()
    if projection.empty:
        return pd.DataFrame(columns=columns)

    def join_unique(series: pd.Series) -> str:
        values = [str(v).strip() for v in series.fillna("").astype(str) if str(v).strip()]
        return "|".join(sorted(set(values)))

    meta = mapping_status.copy()
    for col in keys + ["sector_code_9th", "ninth_fuel_code", "esto_flow", "esto_product"]:
        if col not in meta.columns:
            meta[col] = ""
        meta[col] = meta[col].fillna("").astype(str).str.strip()
    meta = (
        meta.groupby(keys, as_index=False)
        .agg(
            ninth_sector=("sector_code_9th", join_unique),
            ninth_fuel=("ninth_fuel_code", join_unique),
            esto_flow=("esto_flow", join_unique),
            esto_product=("esto_product", join_unique),
        )
    )

    out = projection.merge(meta, on=keys, how="left")
    out["year"] = pd.to_numeric(out["year"], errors="coerce").astype("Int64")
    out["value_pj"] = pd.to_numeric(out.get("value", pd.Series(index=out.index)), errors="coerce")
    out = out[columns]
    return out.sort_values(
        ["scenario", "year", "esto_flow", "esto_product", "ninth_sector", "ninth_fuel"],
        kind="mergesort",
    ).reset_index(drop=True)


def build_ninth_balance_esto_long_table(ninth_balance: pd.DataFrame) -> pd.DataFrame:
    """Return 9th rows with the same columns as build_simple_leap_balance_table."""
    columns = ["scenario", "year", "leap_sector", "esto_flow", "leap_fuel", "esto_product", "value_pj"]
    if ninth_balance.empty:
        return pd.DataFrame(columns=columns)
    out = ninth_balance.rename(columns={"ninth_sector": "leap_sector", "ninth_fuel": "leap_fuel"}).copy()
    for col in columns:
        if col not in out.columns:
            out[col] = ""
    out["year"] = pd.to_numeric(out["year"], errors="coerce").astype("Int64")
    out["value_pj"] = pd.to_numeric(out["value_pj"], errors="coerce")
    return out[columns].sort_values(
        ["scenario", "year", "leap_sector", "leap_fuel", "esto_flow", "esto_product"],
        kind="mergesort",
    ).reset_index(drop=True)


def build_merged_esto_axis_balance_table(
    *,
    simple_leap_balance: pd.DataFrame,
    simple_ninth_balance: pd.DataFrame,
    comparison_long: pd.DataFrame,
    mapping_status: pd.DataFrame,
) -> pd.DataFrame:
    """Return LEAP, 9th, and ESTO values side-by-side on the ESTO pair axis."""
    columns = [
        "scenario",
        "year",
        "esto_flow",
        "esto_product",
        "leap_sector",
        "leap_fuel",
        "ninth_sector",
        "ninth_fuel",
        "value_pj_leap",
        "value_pj_9th",
        "value_pj_esto",
    ]
    key_cols = ["scenario", "year", "esto_flow", "esto_product"]

    def join_unique(series: pd.Series) -> str:
        values = [str(v).strip() for v in series.fillna("").astype(str) if str(v).strip()]
        return "|".join(sorted(set(values)))

    def _aggregate_simple(
        frame: pd.DataFrame,
        *,
        label_cols: list[str],
        value_col: str,
    ) -> pd.DataFrame:
        out_cols = [*key_cols, *label_cols, value_col]
        if frame.empty:
            return pd.DataFrame(columns=out_cols)
        working = frame.copy()
        for col in [*key_cols, *label_cols]:
            if col not in working.columns:
                working[col] = ""
            working[col] = working[col].fillna("").astype(str).str.strip()
        working["year"] = pd.to_numeric(working["year"], errors="coerce").astype("Int64")
        working["value_pj"] = pd.to_numeric(working.get("value_pj", pd.Series(index=working.index)), errors="coerce")
        grouped = (
            working.groupby(key_cols, dropna=False)
            .agg(
                **{col: (col, join_unique) for col in label_cols},
                **{value_col: ("value_pj", "sum")},
            )
            .reset_index()
        )
        return grouped[out_cols]

    leap = _aggregate_simple(
        simple_leap_balance,
        label_cols=["leap_sector", "leap_fuel"],
        value_col="value_pj_leap",
    )
    ninth = _aggregate_simple(
        simple_ninth_balance,
        label_cols=["ninth_sector", "ninth_fuel"],
        value_col="value_pj_9th",
    )

    esto_cols = [*key_cols, "value_pj_esto"]
    if comparison_long.empty or mapping_status.empty:
        esto = pd.DataFrame(columns=esto_cols)
    else:
        keys = ["sheet", "measure", "fuel_label"]
        base_sources = {"base", "base_estimated", "base_mixed"}
        base = comparison_long.copy()
        source = base.get("source", pd.Series("", index=base.index)).fillna("").astype(str).str.strip()
        base = base[source.isin(base_sources)].copy()
        if base.empty:
            esto = pd.DataFrame(columns=esto_cols)
        else:
            meta = mapping_status.copy()
            for col in keys + ["esto_flow", "esto_product"]:
                if col not in meta.columns:
                    meta[col] = ""
                meta[col] = meta[col].fillna("").astype(str).str.strip()
            meta = (
                meta.groupby(keys, as_index=False)
                .agg(
                    esto_flow=("esto_flow", join_unique),
                    esto_product=("esto_product", join_unique),
                )
            )
            for col in keys + ["scenario"]:
                if col not in base.columns:
                    base[col] = ""
                base[col] = base[col].fillna("").astype(str).str.strip()
            base = base.merge(meta, on=keys, how="left")
            base["year"] = pd.to_numeric(base["year"], errors="coerce").astype("Int64")
            base["value_pj_esto"] = pd.to_numeric(base.get("value", pd.Series(index=base.index)), errors="coerce")
            for col in ["esto_flow", "esto_product"]:
                base[col] = base[col].fillna("").astype(str).str.strip()
            esto = (
                base.groupby(key_cols, dropna=False)["value_pj_esto"]
                .sum()
                .reset_index()
            )
            esto = esto[esto_cols]

    merged = leap.merge(ninth, on=key_cols, how="outer")
    merged = merged.merge(esto, on=key_cols, how="outer")
    for col in columns:
        if col not in merged.columns:
            merged[col] = pd.NA if col.startswith("value_pj_") else ""
    return (
        merged[columns]
        .sort_values(["scenario", "year", "esto_flow", "esto_product"], kind="mergesort")
        .reset_index(drop=True)
    )


def build_simple_balance_duplicate_diagnostics(
    *,
    simple_leap_balance: pd.DataFrame,
    simple_ninth_balance: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Flag exact duplicate rows and source-to-ESTO fan-out in simple balance outputs."""
    detail_columns = [
        "dataset",
        "check",
        "severity",
        "scenario",
        "year",
        "leap_sector",
        "leap_fuel",
        "ninth_sector",
        "ninth_fuel",
        "esto_flow",
        "esto_product",
        "row_count",
        "source_key_count",
        "esto_pair_count",
        "distinct_value_count",
        "value_pj_sum",
        "value_pj_min",
        "value_pj_max",
        "esto_pairs",
        "source_keys",
        "note",
    ]
    summary_columns = [
        "dataset",
        "check",
        "severity",
        "group_count",
        "row_count",
        "note",
    ]
    details: list[pd.DataFrame] = []
    summaries: list[dict[str, object]] = []

    def join_unique(series: pd.Series) -> str:
        values = [str(v).strip() for v in series.fillna("").astype(str) if str(v).strip()]
        return "|".join(sorted(set(values)))

    def _prepare(frame: pd.DataFrame, required: list[str]) -> pd.DataFrame:
        working = frame.copy()
        for col in required:
            if col not in working.columns:
                working[col] = ""
        for col in required:
            if col == "year":
                working[col] = pd.to_numeric(working[col], errors="coerce").astype("Int64")
            elif col == "value_pj":
                working[col] = pd.to_numeric(working[col], errors="coerce")
            else:
                working[col] = working[col].fillna("").astype(str).str.strip()
        return working

    def _blank_detail() -> pd.DataFrame:
        return pd.DataFrame(columns=detail_columns)

    def _append_summary(dataset: str, check: str, severity: str, frame: pd.DataFrame, note: str) -> None:
        summaries.append(
            {
                "dataset": dataset,
                "check": check,
                "severity": severity,
                "group_count": int(len(frame)),
                "row_count": int(frame["row_count"].sum()) if "row_count" in frame.columns and not frame.empty else 0,
                "note": note,
            }
        )

    def _exact_duplicates(dataset: str, frame: pd.DataFrame, descriptor_cols: list[str]) -> pd.DataFrame:
        check = "exact_duplicate_descriptor_rows"
        if frame.empty:
            result = _blank_detail()
            _append_summary(dataset, check, "error", result, "Same descriptor columns appear more than once.")
            return result
        key_cols = [col for col in descriptor_cols if col in frame.columns]
        grouped = (
            frame.groupby(key_cols, dropna=False)
            .agg(
                row_count=("value_pj", "size"),
                distinct_value_count=("value_pj", lambda s: int(pd.to_numeric(s, errors="coerce").nunique(dropna=True))),
                value_pj_sum=("value_pj", "sum"),
                value_pj_min=("value_pj", "min"),
                value_pj_max=("value_pj", "max"),
            )
            .reset_index()
        )
        grouped = grouped[grouped["row_count"].gt(1)].copy()
        if grouped.empty:
            result = _blank_detail()
        else:
            grouped["dataset"] = dataset
            grouped["check"] = check
            grouped["severity"] = "error"
            grouped["source_key_count"] = 1
            grouped["esto_pair_count"] = 1
            grouped["esto_pairs"] = grouped.get("esto_flow", "").astype(str) + " | " + grouped.get("esto_product", "").astype(str)
            if dataset == "leap":
                grouped["source_keys"] = grouped.get("leap_sector", "").astype(str) + " | " + grouped.get("leap_fuel", "").astype(str)
            else:
                grouped["source_keys"] = grouped.get("ninth_sector", "").astype(str) + " | " + grouped.get("ninth_fuel", "").astype(str)
            grouped["note"] = "Duplicate descriptor rows will double-count values unless intentionally duplicated upstream."
            result = grouped.reindex(columns=detail_columns)
        _append_summary(dataset, check, "error", result, "Same descriptor columns appear more than once.")
        return result

    def _source_fanout(
        dataset: str,
        frame: pd.DataFrame,
        source_cols: list[str],
    ) -> pd.DataFrame:
        check = "source_maps_to_multiple_esto_pairs"
        if frame.empty:
            result = _blank_detail()
            _append_summary(dataset, check, "warning", result, "One source key maps to multiple ESTO flow/product pairs.")
            return result
        working = frame.copy()
        working["esto_pair"] = working["esto_flow"].astype(str) + " | " + working["esto_product"].astype(str)
        grouped = (
            working.groupby(source_cols, dropna=False)
            .agg(
                row_count=("value_pj", "size"),
                esto_pair_count=("esto_pair", lambda s: int(s.nunique(dropna=True))),
                distinct_value_count=("value_pj", lambda s: int(pd.to_numeric(s, errors="coerce").nunique(dropna=True))),
                value_pj_sum=("value_pj", "sum"),
                value_pj_min=("value_pj", "min"),
                value_pj_max=("value_pj", "max"),
                esto_flow=("esto_flow", join_unique),
                esto_product=("esto_product", join_unique),
                esto_pairs=("esto_pair", join_unique),
            )
            .reset_index()
        )
        grouped = grouped[grouped["esto_pair_count"].gt(1)].copy()
        if grouped.empty:
            result = _blank_detail()
        else:
            grouped["dataset"] = dataset
            grouped["check"] = check
            grouped["severity"] = "warning"
            grouped["source_key_count"] = 1
            if dataset == "leap":
                grouped["source_keys"] = grouped.get("leap_sector", "").astype(str) + " | " + grouped.get("leap_fuel", "").astype(str)
            else:
                grouped["source_keys"] = grouped.get("ninth_sector", "").astype(str) + " | " + grouped.get("ninth_fuel", "").astype(str)
            grouped["note"] = (
                "One source key fans out to multiple ESTO categories. "
                "This may be valid one-to-many mapping, but repeated values can indicate faulty mapping."
            )
            result = grouped.reindex(columns=detail_columns)
        _append_summary(dataset, check, "warning", result, "One source key maps to multiple ESTO flow/product pairs.")
        return result

    def _esto_multi_source(
        dataset: str,
        frame: pd.DataFrame,
        source_cols: list[str],
    ) -> pd.DataFrame:
        check = "esto_pair_has_multiple_source_keys"
        if frame.empty:
            result = _blank_detail()
            _append_summary(dataset, check, "info", result, "Multiple source keys contribute to one ESTO pair.")
            return result
        working = frame.copy()
        source_label_cols = [col for col in source_cols if col not in {"scenario", "year"}]
        working["source_key"] = working[source_label_cols].fillna("").astype(str).agg(" | ".join, axis=1)
        grouped = (
            working.groupby(["scenario", "year", "esto_flow", "esto_product"], dropna=False)
            .agg(
                row_count=("value_pj", "size"),
                source_key_count=("source_key", lambda s: int(s.nunique(dropna=True))),
                distinct_value_count=("value_pj", lambda s: int(pd.to_numeric(s, errors="coerce").nunique(dropna=True))),
                value_pj_sum=("value_pj", "sum"),
                value_pj_min=("value_pj", "min"),
                value_pj_max=("value_pj", "max"),
                source_keys=("source_key", join_unique),
            )
            .reset_index()
        )
        grouped = grouped[grouped["source_key_count"].gt(1)].copy()
        if grouped.empty:
            result = _blank_detail()
        else:
            grouped["dataset"] = dataset
            grouped["check"] = check
            grouped["severity"] = "info"
            grouped["esto_pair_count"] = 1
            grouped["esto_pairs"] = grouped["esto_flow"].astype(str) + " | " + grouped["esto_product"].astype(str)
            grouped["note"] = (
                "Multiple source keys aggregate into one ESTO category. "
                "This can be expected, but it is useful for audit."
            )
            result = grouped.reindex(columns=detail_columns)
        _append_summary(dataset, check, "info", result, "Multiple source keys contribute to one ESTO pair.")
        return result

    leap = _prepare(
        simple_leap_balance,
        ["scenario", "year", "leap_sector", "leap_fuel", "esto_flow", "esto_product", "value_pj"],
    )
    ninth = _prepare(
        simple_ninth_balance,
        ["scenario", "year", "ninth_sector", "ninth_fuel", "esto_flow", "esto_product", "value_pj"],
    )

    details.append(
        _exact_duplicates(
            "leap",
            leap,
            ["scenario", "year", "leap_sector", "leap_fuel", "esto_flow", "esto_product"],
        )
    )
    details.append(
        _source_fanout(
            "leap",
            leap,
            ["scenario", "year", "leap_sector", "leap_fuel"],
        )
    )
    details.append(
        _esto_multi_source(
            "leap",
            leap,
            ["scenario", "year", "leap_sector", "leap_fuel"],
        )
    )
    details.append(
        _exact_duplicates(
            "9th",
            ninth,
            ["scenario", "year", "ninth_sector", "ninth_fuel", "esto_flow", "esto_product"],
        )
    )
    details.append(
        _source_fanout(
            "9th",
            ninth,
            ["scenario", "year", "ninth_sector", "ninth_fuel"],
        )
    )
    details.append(
        _esto_multi_source(
            "9th",
            ninth,
            ["scenario", "year", "ninth_sector", "ninth_fuel"],
        )
    )

    details = [detail for detail in details if not detail.empty and not detail.isna().all(axis=None)]
    detail_df = pd.concat(details, ignore_index=True) if details else _blank_detail()
    if not detail_df.empty:
        detail_df = detail_df.sort_values(
            ["severity", "dataset", "check", "scenario", "year", "esto_flow", "esto_product"],
            kind="mergesort",
        ).reset_index(drop=True)
    summary_df = pd.DataFrame(summaries, columns=summary_columns)
    return summary_df, detail_df


def build_mapped_ninth_to_esto_balance_rows(
    *,
    comparison_long: pd.DataFrame,
    mapping_status: pd.DataFrame,
) -> pd.DataFrame:
    """Return compact 9th comparator rows on the ESTO flow/product axis."""
    columns = [
        "scenario",
        "year",
        "ninth_sector",
        "esto_flow",
        "ninth_fuel",
        "esto_product",
        "value_pj",
        "subtotal",
    ]
    if comparison_long.empty or mapping_status.empty:
        return pd.DataFrame(columns=columns)

    keys = ["sheet", "measure", "fuel_label"]
    projection = comparison_long.copy()
    projection = projection[projection.get("source", "").fillna("").astype(str).str.strip().eq("projection")].copy()
    if projection.empty:
        return pd.DataFrame(columns=columns)

    def join_unique(series: pd.Series) -> str:
        values = [str(v).strip() for v in series.fillna("").astype(str) if str(v).strip()]
        return "|".join(sorted(set(values)))

    meta = mapping_status.copy()
    for col in keys + ["sector_code_9th", "ninth_fuel_code", "esto_flow", "esto_product"]:
        if col not in meta.columns:
            meta[col] = ""
        meta[col] = meta[col].fillna("").astype(str).str.strip()
    meta = (
        meta.groupby(keys, as_index=False)
        .agg(
            ninth_sector=("sector_code_9th", join_unique),
            ninth_fuel=("ninth_fuel_code", join_unique),
            esto_flow=("esto_flow", join_unique),
            esto_product=("esto_product", join_unique),
        )
    )

    out = projection.merge(meta, on=keys, how="left")
    out["year"] = pd.to_numeric(out["year"], errors="coerce").astype("Int64")
    out["value_pj"] = pd.to_numeric(out.get("value", pd.Series(index=out.index)), errors="coerce")
    out["subtotal"] = False
    out = out[columns]
    return out.sort_values(["scenario", "year", "esto_flow", "esto_product"], kind="mergesort").reset_index(drop=True)


def build_mapping_lineage_audit_table(
    *,
    pre_group_leap_mapped: pd.DataFrame,
    pre_group_incomplete_rows: pd.DataFrame,
    comparison_long_full: pd.DataFrame,
    mapping_status: pd.DataFrame,
    mapped_ninth_to_esto_balance_rows: pd.DataFrame | None = None,
    target_years: Sequence[int] = (2022, 2023, 2060),
) -> pd.DataFrame:
    """
    Build a long-format mapping lineage audit table filtered to target_years.

    Each row documents one source observation feeding into an ESTO-pair comparison:

    dataset values:
      LEAP          — a pre-aggregation LEAP balance row that mapped to an ESTO pair.
                      The leap_sector_name_full_path and raw_leap_fuel_name columns show
                      exactly which LEAP workbook row contributed to the aggregated value.
                      is_subtotal and esto_pair_is_subtotal flag the source mapping's
                      subtotal relationships.
      LEAP_unmapped — a LEAP balance row that could not be matched to any ESTO pair.
                      esto_flow / esto_product are empty; the row appears here so you can
                      see what data was silently excluded from comparisons.
      9th           — a 9th-projection row mapped onto the ESTO-pair axis.
      ESTO          — the ESTO base-year value for this ESTO pair.

    This table is the primary tool for tracing why LEAP, ESTO, and 9th values for a
    given flow/product pair differ. Sort or filter on (esto_flow, esto_product) to see
    everything that was aggregated to produce a given chart point.
    """
    target_year_set = {int(y) for y in target_years}

    OUTPUT_COLUMNS = [
        "dataset",
        "scenario",
        "year",
        "esto_flow",
        "esto_product",
        "source_sector",
        "source_fuel",
        "value_pj",
        "is_subtotal",
        "esto_pair_is_subtotal",
        "ninth_pair_is_subtotal",
        "remove_row",
        "remove_row_reason",
        "pair_mapping_cardinality",
        "subtotal_alignment",
        "sheet",
        "measure",
        "fuel_label",
        "chart_group_key",
    ]

    def _col(df: pd.DataFrame, name: str, default: object = "") -> pd.Series:
        if name in df.columns:
            return df[name]
        return pd.Series(default, index=df.index)

    parts: list[pd.DataFrame] = []

    # ---- LEAP mapped rows (pre-aggregation) ----
    if not pre_group_leap_mapped.empty:
        leap = pre_group_leap_mapped.copy()
        leap["year"] = pd.to_numeric(_col(leap, "year", pd.NA), errors="coerce").astype("Int64")
        leap = leap[leap["year"].isin(target_year_set)].copy()
        if not leap.empty:
            # Prefer the original workbook identifiers for maximum traceability.
            # Fall back to normalised names if originals are absent.
            source_sector = _col(leap, "leap_sector_name_full_path")
            if source_sector.fillna("").astype(str).str.strip().eq("").all():
                source_sector = _col(leap, "leap_sector_name")
            source_fuel = _col(leap, "raw_leap_fuel_name")
            if source_fuel.fillna("").astype(str).str.strip().eq("").all():
                source_fuel = _col(leap, "leap_fuel_name")

            parts.append(pd.DataFrame({
                "dataset": "LEAP",
                "scenario": _col(leap, "scenario").fillna("").astype(str),
                "year": leap["year"],
                "esto_flow": _col(leap, "esto_flow").fillna("").astype(str),
                "esto_product": _col(leap, "esto_product").fillna("").astype(str),
                "source_sector": source_sector.fillna("").astype(str),
                "source_fuel": source_fuel.fillna("").astype(str),
                "value_pj": pd.to_numeric(_col(leap, "value_petajoule", pd.NA), errors="coerce"),
                "is_subtotal": _col(leap, "is_subtotal", False).fillna(False).astype(bool),
                "esto_pair_is_subtotal": _col(leap, "esto_pair_is_subtotal", False).fillna(False).astype(bool),
                "ninth_pair_is_subtotal": False,
                "remove_row": _col(leap, "remove_row", False).fillna(False).astype(bool),
                "remove_row_reason": _col(leap, "remove_row_reason", "").fillna("").astype(str),
                "pair_mapping_cardinality": _col(leap, "pair_mapping_cardinality", "").fillna("").astype(str),
                "subtotal_alignment": _col(leap, "subtotal_alignment", "").fillna("").astype(str),
            }))

    # ---- LEAP unmapped rows (no ESTO pair found) ----
    if not pre_group_incomplete_rows.empty:
        unmapped = pre_group_incomplete_rows.copy()
        unmapped["year"] = pd.to_numeric(_col(unmapped, "year", pd.NA), errors="coerce").astype("Int64")
        unmapped = unmapped[unmapped["year"].isin(target_year_set)].copy()
        if not unmapped.empty:
            source_sector = _col(unmapped, "leap_sector_name_full_path")
            if source_sector.fillna("").astype(str).str.strip().eq("").all():
                source_sector = _col(unmapped, "leap_sector_name")
            source_fuel = _col(unmapped, "raw_leap_fuel_name")
            if source_fuel.fillna("").astype(str).str.strip().eq("").all():
                source_fuel = _col(unmapped, "leap_fuel_name")

            parts.append(pd.DataFrame({
                "dataset": "LEAP_unmapped",
                "scenario": _col(unmapped, "scenario").fillna("").astype(str),
                "year": unmapped["year"],
                "esto_flow": "",
                "esto_product": "",
                "source_sector": source_sector.fillna("").astype(str),
                "source_fuel": source_fuel.fillna("").astype(str),
                "value_pj": pd.to_numeric(_col(unmapped, "value_petajoule", pd.NA), errors="coerce"),
                "is_subtotal": _col(unmapped, "is_subtotal", False).fillna(False).astype(bool),
                "esto_pair_is_subtotal": False,
                "ninth_pair_is_subtotal": False,
                "remove_row": _col(unmapped, "remove_row", False).fillna(False).astype(bool),
                "remove_row_reason": _col(unmapped, "remove_row_reason", "").fillna("").astype(str),
                "pair_mapping_cardinality": _col(unmapped, "pair_mapping_cardinality", "").fillna("").astype(str),
                "subtotal_alignment": "",
            }))

    # ---- 9th projection rows ----
    if mapped_ninth_to_esto_balance_rows is not None and not mapped_ninth_to_esto_balance_rows.empty:
        ninth = mapped_ninth_to_esto_balance_rows.copy()
        ninth["year"] = pd.to_numeric(_col(ninth, "year", pd.NA), errors="coerce").astype("Int64")
        ninth = ninth[ninth["year"].isin(target_year_set)].copy()
        if not ninth.empty:
            parts.append(pd.DataFrame({
                "dataset": "9th",
                "scenario": _col(ninth, "scenario").fillna("").astype(str),
                "year": ninth["year"],
                "esto_flow": _col(ninth, "esto_flow").fillna("").astype(str),
                "esto_product": _col(ninth, "esto_product").fillna("").astype(str),
                "source_sector": _col(ninth, "ninth_sector").fillna("").astype(str),
                "source_fuel": _col(ninth, "ninth_fuel").fillna("").astype(str),
                "value_pj": pd.to_numeric(_col(ninth, "value_pj", pd.NA), errors="coerce"),
                "is_subtotal": _col(ninth, "subtotal", False).fillna(False).astype(bool),
                "esto_pair_is_subtotal": False,
                "ninth_pair_is_subtotal": False,
                "remove_row": False,
                "remove_row_reason": "",
                "pair_mapping_cardinality": "",
                "subtotal_alignment": "",
                "sheet": _col(ninth, "sheet").fillna("").astype(str),
                "measure": _col(ninth, "measure").fillna("").astype(str),
                "fuel_label": _col(ninth, "fuel_label").fillna("").astype(str),
                "chart_group_key": _col(ninth, "chart_group_key").fillna("").astype(str),
            }))
    elif not comparison_long_full.empty and not mapping_status.empty:
        keys = ["sheet", "measure", "fuel_label"]
        proj = comparison_long_full.copy()
        proj["source"] = _col(proj, "source", "").fillna("").astype(str).str.strip()
        proj = proj[proj["source"] == "projection"].copy()
        proj["year"] = pd.to_numeric(_col(proj, "year", pd.NA), errors="coerce").astype("Int64")
        proj = proj[proj["year"].isin(target_year_set)].copy()
        if not proj.empty:
            meta = mapping_status.copy()
            for col in keys + ["sector_code_9th", "ninth_fuel_code", "esto_flow", "esto_product"]:
                if col not in meta.columns:
                    meta[col] = ""
                meta[col] = meta[col].fillna("").astype(str).str.strip()

            def _join_unique(series: pd.Series) -> str:
                vals = [str(v).strip() for v in series.fillna("").astype(str) if str(v).strip()]
                return "|".join(sorted(set(vals)))

            def _join_pipe_tokens_unique(series: pd.Series) -> str:
                tokens: list[str] = []
                for value in series.fillna("").astype(str):
                    for token in value.split("|"):
                        token = token.strip()
                        if token:
                            tokens.append(token)
                return "|".join(sorted(set(tokens)))

            meta_agg = (
                meta.groupby(keys, as_index=False)
                .agg(
                    ninth_sector=("sector_code_9th", _join_pipe_tokens_unique),
                    ninth_fuel=("ninth_fuel_code", _join_unique),
                    esto_flow=("esto_flow", _join_unique),
                    esto_product=("esto_product", _join_unique),
                )
            )
            for col in keys:
                if col not in proj.columns:
                    proj[col] = ""
                proj[col] = proj[col].fillna("").astype(str).str.strip()
            proj = proj.merge(meta_agg, on=keys, how="left")
            for col in ["ninth_sector", "ninth_fuel", "esto_flow", "esto_product"]:
                proj[col] = proj[col].fillna("").astype(str)

            parts.append(pd.DataFrame({
                "dataset": "9th",
                "scenario": _col(proj, "scenario").fillna("").astype(str),
                "year": proj["year"],
                "esto_flow": proj["esto_flow"],
                "esto_product": proj["esto_product"],
                "source_sector": proj["ninth_sector"],
                "source_fuel": proj["ninth_fuel"],
                "value_pj": pd.to_numeric(_col(proj, "value", pd.NA), errors="coerce"),
                "is_subtotal": False,
                "esto_pair_is_subtotal": False,
                "ninth_pair_is_subtotal": False,
                "remove_row": False,
                "remove_row_reason": "",
                "pair_mapping_cardinality": "",
                "subtotal_alignment": "",
            }))

    # ---- ESTO base-year rows ----
    if not comparison_long_full.empty and not mapping_status.empty:
        keys = ["sheet", "measure", "fuel_label"]
        base = comparison_long_full.copy()
        base["source"] = _col(base, "source", "").fillna("").astype(str).str.strip()
        base = base[base["source"] == "base"].copy()
        base["year"] = pd.to_numeric(_col(base, "year", pd.NA), errors="coerce").astype("Int64")
        base = base[base["year"].isin(target_year_set)].copy()
        if not base.empty:
            meta = mapping_status.copy()
            for col in keys + ["esto_flow", "esto_product"]:
                if col not in meta.columns:
                    meta[col] = ""
                meta[col] = meta[col].fillna("").astype(str).str.strip()
            meta_deduped = meta[keys + ["esto_flow", "esto_product"]].drop_duplicates()
            for col in keys:
                if col not in base.columns:
                    base[col] = ""
                base[col] = base[col].fillna("").astype(str).str.strip()
            base = base.merge(meta_deduped, on=keys, how="left")
            for col in ["esto_flow", "esto_product"]:
                base[col] = base[col].fillna("").astype(str)

            esto_rows = pd.DataFrame({
                "dataset": "ESTO",
                "scenario": "ESTO",
                "year": base["year"],
                "esto_flow": base["esto_flow"],
                "esto_product": base["esto_product"],
                "source_sector": "",
                "source_fuel": "",
                "value_pj": pd.to_numeric(_col(base, "value", pd.NA), errors="coerce"),
                "is_subtotal": False,
                "esto_pair_is_subtotal": False,
                "ninth_pair_is_subtotal": False,
                "remove_row": False,
                "remove_row_reason": "",
                "pair_mapping_cardinality": "",
                "subtotal_alignment": "",
            })
            # One ESTO value per pair per year — deduplicate after scenario collapse.
            esto_rows = esto_rows.drop_duplicates(
                subset=["year", "esto_flow", "esto_product"]
            )
            parts.append(esto_rows)

    if not parts:
        return pd.DataFrame(columns=OUTPUT_COLUMNS)

    combined = pd.concat(parts, ignore_index=True, sort=False)
    for col in OUTPUT_COLUMNS:
        if col not in combined.columns:
            combined[col] = ""
    combined["year"] = pd.to_numeric(combined["year"], errors="coerce").astype("Int64")
    combined["value_pj"] = pd.to_numeric(combined["value_pj"], errors="coerce")
    for bool_col in ["is_subtotal", "esto_pair_is_subtotal", "ninth_pair_is_subtotal", "remove_row"]:
        combined[bool_col] = combined[bool_col].fillna(False).astype(bool)

    for str_col in [
        "dataset", "scenario", "esto_flow", "esto_product", "source_sector", "source_fuel",
        "remove_row_reason", "pair_mapping_cardinality", "subtotal_alignment",
    ]:
        combined[str_col] = combined[str_col].fillna("").astype(str)

    return (
        combined[OUTPUT_COLUMNS]
        .sort_values(
            ["esto_flow", "esto_product", "dataset", "year", "source_sector", "source_fuel", "scenario"],
            kind="mergesort",
        )
        .reset_index(drop=True)
    )


def _rows_to_issue_records(
    frame: pd.DataFrame,
    *,
    reason: str,
    details: str = "",
) -> pd.DataFrame:
    if frame.empty:
        return pd.DataFrame(
            columns=[
                "reason",
                "details",
                "scenario",
                "year",
                "source_sheet",
                "leap_sector_name_full_path",
                "leap_flow",
                "leap_flow_name",
                "leap_product",
                "leap_product_name",
                "mapping_failed",
                "mapping_key_sector",
                "mapping_key_fuel",
                "mapping_candidate_rule",
                "esto_flow",
                "esto_product",
                "value_petajoule",
            ]
        )

    def text_col(name: str, default: str = "") -> pd.Series:
        if name in frame.columns:
            return frame[name].fillna("").astype(str)
        return pd.Series(default, index=frame.index, dtype="string").astype(str)

    def numeric_col(name: str) -> pd.Series:
        if name in frame.columns:
            values = frame[name]
        else:
            values = pd.Series(index=frame.index, dtype="float64")
        return pd.to_numeric(values, errors="coerce")

    out = pd.DataFrame(
        {
            "reason": reason,
            "details": details,
            "scenario": text_col("scenario"),
            "year": numeric_col("year").astype("Int64"),
            "source_sheet": text_col("source_sheet"),
            "leap_sector_name_full_path": text_col("leap_sector_name_full_path"),
            "leap_flow": text_col("leap_sector"),
            "leap_flow_name": text_col("leap_sector_name"),
            "leap_product": text_col("leap_fuel"),
            "leap_product_name": text_col("leap_fuel_name"),
            "mapping_failed": text_col("mapping_failed"),
            "mapping_key_sector": text_col("mapping_key_sector"),
            "mapping_key_fuel": text_col("mapping_key_fuel"),
            "mapping_candidate_rule": text_col("mapping_candidate_rule"),
            "esto_flow": text_col("esto_flow"),
            "esto_product": text_col("esto_product"),
            "value_petajoule": numeric_col("value_petajoule"),
        }
    )
    return out


def write_balance_missing_mapping_candidates(
    *,
    runtime_issues: pd.DataFrame,
    output_path: Path | str,
    mapping_workbook_path: Path | str | None = None,
) -> str:
    """Write missing LEAP mapping rows in the same shape as the mapping workbook."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    mapping_workbook_path = Path(mapping_workbook_path) if mapping_workbook_path is not None else None

    def _empty_frame(columns: list[str]) -> pd.DataFrame:
        return pd.DataFrame(columns=columns)

    def _text_col(frame: pd.DataFrame, name: str) -> pd.Series:
        if name not in frame.columns:
            return pd.Series("", index=frame.index, dtype="string")
        return frame[name].fillna("").astype(str).str.strip()

    def _join_unique(values: pd.Series) -> str:
        cleaned = [str(value).strip() for value in values.fillna("").astype(str) if str(value).strip()]
        return "|".join(sorted(set(cleaned)))

    def _year_span(values: pd.Series) -> str:
        years = pd.to_numeric(values, errors="coerce").dropna().astype(int)
        if years.empty:
            return ""
        unique_years = sorted(set(years.tolist()))
        if len(unique_years) == 1:
            return str(unique_years[0])
        return f"{unique_years[0]}-{unique_years[-1]}"

    def _norm(value: object) -> str:
        return " ".join(str(value or "").strip().lower().split())

    def _truthy(value: object) -> bool:
        return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on", "t"}

    def _removed_only_source_keys(sheet_name: str) -> set[tuple[str, str]]:
        if mapping_workbook_path is None or not mapping_workbook_path.exists():
            return set()
        try:
            mapping = pd.read_excel(mapping_workbook_path, sheet_name=sheet_name, dtype=str).fillna("")
        except Exception:
            return set()
        for col in ["leap_sector_name_full_path", "raw_leap_fuel_name", "remove_row"]:
            if col not in mapping.columns:
                mapping[col] = ""
        mapping["source_key"] = list(
            zip(
                mapping["leap_sector_name_full_path"].map(_norm),
                mapping["raw_leap_fuel_name"].map(_norm),
            )
        )
        removed_keys = {
            key for key in mapping.loc[mapping["remove_row"].map(_truthy), "source_key"] if key[0] and key[1]
        }
        active_keys = {
            key for key in mapping.loc[~mapping["remove_row"].map(_truthy), "source_key"] if key[0] and key[1]
        }
        return removed_keys - active_keys

    removed_only_keys = {
        "esto": _removed_only_source_keys("leap_combined_esto"),
        "ninth": _removed_only_source_keys("leap_combined_ninth"),
    }

    def _candidate_rows(issues: pd.DataFrame, *, target: str) -> pd.DataFrame:
        columns = BALANCE_ESTO_MAPPING_COLUMNS if target == "esto" else BALANCE_NINTH_MAPPING_COLUMNS
        if issues.empty:
            return _empty_frame(columns)

        work = issues.copy()
        for col in [
            "reason",
            "details",
            "scenario",
            "year",
            "source_sheet",
            "leap_sector_name_full_path",
            "leap_flow_name",
            "leap_product_name",
            "mapping_failed",
            "mapping_key_sector",
            "mapping_key_fuel",
            "value_petajoule",
        ]:
            if col not in work.columns:
                work[col] = ""

        reason = _text_col(work, "reason").str.lower()
        mapping_failed = _text_col(work, "mapping_failed").str.lower()
        if target == "esto":
            target_mask = reason.str.contains("esto|mapping", regex=True) & (
                mapping_failed.eq("")
                | mapping_failed.str.contains("esto", regex=False)
                | reason.str.contains("missing_esto|esto_pair", regex=True)
            )
        else:
            target_mask = reason.str.contains("ninth|mapping", regex=True) & (
                mapping_failed.str.contains("ninth", regex=False)
                | reason.str.contains("missing_ninth|ninth_pair", regex=True)
            )
        work = work.loc[target_mask].copy()
        if work.empty:
            return _empty_frame(columns)

        mapping_key_sector = _text_col(work, "mapping_key_sector")
        mapping_key_fuel = _text_col(work, "mapping_key_fuel")
        full_path = _text_col(work, "leap_sector_name_full_path")
        fuel_name = _text_col(work, "leap_product_name")
        work["leap_sector_name_full_path"] = mapping_key_sector.where(mapping_key_sector.ne(""), full_path)
        work["raw_leap_fuel_name"] = mapping_key_fuel.where(mapping_key_fuel.ne(""), fuel_name)
        work["leap_sector_name_original"] = _text_col(work, "leap_flow_name")
        missing_original = work["leap_sector_name_original"].eq("")
        work.loc[missing_original, "leap_sector_name_original"] = (
            work.loc[missing_original, "leap_sector_name_full_path"].astype(str).str.split("/").str[-1]
        )
        work["value_petajoule"] = pd.to_numeric(work["value_petajoule"], errors="coerce").fillna(0.0)

        grouped = (
            work.groupby(
                ["leap_sector_name_original", "leap_sector_name_full_path", "raw_leap_fuel_name"],
                dropna=False,
            )
            .agg(
                _source_sheet=("source_sheet", _join_unique),
                details=("details", _join_unique),
                reasons=("reason", _join_unique),
                scenarios=("scenario", _join_unique),
                year_span=("year", _year_span),
                rows=("reason", "size"),
                value=("value_petajoule", "sum"),
            )
            .reset_index()
        )
        grouped["workflow_issue_detail"] = grouped.apply(
            lambda row: "; ".join(
                part
                for part in [
                    f"reasons={row['reasons']}" if row["reasons"] else "",
                    f"rows={int(row['rows'])}",
                    f"scenarios={row['scenarios']}" if row["scenarios"] else "",
                    f"years={row['year_span']}" if row["year_span"] else "",
                    row["details"],
                ]
                if part
            ),
            axis=1,
        )

        out = _empty_frame(columns)
        out["leap_sector_name_original"] = grouped["leap_sector_name_original"]
        out["leap_sector_name_full_path"] = grouped["leap_sector_name_full_path"]
        out["raw_leap_fuel_name"] = grouped["raw_leap_fuel_name"]
        out["value"] = grouped["value"]
        for bool_col in ["leap_is_subtotal", "subtotal_mismatch_is_ok", "remove_row"]:
            out[bool_col] = False
        if target == "esto":
            out["esto_flow"] = ""
            out["esto_product"] = ""
            out["esto_pair_is_subtotal"] = False
            out["esto_pair_abs_sum"] = ""
        else:
            out["ninth_sector"] = ""
            out["ninth_fuel"] = ""
            out["ninth_pair_is_subtotal"] = False
            out["ninth_pair_abs_sum"] = ""
        out["pair_mapping_cardinality"] = ""
        out["subtotal_alignment"] = ""
        out["remove_row_reason"] = ""
        removed_only = set(removed_only_keys.get(target, set()))
        if removed_only:
            source_key = list(
                zip(
                    out["leap_sector_name_full_path"].map(_norm),
                    out["raw_leap_fuel_name"].map(_norm),
                )
            )
            removed_mask = pd.Series([key in removed_only for key in source_key], index=out.index)
            sheet_name = "leap_combined_esto" if target == "esto" else "leap_combined_ninth"
            out.loc[removed_mask, "remove_row"] = True
            out.loc[removed_mask, "remove_row_reason"] = (
                f"this row exists in the {sheet_name} mapping but has its remove_row set to true "
                "so it is not available"
            )
        out = out[columns]
        return (
            out.assign(_sort_abs_value=pd.to_numeric(out["value"], errors="coerce").abs())
            .sort_values(
                ["_sort_abs_value", "leap_sector_name_full_path", "raw_leap_fuel_name"],
                ascending=[False, True, True],
                kind="mergesort",
            )
            .drop(columns=["_sort_abs_value"])
            .reset_index(drop=True)
        )

    issues = pd.DataFrame() if runtime_issues is None or runtime_issues.empty else runtime_issues.copy()
    esto_candidates = _candidate_rows(issues, target="esto")
    ninth_candidates = _candidate_rows(issues, target="ninth")
    summary = pd.DataFrame(
        [
            {"sheet": "esto_missing_paths", "candidate_rows": int(len(esto_candidates))},
            {"sheet": "ninth_missing_paths", "candidate_rows": int(len(ninth_candidates))},
        ]
    )
    with pd.ExcelWriter(output_path) as writer:
        esto_candidates.to_excel(writer, sheet_name="esto_missing_paths", index=False)
        ninth_candidates.to_excel(writer, sheet_name="ninth_missing_paths", index=False)
        summary.to_excel(writer, sheet_name="summary", index=False)
    return str(output_path)


def write_runtime_missing_pair_summary(
    *,
    runtime_issues: pd.DataFrame,
    output_path: Path | str,
) -> str:
    """Write compact non-zero missing-pair summaries from runtime issue records."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    issues = pd.DataFrame() if runtime_issues is None or runtime_issues.empty else runtime_issues.copy()

    def text_col(frame: pd.DataFrame, name: str) -> pd.Series:
        if name not in frame.columns:
            return pd.Series("", index=frame.index, dtype="string")
        return frame[name].fillna("").astype(str).str.strip()

    def pair_summary(frame: pd.DataFrame, sector_col: str, fuel_col: str, out_sector_col: str, out_fuel_col: str) -> pd.DataFrame:
        columns = [out_sector_col, out_fuel_col, "value_petajoule"]
        if frame.empty:
            return pd.DataFrame(columns=columns)
        work = frame.copy()
        work[out_sector_col] = text_col(work, sector_col)
        work[out_fuel_col] = text_col(work, fuel_col)
        work["value_petajoule"] = pd.to_numeric(work.get("value_petajoule"), errors="coerce").fillna(0.0)
        work = work[work[out_sector_col].ne("") & work[out_fuel_col].ne("") & work["value_petajoule"].ne(0)].copy()
        if work.empty:
            return pd.DataFrame(columns=columns)
        out = (
            work.groupby([out_sector_col, out_fuel_col], dropna=False)["value_petajoule"]
            .sum()
            .reset_index()
        )
        out = out[out["value_petajoule"].ne(0)].copy()
        if out.empty:
            return pd.DataFrame(columns=columns)
        return (
            out.assign(_sort_abs_value=out["value_petajoule"].abs())
            .sort_values(["_sort_abs_value", out_sector_col, out_fuel_col], ascending=[False, True, True], kind="mergesort")
            .drop(columns=["_sort_abs_value"])
            .reset_index(drop=True)
        )

    if issues.empty:
        leap_missing = pd.DataFrame(columns=["leap_sector_name_full_path", "raw_leap_fuel_name", "value_petajoule"])
        ninth_missing = pd.DataFrame(columns=["leap_sector_name_full_path", "raw_leap_fuel_name", "value_petajoule"])
        esto_missing = pd.DataFrame(columns=["leap_sector_name_full_path", "raw_leap_fuel_name", "value_petajoule"])
    else:
        reason = text_col(issues, "reason").str.lower()
        mapping_failed = text_col(issues, "mapping_failed").str.lower()
        mapping_key_sector = text_col(issues, "mapping_key_sector")
        mapping_key_fuel = text_col(issues, "mapping_key_fuel")
        issues["summary_leap_sector"] = mapping_key_sector.where(
            mapping_key_sector.ne(""),
            text_col(issues, "leap_sector_name_full_path"),
        )
        issues["summary_leap_fuel"] = mapping_key_fuel.where(
            mapping_key_fuel.ne(""),
            text_col(issues, "leap_product_name"),
        )

        leap_mask = reason.eq("flow_not_in_structure_config")
        ninth_mask = reason.str.contains("ninth|mapping", regex=True) & mapping_failed.str.contains("ninth", regex=False)
        esto_mask = reason.str.contains("esto|mapping", regex=True) & (
            mapping_failed.str.contains("esto", regex=False)
            | reason.str.contains("missing_esto|esto_pair", regex=True)
        )

        leap_missing = pair_summary(
            issues.loc[leap_mask],
            "summary_leap_sector",
            "summary_leap_fuel",
            "leap_sector_name_full_path",
            "raw_leap_fuel_name",
        )
        ninth_missing = pair_summary(
            issues.loc[ninth_mask],
            "summary_leap_sector",
            "summary_leap_fuel",
            "leap_sector_name_full_path",
            "raw_leap_fuel_name",
        )
        esto_missing = pair_summary(
            issues.loc[esto_mask],
            "summary_leap_sector",
            "summary_leap_fuel",
            "leap_sector_name_full_path",
            "raw_leap_fuel_name",
        )

    summary = pd.DataFrame(
        [
            {"sheet": "missing_leap_pairs_nonzero", "pair_rows": int(len(leap_missing))},
            {"sheet": "missing_ninth_pairs_nonzero", "pair_rows": int(len(ninth_missing))},
            {"sheet": "missing_esto_pairs_nonzero", "pair_rows": int(len(esto_missing))},
        ]
    )

    with pd.ExcelWriter(output_path) as writer:
        leap_missing.to_excel(writer, sheet_name="missing_leap_pairs_nonzero", index=False)
        ninth_missing.to_excel(writer, sheet_name="missing_ninth_pairs_nonzero", index=False)
        esto_missing.to_excel(writer, sheet_name="missing_esto_pairs_nonzero", index=False)
        summary.to_excel(writer, sheet_name="summary", index=False)
    return str(output_path)


def write_dashboard_comparator_pair_coverage(
    *,
    mapping_status: pd.DataFrame,
    dashboard_exposure: pd.DataFrame | None = None,
    chart_group_exposure: pd.DataFrame | None = None,
    all_chart_groups: pd.DataFrame | None = None,
    base_df: pd.DataFrame,
    ninth_df: pd.DataFrame,
    output_path: Path | str,
    base_economy: str,
    projection_economy: str,
    base_year: int,
    projection_years: Sequence[int],
    scenarios: Sequence[str],
    runtime_issues: pd.DataFrame | None = None,
    chart_navigation_guide_path: Path | str | None = None,
    mapping_workbook_path: Path | str | None = None,
    mapping_sheet_name: str = "leap_combined_esto",
    fail_on_unsafe_esto_dashboard_pairs: bool = True,
) -> str:
    """Write compact coverage of raw ESTO/9th pairs represented by dashboard mappings."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    def clean(value: object) -> str:
        text = str(value or "").strip()
        return "" if text.lower() in {"", "nan", "none", "x"} else text

    def norm(value: object) -> str:
        return " ".join(clean(value).lower().split())

    def first_non_empty(row: pd.Series, columns: list[str]) -> str:
        for col in reversed(columns):
            value = clean(row.get(col, ""))
            if value:
                return value
        return ""

    def flow_code(value: object) -> str:
        match = re.match(r"^\s*(\d+(?:[._]\d+)*)", str(value or "").strip())
        return match.group(1).replace("_", ".") if match else ""

    def truthy(value: object) -> bool:
        return str(value or "").strip().lower() in {"1", "true", "t", "yes", "y"}

    def flow_is_under(flow: object, possible_parent: object) -> bool:
        child_code = flow_code(flow)
        parent_code = flow_code(possible_parent)
        if not child_code or not parent_code:
            return norm(flow) == norm(possible_parent)
        return child_code == parent_code or child_code.startswith(parent_code + ".")

    def sort_pair_values(frame: pd.DataFrame, pair_cols: list[str]) -> pd.DataFrame:
        if frame.empty:
            return frame
        return (
            frame.assign(_sort_abs_value=pd.to_numeric(frame["value_petajoule"], errors="coerce").abs())
            .sort_values(["_sort_abs_value", *pair_cols], ascending=[False, True, True], kind="mergesort")
            .drop(columns=["_sort_abs_value"])
            .reset_index(drop=True)
        )

    def sort_duplicate_pair_values(frame: pd.DataFrame, pair_cols: list[str]) -> pd.DataFrame:
        if frame.empty:
            return frame
        return (
            frame.assign(
                _sort_row_count=pd.to_numeric(frame["dashboard_row_count"], errors="coerce").fillna(0),
                _sort_abs_value=pd.to_numeric(frame["value_petajoule"], errors="coerce").abs(),
            )
            .sort_values(
                ["_sort_row_count", "_sort_abs_value", *pair_cols],
                ascending=[False, False, *([True] * len(pair_cols))],
                kind="mergesort",
            )
            .drop(columns=["_sort_row_count", "_sort_abs_value"])
            .reset_index(drop=True)
        )

    def sort_chart_groups(frame: pd.DataFrame) -> pd.DataFrame:
        if frame.empty:
            return frame
        for col in ["dashboard_path", "chart_group_key", "sheet", "measure", "fuel_label", "chart_file", "chart_group_id"]:
            if col not in frame.columns:
                frame[col] = ""
        return (
            frame.sort_values(
                ["dashboard_path", "chart_group_key", "sheet", "measure", "fuel_label", "chart_file", "chart_group_id"],
                kind="mergesort",
            ).reset_index(drop=True)
        )

    template = _load_dashboard_template_allowlist(chart_navigation_guide_path)
    template_exact_esto_pairs: set[tuple[str, str]] = set()
    template_total_flows: set[str] = set()
    template_aggregate_flows: set[str] = set()
    if template:
        template_aggregate_groups: list[dict[str, object]] = []
        template_measure_default = str((template.get("defaults") or {}).get("measure", "")).strip()
        def _walk_template(node: dict[str, Any]) -> None:
            for aggregate in _dashboard_template_aggregate_specs(node, default_measure=template_measure_default):
                source_flows = [clean(flow) for flow in list(aggregate.get("source_flows", []) or []) if clean(flow)]
                if source_flows:
                    template_aggregate_groups.append(aggregate)
                for flow in source_flows:
                    template_aggregate_flows.add(flow)
            for spec in _dashboard_template_graph_specs(node, default_measure=template_measure_default):
                for flow in list(spec.get("esto_flows", []) or [spec.get("esto_flow", "")]):
                    flow_text = clean(flow)
                    for product in list(spec.get("products", []) or []):
                        product_text = clean(product)
                        if flow_text and product_text:
                            template_exact_esto_pairs.add((norm(flow_text), norm(product_text)))
                            if _product_is_total(product_text):
                                template_total_flows.add(flow_text)
            for key, child in node.items():
                if _dashboard_template_is_reserved_key(key) or not isinstance(child, dict):
                    continue
                _walk_template(child)

        _walk_template(template)
    else:
        template_aggregate_groups = []

    def runtime_leap_missing_pairs() -> pd.DataFrame:
        columns = ["leap_sector_name_full_path", "raw_leap_fuel_name", "value_petajoule"]
        if runtime_issues is None or runtime_issues.empty:
            return pd.DataFrame(columns=columns)
        issues = runtime_issues.copy()
        for col in [
            "reason",
            "leap_sector_name_full_path",
            "leap_product_name",
            "mapping_key_sector",
            "mapping_key_fuel",
            "value_petajoule",
        ]:
            if col not in issues.columns:
                issues[col] = ""
        reason = issues["reason"].fillna("").astype(str).str.strip().str.lower()
        issues = issues[reason.eq("flow_not_in_structure_config")].copy()
        if issues.empty:
            return pd.DataFrame(columns=columns)
        mapping_key_sector = issues["mapping_key_sector"].fillna("").astype(str).str.strip()
        mapping_key_fuel = issues["mapping_key_fuel"].fillna("").astype(str).str.strip()
        issues["leap_sector_name_full_path"] = mapping_key_sector.where(
            mapping_key_sector.ne(""),
            issues["leap_sector_name_full_path"].fillna("").astype(str).str.strip(),
        )
        issues["raw_leap_fuel_name"] = mapping_key_fuel.where(
            mapping_key_fuel.ne(""),
            issues["leap_product_name"].fillna("").astype(str).str.strip(),
        )
        issues["value_petajoule"] = pd.to_numeric(issues["value_petajoule"], errors="coerce").fillna(0.0)
        issues = issues[
            issues["leap_sector_name_full_path"].ne("")
            & issues["raw_leap_fuel_name"].ne("")
            & issues["value_petajoule"].ne(0)
        ].copy()
        if issues.empty:
            return pd.DataFrame(columns=columns)
        exact_pairs = set(meta_esto_mapped) | set(template_exact_esto_pairs)
        total_flows = set(template_total_flows)
        total_flows.update(
            str(row.esto_flow).strip()
            for row in meta[["esto_flow", "esto_product", "fuel_label"]].itertuples(index=False)
            if clean(row.esto_flow) and (_product_is_total(row.esto_product) or norm(row.fuel_label) == "total")
        )
        aggregate_flows = set(template_aggregate_flows)

        def _covered_by_dashboard(row: pd.Series) -> bool:
            flow = clean(row.get("esto_flow", ""))
            product = clean(row.get("esto_product", ""))
            if not flow or not product:
                return False
            if (norm(flow), norm(product)) in exact_pairs:
                return True
            if any(flow_is_under(flow, covered_flow) for covered_flow in total_flows):
                return True
            if any(flow_is_under(flow, covered_flow) for covered_flow in aggregate_flows):
                return True
            return False

        issues = issues[~issues.apply(_covered_by_dashboard, axis=1)].copy()
        if issues.empty:
            return pd.DataFrame(columns=columns)
        out = (
            issues.groupby(["leap_sector_name_full_path", "raw_leap_fuel_name"], as_index=False)["value_petajoule"]
            .sum()
        )
        out = out[out["value_petajoule"].ne(0)].copy()
        return sort_pair_values(out, ["leap_sector_name_full_path", "raw_leap_fuel_name"])

    meta = pd.DataFrame() if mapping_status is None or mapping_status.empty else mapping_status.copy()
    for col in [
        "sheet",
        "page_key",
        "page_label",
        "chart_group_key",
        "chart_group_label",
        "measure",
        "fuel_label",
        "sector_code_9th",
        "ninth_fuel_code",
        "esto_flow",
        "esto_product",
    ]:
        if col not in meta.columns:
            meta[col] = ""
        meta[col] = meta[col].fillna("").astype(str).str.strip()
    meta = _backfill_dashboard_hierarchy(meta)
    meta_esto_mapped = {
        (norm(row["esto_flow"]), norm(row["esto_product"]))
        for _, row in meta.iterrows()
        if norm(row["esto_flow"]) and norm(row["esto_product"])
    }

    # Raw ninth pairs with any non-zero value in the requested scenario/year/economy slice.
    ninth_pairs = pd.DataFrame(columns=["ninth_sector", "ninth_fuel", "value_petajoule"])
    if ninth_df is not None and not ninth_df.empty:
        work = ninth_df.copy()
        for col in ["scenarios", "economy", "sectors", "sub1sectors", "sub2sectors", "sub3sectors", "sub4sectors", "fuels", "subfuels"]:
            if col not in work.columns:
                work[col] = ""
        for col in ["subtotal_layout", "subtotal_results"]:
            if col not in work.columns:
                work[col] = False
            work[col] = work[col].fillna(False).map(truthy)
        year_cols = [str(year) for year in projection_years if str(year) in work.columns]
        scenario_set = {str(s).strip().lower() for s in scenarios}
        work = work[
            work["economy"].fillna("").astype(str).str.strip().eq(str(projection_economy))
            & work["scenarios"].fillna("").astype(str).str.strip().str.lower().isin(scenario_set)
        ].copy()
        if year_cols and not work.empty:
            work["ninth_sector"] = work.apply(
                lambda row: first_non_empty(row, ["sectors", "sub1sectors", "sub2sectors", "sub3sectors", "sub4sectors"]),
                axis=1,
            )
            work["ninth_fuel"] = work.apply(lambda row: first_non_empty(row, ["fuels", "subfuels"]), axis=1)
            values = work[year_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
            work["value_petajoule"] = values.sum(axis=1)
            work = work[work["ninth_sector"].ne("") & work["ninth_fuel"].ne("") & work["value_petajoule"].ne(0)].copy()
            ninth_pairs = (
                work.groupby(["ninth_sector", "ninth_fuel"], as_index=False)
                .agg(
                    value_petajoule=("value_petajoule", "sum"),
                    subtotal_layout=("subtotal_layout", "max"),
                    subtotal_results=("subtotal_results", "max"),
                )
                .pipe(sort_pair_values, ["ninth_sector", "ninth_fuel"])
            )
            if not ninth_pairs.empty:
                ninth_pairs["ninth_pair_is_subtotal"] = (
                    ninth_pairs["subtotal_layout"].fillna(False).astype(bool)
                    | ninth_pairs["subtotal_results"].fillna(False).astype(bool)
                )
                ninth_pairs["subtotal_alignment"] = ninth_pairs["ninth_pair_is_subtotal"].map(
                    lambda flag: "subtotal" if flag else "non_subtotal"
                )
    for col, default in [
        ("subtotal_layout", False),
        ("subtotal_results", False),
        ("ninth_pair_is_subtotal", False),
        ("subtotal_alignment", ""),
    ]:
        if col not in ninth_pairs.columns:
            ninth_pairs[col] = default

    # Raw ESTO pairs with non-zero base-year value.
    esto_pairs = pd.DataFrame(columns=["esto_flow", "esto_product", "value_petajoule"])
    if base_df is not None and not base_df.empty:
        work = base_df.copy()
        for col in ["economy", "flows", "products", str(base_year)]:
            if col not in work.columns:
                work[col] = ""
        if "is_subtotal" not in work.columns:
            work["is_subtotal"] = False
        work["is_subtotal"] = work["is_subtotal"].fillna(False).map(truthy)
        work = work[work["economy"].fillna("").astype(str).str.strip().eq(str(base_economy))].copy()
        work["esto_flow"] = work["flows"].fillna("").astype(str).str.strip()
        work["esto_product"] = work["products"].fillna("").astype(str).str.strip()
        work["value_petajoule"] = pd.to_numeric(work[str(base_year)], errors="coerce").fillna(0.0)
        work = work[work["esto_flow"].ne("") & work["esto_product"].ne("") & work["value_petajoule"].ne(0)].copy()
        esto_pairs = (
            work.groupby(["esto_flow", "esto_product"], as_index=False)
            .agg(
                value_petajoule=("value_petajoule", "sum"),
                esto_pair_is_subtotal=("is_subtotal", "max"),
            )
            .pipe(sort_pair_values, ["esto_flow", "esto_product"])
        )
        if not esto_pairs.empty and "subtotal_alignment" not in esto_pairs.columns:
            esto_pairs["subtotal_alignment"] = esto_pairs["esto_pair_is_subtotal"].map(
                lambda flag: "subtotal" if bool(flag) else "non_subtotal"
            )
    for col, default in [
        ("esto_pair_is_subtotal", False),
        ("subtotal_alignment", ""),
    ]:
        if col not in esto_pairs.columns:
            esto_pairs[col] = default

    def duplicate_dashboard_pairs(pair_cols: list[str], raw_pairs: pd.DataFrame, value_cols: list[str]) -> pd.DataFrame:
        valid = meta.copy()
        for col in pair_cols:
            valid = valid[valid[col].ne("")]
        if valid.empty:
            return pd.DataFrame(columns=[*pair_cols, "dashboard_row_count", "value_petajoule", "dashboard_rows"])
        valid["dashboard_row"] = valid[["page_label", "chart_group_label", "measure", "fuel_label"]].fillna("").astype(str).agg(" | ".join, axis=1)
        grouped = (
            valid.groupby(pair_cols, as_index=False)
            .agg(
                dashboard_row_count=("dashboard_row", "nunique"),
                dashboard_rows=("dashboard_row", lambda s: " || ".join(sorted(set(s))[:20])),
            )
        )
        grouped = grouped[grouped["dashboard_row_count"].gt(1)].copy()
        if grouped.empty:
            return pd.DataFrame(columns=[*pair_cols, "dashboard_row_count", "value_petajoule", "dashboard_rows"])
        value_lookup = raw_pairs.copy()
        if value_lookup.empty:
            grouped["value_petajoule"] = 0.0
        else:
            grouped = grouped.merge(value_lookup[[*value_cols, "value_petajoule"]], left_on=pair_cols, right_on=value_cols, how="left")
            drop_cols = [col for col in value_cols if col not in pair_cols]
            if drop_cols:
                grouped = grouped.drop(columns=drop_cols)
            grouped["value_petajoule"] = pd.to_numeric(grouped["value_petajoule"], errors="coerce").fillna(0.0)
        return sort_duplicate_pair_values(
            grouped[[*pair_cols, "dashboard_row_count", "value_petajoule", "dashboard_rows"]],
            pair_cols,
        )

    def bool_is_true(value: object) -> bool:
        return str(value or "").strip().lower() in {"1", "true", "t", "yes", "y"}

    def active_mapping_pairs(sheet_name: str, left_col: str, right_col: str) -> set[tuple[str, str]]:
        if mapping_workbook_path is None:
            return set()
        try:
            mapping = pd.read_excel(mapping_workbook_path, sheet_name=sheet_name, dtype=str).fillna("")
        except Exception:
            return set()
        for col in [left_col, right_col, "remove_row"]:
            if col not in mapping.columns:
                mapping[col] = ""
            mapping[col] = mapping[col].fillna("").astype(str).str.strip()
        mapping = mapping[
            ~mapping["remove_row"].map(bool_is_true)
            & mapping[left_col].ne("")
            & mapping[right_col].ne("")
        ].copy()
        return {
            (norm(row[left_col]), norm(row[right_col]))
            for _, row in mapping.iterrows()
            if norm(row[left_col]) and norm(row[right_col])
        }

    def value_for_esto_pair(flow: object, product: object) -> float:
        if esto_pairs.empty:
            return 0.0
        pair_key = (norm(flow), norm(product))
        mask = esto_pairs.apply(lambda row: (norm(row["esto_flow"]), norm(row["esto_product"])) == pair_key, axis=1)
        values = pd.to_numeric(esto_pairs.loc[mask, "value_petajoule"], errors="coerce").fillna(0.0)
        return float(values.sum()) if not values.empty else 0.0

    def dashboard_exposure_rows() -> pd.DataFrame:
        source = dashboard_exposure if dashboard_exposure is not None and not dashboard_exposure.empty else meta
        exposure = source.copy()
        for col in [
            "sheet",
            "page_key",
            "page_label",
            "chart_group_key",
            "chart_group_label",
            "measure",
            "fuel_label",
            "source",
            "esto_flow",
            "esto_product",
            "chart_group_id",
            "dashboard_path",
            "chart_file",
            "section_id",
            "section_label",
            "entry_kind",
        ]:
            if col not in exposure.columns:
                exposure[col] = ""
            exposure[col] = exposure[col].fillna("").astype(str).str.strip()
        exposure = _backfill_dashboard_hierarchy(exposure)
        exposure = exposure[exposure["esto_flow"].ne("") & exposure["esto_product"].ne("")].copy()
        exposure["dashboard_row"] = exposure[["page_label", "chart_group_label", "measure", "fuel_label"]].fillna("").astype(str).agg(" | ".join, axis=1)
        return exposure.drop_duplicates(
            [
                "source",
                "sheet",
                "chart_group_key",
                "measure",
                "fuel_label",
                "esto_flow",
                "esto_product",
                "dashboard_row",
                "chart_group_id",
                "dashboard_path",
                "chart_file",
                "section_id",
                "section_label",
                "entry_kind",
            ]
        ).reset_index(drop=True)

    exposure_rows = dashboard_exposure_rows()

    def _clean_chart_group_frame(frame: pd.DataFrame | None, *, include_exposed_flag: bool = False) -> pd.DataFrame:
        columns = [
            "chart_group_id",
            "dashboard_path",
            "sheet",
            "page_key",
            "page_label",
            "chart_group_key",
            "chart_group_label",
            "measure",
            "fuel_label",
            "chart_file",
            "section_id",
            "section_label",
            "entry_kind",
        ]
        if include_exposed_flag:
            columns.append("exposed_in_dashboard")
        if frame is None or frame.empty:
            return pd.DataFrame(columns=columns)
        out = frame.copy()
        rename_map = {"fuel": "fuel_label"}
        out = out.rename(columns=rename_map)
        for col in columns:
            if col not in out.columns:
                out[col] = ""
        string_cols = [col for col in columns if col != "exposed_in_dashboard"]
        for col in string_cols:
            out[col] = out[col].fillna("").astype(str).str.strip()
        if include_exposed_flag:
            out["exposed_in_dashboard"] = out["exposed_in_dashboard"].fillna(False).astype(bool)
        out = out[out["chart_group_id"].ne("")].copy()
        return sort_chart_groups(out[columns].drop_duplicates())

    def chart_group_exposure_sheet() -> pd.DataFrame:
        columns = [
            "chart_group_id",
            "dashboard_path",
            "sheet",
            "page_key",
            "page_label",
            "chart_group_key",
            "chart_group_label",
            "measure",
            "fuel_label",
            "chart_file",
            "section_id",
            "section_label",
            "entry_kind",
        ]
        cleaned = _clean_chart_group_frame(chart_group_exposure)
        if not cleaned.empty:
            return cleaned
        if exposure_rows.empty:
            return pd.DataFrame(columns=columns)
        available = exposure_rows.copy()
        if available["chart_group_id"].eq("").all():
            return pd.DataFrame(columns=columns)
        out = available[available["chart_group_id"].ne("")][columns].drop_duplicates().copy()
        return sort_chart_groups(out)

    def all_chart_groups_sheet() -> pd.DataFrame:
        columns = [
            "chart_group_id",
            "dashboard_path",
            "sheet",
            "measure",
            "fuel_label",
            "chart_file",
            "section_id",
            "section_label",
            "entry_kind",
            "exposed_in_dashboard",
        ]
        cleaned = _clean_chart_group_frame(all_chart_groups, include_exposed_flag=True)
        if not cleaned.empty:
            return cleaned[columns]
        return pd.DataFrame(columns=columns)

    exposure_ninth_mapped = {
        (norm(row.sector_code_9th), norm(row.ninth_fuel_code))
        for row in exposure_rows.itertuples(index=False)
        if hasattr(row, "sector_code_9th")
        and hasattr(row, "ninth_fuel_code")
        and norm(getattr(row, "sector_code_9th", ""))
        and norm(getattr(row, "ninth_fuel_code", ""))
    }
    if not exposure_ninth_mapped:
        exposure_ninth_mapped = {
            (norm(row["sector_code_9th"]), norm(row["ninth_fuel_code"]))
            for _, row in meta.iterrows()
            if norm(row["sector_code_9th"]) and norm(row["ninth_fuel_code"])
        }
    exposure_ninth_mapped |= active_mapping_pairs("leap_combined_ninth", "ninth_sector", "ninth_fuel")

    exposure_esto_mapped = {
        (norm(row.esto_flow), norm(row.esto_product))
        for row in exposure_rows.itertuples(index=False)
        if norm(getattr(row, "esto_flow", "")) and norm(getattr(row, "esto_product", ""))
    }
    if not exposure_esto_mapped:
        exposure_esto_mapped = {
            (norm(row["esto_flow"]), norm(row["esto_product"]))
            for _, row in meta.iterrows()
            if norm(row["esto_flow"]) and norm(row["esto_product"])
        }
    exposure_esto_mapped |= active_mapping_pairs(mapping_sheet_name, "esto_flow", "esto_product")

    missing_ninth = ninth_pairs[
        ~ninth_pairs.apply(lambda row: (norm(row["ninth_sector"]), norm(row["ninth_fuel"])) in exposure_ninth_mapped, axis=1)
    ].copy() if not ninth_pairs.empty else ninth_pairs.copy()
    missing_ninth = sort_pair_values(missing_ninth, ["ninth_sector", "ninth_fuel"])

    missing_esto = esto_pairs[
        ~esto_pairs.apply(lambda row: (norm(row["esto_flow"]), norm(row["esto_product"])) in exposure_esto_mapped, axis=1)
    ].copy() if not esto_pairs.empty else esto_pairs.copy()
    missing_esto = sort_pair_values(missing_esto, ["esto_flow", "esto_product"])

    def duplicate_esto_dashboard_exposure() -> pd.DataFrame:
        columns = [
            "issue",
            "measure",
            "esto_flow",
            "esto_product",
            "component_id",
            "mapping_type",
            "chart_group_count",
            "dashboard_row_count",
            "value_petajoule",
            "chart_group_ids",
            "chart_files",
            "dashboard_paths",
            "dashboard_rows",
        ]
        if exposure_rows.empty:
            return pd.DataFrame(columns=columns)
        exposure = exposure_rows.copy()
        base_exposure = exposure[exposure["source"].str.lower().eq("base")].copy()
        if not base_exposure.empty:
            exposure = base_exposure
        if exposure["chart_group_id"].eq("").all():
            exposure["chart_group_id"] = exposure["dashboard_row"]
        if "chart_file" not in exposure.columns:
            exposure["chart_file"] = ""
        if "dashboard_path" not in exposure.columns:
            exposure["dashboard_path"] = ""
        grouped = (
            exposure.groupby(["measure", "esto_flow", "esto_product"], as_index=False)
            .agg(
                chart_group_count=("chart_group_id", "nunique"),
                dashboard_row_count=("dashboard_row", "nunique"),
                chart_group_ids=("chart_group_id", lambda s: " || ".join(sorted(set(str(v) for v in s if str(v).strip()))[:50])),
                chart_files=("chart_file", lambda s: " || ".join(sorted(set(str(v) for v in s if str(v).strip()))[:50])),
                dashboard_paths=("dashboard_path", lambda s: " || ".join(sorted(set(str(v) for v in s if str(v).strip()))[:50])),
                dashboard_rows=("dashboard_row", lambda s: " || ".join(sorted(set(str(v) for v in s if str(v).strip()))[:50])),
            )
        )
        grouped = grouped[grouped["chart_group_count"].gt(1)].copy()
        if grouped.empty:
            return pd.DataFrame(columns=columns)
        grouped["issue"] = "duplicate_final_esto_pair_exposure"
        grouped["value_petajoule"] = grouped.apply(lambda row: value_for_esto_pair(row["esto_flow"], row["esto_product"]), axis=1)
        grouped["component_id"] = grouped.apply(
            lambda row: esto_pair_component_lookup.get((norm(row["esto_flow"]), norm(row["esto_product"])), {}).get("component_id", ""),
            axis=1,
        )
        grouped["mapping_type"] = grouped.apply(
            lambda row: esto_pair_component_lookup.get((norm(row["esto_flow"]), norm(row["esto_product"])), {}).get("mapping_type", ""),
            axis=1,
        )
        return sort_duplicate_pair_values(grouped[columns], ["measure", "esto_flow", "esto_product"])

    def mapping_components() -> list[dict[str, object]]:
        if mapping_workbook_path is None:
            return []
        try:
            mapping = pd.read_excel(mapping_workbook_path, sheet_name=mapping_sheet_name, dtype=str).fillna("")
        except Exception as exc:  # pragma: no cover - surfaced in workflow output.
            raise RuntimeError(f"Could not read ESTO mapping sheet {mapping_sheet_name!r} from {mapping_workbook_path}: {exc}") from exc
        required = ["leap_sector_name_full_path", "raw_leap_fuel_name", "esto_flow", "esto_product", "remove_row"]
        for col in required:
            if col not in mapping.columns:
                mapping[col] = ""
            mapping[col] = mapping[col].fillna("").astype(str).str.strip()
        mapping = mapping[
            ~mapping["remove_row"].map(bool_is_true)
            & mapping["leap_sector_name_full_path"].ne("")
            & mapping["raw_leap_fuel_name"].ne("")
            & mapping["esto_flow"].ne("")
            & mapping["esto_product"].ne("")
        ].copy()
        if mapping.empty:
            return []
        mapping["_leap_pair_norm"] = mapping.apply(
            lambda row: (norm(row["leap_sector_name_full_path"]), norm(row["raw_leap_fuel_name"])),
            axis=1,
        )
        mapping["_esto_pair_norm"] = mapping.apply(
            lambda row: (norm(row["esto_flow"]), norm(row["esto_product"])),
            axis=1,
        )
        mapping["_leap_pair"] = mapping.apply(
            lambda row: (clean(row["leap_sector_name_full_path"]), clean(row["raw_leap_fuel_name"])),
            axis=1,
        )
        mapping["_esto_pair"] = mapping.apply(lambda row: (clean(row["esto_flow"]), clean(row["esto_product"])), axis=1)

        parent: dict[tuple[str, str, str], tuple[str, str, str]] = {}

        def find(node: tuple[str, str, str]) -> tuple[str, str, str]:
            parent.setdefault(node, node)
            while parent[node] != node:
                parent[node] = parent[parent[node]]
                node = parent[node]
            return node

        def union(left: tuple[str, str, str], right: tuple[str, str, str]) -> None:
            left_root = find(left)
            right_root = find(right)
            if left_root != right_root:
                parent[right_root] = left_root

        for _, row in mapping.iterrows():
            leap_pair_norm = row["_leap_pair_norm"]
            esto_pair_norm = row["_esto_pair_norm"]
            leap_node = ("leap", leap_pair_norm[0], leap_pair_norm[1])
            esto_node = ("esto", esto_pair_norm[0], esto_pair_norm[1])
            union(leap_node, esto_node)

        components_by_root: dict[tuple[str, str, str], dict[str, set[tuple[str, str]]]] = {}
        for _, row in mapping.iterrows():
            leap_pair_norm = row["_leap_pair_norm"]
            root = find(("leap", leap_pair_norm[0], leap_pair_norm[1]))
            component = components_by_root.setdefault(
                root,
                {"leap": set(), "leap_norm": set(), "esto": set(), "esto_norm": set()},
            )
            component["leap"].add(row["_leap_pair"])
            component["leap_norm"].add(row["_leap_pair_norm"])
            component["esto"].add(row["_esto_pair"])
            component["esto_norm"].add(row["_esto_pair_norm"])

        out: list[dict[str, object]] = []
        for idx, component in enumerate(components_by_root.values(), start=1):
            leap_count = len(component["leap_norm"])
            esto_count = len(component["esto_norm"])
            if leap_count == 1 and esto_count == 1:
                mapping_type = "one_to_one"
            elif leap_count > 1 and esto_count == 1:
                mapping_type = "many_to_one"
            elif leap_count == 1 and esto_count > 1:
                mapping_type = "one_to_many"
            else:
                mapping_type = "many_to_many"
            out.append({"component_id": f"component_{idx:05d}", "mapping_type": mapping_type, **component})
        return out

    components = mapping_components()
    esto_pair_component_lookup = {
        pair_norm: {"component_id": str(component["component_id"]), "mapping_type": str(component["mapping_type"])}
        for component in components
        for pair_norm in component["esto_norm"]
    }

    def allowed_many_to_one_esto_dashboard_exposure(all_duplicate_esto: pd.DataFrame) -> pd.DataFrame:
        if all_duplicate_esto.empty:
            return all_duplicate_esto.copy()
        allowed = all_duplicate_esto[all_duplicate_esto["mapping_type"].astype(str).eq("many_to_one")].copy()
        if allowed.empty:
            return allowed
        allowed["issue"] = "allowed_many_to_one_esto_exposure"
        return sort_duplicate_pair_values(
            allowed,
            ["measure", "esto_flow", "esto_product"],
        )

    def split_esto_mapping_components() -> pd.DataFrame:
        columns = [
            "issue",
            "component_id",
            "mapping_type",
            "chart_group_id",
            "dashboard_path",
            "chart_file",
            "dashboard_row",
            "esto_pair_count",
            "leap_pair_count",
            "exposed_esto_pairs",
            "missing_esto_pairs_from_row",
            "component_esto_pairs",
            "component_leap_pairs",
            "value_petajoule",
        ]
        if exposure_rows.empty or not components:
            return pd.DataFrame(columns=columns)
        exposure = exposure_rows.copy()
        if "mapping_record_type" in exposure.columns:
            direct_exposure = exposure[~exposure["mapping_record_type"].astype(str).str.strip().eq("derived_total")].copy()
            if not direct_exposure.empty:
                exposure = direct_exposure
        exposure = exposure[exposure["chart_group_id"].ne("")].copy()
        if exposure.empty:
            return pd.DataFrame(columns=columns)
        group_pairs: dict[str, dict[str, object]] = {}
        for row in exposure.itertuples(index=False):
            group_id = str(getattr(row, "chart_group_id", "")).strip() or str(row.dashboard_row).strip()
            if not group_id:
                continue
            info = group_pairs.setdefault(
                group_id,
                {
                    "pairs": set(),
                    "dashboard_row": str(row.dashboard_row).strip(),
                    "dashboard_path": str(getattr(row, "dashboard_path", "")).strip(),
                    "chart_file": str(getattr(row, "chart_file", "")).strip(),
                },
            )
            info["pairs"].add((norm(row.esto_flow), norm(row.esto_product)))

        rows: list[dict[str, object]] = []
        for component in components:
            if component["mapping_type"] not in {"one_to_many", "many_to_many"}:
                continue
            component_pairs_norm = set(component["esto_norm"])
            component_pairs_display = sorted(component["esto"], key=lambda item: (norm(item[0]), norm(item[1])))
            leap_pairs_display = sorted(component["leap"], key=lambda item: (norm(item[0]), norm(item[1])))
            component_products_norm = {norm(product) for _, product in component_pairs_display if norm(product)}
            component_flows_norm = {norm(flow) for flow, _ in component_pairs_display if norm(flow)}
            if len(component_products_norm) == 1:
                only_product_norm = next(iter(component_products_norm))
                if _product_is_total(only_product_norm) and any(
                    component_flows_norm.issubset(set(group["source_flows_norm"]))
                    and str(group["fuel_norm"]) == "total"
                    for group in template_aggregate_groups
                ):
                    continue
            for chart_group_id, group_info in sorted(group_pairs.items()):
                exposed_pairs = set(group_info.get("pairs", set()))
                touched = exposed_pairs & component_pairs_norm
                if not touched or touched == component_pairs_norm:
                    continue
                missing = component_pairs_norm - touched
                rows.append(
                    {
                        "issue": "linked_esto_component_split_across_dashboard_rows",
                        "component_id": component["component_id"],
                        "mapping_type": component["mapping_type"],
                        "chart_group_id": chart_group_id,
                        "dashboard_path": str(group_info.get("dashboard_path", "")),
                        "chart_file": str(group_info.get("chart_file", "")),
                        "dashboard_row": str(group_info.get("dashboard_row", "")),
                        "esto_pair_count": len(component_pairs_norm),
                        "leap_pair_count": len(component["leap_norm"]),
                        "exposed_esto_pairs": " || ".join(f"{flow} | {product}" for flow, product in sorted(touched)),
                        "missing_esto_pairs_from_row": " || ".join(f"{flow} | {product}" for flow, product in sorted(missing)),
                        "component_esto_pairs": " || ".join(f"{flow} | {product}" for flow, product in component_pairs_display),
                        "component_leap_pairs": " || ".join(f"{sector} | {fuel}" for sector, fuel in leap_pairs_display[:50]),
                        "value_petajoule": sum(value_for_esto_pair(flow, product) for flow, product in component_pairs_display),
                    }
                )
        if not rows:
            return pd.DataFrame(columns=columns)
        out = pd.DataFrame(rows, columns=columns)
        return (
            out.assign(
                _sort_abs_value=pd.to_numeric(out["value_petajoule"], errors="coerce").abs(),
                _sort_pair_count=pd.to_numeric(out["esto_pair_count"], errors="coerce").fillna(0),
            )
            .sort_values(
                ["_sort_abs_value", "_sort_pair_count", "mapping_type", "component_id", "dashboard_path", "dashboard_row"],
                ascending=[False, False, True, True, True, True],
                kind="mergesort",
            )
            .drop(columns=["_sort_abs_value", "_sort_pair_count"])
            .reset_index(drop=True)
        )

    def parent_child_esto_exposure_warnings() -> pd.DataFrame:
        columns = [
            "issue",
            "measure",
            "esto_product",
            "parent_esto_flow",
            "child_esto_flow",
            "parent_dashboard_rows",
            "child_dashboard_rows",
            "value_petajoule",
        ]
        if exposure_rows.empty:
            return pd.DataFrame(columns=columns)
        exposure = exposure_rows.copy()
        base_exposure = exposure[exposure["source"].str.lower().eq("base")].copy()
        if not base_exposure.empty:
            exposure = base_exposure
        pair_rows = (
            exposure.groupby(["measure", "esto_flow", "esto_product"])["dashboard_row"]
            .agg(lambda s: " || ".join(sorted(set(str(v) for v in s if str(v).strip()))[:50]))
            .reset_index()
        )
        rows: list[dict[str, object]] = []
        for (measure, product), group in pair_rows.groupby(["measure", "esto_product"]):
            flow_rows = list(group.itertuples(index=False))
            for parent in flow_rows:
                parent_flow = str(parent.esto_flow)
                if not flow_code(parent_flow):
                    continue
                for child in flow_rows:
                    child_flow = str(child.esto_flow)
                    if parent_flow == child_flow:
                        continue
                    if flow_is_under(child_flow, parent_flow):
                        rows.append(
                            {
                                "issue": "parent_child_esto_flow_exposure_warning",
                                "measure": measure,
                                "esto_product": product,
                                "parent_esto_flow": parent_flow,
                                "child_esto_flow": child_flow,
                                "parent_dashboard_rows": parent.dashboard_row,
                                "child_dashboard_rows": child.dashboard_row,
                                "value_petajoule": value_for_esto_pair(child_flow, product),
                            }
                        )
        if not rows:
            return pd.DataFrame(columns=columns)
        out = pd.DataFrame(rows, columns=columns).drop_duplicates().reset_index(drop=True)
        return (
            out.assign(_sort_abs_value=pd.to_numeric(out["value_petajoule"], errors="coerce").abs())
            .sort_values(
                ["_sort_abs_value", "measure", "esto_product", "parent_esto_flow", "child_esto_flow"],
                ascending=[False, True, True, True, True],
                kind="mergesort",
            )
            .drop(columns=["_sort_abs_value"])
            .reset_index(drop=True)
        )

    duplicate_ninth = duplicate_dashboard_pairs(["sector_code_9th", "ninth_fuel_code"], ninth_pairs.rename(columns={"ninth_sector": "sector_code_9th", "ninth_fuel": "ninth_fuel_code"}), ["sector_code_9th", "ninth_fuel_code"])
    all_duplicate_esto = duplicate_esto_dashboard_exposure()
    allowed_many_to_one_esto = allowed_many_to_one_esto_dashboard_exposure(all_duplicate_esto)
    duplicate_esto = all_duplicate_esto[
        ~all_duplicate_esto["mapping_type"].astype(str).eq("many_to_one")
    ].copy()
    split_components = split_esto_mapping_components()
    parent_child_warnings = parent_child_esto_exposure_warnings()
    missing_leap = runtime_leap_missing_pairs()
    chart_group_exposure = chart_group_exposure_sheet()
    all_chart_groups_sheet_df = all_chart_groups_sheet()

    summary_order = [
        "all_chart_groups",
        "chart_group_exposure",
        "missing_leap_pairs_nonzero",
        "missing_ninth_dashboard_pairs",
        "missing_esto_dashboard_pairs",
        "duplicate_ninth_dashboard_pairs",
        "duplicate_esto_dashboard_pairs",
        "allowed_many_to_one_esto",
        "split_esto_mapping_components",
        "parent_child_esto_warnings",
    ]
    summary = (
        pd.DataFrame(
            [
                {"sheet": "all_chart_groups", "pair_rows": int(len(all_chart_groups_sheet_df))},
                {"sheet": "chart_group_exposure", "pair_rows": int(len(chart_group_exposure))},
                {"sheet": "missing_leap_pairs_nonzero", "pair_rows": int(len(missing_leap))},
                {"sheet": "missing_ninth_dashboard_pairs", "pair_rows": int(len(missing_ninth))},
                {"sheet": "missing_esto_dashboard_pairs", "pair_rows": int(len(missing_esto))},
                {"sheet": "duplicate_ninth_dashboard_pairs", "pair_rows": int(len(duplicate_ninth))},
                {"sheet": "duplicate_esto_dashboard_pairs", "pair_rows": int(len(duplicate_esto))},
                {"sheet": "allowed_many_to_one_esto", "pair_rows": int(len(allowed_many_to_one_esto))},
                {"sheet": "split_esto_mapping_components", "pair_rows": int(len(split_components))},
                {"sheet": "parent_child_esto_warnings", "pair_rows": int(len(parent_child_warnings))},
            ]
        )
        .assign(_sort_order=lambda df: df["sheet"].map({name: i for i, name in enumerate(summary_order)}).fillna(len(summary_order)))
        .sort_values(["_sort_order", "sheet"], kind="mergesort")
        .drop(columns=["_sort_order"])
        .reset_index(drop=True)
    )

    with pd.ExcelWriter(output_path) as writer:
        all_chart_groups_sheet_df.to_excel(writer, sheet_name="all_chart_groups", index=False)
        chart_group_exposure.to_excel(writer, sheet_name="chart_group_exposure", index=False)
        missing_leap.to_excel(writer, sheet_name="missing_leap_pairs_nonzero", index=False)
        missing_ninth.to_excel(writer, sheet_name="missing_ninth_dashboard_pairs", index=False)
        missing_esto.to_excel(writer, sheet_name="missing_esto_dashboard_pairs", index=False)
        duplicate_ninth.to_excel(writer, sheet_name="duplicate_ninth_dashboard_pairs", index=False)
        duplicate_esto.to_excel(writer, sheet_name="duplicate_esto_dashboard_pairs", index=False)
        allowed_many_to_one_esto.to_excel(writer, sheet_name="allowed_many_to_one_esto", index=False)
        split_components.to_excel(writer, sheet_name="split_esto_mapping_components", index=False)
        parent_child_warnings.to_excel(writer, sheet_name="parent_child_esto_warnings", index=False)
        summary.to_excel(writer, sheet_name="summary", index=False)
    unsafe_rows = int(len(duplicate_esto) + len(split_components))
    if fail_on_unsafe_esto_dashboard_pairs and unsafe_rows:
        raise RuntimeError(
            "Unsafe ESTO dashboard exposure remains. "
            f"See {output_path} sheets duplicate_esto_dashboard_pairs and split_esto_mapping_components. "
            f"Rows: {unsafe_rows}"
        )
    return str(output_path)


def attach_chart_groups_to_dashboard_exposure(
    dashboard_exposure: pd.DataFrame,
    chart_group_exposure_path: Path | str | None,
    all_chart_groups_path: Path | str | None = None,
) -> pd.DataFrame:
    """Attach rendered chart-group metadata to the per-line dashboard exposure ledger."""
    exposure = pd.DataFrame() if dashboard_exposure is None else dashboard_exposure.copy()
    for col in [
        "chart_group_id",
        "dashboard_path",
        "chart_file",
        "page_key",
        "page_label",
        "chart_group_key",
        "chart_group_label",
        "section_id",
        "section_label",
        "entry_kind",
        "all_chart_group_id",
        "all_dashboard_path",
        "all_chart_file",
        "all_page_key",
        "all_page_label",
        "all_chart_group_key",
        "all_chart_group_label",
        "all_section_id",
        "all_section_label",
        "all_entry_kind",
        "chart_exposed_in_dashboard",
    ]:
        if col not in exposure.columns:
            exposure[col] = ""
    if exposure.empty:
        return exposure

    def _load_chart_groups(path_like: Path | str | None, *, include_exposed: bool = False) -> pd.DataFrame:
        if not path_like:
            return pd.DataFrame()
        chart_path = Path(path_like)
        if not chart_path.exists():
            return pd.DataFrame()
        try:
            chart_groups = pd.read_csv(chart_path, dtype=str).fillna("")
        except Exception:
            return pd.DataFrame()
        if chart_groups.empty:
            return pd.DataFrame()
        chart_groups = chart_groups.rename(columns={"fuel": "fuel_label"})
        needed = [
            "sheet",
            "page_key",
            "page_label",
            "chart_group_key",
            "chart_group_label",
            "measure",
            "fuel_label",
            "chart_group_id",
            "dashboard_path",
            "chart_file",
            "section_id",
            "section_label",
            "entry_kind",
        ]
        if include_exposed:
            needed.append("exposed_in_dashboard")
        for col in needed:
            if col not in chart_groups.columns:
                chart_groups[col] = ""
            chart_groups[col] = chart_groups[col].fillna("").astype(str).str.strip()
        if include_exposed:
            chart_groups["exposed_in_dashboard"] = (
                chart_groups["exposed_in_dashboard"].fillna("").astype(str).str.strip().str.lower().isin({"1", "true", "yes", "y", "t"})
            )
        return chart_groups[needed].drop_duplicates().copy()

    chart_groups = _load_chart_groups(chart_group_exposure_path, include_exposed=False)
    all_chart_groups = _load_chart_groups(all_chart_groups_path, include_exposed=True)

    exposure = _backfill_dashboard_hierarchy(exposure)
    for col in ["sheet", "page_key", "page_label", "chart_group_key", "chart_group_label", "measure", "fuel_label"]:
        if col not in exposure.columns:
            exposure[col] = ""
        exposure[col] = exposure[col].fillna("").astype(str).str.strip()

    preferred_join_cols = ["chart_group_key", "measure", "fuel_label"]
    fallback_join_cols = ["sheet", "measure", "fuel_label"]
    metadata_cols = [
        "chart_group_id",
        "dashboard_path",
        "chart_file",
        "page_key",
        "page_label",
        "chart_group_label",
        "section_id",
        "section_label",
        "entry_kind",
    ]

    def _one_metadata_row_per_key(frame: pd.DataFrame, join_cols: list[str], value_cols: list[str]) -> pd.DataFrame:
        """Keep chart metadata joins row-preserving when a coarse fallback key is duplicated."""
        if frame.empty:
            return frame.copy()
        keep_cols = [*join_cols, *value_cols]
        work = frame[keep_cols].drop_duplicates().copy()
        if work.empty:
            return work
        return work.groupby(join_cols, as_index=False, dropna=False).first()

    merged = exposure.drop(
        columns=metadata_cols
        + [
            "all_chart_group_id",
            "all_dashboard_path",
            "all_chart_file",
            "all_page_key",
            "all_page_label",
            "all_chart_group_key",
            "all_chart_group_label",
            "all_section_id",
            "all_section_label",
            "all_entry_kind",
            "chart_exposed_in_dashboard",
        ],
        errors="ignore",
    ).copy()
    if not chart_groups.empty:
        chart_groups = _backfill_dashboard_hierarchy(chart_groups)
        preferred_meta = _one_metadata_row_per_key(chart_groups, preferred_join_cols, metadata_cols)
        merged = merged.merge(
            preferred_meta,
            on=preferred_join_cols,
            how="left",
        )
        fallback_needed = merged["chart_group_id"].fillna("").astype(str).str.strip().eq("")
        if fallback_needed.any():
            fallback_meta = _one_metadata_row_per_key(chart_groups, fallback_join_cols, metadata_cols)
            fallback = merged.loc[fallback_needed].drop(columns=metadata_cols, errors="ignore").merge(
                fallback_meta,
                on=fallback_join_cols,
                how="left",
            )
            merged.loc[fallback_needed, fallback.columns] = fallback.to_numpy()
    else:
        for col in metadata_cols:
            if col not in merged.columns:
                merged[col] = ""
    for col in metadata_cols:
        if col not in merged.columns:
            merged[col] = ""
        merged[col] = merged[col].fillna("").astype(str).str.strip()

    if not all_chart_groups.empty:
        all_metadata_cols = [
            "all_chart_group_id",
            "all_dashboard_path",
            "all_chart_file",
            "all_page_key",
            "all_page_label",
            "all_chart_group_key",
            "all_chart_group_label",
            "all_section_id",
            "all_section_label",
            "all_entry_kind",
            "chart_exposed_in_dashboard",
        ]
        all_join = all_chart_groups.rename(
            columns={
                "chart_group_id": "all_chart_group_id",
                "dashboard_path": "all_dashboard_path",
                "chart_file": "all_chart_file",
                "section_id": "all_section_id",
                "section_label": "all_section_label",
                "entry_kind": "all_entry_kind",
                "exposed_in_dashboard": "chart_exposed_in_dashboard",
            }
        )
        for source_col, target_col in [
            ("page_key", "all_page_key"),
            ("page_label", "all_page_label"),
            ("chart_group_key", "all_chart_group_key"),
            ("chart_group_label", "all_chart_group_label"),
        ]:
            all_join[target_col] = all_join.get(source_col, "").fillna("").astype(str).str.strip()
        all_meta = _one_metadata_row_per_key(all_join, preferred_join_cols, all_metadata_cols)
        merged = merged.merge(
            all_meta,
            on=preferred_join_cols,
            how="left",
        )
        for col in all_metadata_cols[:-1]:
            if col not in merged.columns:
                merged[col] = ""
            merged[col] = merged[col].fillna("").astype(str).str.strip()
        merged["chart_exposed_in_dashboard"] = merged["chart_exposed_in_dashboard"].fillna(False).astype(bool)
    else:
        for col in [
            "all_chart_group_id",
            "all_dashboard_path",
            "all_chart_file",
            "all_section_id",
            "all_section_label",
            "all_entry_kind",
        ]:
            if col not in merged.columns:
                merged[col] = ""
            merged[col] = merged[col].fillna("").astype(str).str.strip()
        merged["chart_exposed_in_dashboard"] = False
    return merged


def attach_chart_groups_to_mapping_lineage_audit(
    mapping_lineage_audit: pd.DataFrame,
    chart_group_exposure_path: Path | str | None,
    all_chart_groups_path: Path | str | None = None,
) -> pd.DataFrame:
    """Filter mapping lineage audit rows to rendered charts and attach chart metadata."""
    audit = pd.DataFrame() if mapping_lineage_audit is None else mapping_lineage_audit.copy()
    chart_cols = [
        "chart_group_id",
        "dashboard_path",
        "chart_file",
        "page_key",
        "page_label",
        "chart_group_key",
        "chart_group_label",
        "section_id",
        "section_label",
        "entry_kind",
        "all_chart_group_id",
        "all_dashboard_path",
        "all_chart_file",
        "all_page_key",
        "all_page_label",
        "all_chart_group_key",
        "all_chart_group_label",
        "all_section_id",
        "all_section_label",
        "all_entry_kind",
        "chart_exposed_in_dashboard",
    ]
    for col in chart_cols:
        if col not in audit.columns:
            audit[col] = False if col == "chart_exposed_in_dashboard" else ""
    if audit.empty:
        leading = [
            "chart_group_id",
            "dashboard_path",
            "chart_file",
            "chart_group_key",
            "chart_group_label",
            "dataset",
            "scenario",
            "year",
            "esto_flow",
            "esto_product",
            "source_sector",
            "source_fuel",
            "value_pj",
        ]
        return audit.reindex(columns=[*leading, *[c for c in audit.columns if c not in leading]])

    working = audit.copy()
    for col in ["esto_flow", "esto_product", "source_sector", "source_fuel"]:
        if col not in working.columns:
            working[col] = ""
        working[col] = working[col].fillna("").astype(str).str.strip()

    # For 9th lineage rows, the charted fuel is the ESTO product label, not the
    # raw 9th fuel code. For LEAP/ESTO rows this is also the intended chart fuel.
    working["fuel_label"] = working["esto_product"].map(lambda value: _strip_esto_code_prefix(value) or _clean_token(value))

    def _lineage_measure_rows(frame: pd.DataFrame) -> pd.DataFrame:
        if frame.empty:
            return frame.copy()
        base = frame.copy()
        if "sheet" not in base.columns:
            base["sheet"] = ""
        base["sheet"] = base["sheet"].fillna("").astype(str).str.strip()
        missing_sheet = base["sheet"].eq("")
        base.loc[missing_sheet, "sheet"] = base.loc[missing_sheet, "esto_flow"].map(_sheet_key_from_esto_flow)
        value = pd.to_numeric(base.get("value_pj", pd.Series(index=base.index)), errors="coerce")
        flow_text = base["esto_flow"].fillna("").astype(str).str.strip()
        is_transformation = flow_text.str.startswith("09") | flow_text.str.startswith("08.")
        normal = base[~is_transformation].copy()
        normal["measure"] = "Energy balance (PJ)"
        transform = base[is_transformation].copy()
        if transform.empty:
            return normal
        tvalue = pd.to_numeric(transform.get("value_pj", pd.Series(index=transform.index)), errors="coerce")
        input_rows = transform[tvalue.le(0) | tvalue.isna()].copy()
        output_rows = transform[tvalue.ge(0) | tvalue.isna()].copy()
        input_rows["measure"] = TRANSFORMATION_INPUT_MEASURE
        output_rows["measure"] = TRANSFORMATION_OUTPUT_MEASURE
        return pd.concat([normal, input_rows, output_rows], ignore_index=True, sort=False)

    working = _lineage_measure_rows(working)

    attached = attach_chart_groups_to_dashboard_exposure(
        working,
        chart_group_exposure_path,
        all_chart_groups_path,
    )

    attached = attached.drop(columns=["sheet", "measure", "fuel_label"], errors="ignore")
    leading = [
        "chart_group_id",
        "dashboard_path",
        "chart_file",
        "chart_group_key",
        "chart_group_label",
        "section_id",
        "section_label",
        "entry_kind",
        "dataset",
        "scenario",
        "year",
        "esto_flow",
        "esto_product",
        "source_sector",
        "source_fuel",
        "value_pj",
    ]
    attached = attached.reindex(columns=[*leading, *[c for c in attached.columns if c not in leading]])
    if not attached.empty:
        attached = attached.sort_values(
            ["dashboard_path", "chart_group_id", "dataset", "scenario", "year", "esto_flow", "esto_product", "source_sector", "source_fuel"],
            kind="mergesort",
        ).reset_index(drop=True)
    return attached


def _chart_kind_from_entry(value: object) -> str:
    text = _clean_token(value).lower()
    if text == "aggregate":
        return "aggregate_total"
    if text == "direct":
        return "by_fuel"
    return text or "by_fuel"


def _output_existing_columns(frame: pd.DataFrame, columns: Sequence[str]) -> pd.DataFrame:
    out = frame.copy()
    keep = [col for col in columns if col in out.columns]
    return out[keep].copy()


def _with_simplified_dashboard_context(frame: pd.DataFrame) -> pd.DataFrame:
    out = pd.DataFrame() if frame is None else frame.copy()
    out = _add_dashboard_context_aliases(out)
    if "entry_kind" in out.columns:
        entry_kind = out["entry_kind"].fillna("").astype(str)
        if "chart_kind" not in out.columns:
            out["chart_kind"] = entry_kind.map(_chart_kind_from_entry)
        else:
            out["chart_kind"] = out["chart_kind"].fillna("").astype(str).str.strip()
            missing = out["chart_kind"].eq("")
            out.loc[missing, "chart_kind"] = entry_kind.loc[missing].map(_chart_kind_from_entry)
    return out


def simplify_mapping_lineage_audit_output(mapping_lineage_audit: pd.DataFrame) -> pd.DataFrame:
    """Return a human-facing mapping-lineage table without renderer bookkeeping columns."""
    out = _with_simplified_dashboard_context(mapping_lineage_audit)
    if out.empty:
        columns = [
            "dashboard_path",
            "chart_file",
            "dashboard_page_key",
            "dashboard_page_label",
            "dashboard_section_key",
            "dashboard_section_label",
            "chart_kind",
            "esto_flow_group_key",
            "esto_flow_group_label",
            "dataset",
            "scenario",
            "year",
            "esto_flow",
            "esto_product",
            "source_sector",
            "source_fuel",
            "value_pj",
            "is_subtotal",
            "esto_pair_is_subtotal",
            "ninth_pair_is_subtotal",
        ]
        return pd.DataFrame(columns=columns)

    for col in ["chart_group_id", "chart_file", "dashboard_path", "dashboard_section_label", "esto_flow"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()

    group_cols = ["chart_group_id"]
    if out["chart_group_id"].eq("").all():
        group_cols = ["chart_file"] if out["chart_file"].ne("").any() else ["dashboard_path", "dashboard_section_label"]

    group_context = []
    for values, group in out.groupby(group_cols, dropna=False):
        key_values = values if isinstance(values, tuple) else (values,)
        key_record = dict(zip(group_cols, key_values, strict=False))
        flows = [flow for flow in group["esto_flow"].dropna().astype(str).str.strip().unique() if flow]
        label = next((text for text in group["dashboard_section_label"].dropna().astype(str).str.strip() if text), "")
        group_context.append(
            {
                **key_record,
                "esto_flow_group_key": _esto_flow_group_key(flows),
                "esto_flow_group_label": _esto_flow_group_label(flows, fallback_label=label),
            }
        )
    context = pd.DataFrame(group_context)
    if not context.empty:
        out = out.drop(columns=["esto_flow_group_key", "esto_flow_group_label"], errors="ignore").merge(
            context,
            on=group_cols,
            how="left",
        )

    out["esto_flow_key"] = out["esto_flow"].map(_esto_flow_key)
    columns = [
        "dashboard_path",
        "chart_file",
        "dashboard_page_key",
        "dashboard_page_label",
        "dashboard_section_key",
        "dashboard_section_label",
        "chart_kind",
        "esto_flow_group_key",
        "esto_flow_group_label",
        "esto_flow_key",
        "dataset",
        "scenario",
        "year",
        "esto_flow",
        "esto_product",
        "source_sector",
        "source_fuel",
        "value_pj",
        "is_subtotal",
        "esto_pair_is_subtotal",
        "ninth_pair_is_subtotal",
    ]
    sort_cols = [col for col in ["dashboard_path", "chart_file", "dataset", "scenario", "year", "esto_flow", "esto_product", "source_sector", "source_fuel"] if col in out.columns]
    out = _output_existing_columns(out, columns)
    if sort_cols:
        out = out.sort_values([col for col in sort_cols if col in out.columns], kind="mergesort").reset_index(drop=True)
    return out


def simplify_chart_line_mapping_ledger_output(chart_line_mapping_ledger: pd.DataFrame) -> pd.DataFrame:
    """Return a compact per-chart-line ledger for human review."""
    out = _with_simplified_dashboard_context(chart_line_mapping_ledger)
    if not out.empty:
        out = _add_esto_flow_context_columns(out)
    rename_map = {
        "sector_code_9th": "source_sector",
        "ninth_fuel_code": "source_fuel",
    }
    out = out.rename(columns={k: v for k, v in rename_map.items() if k in out.columns and v not in out.columns})
    columns = [
        "dashboard_path",
        "chart_file",
        "dashboard_page_key",
        "dashboard_page_label",
        "dashboard_section_key",
        "dashboard_section_label",
        "chart_kind",
        "esto_flow_group_key",
        "esto_flow_group_label",
        "esto_flow_key",
        "economy",
        "scenario",
        "year",
        "measure",
        "fuel_label",
        "source",
        "value",
        "ninth_pairs_label",
        "is_total_line",
        "mapping_record_type",
        "esto_flow",
        "esto_product",
        "source_sector",
        "source_fuel",
        "exact_comparator_key",
        "duplicate_exact_comparator_key_count",
        "component_included_in_total",
        "total_component_bucket_count",
        "mapping_source",
        "mapping_note",
    ]
    return _output_existing_columns(out, columns)


def simplify_chart_total_component_ledger_output(chart_total_component_ledger: pd.DataFrame) -> pd.DataFrame:
    """Return a compact total-component ledger focused on included members."""
    out = _with_simplified_dashboard_context(chart_total_component_ledger)
    rename_map = {
        "sector_code_9th": "source_sector",
        "ninth_fuel_code": "source_fuel",
    }
    out = out.rename(columns={k: v for k, v in rename_map.items() if k in out.columns and v not in out.columns})
    columns = [
        "dashboard_page_key",
        "dashboard_page_label",
        "dashboard_section_key",
        "dashboard_section_label",
        "chart_kind",
        "economy",
        "scenario",
        "source",
        "year",
        "measure",
        "sheet",
        "member_fuel_label",
        "member_value",
        "component_included_in_total",
        "exact_comparator_key",
        "duplicate_exact_comparator_key_count",
        "source_sector",
        "source_fuel",
        "projection_parent_sector_code",
        "sector_depth",
        "fuel_depth",
        "is_leaf_level",
    ]
    return _output_existing_columns(out, columns)


def write_ninth_mapping_data_coverage(
    *,
    ninth_df: pd.DataFrame,
    ninth_mapping_pairs: pd.DataFrame,
    output_path: Path | str,
    projection_economy: str,
    scenarios: Sequence[str],
    years: Sequence[int],
) -> str:
    """Write 9th data sector/fuel paths missing from the explicit LEAP-to-9th mapping sheet."""
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    sector_cols = ["sectors", "sub1sectors", "sub2sectors", "sub3sectors", "sub4sectors"]
    fuel_cols = ["fuels", "subfuels"]
    output_cols = [
        "ninth_sector",
        "ninth_fuel",
        "esto_flow",
        "esto_product",
        "sector_path",
        "fuel_path",
        "scenario_count",
        "scenarios",
        "year_min",
        "year_max",
        "value_abs_sum",
        "value_sum",
        "source_rows",
        "exact_mapping_found",
        "covered_by_parent_mapping",
        "covering_mapping_pairs",
    ]
    hierarchy_output_cols = [
        "ninth_sector",
        "ninth_sector_level",
        "ninth_fuel",
        "ninth_fuel_level",
        "esto_flow",
        "esto_product",
        "source_sector_paths",
        "source_fuel_paths",
        "scenario_count",
        "scenarios",
        "year_min",
        "year_max",
        "value_abs_sum",
        "value_sum",
        "source_rows",
        "exact_mapping_found",
        "target_mapping_pairs",
    ]

    if ninth_df is None or ninth_df.empty:
        empty = pd.DataFrame(columns=output_cols)
        empty_hierarchy = pd.DataFrame(columns=hierarchy_output_cols)
        with pd.ExcelWriter(output_path) as writer:
            empty.to_excel(writer, sheet_name="missing_exact_9th_pairs", index=False)
            empty.to_excel(writer, sheet_name="missing_uncovered_9th_pairs", index=False)
            empty.to_excel(writer, sheet_name="all_9th_data_paths", index=False)
            empty_hierarchy.to_excel(writer, sheet_name="missing_hierarchy_pairs", index=False)
            empty_hierarchy.to_excel(writer, sheet_name="all_hierarchy_pairs", index=False)
        return str(output_path)

    def clean(value: object) -> str:
        text = str(value or "").strip()
        return "" if text.lower() in {"", "nan", "none", "x"} else text

    def norm(value: object) -> str:
        return clean(value).lower()

    def join_path(row: pd.Series, cols: list[str]) -> str:
        return "/".join(clean(row.get(col, "")) for col in cols if clean(row.get(col, "")))

    def deepest(row: pd.Series, cols: list[str]) -> str:
        values = [clean(row.get(col, "")) for col in cols if clean(row.get(col, ""))]
        return values[-1] if values else ""

    working = ninth_df.copy()
    if "economy" in working.columns:
        working = working[working["economy"].fillna("").astype(str).str.strip().eq(str(projection_economy).strip())]
    scenario_values = {str(value).strip().lower() for value in scenarios if str(value).strip()}
    if "scenarios" in working.columns and scenario_values:
        working = working[working["scenarios"].fillna("").astype(str).str.strip().str.lower().isin(scenario_values)]
    for subtotal_col in ["subtotal_results", "subtotal_layout"]:
        if subtotal_col in working.columns:
            working = working[~working[subtotal_col].fillna("").astype(str).str.strip().str.lower().isin({"true", "1", "yes", "y"})]

    year_cols = []
    year_lookup: dict[object, int] = {}
    requested_years = {int(year) for year in years}
    for col in working.columns:
        text = str(col).strip()
        if text.isdigit() and int(text) in requested_years:
            year_cols.append(col)
            year_lookup[col] = int(text)
    if not year_cols:
        all_data = pd.DataFrame(columns=output_cols)
        all_hierarchy = pd.DataFrame(columns=hierarchy_output_cols)
    else:
        values = working[year_cols].apply(pd.to_numeric, errors="coerce").fillna(0.0)
        data_mask = values.ne(0).any(axis=1)
        working = working.loc[data_mask].copy()
        values = values.loc[data_mask]
        if working.empty:
            all_data = pd.DataFrame(columns=output_cols)
            all_hierarchy = pd.DataFrame(columns=hierarchy_output_cols)
        else:
            working["ninth_sector"] = working.apply(lambda row: deepest(row, sector_cols), axis=1)
            working["ninth_fuel"] = working.apply(lambda row: deepest(row, fuel_cols), axis=1)
            working["sector_path"] = working.apply(lambda row: join_path(row, sector_cols), axis=1)
            working["fuel_path"] = working.apply(lambda row: join_path(row, fuel_cols), axis=1)
            working["value_abs_sum"] = values.abs().sum(axis=1)
            working["value_sum"] = values.sum(axis=1)
            working = working[working["ninth_sector"].ne("") & working["ninth_fuel"].ne("")].copy()

            scen_col = "scenarios" if "scenarios" in working.columns else None
            group_cols = ["ninth_sector", "ninth_fuel", "sector_path", "fuel_path"]
            grouped = (
                working.groupby(group_cols, dropna=False)
                .agg(
                    scenario_count=(scen_col, "nunique") if scen_col else ("ninth_sector", "size"),
                    scenarios=(scen_col, lambda values: "|".join(sorted(set(str(v).strip() for v in values if str(v).strip()))))
                    if scen_col
                    else ("ninth_sector", lambda _: ""),
                    value_abs_sum=("value_abs_sum", "sum"),
                    value_sum=("value_sum", "sum"),
                    source_rows=("ninth_sector", "size"),
                )
                .reset_index()
            )
            grouped["year_min"] = min(requested_years)
            grouped["year_max"] = max(requested_years)
            all_data = grouped

            hierarchy_rows: list[dict[str, object]] = []
            for _, row in working.iterrows():
                sector_values = [(col, clean(row.get(col, ""))) for col in sector_cols if clean(row.get(col, ""))]
                fuel_values = [(col, clean(row.get(col, ""))) for col in fuel_cols if clean(row.get(col, ""))]
                if not sector_values or not fuel_values:
                    continue
                for sector_level, sector_code in sector_values:
                    for fuel_level, fuel_code in fuel_values:
                        hierarchy_rows.append(
                            {
                                "ninth_sector": sector_code,
                                "ninth_sector_level": sector_level,
                                "ninth_fuel": fuel_code,
                                "ninth_fuel_level": fuel_level,
                                "sector_path": row["sector_path"],
                                "fuel_path": row["fuel_path"],
                                "scenario": row.get("scenarios", ""),
                                "value_abs_sum": row["value_abs_sum"],
                                "value_sum": row["value_sum"],
                            }
                        )
            if hierarchy_rows:
                hierarchy_raw = pd.DataFrame(hierarchy_rows)

                def join_unique(values: pd.Series) -> str:
                    return "|".join(sorted(set(str(value).strip() for value in values if str(value).strip())))

                all_hierarchy = (
                    hierarchy_raw.groupby(
                        ["ninth_sector", "ninth_sector_level", "ninth_fuel", "ninth_fuel_level"],
                        dropna=False,
                    )
                    .agg(
                        source_sector_paths=("sector_path", join_unique),
                        source_fuel_paths=("fuel_path", join_unique),
                        scenario_count=("scenario", "nunique"),
                        scenarios=("scenario", join_unique),
                        value_abs_sum=("value_abs_sum", "sum"),
                        value_sum=("value_sum", "sum"),
                        source_rows=("ninth_sector", "size"),
                    )
                    .reset_index()
                )
                all_hierarchy["year_min"] = min(requested_years)
                all_hierarchy["year_max"] = max(requested_years)
            else:
                all_hierarchy = pd.DataFrame(columns=hierarchy_output_cols)

    pairs = pd.DataFrame() if ninth_mapping_pairs is None else ninth_mapping_pairs.copy()
    pairs = pairs.rename(
        columns={
            "leap_sector_name_full_path": "leap_mapping_sector",
            "raw_leap_fuel_name": "leap_mapping_fuel",
        }
    )
    if "remove_row" in pairs.columns:
        remove_mask = pairs["remove_row"].fillna("").astype(str).str.strip().str.lower().isin(
            {"1", "true", "yes", "y", "on", "t"}
        )
        pairs = pairs.loc[~remove_mask].copy()
    for col in ["ninth_sector", "ninth_fuel", "esto_flow", "esto_product", "leap_mapping_sector", "leap_mapping_fuel"]:
        if col not in pairs.columns:
            pairs[col] = ""
        pairs[col] = pairs[col].fillna("").astype(str).str.strip()
    pairs = pairs[pairs["ninth_sector"].ne("") & pairs["ninth_fuel"].ne("")].copy()
    exact_pairs = {(norm(row["ninth_sector"]), norm(row["ninth_fuel"])) for _, row in pairs.iterrows()}
    mapped_pairs_by_norm: dict[tuple[str, str], list[str]] = defaultdict(list)
    for _, row in pairs.iterrows():
        key = (norm(row["ninth_sector"]), norm(row["ninth_fuel"]))
        if clean(row["leap_mapping_sector"]) or clean(row["leap_mapping_fuel"]):
            target = (
                f"{clean(row['leap_mapping_sector'])} + {clean(row['leap_mapping_fuel'])} "
                f"-> {clean(row['ninth_sector'])} + {clean(row['ninth_fuel'])}"
            )
        else:
            target = (
                f"{clean(row['ninth_sector'])} + {clean(row['ninth_fuel'])} "
                f"-> {clean(row['esto_flow'])} + {clean(row['esto_product'])}"
            )
        if target not in mapped_pairs_by_norm[key]:
            mapped_pairs_by_norm[key].append(target)

    if all_data.empty:
        all_data = pd.DataFrame(columns=output_cols)
    else:
        def coverage(row: pd.Series) -> tuple[bool, str]:
            sector_tokens = [norm(part) for part in str(row.get("sector_path", "")).split("/") if norm(part)]
            fuel_tokens = [norm(part) for part in str(row.get("fuel_path", "")).split("/") if norm(part)]
            hits: list[str] = []
            for sector in sector_tokens:
                for fuel in fuel_tokens:
                    hits.extend(mapped_pairs_by_norm.get((sector, fuel), []))
            hits = sorted(set(hits))
            return bool(hits), " | ".join(hits)

        exact = all_data.apply(lambda row: (norm(row["ninth_sector"]), norm(row["ninth_fuel"])) in exact_pairs, axis=1)
        coverage_result = all_data.apply(coverage, axis=1)
        all_data["exact_mapping_found"] = exact.astype(bool)
        all_data["covered_by_parent_mapping"] = coverage_result.map(lambda item: item[0]).astype(bool)
        all_data["covering_mapping_pairs"] = coverage_result.map(lambda item: item[1])
        all_data["esto_flow"] = ""
        all_data["esto_product"] = ""
        all_data = all_data[output_cols].sort_values(
            ["exact_mapping_found", "covered_by_parent_mapping", "value_abs_sum", "ninth_sector", "ninth_fuel"],
            ascending=[True, True, False, True, True],
            kind="mergesort",
        )

    if all_hierarchy.empty:
        all_hierarchy = pd.DataFrame(columns=hierarchy_output_cols)
    else:
        all_hierarchy["exact_mapping_found"] = all_hierarchy.apply(
            lambda row: (norm(row["ninth_sector"]), norm(row["ninth_fuel"])) in exact_pairs,
            axis=1,
        ).astype(bool)
        all_hierarchy["target_mapping_pairs"] = all_hierarchy.apply(
            lambda row: " | ".join(
                sorted(set(mapped_pairs_by_norm.get((norm(row["ninth_sector"]), norm(row["ninth_fuel"])), [])))
            ),
            axis=1,
        )
        all_hierarchy["esto_flow"] = ""
        all_hierarchy["esto_product"] = ""
        all_hierarchy = all_hierarchy[hierarchy_output_cols].sort_values(
            ["exact_mapping_found", "value_abs_sum", "ninth_sector", "ninth_fuel"],
            ascending=[True, False, True, True],
            kind="mergesort",
        )

    missing_exact = all_data[~all_data["exact_mapping_found"]].copy() if not all_data.empty else all_data.copy()
    missing_uncovered = (
        all_data[~all_data["covered_by_parent_mapping"]].copy() if not all_data.empty else all_data.copy()
    )
    missing_hierarchy = (
        all_hierarchy[~all_hierarchy["exact_mapping_found"]].copy()
        if not all_hierarchy.empty
        else all_hierarchy.copy()
    )
    summary = pd.DataFrame(
        [
            {"metric": "all_9th_data_paths", "rows": int(len(all_data))},
            {"metric": "missing_exact_9th_pairs", "rows": int(len(missing_exact))},
            {"metric": "missing_uncovered_9th_pairs", "rows": int(len(missing_uncovered))},
            {"metric": "all_9th_hierarchy_pairs", "rows": int(len(all_hierarchy))},
            {"metric": "missing_9th_hierarchy_pairs", "rows": int(len(missing_hierarchy))},
        ]
    )
    with pd.ExcelWriter(output_path) as writer:
        missing_exact.to_excel(writer, sheet_name="missing_exact_9th_pairs", index=False)
        missing_uncovered.to_excel(writer, sheet_name="missing_uncovered_9th_pairs", index=False)
        all_data.to_excel(writer, sheet_name="all_9th_data_paths", index=False)
        missing_hierarchy.to_excel(writer, sheet_name="missing_hierarchy_pairs", index=False)
        all_hierarchy.to_excel(writer, sheet_name="all_hierarchy_pairs", index=False)
        summary.to_excel(writer, sheet_name="summary", index=False)
    return str(output_path)


def _extract_balance_workbook(
    workbook_path: Path,
    *,
    template_sheet: str,
    mapping_pairs_path: ConfigTableRef,
    codebook_path: Path,
    explicit_pair_mappings_only: bool = False,
    allow_descendant_mapping_expansion: bool = True,
    expected_scenario: str | None = None,
) -> dict[str, Any]:
    chosen_template = _pick_template_sheet(workbook_path, template_sheet)
    extractor = TemplateBalanceExtractor(
        template_sheet=chosen_template,
        mapping_pairs_path=mapping_pairs_path,
        codebook_path=codebook_path,
        reinterpret_fuel_rows_as_parent_sector=False,
        explicit_pair_mappings_only=explicit_pair_mappings_only,
        allow_descendant_mapping_expansion=allow_descendant_mapping_expansion,
    )
    extractor.load_mappings()
    selected_sheets = _list_balance_sheets(workbook_path)
    if not selected_sheets:
        raise ValueError(f"No balance sheets found in workbook: {workbook_path}")
    raw_long, mapped_long, coverage, unit_diag, report = extractor.extract_workbook(
        workbook_path,
        include_zero_values=False,
        sheet_name_filter=selected_sheets,
        convert_units_to_petajoule=True,
    )
    if expected_scenario:
        for frame in (raw_long, mapped_long):
            if frame.empty or "scenario" not in frame.columns:
                continue
            mismatched = frame["scenario"].map(_normalize_scenario).ne(expected_scenario)
            if mismatched.any():
                mismatched_labels = sorted(frame.loc[mismatched, "scenario"].astype(str).unique())
                print(
                    f"[WARN] {workbook_path.name}: {int(mismatched.sum())} row(s) carry an internal "
                    f"Scenario label {mismatched_labels} that disagrees with this workbook's "
                    f"REF/TGT identity ({expected_scenario}). "
                    + (
                        "Trusting the workbook identity (BALANCE_EXPORT_TRUST_FILENAME_SCENARIO=True)."
                        if BALANCE_EXPORT_TRUST_FILENAME_SCENARIO
                        else "Keeping the internal label (BALANCE_EXPORT_TRUST_FILENAME_SCENARIO=False)."
                    )
                )
                if BALANCE_EXPORT_TRUST_FILENAME_SCENARIO:
                    frame.loc[mismatched, "scenario"] = expected_scenario
    return {
        "template_sheet": chosen_template,
        "raw_long": raw_long,
        "mapped_long": mapped_long,
        "coverage": coverage,
        "unit_diag": unit_diag,
        "report": report,
    }


def load_balance_leap_long(
    *,
    ref_workbook_path: Path | str = DEFAULT_REF_WORKBOOK_PATH,
    tgt_workbook_path: Path | str = DEFAULT_TGT_WORKBOOK_PATH,
    template_sheet: str = "EBal|2060",
    mapping_pairs_path: ConfigTableRef = DEFAULT_MAPPING_PAIRS_PATH,
    codebook_path: Path | str = DEFAULT_CODEBOOK_PATH,
    structure_config: dict[str, Any] | None = None,
    known_issues: dict[str, Any] | None = None,
    projection_economy: str = "20_USA",
    explicit_pair_mappings_only: bool = False,
    allow_descendant_mapping_expansion: bool = True,
) -> dict[str, Any]:
    """
    Load LEAP balance exports (REF/TGT), map them, keep fully mapped rows only,
    and return normalized LEAP long-format rows in PJ.
    """
    structure = structure_config or {}
    issues_cfg = known_issues or {}

    ref_path = _resolve(ref_workbook_path)
    tgt_path = _resolve(tgt_workbook_path)
    mapping_pairs = _resolve_config_table_ref(mapping_pairs_path)
    codebook = _resolve(codebook_path)

    for candidate in [ref_path, tgt_path]:
        if not candidate.exists():
            raise FileNotFoundError(f"Missing required input: {candidate}")
    mapping_pairs_file, mapping_pairs_sheet = split_config_table_ref(mapping_pairs)
    if not config_table_exists(mapping_pairs_file, sheet_name=mapping_pairs_sheet):
        raise FileNotFoundError(f"Missing required input: {mapping_pairs}")
    if not config_table_exists(codebook, sheet_name="code_to_name"):
        raise FileNotFoundError(f"Missing required input: {codebook}")

    extracted_ref = _extract_balance_workbook(
        ref_path,
        template_sheet=template_sheet,
        mapping_pairs_path=mapping_pairs,
        codebook_path=codebook,
        explicit_pair_mappings_only=explicit_pair_mappings_only,
        allow_descendant_mapping_expansion=allow_descendant_mapping_expansion,
    )
    extracted_tgt = _extract_balance_workbook(
        tgt_path,
        template_sheet=template_sheet,
        mapping_pairs_path=mapping_pairs,
        codebook_path=codebook,
        explicit_pair_mappings_only=explicit_pair_mappings_only,
        allow_descendant_mapping_expansion=allow_descendant_mapping_expansion,
    )

    combined = pd.concat(
        [extracted_ref["mapped_long"], extracted_tgt["mapped_long"]],
        ignore_index=True,
        sort=False,
    )
    if combined.empty:
        raise RuntimeError("Balance extraction produced no rows.")

    combined["scenario"] = combined.get("scenario", "").map(_normalize_scenario)
    combined["year"] = pd.to_numeric(combined.get("year", pd.NA), errors="coerce").astype("Int64")
    combined["value_petajoule"] = pd.to_numeric(
        combined.get("value_petajoule", combined.get("value", pd.NA)),
        errors="coerce",
    )

    for col in ["leap_sector", "leap_fuel", "esto_flow", "esto_product"]:
        if col not in combined.columns:
            combined[col] = ""
        combined[col] = combined[col].fillna("").astype(str).str.strip()

    combined, override_report = _apply_mapping_overrides(
        combined,
        list(issues_cfg.get("mapping_overrides", []) or []),
    )

    for subtotal_col in ["leap_is_subtotal", "esto_is_subtotal", "ninth_is_subtotal"]:
        if subtotal_col not in combined.columns:
            combined[subtotal_col] = False
        combined[subtotal_col] = combined[subtotal_col].fillna(False).map(_to_bool)
    inferred_subtotal = combined.apply(_infer_subtotal_flag, axis=1)
    combined["leap_is_subtotal"] = combined["leap_is_subtotal"] | inferred_subtotal
    combined["is_subtotal"] = combined["leap_is_subtotal"]

    row_filters = issues_cfg.get("row_filters", {}) or {}
    combined = _apply_balance_row_filters(combined, row_filters)

    if "esto_mapping_found" not in combined.columns:
        combined["esto_mapping_found"] = combined["esto_flow"].ne("") & combined["esto_product"].ne("")
    if "ninth_mapping_found" not in combined.columns:
        combined["ninth_mapping_found"] = combined["leap_sector"].ne("") & combined["leap_fuel"].ne("")
    combined["esto_mapping_found"] = combined["esto_mapping_found"].fillna(False).map(_to_bool)
    combined["ninth_mapping_found"] = combined["ninth_mapping_found"].fillna(False).map(_to_bool)
    missing_estox = (
        ~combined["esto_mapping_found"]
        | combined["esto_flow"].eq("")
        | combined["esto_product"].eq("")
    )
    missing_ninth = (
        ~combined["ninth_mapping_found"]
        | combined["leap_sector"].eq("")
        | combined["leap_fuel"].eq("")
    )
    combined["mapping_failed"] = ""
    combined.loc[missing_estox & missing_ninth, "mapping_failed"] = "ESTO and Ninth"
    combined.loc[missing_estox & ~missing_ninth, "mapping_failed"] = "ESTO"
    combined.loc[~missing_estox & missing_ninth, "mapping_failed"] = "Ninth"
    required_mapping_mask = ~missing_estox & ~missing_ninth
    incomplete_rows = combined[~required_mapping_mask].copy()
    mapped = combined[required_mapping_mask].copy()

    flow_name_map, fuel_name_map = _load_codebook_name_maps(codebook)
    mapped = _build_flow_and_product_labels(
        mapped,
        flow_name_map=flow_name_map,
        fuel_name_map=fuel_name_map,
        label_overrides=issues_cfg.get("label_overrides", {}) or {},
    )

    conflict_rows = pd.DataFrame(columns=mapped.columns)
    agg_keys = ["scenario", "year", "leap_sector", "leap_fuel", "esto_flow", "esto_product"]
    for col in [
        "leap_sector_name_full_path",
        "leap_sector_name_original",
        "mapping_key_sector",
        "mapping_key_fuel",
        "mapping_candidate_rule",
    ]:
        if col not in mapped.columns:
            mapped[col] = ""

    grouped = (
        mapped.groupby(agg_keys, as_index=False)
        .agg(
            leap_value=("value_petajoule", "sum"),
            leap_sector_name=("leap_sector_name", _coalesce_unique),
            leap_fuel_name=("leap_fuel_name", _coalesce_unique),
            leap_sector_name_full_path=("leap_sector_name_full_path", _coalesce_unique),
            leap_sector_name_original=("leap_sector_name_original", _coalesce_unique),
            mapping_key_sector=("mapping_key_sector", _coalesce_unique),
            mapping_key_fuel=("mapping_key_fuel", _coalesce_unique),
            mapping_candidate_rule=("mapping_candidate_rule", _coalesce_unique),
            flow_label=("flow_label", _coalesce_unique),
            fuel_label=("fuel_label", _coalesce_unique),
            is_subtotal=("is_subtotal", "max"),
            leap_is_subtotal=("leap_is_subtotal", "max"),
            esto_is_subtotal=("esto_is_subtotal", "max"),
            ninth_is_subtotal=("ninth_is_subtotal", "max"),
            source_sheet=("source_sheet", lambda s: "|".join(sorted(set([v for v in s.astype(str) if v])))),
            source_workbook=(
                "source_workbook",
                lambda s: "|".join(sorted(set([v for v in s.astype(str) if v]))),
            ),
        )
        .reset_index(drop=True)
    )

    flow_to_sheet = {
        str(k): str(v)
        for k, v in (structure.get("flow_to_sheet", {}) or {}).items()
        if _clean_token(k) and _clean_token(v)
    }
    sheet_catalog = structure.get("sheet_catalog", {}) or {}

    grouped["sheet"] = grouped["leap_sector"].map(flow_to_sheet).fillna("") if flow_to_sheet else ""
    flow_unmapped_rows = grouped[grouped["sheet"].eq("")].copy()
    if grouped["sheet"].eq("").any():
        fallback_sheet = grouped["leap_sector_name"].fillna("").astype(str).str.strip()
        fallback_sheet = fallback_sheet.where(fallback_sheet.ne(""), grouped["leap_sector"].fillna("").astype(str))
        grouped.loc[grouped["sheet"].eq(""), "sheet"] = fallback_sheet[grouped["sheet"].eq("")]

    def _measure_for_sheet(sheet_key: str) -> str:
        cfg = sheet_catalog.get(sheet_key, {}) or {}
        measure = _clean_token(cfg.get("measure", ""))
        return measure or "Energy balance (PJ)"

    grouped["measure"] = grouped["sheet"].map(_measure_for_sheet)

    grouped["economy"] = projection_economy
    grouped["region"] = ""
    grouped["sector_code_9th"] = grouped["leap_sector"]
    grouped["ninth_fuel_code"] = grouped["leap_fuel"]
    grouped["leap_variable"] = "Energy Balance"
    grouped["leap_units"] = "Petajoule"
    grouped["leap_scale_note"] = ""

    leap_long = grouped[
        [
            "economy",
            "scenario",
            "region",
            "sheet",
            "sector_code_9th",
            "flow_label",
            "fuel_label",
            "year",
            "leap_value",
            "leap_variable",
            "leap_units",
            "measure",
            "leap_scale_note",
            "ninth_fuel_code",
            "esto_flow",
            "esto_product",
            "source_sheet",
            "source_workbook",
            "leap_sector",
            "leap_fuel",
            "leap_sector_name",
            "leap_sector_name_original",
            "leap_sector_name_full_path",
            "leap_fuel_name",
            "mapping_key_sector",
            "mapping_key_fuel",
            "mapping_candidate_rule",
            "is_subtotal",
            "leap_is_subtotal",
            "esto_is_subtotal",
            "ninth_is_subtotal",
        ]
    ].rename(columns={"flow_label": "sector_name", "sheet": "sheet_name"})

    leap_long["year"] = pd.to_numeric(leap_long["year"], errors="coerce").astype("Int64")
    leap_long["leap_value"] = pd.to_numeric(leap_long["leap_value"], errors="coerce")
    leap_long = leap_long.sort_values(
        ["scenario", "sheet_name", "fuel_label", "year"],
        kind="mergesort",
    ).reset_index(drop=True)

    map_unique = leap_long[
        [
            "sheet_name",
            "measure",
            "fuel_label",
            "sector_code_9th",
            "ninth_fuel_code",
            "esto_flow",
            "esto_product",
            "leap_is_subtotal",
            "esto_is_subtotal",
            "ninth_is_subtotal",
        ]
    ].drop_duplicates()

    mapping_status = map_unique.rename(columns={"sheet_name": "sheet"}).copy()
    mapping_status["mapped"] = True
    mapping_status["has_any_mapping"] = True
    mapping_status["base_mapping_complete"] = True
    mapping_status["projection_mapping_complete"] = True
    mapping_status["partially_mapped"] = False
    mapping_status["missing_ninth_fuel"] = False
    mapping_status["missing_esto_flow"] = False
    mapping_status["missing_esto_product"] = False
    mapping_status["has_mapping_note"] = False
    mapping_status["mapping_source"] = "balance_table"
    mapping_status["flow_source"] = "balance_table"
    mapping_status["fuel_source"] = "balance_table"
    mapping_status["sector_match_method"] = "balance_flow_code"
    mapping_status["mapping_note"] = ""
    mapping_status["projection_parent_fallback"] = False
    mapping_status["projection_parent_sector_code"] = ""
    mapping_status["comparator_scope"] = "exact"

    issue_frames = [
        _rows_to_issue_records(
            incomplete_rows,
            reason="incomplete_mapping",
            details="No complete explicit LEAP-to-ESTO pair mapping was found.",
        ),
        _rows_to_issue_records(conflict_rows, reason="mapping_conflict_after_aggregation"),
    ]
    if not flow_unmapped_rows.empty:
        issue_frames.append(
            pd.DataFrame(
                {
                    "reason": "flow_not_in_structure_config",
                    "details": "No flow_to_sheet entry found for leap flow code",
                    "scenario": flow_unmapped_rows["scenario"].astype(str),
                    "year": flow_unmapped_rows["year"].astype("Int64"),
                    "source_sheet": flow_unmapped_rows.get("source_sheet", "").astype(str),
                    "leap_sector_name_full_path": flow_unmapped_rows.get(
                        "leap_sector_name_full_path", ""
                    ).astype(str),
                    "leap_flow": flow_unmapped_rows["leap_sector"].astype(str),
                    "leap_flow_name": flow_unmapped_rows.get("leap_sector_name", "").astype(str),
                    "leap_product": flow_unmapped_rows["leap_fuel"].astype(str),
                    "leap_product_name": flow_unmapped_rows.get("leap_fuel_name", "").astype(str),
                    "mapping_failed": "dashboard sheet",
                    "mapping_key_sector": flow_unmapped_rows.get("mapping_key_sector", "").astype(str),
                    "mapping_key_fuel": flow_unmapped_rows.get("mapping_key_fuel", "").astype(str),
                    "mapping_candidate_rule": flow_unmapped_rows.get("mapping_candidate_rule", "").astype(str),
                    "esto_flow": flow_unmapped_rows["esto_flow"].astype(str),
                    "esto_product": flow_unmapped_rows["esto_product"].astype(str),
                    "value_petajoule": pd.to_numeric(flow_unmapped_rows["leap_value"], errors="coerce"),
                }
            )
        )
    issue_frames = [frame for frame in issue_frames if not frame.empty and not frame.isna().all(axis=None)]
    issues_df = pd.concat(issue_frames, ignore_index=True, sort=False) if issue_frames else pd.DataFrame()

    return {
        "leap_long": leap_long,
        "mapping_status": mapping_status,
        "issues": issues_df,
        "override_report": override_report,
        "unit_diagnostics": pd.concat(
            [extracted_ref["unit_diag"], extracted_tgt["unit_diag"]],
            ignore_index=True,
            sort=False,
        ),
        "coverage": pd.concat(
            [extracted_ref["coverage"], extracted_tgt["coverage"]],
            ignore_index=True,
            sort=False,
        ),
        "matching_diagnostics": pd.concat(
            [
                extracted_ref["report"].get("matching_diagnostics", pd.DataFrame()),
                extracted_tgt["report"].get("matching_diagnostics", pd.DataFrame()),
            ],
            ignore_index=True,
            sort=False,
        ),
        "extraction_summary": {
            "ref": extracted_ref["report"].get("summary", {}),
            "tgt": extracted_tgt["report"].get("summary", {}),
            "selected_template_ref": extracted_ref["template_sheet"],
            "selected_template_tgt": extracted_tgt["template_sheet"],
            "leap_rows_after_filters": int(len(leap_long)),
            "mapping_rows": int(len(mapping_status)),
            "issue_rows": int(len(issues_df)),
        },
    }


def _build_mapping_status_with_availability(
    mapping_status: pd.DataFrame,
    comparison_long: pd.DataFrame,
) -> pd.DataFrame:
    if mapping_status.empty:
        return mapping_status.copy()

    out = mapping_status.copy()
    out = _backfill_dashboard_hierarchy(out)
    keys = ["chart_group_key", "measure", "fuel_label"]

    comp = comparison_long.copy()
    if comp.empty:
        out["has_leap"] = False
        out["has_base"] = False
        out["has_projection"] = False
        return out

    comp = _backfill_dashboard_hierarchy(comp)
    comp["value_num"] = pd.to_numeric(comp["value"], errors="coerce")
    comp["has_value"] = comp["value_num"].notna()

    availability = (
        comp.groupby(keys + ["source"], as_index=False)["has_value"]
        .max()
        .pivot_table(index=keys, columns="source", values="has_value", aggfunc="max")
        .reset_index()
    )
    if hasattr(availability.columns, "name"):
        availability.columns.name = None

    source_aliases = {
        "leap": "has_leap",
        "base": "has_base",
        "projection": "has_projection",
    }
    for source_col, target_col in source_aliases.items():
        if source_col not in availability.columns:
            availability[source_col] = False
        availability[target_col] = availability[source_col].astype("boolean").fillna(False).astype(bool)

    out = out.drop(columns=["has_leap", "has_base", "has_projection"], errors="ignore")
    out = out.merge(
        availability[keys + ["has_leap", "has_base", "has_projection"]],
        on=keys,
        how="left",
    )
    for col in ["has_leap", "has_base", "has_projection"]:
        out[col] = out[col].astype("boolean").fillna(False).astype(bool)
    return out


def build_balance_comparison(
    *,
    leap_long: pd.DataFrame,
    mapping_status: pd.DataFrame,
    base_year: int,
    projection_years: Sequence[int],
    base_economy: str,
    projection_economy: str,
    scenario_map: dict[str, str],
    sheet_map_path: Path | str = DEFAULT_SHEET_MAP_PATH,
    backup_mappings_path: Path | str = DEFAULT_BACKUP_MAPPINGS_PATH,
    codebook_path: Path | str = DEFAULT_CODEBOOK_PATH,
    canonical_pairs_path: ConfigTableRef = DEFAULT_MAPPING_PAIRS_PATH,
    explicit_mappings_path: Path | str = DEFAULT_EXPLICIT_MAPPINGS_PATH,
    explicit_reassignments_path: Path | str = DEFAULT_EXPLICIT_REASSIGNMENTS_PATH,
    synthetic_reference_rows_path: Path | str = DEFAULT_SYNTHETIC_REFERENCE_ROWS_PATH,
    esto_table_path: Path | str = DEFAULT_BASE_TABLE_PATH,
    projection_table_path: Path | str = DEFAULT_PROJECTION_TABLE_PATH,
    chart_navigation_guide_path: Path | str | None = None,
    known_issues: dict[str, Any] | None = None,
    base_df: pd.DataFrame | None = None,
    ninth_df: pd.DataFrame | None = None,
) -> dict[str, Any]:
    """
    Build comparison_long/wide and updated mapping_status using LEAP balance rows.
    """
    if leap_long.empty:
        raise RuntimeError("leap_long is empty; cannot build comparison outputs.")

    mapping_inputs: dict[str, Any] | None = None
    reassignment_status = pd.DataFrame()
    synthetic_reference_status = pd.DataFrame()

    if base_df is None or ninth_df is None:
        mapping_inputs = load_mapping_inputs(
            sheet_map_path=_resolve(sheet_map_path),
            backup_mappings_path=_resolve(backup_mappings_path),
            codebook_path=_resolve(codebook_path),
            canonical_pairs_path=_resolve_config_table_ref(canonical_pairs_path),
            explicit_mappings_path=_resolve(explicit_mappings_path),
            explicit_reassignments_path=_resolve(explicit_reassignments_path),
        )
        base_df, ninth_df, reassignment_status, synthetic_reference_status = load_reference_tables(
            esto_table_path=_resolve(esto_table_path),
            projection_table_path=_resolve(projection_table_path),
            explicit_reassignments=mapping_inputs["explicit_reassignments"],
            explicit_mappings=mapping_inputs["explicit_mappings"],
            canonical_pairs=mapping_inputs["canonical_pairs"],
            synthetic_reference_rows_path=_resolve(synthetic_reference_rows_path),
            drop_all_zero_base_rows=True,
            drop_all_zero_projection_rows=False,
        )

    scenario_to_projection = {
        _normalize_scenario(k): str(v).strip().lower()
        for k, v in (scenario_map or {}).items()
        if _clean_token(k) and _clean_token(v)
    }

    rows: list[dict[str, Any]] = []
    base_cache: dict[tuple[str, str], float] = {}
    projection_cache: dict[tuple[str, str, str], pd.Series] = {}

    leap_working = leap_long.copy()
    leap_working["scenario"] = leap_working["scenario"].map(_normalize_scenario)
    leap_working["year"] = pd.to_numeric(leap_working["year"], errors="coerce").astype("Int64")
    leap_working["leap_value"] = pd.to_numeric(leap_working["leap_value"], errors="coerce")

    # LEAP rows are year-specific and should be carried through directly.
    for subtotal_col in ["leap_is_subtotal", "esto_is_subtotal", "ninth_is_subtotal"]:
        if subtotal_col not in leap_working.columns:
            leap_working[subtotal_col] = False
        leap_working[subtotal_col] = leap_working[subtotal_col].fillna(False).astype(bool)

    leap_rows = leap_working[
        [
            "scenario",
            "sheet_name",
            "measure",
            "fuel_label",
            "year",
            "leap_value",
            "leap_is_subtotal",
            "esto_is_subtotal",
            "ninth_is_subtotal",
        ]
    ].rename(columns={"sheet_name": "sheet", "leap_value": "value"})
    leap_rows["economy"] = projection_economy
    leap_rows["source"] = "leap"
    rows.extend(
        leap_rows[
            [
                "economy",
                "scenario",
                "sheet",
                "measure",
                "fuel_label",
                "source",
                "year",
                "value",
                "leap_is_subtotal",
                "esto_is_subtotal",
                "ninth_is_subtotal",
            ]
        ].to_dict("records")
    )

    # Base/projection comparators should be added once per unique mapping group,
    # not once per LEAP year-row.
    mapping_group_cols = [
        "scenario",
        "sheet_name",
        "measure",
        "fuel_label",
        "sector_code_9th",
        "ninth_fuel_code",
        "esto_flow",
        "esto_product",
        "leap_is_subtotal",
        "esto_is_subtotal",
        "ninth_is_subtotal",
    ]
    mapping_groups = (
        leap_working[mapping_group_cols]
        .drop_duplicates()
        .rename(columns={"sheet_name": "sheet"})
        .reset_index(drop=True)
    )

    for row in mapping_groups.itertuples(index=False):
        row_dict = row._asdict()
        scenario = _normalize_scenario(row_dict.get("scenario", ""))
        sheet = _clean_token(row_dict.get("sheet", ""))
        measure = _clean_token(row_dict.get("measure", ""))
        fuel_label = _clean_token(row_dict.get("fuel_label", ""))
        sector_code = _clean_token(row_dict.get("sector_code_9th", ""))
        ninth_fuel_code = _clean_token(row_dict.get("ninth_fuel_code", ""))
        esto_flow = _clean_token(row_dict.get("esto_flow", ""))
        esto_product = _clean_token(row_dict.get("esto_product", ""))
        leap_is_subtotal = bool(row_dict.get("leap_is_subtotal", False))
        esto_is_subtotal = bool(row_dict.get("esto_is_subtotal", False))
        ninth_is_subtotal = bool(row_dict.get("ninth_is_subtotal", False))

        base_key = (esto_flow, esto_product)
        if base_key not in base_cache:
            base_cache[base_key] = pull_base_year_value(
                base_df,
                base_year=base_year,
                economy_code=base_economy,
                esto_flow=esto_flow,
                esto_product=esto_product,
                value_sign_role="",
            )
        rows.append(
            {
                "economy": base_economy,
                "scenario": scenario,
                "sheet": sheet,
                "measure": measure,
                "fuel_label": fuel_label,
                "source": "base",
                "year": int(base_year),
                "value": float(base_cache[base_key]) if pd.notna(base_cache[base_key]) else float("nan"),
                "leap_is_subtotal": leap_is_subtotal,
                "esto_is_subtotal": esto_is_subtotal,
                "ninth_is_subtotal": ninth_is_subtotal,
            }
        )

        projection_scenario = scenario_to_projection.get(scenario, scenario.lower())
        projection_key = (sector_code, ninth_fuel_code, projection_scenario)
        if projection_key not in projection_cache:
            projection_cache[projection_key] = pull_projection_series(
                ninth_df,
                sector_code=sector_code,
                fuel_code=ninth_fuel_code,
                economy_code=projection_economy,
                scenario=projection_scenario,
                projection_years=projection_years,
                value_sign_role="",
            )
        for proj_year, proj_value in projection_cache[projection_key].items():
            rows.append(
                {
                    "economy": projection_economy,
                    "scenario": scenario,
                    "sheet": sheet,
                    "measure": measure,
                    "fuel_label": fuel_label,
                    "source": "projection",
                    "year": int(proj_year),
                    "value": float(proj_value) if pd.notna(proj_value) else float("nan"),
                    "leap_is_subtotal": leap_is_subtotal,
                    "esto_is_subtotal": esto_is_subtotal,
                    "ninth_is_subtotal": ninth_is_subtotal,
                }
            )

    comparison_long = pd.DataFrame(rows)
    comparison_long["year"] = pd.to_numeric(comparison_long["year"], errors="coerce").astype("Int64")
    comparison_long["value"] = pd.to_numeric(comparison_long["value"], errors="coerce")
    if not template_groups.empty and "template_is_multi_flow" in template_groups.columns:
        multi_template_groups = template_groups[template_groups["template_is_multi_flow"].fillna(False).astype(bool)].copy()
        if not multi_template_groups.empty:
            covered_cols = [
                "chart_group_key",
                "measure",
                "fuel_label",
                "esto_flow_key",
            ]
            for col in covered_cols:
                if col not in multi_template_groups.columns:
                    multi_template_groups[col] = ""
                if col not in comparison_long.columns:
                    comparison_long[col] = ""
                multi_template_groups[col] = multi_template_groups[col].fillna("").astype(str).str.strip()
                comparison_long[col] = comparison_long[col].fillna("").astype(str).str.strip()
            covered = set(
                tuple(row)
                for row in multi_template_groups[covered_cols]
                .drop_duplicates()
                .itertuples(index=False, name=None)
            )
            if covered:
                direct_duplicate_mask = (
                    ~comparison_long["sheet"].fillna("").astype(str).str.startswith("template__")
                    & comparison_long[covered_cols].apply(lambda row: tuple(row) in covered, axis=1)
                )
                comparison_long = comparison_long.loc[~direct_duplicate_mask].copy()

    for subtotal_col in ["leap_is_subtotal", "esto_is_subtotal", "ninth_is_subtotal"]:
        if subtotal_col not in comparison_long.columns:
            comparison_long[subtotal_col] = False
        comparison_long[subtotal_col] = comparison_long[subtotal_col].fillna(False).astype(bool)
    flag_meta = (
        comparison_long.groupby(["scenario", "sheet", "measure", "fuel_label"], dropna=False)[
            ["leap_is_subtotal", "esto_is_subtotal", "ninth_is_subtotal"]
        ]
        .max()
        .reset_index()
    )
    comparison_long = comparison_long.drop(columns=["leap_is_subtotal", "esto_is_subtotal", "ninth_is_subtotal"])

    group_cols = [
        "economy",
        "scenario",
        "sheet",
        "page_key",
        "page_label",
        "chart_group_key",
        "chart_group_label",
        "dashboard_page_key",
        "dashboard_page_label",
        "dashboard_section_key",
        "dashboard_section_label",
        "chart_kind",
        "esto_flow_key",
        "esto_flow_group_key",
        "esto_flow_group_label",
        "measure",
        "fuel_label",
        "source",
        "year",
    ]
    comparison_long = _backfill_dashboard_hierarchy(comparison_long, sheet_catalog=hierarchy_sheet_catalog)
    comparison_long = _add_dashboard_context_aliases(_add_esto_flow_context_columns(comparison_long))
    comparison_long = _collapse_template_multi_flow_comparison_rows(comparison_long, template_groups)
    comparison_long = _add_leap_parent_transfer_rows_to_template(comparison_long)
    comparison_long = _fill_template_transfer_base_values(
        comparison_long,
        base_df=base_df,
        base_year=base_year,
        base_economy=base_economy,
    )
    comparison_long = (
        comparison_long.groupby(group_cols, as_index=False)["value"]
        .sum(min_count=1)
        .sort_values(group_cols, kind="mergesort")
        .reset_index(drop=True)
    )
    comparison_long = _fill_template_transfer_base_values(
        comparison_long,
        base_df=base_df,
        base_year=base_year,
        base_economy=base_economy,
    )
    comparison_long = comparison_long.merge(
        flag_meta,
        on=["scenario", "sheet", "measure", "fuel_label"],
        how="left",
    )
    for subtotal_col in ["leap_is_subtotal", "esto_is_subtotal", "ninth_is_subtotal"]:
        comparison_long[subtotal_col] = comparison_long[subtotal_col].fillna(False).astype(bool)

    comparison_wide = (
        comparison_long.pivot_table(
            index=[
                "economy",
                "scenario",
                "sheet",
                "page_key",
                "page_label",
                "chart_group_key",
                "chart_group_label",
                "dashboard_page_key",
                "dashboard_page_label",
                "dashboard_section_key",
                "dashboard_section_label",
                "chart_kind",
                "esto_flow_key",
                "esto_flow_group_key",
                "esto_flow_group_label",
                "measure",
                "fuel_label",
                "year",
            ],
            columns="source",
            values="value",
            aggfunc="first",
        )
        .reset_index()
    )
    if hasattr(comparison_wide.columns, "name"):
        comparison_wide.columns.name = None

    mapping_status_out = _build_mapping_status_with_availability(mapping_status, comparison_long)

    issues_cfg = known_issues or {}
    for gap in list(issues_cfg.get("known_gaps", []) or []):
        sheet = _clean_token(gap.get("sheet", ""))
        fuel = _clean_token(gap.get("fuel_label", ""))
        note = _clean_token(gap.get("note", ""))
        if not (sheet and fuel and note):
            continue
        mask = mapping_status_out["sheet"].eq(sheet) & mapping_status_out["fuel_label"].eq(fuel)
        if not mask.any():
            continue
        current = mapping_status_out.loc[mask, "mapping_note"].fillna("").astype(str).str.strip()
        mapping_status_out.loc[mask, "mapping_note"] = current.where(current.ne(""), note)

    return {
        "comparison_long": comparison_long,
        "comparison_wide": comparison_wide,
        "mapping_status": mapping_status_out,
        "base_df": base_df,
        "ninth_df": ninth_df,
        "reassignment_status": reassignment_status,
        "synthetic_reference_status": synthetic_reference_status,
        "mapping_inputs": mapping_inputs,
    }


def _flatten_page_tree(tree: list[dict[str, Any]]) -> list[dict[str, Any]]:
    flat: list[dict[str, Any]] = []

    def _walk(node: dict[str, Any], path: list[str]) -> None:
        label = _clean_token(node.get("label", ""))
        node_id = _clean_token(node.get("id", ""))
        current_path = path + ([label] if label else [])
        flat.append(
            {
                "id": node_id,
                "label": label,
                "path": current_path,
                "children": list(node.get("children", []) or []),
            }
        )
        for child in list(node.get("children", []) or []):
            if isinstance(child, dict):
                _walk(child, current_path)

    for root in tree:
        if isinstance(root, dict):
            _walk(root, [])
    return flat


def _page_filename_from_path(path_parts: Sequence[str]) -> str:
    if not path_parts:
        return "index.html"
    token = "__".join(_safe_token(part.replace("\\", "_")) for part in path_parts)
    return f"node__{token}.html"


def _display_balance_path_part(part: object, *, is_top: bool = False) -> str:
    text = str(part or "").strip()
    if is_top:
        return BALANCE_DASHBOARD_TOP_LABELS.get(text, text)
    return text


def _section_id_from_path(path: Sequence[str]) -> str:
    token = "__".join(_safe_token(str(part).replace("\\", "_")) for part in path if str(part).strip())
    return f"sec-{token or 'section'}"


def _chart_file_lookup(comparison_long: pd.DataFrame, *, hide_leap_only_charts: bool) -> dict[tuple[str, str, str], str]:
    render_long = _prepare_render_long(comparison_long)
    chart_lookup: dict[tuple[str, str, str], str] = {}

    for (sheet, measure, fuel), sub in render_long.groupby(["sheet", "measure", "fuel_label"], dropna=False):
        values = pd.to_numeric(sub["value"], errors="coerce").fillna(0.0)
        if not values.ne(0).any():
            continue
        force_show_chart = (
            bool(sub["force_show_chart"].fillna(False).astype(bool).any())
            if "force_show_chart" in sub.columns
            else False
        )
        if hide_leap_only_charts:
            non_leap_sources = {
                str(src).strip()
                for src in sub["source"].dropna().astype(str)
                if str(src).strip() and str(src).strip() != "leap"
            }
            if not non_leap_sources and not force_show_chart:
                continue

        sheet_text = str(sheet)
        measure_text = str(measure or "").strip()
        file_sheet = f"{sheet_text}__{measure_text}" if measure_text else sheet_text
        sheet_slug = _safe_token(file_sheet.replace("\\", "_"))
        fuel_slug = _safe_token(str(fuel))
        chart_lookup[(sheet_text, measure_text, str(fuel))] = f"{sheet_slug}__{fuel_slug}.html"

    return chart_lookup


def _collect_node_chart_entries(
    *,
    comparison_long: pd.DataFrame,
    sheet_catalog: dict[str, Any],
    chart_lookup: dict[tuple[str, str, str], str],
) -> tuple[dict[tuple[str, ...], list[dict[str, str]]], dict[str, list[dict[str, str]]]]:
    sheet_charts: dict[str, list[dict[str, str]]] = defaultdict(list)

    comp = comparison_long.copy()
    comp = _backfill_dashboard_hierarchy(comp, sheet_catalog=sheet_catalog)
    comp["sheet"] = comp["sheet"].astype(str)
    comp["measure"] = comp.get("measure", "").fillna("").astype(str)
    comp["fuel_label"] = comp["fuel_label"].astype(str)

    for (sheet, measure, fuel), _ in comp.groupby(["sheet", "measure", "fuel_label"], dropna=False):
        key = (str(sheet), str(measure or "").strip(), str(fuel))
        file_name = chart_lookup.get(key)
        if not file_name:
            continue
        sheet_charts[str(sheet)].append(
            {
                "sheet": str(sheet),
                "measure": str(measure or "").strip(),
                "fuel": str(fuel),
                "file": file_name,
            }
        )

    node_entries: dict[tuple[str, ...], list[dict[str, str]]] = defaultdict(list)
    for sheet_key, entries in sheet_charts.items():
        cfg = sheet_catalog.get(sheet_key, {}) or {}
        path = tuple(str(v).strip() for v in list(cfg.get("path", []) or []) if str(v).strip())
        if not path:
            path = (sheet_key,)
        display_sheet = _clean_token(cfg.get("display_label", "")) or (path[-1] if path else sheet_key)
        display_path = [
            _display_balance_path_part(part, is_top=idx == 0)
            for idx, part in enumerate(path)
        ]
        hierarchy = _dashboard_hierarchy_from_path(
            path,
            entry_kind="direct",
            measure=(entries[0].get("measure", "") if entries else ""),
            fallback_label=sheet_key,
        )
        for entry in entries:
            tagged = dict(entry)
            tagged.update(hierarchy)
            tagged["display_sheet"] = display_sheet
            tagged["path_label"] = " > ".join(display_path)
            node_entries[path].append(tagged)

    return node_entries, sheet_charts


def _load_dashboard_template_allowlist(path: Path | str | None) -> dict[str, Any]:
    """
    Read the optional dashboard navigation template as an authoring allowlist.

    When supplied, this config is the source of truth for which dashboard paths
    and fuels may appear in navigation. The renderer still writes the expanded
    chart hierarchy to outputs, but it should not invent menu entries that are
    absent from the checked-in template.
    """
    if path is None:
        return {}
    resolved = _resolve(path)
    if not resolved.exists():
        return {}
    try:
        template = json.loads(resolved.read_text(encoding="utf-8"))
    except json.JSONDecodeError as exc:
        raise ValueError(f"Invalid dashboard navigation template JSON at {resolved}: {exc}") from exc
    if not isinstance(template, dict):
        raise ValueError(f"Dashboard navigation template must be a JSON object: {resolved}")
    _validate_dashboard_template_allowlist(template, resolved)
    return template


TEMPLATE_RESERVED_KEYS = {
    "defaults",
    "about_page",
    "aggregate",
    "aggregate_graphs",
    "graphs",
    "by_fuel_graphs",
    "esto_flow",
    "_comments",
    "comments",
}


def _dashboard_template_is_reserved_key(key: object) -> bool:
    """Return True for template metadata keys that should not create pages."""
    text = str(key or "").strip()
    return text in TEMPLATE_RESERVED_KEYS or text.lower().startswith("note")


def _dashboard_template_about_page(template: dict[str, Any]) -> dict[str, Any]:
    about = template.get("about_page", {}) if isinstance(template, dict) else {}
    return about if isinstance(about, dict) else {}


def _as_clean_list(value: object) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    if isinstance(value, tuple):
        return [str(item).strip() for item in value if str(item).strip()]
    text = str(value or "").strip()
    return [text] if text else []


def _template_measure_list(
    raw_spec: dict[str, Any],
    default_measure: str = "",
    *,
    use_default: bool = True,
) -> list[str]:
    measures_raw = raw_spec.get("measures", raw_spec.get("measure", ""))
    measures = _as_clean_list(measures_raw)
    if use_default and not measures and str(default_measure or "").strip():
        measures = [str(default_measure).strip()]
    return measures


def _dashboard_template_aggregate_specs(
    node: dict[str, Any],
    *,
    default_measure: str = "",
) -> list[dict[str, Any]]:
    """Normalize V2 aggregate_graphs and legacy aggregate declarations."""
    if not isinstance(node, dict):
        return []
    raw_specs: list[dict[str, Any]] = []
    for key, value in node.items():
        if key == "aggregate_graphs" and isinstance(value, dict):
            raw_specs.append(value)
        elif key == "aggregate" and isinstance(value, dict):
            raw_specs.append(
                {
                    "fuels": value.get("fuel", value.get("fuels", "Total")),
                    "esto_flows": value.get("source_esto_flows", value.get("esto_flows", [])),
                    "measures": value.get("measures", value.get("measure", "")),
                }
            )

    normalized: list[dict[str, Any]] = []
    for idx, raw_spec in enumerate(raw_specs, start=1):
        flows = _as_clean_list(raw_spec.get("esto_flows", raw_spec.get("source_esto_flows", [])))
        fuel = str(raw_spec.get("fuels", raw_spec.get("fuel", "Total")) or "Total").strip() or "Total"
        measures = _template_measure_list(raw_spec, default_measure)
        if not flows:
            continue
        normalized.append(
            {
                "aggregate_id": f"aggregate_{idx}",
                "fuel": fuel,
                "fuels": fuel,
                "source_flows": flows,
                "esto_flows": flows,
                "source_flows_norm": {" ".join(flow.lower().split()) for flow in flows if flow},
                "fuel_norm": " ".join(fuel.lower().split()),
                "measures": measures,
                "use_esto_to_ninth_mapping": _to_bool(
                    raw_spec.get("use_esto_to_ninth_mapping", False),
                    default=False,
                ),
            }
        )
    return normalized


def _validate_dashboard_template_allowlist(template: dict[str, Any], resolved: Path) -> None:
    def _walk(node: dict[str, Any], path: tuple[str, ...] = ()) -> None:
        if "fuels" in node:
            label = " > ".join(path) or "<root>"
            raise ValueError(
                "Dashboard navigation template uses legacy 'fuels' format at "
                f"{label} in {resolved}. Convert this node to the 'graphs' format."
            )
        graphs = node.get("graphs")
        if graphs is not None and not isinstance(graphs, dict):
            label = " > ".join(path) or "<root>"
            raise ValueError(f"Template 'graphs' must be an object at {label} in {resolved}")
        aggregate_graphs = node.get("aggregate_graphs")
        if aggregate_graphs is not None and not isinstance(aggregate_graphs, dict):
            label = " > ".join(path) or "<root>"
            raise ValueError(f"Template 'aggregate_graphs' must be an object at {label} in {resolved}")
        by_fuel_graphs = node.get("by_fuel_graphs")
        if by_fuel_graphs is not None and not isinstance(by_fuel_graphs, dict):
            label = " > ".join(path) or "<root>"
            raise ValueError(f"Template 'by_fuel_graphs' must be an object at {label} in {resolved}")
        for key, child in node.items():
            if _dashboard_template_is_reserved_key(key) or not isinstance(child, dict):
                continue
            _walk(child, (*path, str(key).strip()))

    _walk(template)


def _dashboard_template_graph_specs(
    node: dict[str, Any],
    *,
    default_measure: str = "",
) -> list[dict[str, Any]]:
    """Normalize V2 by_fuel_graphs and legacy graph declarations."""
    if not isinstance(node, dict):
        return []
    common_flows = _as_clean_list(node.get("esto_flow", ""))
    normalized: list[dict[str, Any]] = []

    def _normalize_one(graph_id: str, raw_spec: dict[str, Any], fallback_flows: list[str]) -> None:
        if not isinstance(raw_spec, dict):
            raise ValueError(f"Dashboard template graph {graph_id!r} must be an object")
        flows = _as_clean_list(raw_spec.get("esto_flows", raw_spec.get("esto_flow", ""))) or fallback_flows
        measures = _template_measure_list(raw_spec, default_measure, use_default=False)
        products_raw = raw_spec.get("products", raw_spec.get("fuels", []))
        include_all = isinstance(products_raw, str) and str(products_raw).strip().lower() == "all"
        if include_all:
            products: list[str] = []
        else:
            if isinstance(products_raw, list):
                products = [str(value).strip() for value in products_raw if str(value).strip()]
            elif isinstance(products_raw, str) and str(products_raw).strip():
                products = [str(products_raw).strip()]
            else:
                products = []
        exclude_products = [
            str(value).strip()
            for value in list(raw_spec.get("exclude_products", []) or [])
            if str(value).strip()
        ]
        products_norm = {norm_text for norm_text in [" ".join(product.lower().split()) for product in products] if norm_text}
        exclude_products_norm = {
            norm_text
            for norm_text in [" ".join(product.lower().split()) for product in exclude_products]
            if norm_text
        }
        for measure in measures or [""]:
            normalized.append(
                {
                    "graph_id": graph_id,
                    "esto_flow": flows[0] if flows else "",
                    "esto_flows": flows,
                    "esto_flows_norm": {" ".join(flow.lower().split()) for flow in flows if flow},
                    "measure": measure,
                    "include_all_products": include_all,
                    "products": products,
                    "products_norm": products_norm,
                    "exclude_products": exclude_products,
                    "exclude_products_norm": exclude_products_norm,
                    "is_multi_flow": len(flows) > 1,
                    "use_esto_to_ninth_mapping": _to_bool(
                        raw_spec.get("use_esto_to_ninth_mapping", False),
                        default=False,
                    ),
                }
            )

    for key, value in node.items():
        if key == "by_fuel_graphs" and isinstance(value, dict):
            _normalize_one("by_fuel_graphs", value, common_flows)
        elif key == "graphs" and isinstance(value, dict):
            for graph_name, raw_spec in value.items():
                graph_id = str(graph_name).strip()
                if not graph_id:
                    continue
                _normalize_one(graph_id, raw_spec, common_flows)
    return normalized


def build_esto_axis_structure_from_dashboard_template(
    chart_navigation_guide_path: Path | str | None,
) -> dict[str, Any]:
    """
    Build the ESTO-axis page structure from the checked-in dashboard template.

    The template is the authoring source of truth for dashboard hierarchy.  This
    structure object is only an in-memory adapter for the existing renderer and
    conversion code; workflows should write the resolved version as an output
    diagnostic, not maintain a second checked-in hierarchy file.
    """
    template = _load_dashboard_template_allowlist(chart_navigation_guide_path)
    if not template:
        return {"page_tree": [], "sheet_catalog": {}, "esto_flow_to_sheet": {}, "empty_page_notice": DEFAULT_EMPTY_PAGE_NOTICE}

    top_label_inverse = {v: k for k, v in BALANCE_DASHBOARD_TOP_LABELS.items()}
    measure_default = str((template.get("defaults") or {}).get("measure", "")).strip() or "Energy balance (PJ)"
    page_paths: list[list[str]] = []
    sheet_catalog: dict[str, dict[str, Any]] = {}
    esto_flow_to_sheet: dict[str, str] = {}

    def _raw_path(display_path: tuple[str, ...]) -> tuple[str, ...]:
        if not display_path:
            return display_path
        return (top_label_inverse.get(display_path[0], display_path[0]), *display_path[1:])

    def _remember_path(display_path: tuple[str, ...]) -> tuple[str, ...]:
        raw_path = _raw_path(display_path)
        if raw_path:
            page_paths.append(list(raw_path))
        return raw_path

    def _remember_flow(flow: object, display_path: tuple[str, ...]) -> None:
        flow_text = str(flow or "").strip()
        if not flow_text:
            return
        raw_path = _raw_path(display_path)
        sheet_key = _sheet_key_from_esto_flow(flow_text)
        esto_flow_to_sheet.setdefault(flow_text, sheet_key)
        existing = sheet_catalog.get(sheet_key)
        if existing is not None and len(existing.get("path", []) or []) >= len(raw_path):
            return
        hierarchy = _dashboard_hierarchy_from_path(
            raw_path if raw_path else _esto_flow_path(flow_text, _esto_flow_label_lookup([flow_text])),
            entry_kind="direct",
            measure=measure_default,
            fallback_label=flow_text,
        )
        sheet_catalog[sheet_key] = {
            "display_label": raw_path[-1] if raw_path else (_strip_esto_code_prefix(flow_text) or flow_text),
            "path": list(raw_path) if raw_path else _esto_flow_path(flow_text, _esto_flow_label_lookup([flow_text])),
            "measure": measure_default,
            "sort_order": len(sheet_catalog),
            **hierarchy,
        }

    def _walk(node: dict[str, Any], display_path: tuple[str, ...] = ()) -> None:
        if display_path:
            _remember_path(display_path)
        for aggregate in _dashboard_template_aggregate_specs(node, default_measure=measure_default):
            for flow in list(aggregate.get("source_flows", []) or []):
                _remember_flow(flow, display_path)
        for spec in _dashboard_template_graph_specs(node, default_measure=measure_default):
            for flow in list(spec.get("esto_flows", []) or [spec.get("esto_flow", "")]):
                _remember_flow(flow, display_path)
        for key, child in node.items():
            if _dashboard_template_is_reserved_key(key) or not isinstance(child, dict):
                continue
            label = str(key).strip()
            if label:
                _walk(child, (*display_path, label))

    _walk(template)
    unique_paths: list[tuple[str, ...]] = []
    seen_paths: set[tuple[str, ...]] = set()
    for path in page_paths:
        clean_path = tuple(path)
        if not clean_path or clean_path in seen_paths:
            continue
        seen_paths.add(clean_path)
        unique_paths.append(clean_path)
    return {
        "page_tree": _build_page_tree_from_paths([list(path) for path in unique_paths]),
        "sheet_catalog": sheet_catalog,
        "esto_flow_to_sheet": esto_flow_to_sheet,
        "empty_page_notice": DEFAULT_EMPTY_PAGE_NOTICE,
    }


def _select_all_graph_products(
    *,
    esto_flow: str,
    measure: str = "",
    exclude_products_norm: set[str],
    mapping_status: pd.DataFrame,
    chart_comparison_long: "pd.DataFrame | None" = None,
) -> tuple[list[str], list[str]]:
    """
    Return (selected, available) ESTO products for a fuels:'all' graph spec.

    'available' is the set of non-total ESTO products for the given flow that
    have any mapped data. For transformation measures (Inputs or Outputs), when
    chart_comparison_long is supplied, available is refined to products whose
    fuel_label appears in the sign-split chart data for that measure — covering
    LEAP or 9th rows that are chartable for that measure even when ESTO
    base-year is zero.

    'selected' = available minus exclude_products.

    This is the single canonical resolver for fuels:'all' graphs. All callers
    must use this function; do not reimplement the logic inline.
    """
    def _norm(v: object) -> str:
        return " ".join(str(v or "").strip().lower().split())

    def _is_total(v: object) -> bool:
        text = _norm(v)
        return text == "total" or bool(re.match(r"^\d+(?:[._]\d+)*\s+total$", text))

    if mapping_status is None or mapping_status.empty:
        return [], []

    ms = mapping_status.copy()
    for col in ["esto_flow", "esto_product"]:
        if col not in ms.columns:
            ms[col] = ""
        ms[col] = ms[col].fillna("").astype(str).str.strip()

    flow_norm = _norm(esto_flow)
    flow_mask = ms["esto_flow"].map(_norm).eq(flow_norm)
    available_from_mapping = sorted({
        p
        for p in ms.loc[flow_mask, "esto_product"].unique()
        if p and not _is_total(p)
    })

    measure_norm = _norm(measure)
    if (
        chart_comparison_long is not None
        and not chart_comparison_long.empty
        and "fuel_label" in ms.columns
    ):
        # Identify which rendered chart rows carry this ESTO flow, then keep
        # products whose display rows have non-zero magnitude in any source.
        ms_fl = ms.copy()
        for col in ["sheet", "fuel_label"]:
            if col not in ms_fl.columns:
                ms_fl[col] = ""
            ms_fl[col] = ms_fl[col].fillna("").astype(str).str.strip()
        flow_sheets = set(ms_fl.loc[flow_mask, "sheet"].unique()) - {""}

        if flow_sheets:
            cdf = chart_comparison_long.copy()
            for col in ["sheet", "fuel_label", "measure"]:
                if col not in cdf.columns:
                    cdf[col] = ""
                cdf[col] = cdf[col].fillna("").astype(str).str.strip()
            if "value" not in cdf.columns:
                cdf["value"] = 0.0
            cdf["value"] = pd.to_numeric(cdf["value"], errors="coerce").fillna(0.0)

            chart_fuel_labels = set(
                cdf.loc[
                    cdf["sheet"].isin(flow_sheets)
                    & cdf["measure"].map(_norm).eq(measure_norm)
                    & cdf["value"].abs().gt(1e-12),
                    "fuel_label",
                ].unique()
            ) - {""}

            # Resolve fuel_labels → esto_products via mapping_status.
            available = sorted({
                str(getattr(row, "esto_product", "")).strip()
                for row in ms_fl.loc[flow_mask].itertuples(index=False)
                if str(getattr(row, "fuel_label", "")).strip() in chart_fuel_labels
                and str(getattr(row, "esto_product", "")).strip()
                and not _is_total(str(getattr(row, "esto_product", "")).strip())
            })
        else:
            available = available_from_mapping
    else:
        available = available_from_mapping

    selected = [p for p in available if _norm(p) not in exclude_products_norm]
    return selected, available


def write_dashboard_graph_fuel_coverage(
    *,
    template: dict[str, Any],
    mapping_status: pd.DataFrame,
    output_path: Path | str,
    default_measure: str = "Energy balance (PJ)",
    chart_comparison_long: "pd.DataFrame | None" = None,
) -> str:
    """
    Write per-graph ESTO product availability and selection diagnostics.

    For each graph in the template, one CSV row is produced containing:
    - dashboard_path: human-readable path (e.g. "Buildings > Residential")
    - node_esto_flow: ESTO flow code attached to the node
    - graph_id: arbitrary graph key from the template
    - measure: chart measure string
    - selection_mode: "all" or "explicit_products"
    - available_esto_products: all non-total products in mapping_status for the flow
    - products_selected: products that will be charted for this graph
    - excluded_products: products explicitly removed via exclude_products
    - available_not_shown: available products absent from the selected set

    For explicit_products graphs, products_selected is the declared list (these
    may include products not present in the current data). For fuels:'all' graphs,
    both sets are derived from mapping_status so the diagnostic is data-driven.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    def _norm(v: object) -> str:
        return " ".join(str(v or "").strip().lower().split())

    def _is_total(v: object) -> bool:
        text = _norm(v)
        return text == "total" or bool(re.match(r"^\d+(?:[._]\d+)*\s+total$", text))

    # Build flow_norm -> sorted list of available non-total products from mapping_status.
    flow_products: dict[str, list[str]] = {}
    if mapping_status is not None and not mapping_status.empty:
        ms = mapping_status.copy()
        for col in ["esto_flow", "esto_product"]:
            if col not in ms.columns:
                ms[col] = ""
            ms[col] = ms[col].fillna("").astype(str).str.strip()
        by_flow: dict[str, set[str]] = defaultdict(set)
        for row in ms[["esto_flow", "esto_product"]].drop_duplicates().itertuples(index=False):
            fn = _norm(row.esto_flow)
            product = row.esto_product
            if fn and product and not _is_total(product):
                by_flow[fn].add(product)
        flow_products = {fn: sorted(products) for fn, products in by_flow.items()}

    rows: list[dict[str, Any]] = []
    measure_default = (
        str((template.get("defaults") or {}).get("measure", default_measure)).strip()
        or default_measure
    )

    def _walk(node: dict[str, Any], path: tuple[str, ...]) -> None:
        specs = _dashboard_template_graph_specs(node, default_measure=measure_default)
        if specs:
            path_str = " > ".join(path) if path else "<root>"
            for spec in specs:
                flow = str(spec.get("esto_flow", "")).strip()
                flows = [str(item).strip() for item in list(spec.get("esto_flows", []) or [flow]) if str(item).strip()]
                measure = str(spec.get("measure", "")).strip() or measure_default
                graph_id = str(spec.get("graph_id", "")).strip()
                include_all = bool(spec.get("include_all_products", False))
                explicit_products = list(spec.get("products", []) or [])
                exclude_products = list(spec.get("exclude_products", []) or [])
                exclude_norm = set(spec.get("exclude_products_norm", set()))

                if include_all:
                    selection_mode = "all"
                    selected_set: set[str] = set()
                    available_set: set[str] = set()
                    for flow_item in flows:
                        selected_part, available_part = _select_all_graph_products(
                            esto_flow=flow_item,
                            measure=measure,
                            exclude_products_norm=exclude_norm,
                            mapping_status=mapping_status,
                            chart_comparison_long=chart_comparison_long,
                        )
                        selected_set.update(selected_part)
                        available_set.update(available_part)
                    selected = sorted(selected_set)
                    available = sorted(available_set)
                else:
                    selection_mode = "explicit_products"
                    available = sorted({product for flow_item in flows for product in flow_products.get(_norm(flow_item), [])})
                    selected = explicit_products

                selected_set = set(selected)
                available_not_shown = [p for p in available if p not in selected_set]
                component_pairs = [
                    f"{flow_item} | {product}"
                    for flow_item in flows
                    for product in selected
                ]

                rows.append({
                    "dashboard_path": path_str,
                    "node_esto_flow": flow,
                    "configured_esto_flows": "; ".join(flows),
                    "graph_id": graph_id,
                    "measure": measure,
                    "selection_mode": selection_mode,
                    "available_esto_products": "; ".join(available),
                    "products_selected": "; ".join(selected),
                    "excluded_products": "; ".join(exclude_products),
                    "available_not_shown": "; ".join(available_not_shown),
                    "component_esto_pairs": " || ".join(component_pairs),
                })

        for key, child in node.items():
            if _dashboard_template_is_reserved_key(key) or not isinstance(child, dict):
                continue
            label = str(key).strip()
            if label:
                _walk(child, (*path, label))

    _walk(template, ())

    columns = [
        "dashboard_path",
        "node_esto_flow",
        "configured_esto_flows",
        "graph_id",
        "measure",
        "selection_mode",
        "available_esto_products",
        "products_selected",
        "excluded_products",
        "available_not_shown",
        "component_esto_pairs",
    ]
    df = pd.DataFrame(rows, columns=columns) if rows else pd.DataFrame(columns=columns)
    if not df.empty:
        df = df.sort_values(
            ["dashboard_path", "node_esto_flow", "graph_id", "measure", "selection_mode"],
            kind="mergesort",
        ).reset_index(drop=True)
    df.to_csv(output_path, index=False)
    return str(output_path)


def _split_transformation_input_output_measures(
    comparison_long: pd.DataFrame,
    sheet_catalog: dict[str, Any],
) -> pd.DataFrame:
    """
    Split ESTO-axis transformation balances into input and output measures.

    Balance rows are signed: negative values are inputs and positive values are
    outputs. Zero values are copied into both measures so a scenario/source with
    a zero baseline remains visible when another source has non-zero values.
    """
    if comparison_long.empty or not sheet_catalog:
        return comparison_long.copy()

    sheet_top_group = {
        str(sheet).strip(): str((cfg or {}).get("path", [""])[0]).strip()
        for sheet, cfg in sheet_catalog.items()
        if str(sheet).strip() and (cfg or {}).get("path")
    }
    transformation_sheets = {
        sheet
        for sheet, top_group in sheet_top_group.items()
        if top_group in TRANSFORMATION_DASHBOARD_TOP_GROUPS
    }
    if not transformation_sheets:
        return comparison_long.copy()

    frame = comparison_long.copy()
    if "measure" not in frame.columns:
        frame["measure"] = ""
    frame["measure"] = frame["measure"].fillna("").astype(str)
    frame["sheet"] = frame["sheet"].fillna("").astype(str)
    frame["value"] = pd.to_numeric(frame["value"], errors="coerce")

    dashboard_top = pd.Series("", index=frame.index, dtype="object")
    for col in ["dashboard_page_label", "page_label"]:
        if col in frame.columns:
            candidate = frame[col].fillna("").astype(str).str.strip()
            dashboard_top = dashboard_top.where(dashboard_top.astype(str).str.strip().ne(""), candidate)

    target = frame["sheet"].isin(transformation_sheets) | dashboard_top.isin(TRANSFORMATION_DASHBOARD_TOP_GROUPS)
    unchanged = frame[~target].copy()
    transform = frame[target].copy()
    if transform.empty:
        return frame

    input_rows = transform[transform["value"].le(0) | transform["value"].isna()].copy()
    output_rows = transform[transform["value"].ge(0) | transform["value"].isna()].copy()
    input_rows["measure"] = TRANSFORMATION_INPUT_MEASURE
    output_rows["measure"] = TRANSFORMATION_OUTPUT_MEASURE
    input_rows["value"] = input_rows["value"].abs()
    output_rows["value"] = output_rows["value"].abs()

    return pd.concat([unchanged, input_rows, output_rows], ignore_index=True, sort=False)


_V2_STYLE_CSS = """
:root {
  color-scheme: light;
  --page-padding-x: clamp(12px, 1.8vw, 24px);
  --page-padding-y: clamp(14px, 1.8vw, 24px);
  --body-font-size: clamp(15px, 0.22vw + 14px, 18px);
  --title-font-size: clamp(24px, 0.75vw + 18px, 34px);
  --section-title-size: clamp(18px, 0.45vw + 14px, 24px);
}
html { background: #f4f6f8; }
body {
  font-family: Segoe UI, Arial, sans-serif;
  margin: 0;
  background: #f4f6f8;
  color: #111;
  font-size: var(--body-font-size);
  line-height: 1.45;
  min-width: 320px;
}
a { color: #0b3d5c; text-decoration: none; }
a:hover { text-decoration: underline; }
.page-shell { width: 100%; max-width: none; margin: 0 auto; padding: 0 var(--page-padding-x) 32px; box-sizing: border-box; }
.page-header {
  position: sticky;
  top: 0;
  z-index: 100;
  margin: 0 0 14px 0;
  padding: var(--page-padding-y) 0 10px 0;
  background: rgba(244, 246, 248, 0.96);
  backdrop-filter: blur(8px);
  border-bottom: 1px solid #d8dee4;
}
.header-main-row { display:flex;flex-wrap:wrap;justify-content:space-between;align-items:flex-start;gap:10px 16px; }
.header-side-controls {
  display:flex;flex-wrap:wrap;align-items:center;justify-content:flex-end;
  gap:8px 10px;flex:1 1 460px;
}
.header-inline-controls {
  display:flex;align-items:center;gap:8px;justify-content:flex-end;
  flex:0 0 auto;flex-wrap:nowrap;margin-left:auto;
}
.header-links { display: flex; flex-wrap: wrap; gap: 8px; }
.header-chip {
  padding: 6px 10px;
  border: 1px solid #c5ccd3;
  border-radius: 6px;
  background: #fff;
  color: #0b3d5c;
  font-size: 13px;
  text-decoration: none;
}
.header-chip[data-current="true"] {
  border-color:#1f6feb;
  box-shadow:0 0 0 2px rgba(31, 111, 235, 0.16);
  font-weight:700;
}
.header-nav-separator {
  color:#6b7280;
  font-weight:700;
  line-height:1.25;
  padding:6px 2px;
}
.header-toggle {
  width: 30px;
  height: 30px;
  border: 1px solid #c5ccd3;
  border-radius: 999px;
  background: #fff;
  color: #0b3d5c;
  cursor: pointer;
}
.header-toggle-row {
  display:flex;
  justify-content:flex-end;
  margin-top:8px;
}
.page-header.is-collapsed .header-collapsible { display: none; }
.page-header.is-collapsed {
  padding-bottom:0;
  background:transparent;
  backdrop-filter:none;
  border-bottom-color:transparent;
}
.page-header.is-collapsed .header-toggle-row { margin-top:0; }
.jump-nav {
  margin-top:8px;
  padding-top:8px;
  border-top:1px solid #d8dee4;
  display:flex;
  flex-wrap:wrap;
  gap:8px 10px;
  align-items:flex-start;
}
.jump-nav-label {
  font-weight:600;
  color:#4b5563;
  font-size:12px;
  white-space:nowrap;
  padding-top:4px;
}
.jump-nav-groups {
  display:flex;
  flex-direction:column;
  gap:6px;
  min-width:0;
  flex:1 1 640px;
}
.jump-nav-row {
  display:flex;
  flex-wrap:wrap;
  gap:6px 8px;
  align-items:center;
  min-width:0;
}
.jump-nav-row[data-level="2"] { padding-left:18px; }
.jump-nav-row[data-level="3"] { padding-left:36px; }
.jump-nav-row[data-level="4"] { padding-left:54px; }
.jump-chip {
  position:relative;
  display:inline-flex;
  align-items:center;
  gap:6px;
  padding:4px 9px;
  border:1px solid #c5ccd3;
  border-radius:999px;
  background:#fff;
  color:#0b3d5c;
  text-decoration:none;
  font-size:12px;
  line-height:1.25;
  box-shadow:0 1px 1px rgba(15, 23, 42, 0.04);
}
.jump-chip::before {
  content:"";
  display:block;
  width:8px;
  height:8px;
  border-radius:999px;
  flex:0 0 auto;
  background:#94a3b8;
}
.jump-chip[data-level="1"] {
  background:#fff4e6;
  border-color:#f2a65a;
  color:#7a3b00;
}
.jump-chip[data-level="1"]::before { background:#f97316; }
.jump-chip[data-level="2"] {
  background:#f5edff;
  border-color:#c69af0;
  color:#4c1d70;
}
.jump-chip[data-level="2"]::before { background:#9333ea; }
.jump-chip[data-level="3"], .jump-chip[data-level="4"] {
  background:#f8fafc;
  border-color:#cbd5e1;
  color:#334155;
}
.jump-chip[data-level="3"]::before, .jump-chip[data-level="4"]::before { background:#94a3b8; }
.jump-chip[data-current="true"] {
  border-color:#1f6feb;
  box-shadow:0 0 0 2px rgba(31, 111, 235, 0.16);
  font-weight:700;
}
.jump-nav-separator {
  color:#6b7280;
  font-weight:700;
  line-height:1.25;
  padding:4px 2px;
}
.section-link-list {
  margin:0 0 14px 0;
  padding:0;
  list-style:none;
  display:flex;
  flex-direction:column;
  gap:5px;
}
.section-link-list li {
  line-height:1.35;
}
.section-link-list li[data-level="2"] { padding-left:18px; }
.section-link-list li[data-level="3"] { padding-left:36px; }
.section-link-list li[data-level="4"] { padding-left:54px; }
.section-link-list a {
  color:#0b3d5c;
  text-decoration:none;
  font-weight:600;
}
.section-card {
  margin: 10px 0 14px 0;
  padding: 10px;
  border: 1px solid #d6dde5;
  border-radius: 8px;
  background: #fff;
}
.section-note {
  margin: 10px 0 14px 0;
  padding: 12px;
  border: 1px dashed #c5ccd3;
  border-radius: 8px;
  background: #fff;
  color: #4b5563;
}
.meta-line { margin-bottom: 8px; font-weight: 600; color: #1f2d3d; }
.meta-subline { margin-top: -4px; margin-bottom: 8px; color: #4b5563; font-size: 12px; }
.dashboard-grid {
  display:grid;
  grid-template-columns:repeat(4, minmax(0, 1fr));
  gap:12px;
  align-items:start;
}
.dashboard-grid.expand-1 { grid-template-columns:minmax(0, 1fr); }
.dashboard-grid.expand-2 { grid-template-columns:repeat(2, minmax(0, 1fr)); }
.dashboard-grid.expand-3 { grid-template-columns:repeat(3, minmax(0, 1fr)); }
.chart-card {
  margin:0;
  padding:10px;
  border:1px solid #d0d7de;
  border-radius:8px;
  background:#fff;
  box-shadow:0 1px 2px rgba(0,0,0,0.05);
}
.chart-caption {
  font-weight:600;
  margin-bottom:4px;
}
.lazy-chart-frame {
  width: 100%;
  height: clamp(380px, 62vh, 1100px);
  border: 1px solid #d0d7de;
  border-radius: 6px;
  background: #fff;
  display: block;
  box-sizing: border-box;
}
.lazy-chart-frame.is-unloaded {
  background:#f8fafc;
}
.chart-load-state {
  min-height:22px;
  margin:4px 0 6px 0;
  color:#64748b;
  font-size:12px;
}
.chart-load-state[data-loaded="true"] { display:none; }
@media (max-width: 720px) {
  .dashboard-grid { grid-template-columns:minmax(0, 1fr); }
  .lazy-chart-frame { height: 420px; }
}
"""


_HEADER_TOGGLE_SCRIPT = """
(function() {
  const pageHeader = document.getElementById('page-header');
  const toggle = document.getElementById('header-toggle');
  if (!pageHeader || !toggle) return;
  const key = 'balance-dashboard-header-collapsed';
  const apply = (collapsed) => {
    pageHeader.classList.toggle('is-collapsed', collapsed);
    toggle.textContent = collapsed ? '▾' : '▴';
    toggle.setAttribute('aria-expanded', collapsed ? 'false' : 'true');
    toggle.setAttribute('aria-label', collapsed ? 'Expand header' : 'Collapse header');
  };
  let collapsed = false;
  try { collapsed = window.localStorage.getItem(key) === 'true'; } catch (err) {}
  apply(collapsed);
  toggle.addEventListener('click', () => {
    collapsed = !pageHeader.classList.contains('is-collapsed');
    apply(collapsed);
    try { window.localStorage.setItem(key, collapsed ? 'true' : 'false'); } catch (err) {}
  });
})();
"""


_CHART_VIRTUALIZATION_SCRIPT = """
(function() {
  const frames = Array.from(document.querySelectorAll('iframe.lazy-chart-frame[data-src]'));
  if (!frames.length) return;

  const LOAD_WINDOW = 50;
  const UNLOAD_DISTANCE = 100;
  let ticking = false;

  const setLoaded = (frame, loaded) => {
    const state = frame.previousElementSibling;
    if (loaded) {
      if (!frame.getAttribute('src')) {
        frame.setAttribute('src', frame.dataset.src);
      }
      frame.classList.remove('is-unloaded');
      if (state && state.classList.contains('chart-load-state')) {
        state.dataset.loaded = 'true';
        state.textContent = '';
      }
      return;
    }
    if (frame.getAttribute('src')) {
      frame.removeAttribute('src');
    }
    frame.classList.add('is-unloaded');
    if (state && state.classList.contains('chart-load-state')) {
      state.dataset.loaded = 'false';
      state.textContent = 'Chart unloaded until nearby';
    }
  };

  const nearestFrameIndex = () => {
    const viewportMid = window.scrollY + window.innerHeight / 2;
    let bestIndex = 0;
    let bestDistance = Number.POSITIVE_INFINITY;
    frames.forEach((frame, index) => {
      const rect = frame.getBoundingClientRect();
      const frameMid = window.scrollY + rect.top + rect.height / 2;
      const distance = Math.abs(frameMid - viewportMid);
      if (distance < bestDistance) {
        bestDistance = distance;
        bestIndex = index;
      }
    });
    return bestIndex;
  };

  const updateFrames = () => {
    ticking = false;
    const current = nearestFrameIndex();
    const half = Math.floor(LOAD_WINDOW / 2);
    let start = Math.max(0, current - half);
    let end = Math.min(frames.length - 1, start + LOAD_WINDOW - 1);
    start = Math.max(0, end - LOAD_WINDOW + 1);

    frames.forEach((frame, index) => {
      if (index >= start && index <= end) {
        setLoaded(frame, true);
      } else if (Math.abs(index - current) > UNLOAD_DISTANCE) {
        setLoaded(frame, false);
      }
    });
  };

  const scheduleUpdate = () => {
    if (ticking) return;
    ticking = true;
    window.requestAnimationFrame(updateFrames);
  };

  frames.forEach((frame) => {
    frame.classList.add('is-unloaded');
    const state = frame.previousElementSibling;
    if (state && state.classList.contains('chart-load-state')) {
      state.dataset.loaded = 'false';
      state.textContent = 'Chart queued';
    }
  });

  window.addEventListener('scroll', scheduleUpdate, { passive: true });
  window.addEventListener('resize', scheduleUpdate);
  window.addEventListener('hashchange', () => window.setTimeout(scheduleUpdate, 80));
  window.setTimeout(updateFrames, 0);
})();
"""


def _as_sort_int(value: object, default: int = 9999) -> int:
    try:
        return int(str(value).strip())
    except (TypeError, ValueError):
        return default


def _build_page_html(
    *,
    title: str,
    top_links: list[tuple[str, str]],
    header_nav_links: list[dict[str, object]],
    child_links: list[dict[str, object]],
    current_file: str,
    page_measure: str,
    chart_entries: list[dict[str, str]],
    empty_notice: str,
    fallback_note: str = "",
) -> str:
    title_href = current_file or "#page-header"
    separator_after = {"Others", "Other transformation"}

    def _header_page_links_html() -> str:
        chips: list[str] = []
        for label, href in top_links:
            current = "true" if href == current_file else "false"
            chips.append(f'<a href="{href}" class="header-chip" data-current="{current}">{label}</a>')
            if label in separator_after:
                chips.append('<span class="header-nav-separator" aria-hidden="true">|</span>')
        return "".join(chips)

    def _jump_nav_html(links: list[dict[str, object]]) -> str:
        if not links:
            return ""
        grouped_links: dict[int, list[dict[str, object]]] = defaultdict(list)
        for link in links:
            level = max(1, min(4, int(link.get("level", 1) or 1)))
            grouped_links[level].append(link)

        def _jump_chip_html(entry: dict[str, object], level: int) -> str:
            label = str(entry.get("label", ""))
            chip = (
                f'<a href="{entry.get("href", "#")}" class="jump-chip" data-level="{level}" '
                f'data-current="{"true" if str(entry.get("href", "")) == current_file else "false"}">{label}</a>'
            )
            if level == 1 and label in separator_after:
                chip += '<span class="jump-nav-separator" aria-hidden="true">|</span>'
            return chip

        group_html = "".join(
            '<div class="jump-nav-row" data-level="{level}">{chips}</div>'.format(
                level=level,
                chips="".join(_jump_chip_html(entry, level) for entry in grouped_links[level]),
            )
            for level in sorted(grouped_links)
        )
        return (
            '<div class="jump-nav">'
            '<span class="jump-nav-label">Sections:</span>'
            f'<div class="jump-nav-groups">{group_html}</div>'
            + '</div>'
        )

    children_html = "".join(
        [
            (
                f'<li data-level="{int(link.get("level", 1) or 1)}">'
                f'<a href="{link.get("href", "#")}">{link.get("label", "")}</a> '
                f'<span style="color:#4b5563;">({int(link.get("count", 0) or 0)} charts)</span>'
                "</li>"
            )
            for link in child_links
        ]
    )
    children_section = ""
    if child_links and not chart_entries:
        children_section = (
            '<section>'
            '<h2 style="margin:0 0 6px 0;font-size:var(--section-title-size);color:#23384d;">Sections</h2>'
            f'<ul class="section-link-list">{children_html}</ul>'
            '</section>'
        )
    section_jump_row = _jump_nav_html(header_nav_links)
    page_measure_html = ""
    if page_measure and page_measure not in {"Energy balance (PJ)", "Overview"}:
        page_measure_html = (
            f'<div style="margin:6px 0 0 0;color:#4b5563;font-size:12px;line-height:1.3;">'
            f'Measure: {page_measure}</div>'
        )
    note_html = ""
    if fallback_note:
        note_html = f'<section class="section-note">{fallback_note}</section>'

    if chart_entries:
        grouped_chart_entries: dict[str, list[dict[str, str]]] = defaultdict(list)
        section_labels: dict[str, str] = {}
        for entry in chart_entries:
            section_id = str(entry.get("section_id", "") or "sec-charts")
            section_label = str(entry.get("section_label", "") or "Charts")
            section_labels.setdefault(section_id, section_label)
            grouped_chart_entries[section_id].append(entry)
        section_blocks: list[str] = []

        def _grid_expand_class(item_count: int) -> str:
            if item_count == 1:
                return " expand-1"
            if item_count == 2:
                return " expand-2"
            if item_count == 3:
                return " expand-3"
            return ""

        def _entry_sort_key(entry: dict[str, str]) -> tuple[object, ...]:
            fuel = str(entry.get("fuel", "")).strip()
            measure = str(entry.get("measure", "")).strip()
            template_order = _as_sort_int(
                entry.get("template_order", ""),
                0 if str(entry.get("entry_kind", "")).strip() == "aggregate" else 1,
            )
            measure_rank = {
                TRANSFORMATION_INPUT_MEASURE: 0,
                TRANSFORMATION_OUTPUT_MEASURE: 1,
            }.get(measure, 2)
            return (
                template_order,
                measure_rank,
                measure.lower(),
                0 if fuel == "Total" else 1,
                str(entry.get("sheet", "")).lower(),
                fuel.lower(),
            )

        def _entry_caption(entry: dict[str, str], entries: list[dict[str, str]]) -> str:
            fuel = str(entry.get("fuel", "")).strip()
            if str(entry.get("entry_kind", "")).strip() == "aggregate":
                return fuel or "Total"
            sheets = {str(item.get("display_sheet", item.get("sheet", ""))).strip() for item in entries}
            if len(sheets) > 1:
                return f'{entry.get("display_sheet", entry.get("sheet", ""))}: {fuel}'
            return fuel

        for section_id, entries in grouped_chart_entries.items():
            measure_sections: list[str] = []
            if entries and all(str(entry.get("entry_kind", "")).strip() == "aggregate" for entry in entries):
                measure_entries = sorted(entries, key=_entry_sort_key)
                cards = "".join(
                    [
                        (
                            '<figure class="chart-card">'
                            f'<figcaption class="chart-caption">{entry.get("measure", "") or _entry_caption(entry, measure_entries)}</figcaption>'
                            f'<div class="meta-subline">{entry.get("path_label", "")}</div>'
                            f'<div class="chart-load-state" data-loaded="false">Chart queued</div>'
                            f'<iframe data-src="../charts/{entry["file"]}" class="lazy-chart-frame is-unloaded" loading="lazy" title="{entry.get("measure", "") or _entry_caption(entry, measure_entries)}"></iframe>'
                            "</figure>"
                        )
                        for entry in measure_entries
                    ]
                )
                measure_sections.append(
                    f'<section class="measure-group" style="margin:6px 0 14px 0;">'
                    + f'<div class="dashboard-grid{_grid_expand_class(len(measure_entries))}">{cards}</div>'
                    + "</section>"
                )
            else:
                measure_groups: dict[str, list[dict[str, str]]] = defaultdict(list)
                for entry in entries:
                    measure_key = str(entry.get("measure", "")).strip() or "Summary (PJ)"
                    measure_groups[measure_key].append(entry)
                for measure_key in sorted(
                    measure_groups,
                    key=lambda value: {
                        TRANSFORMATION_INPUT_MEASURE: (0, value.lower()),
                        TRANSFORMATION_OUTPUT_MEASURE: (1, value.lower()),
                    }.get(value, (2, value.lower())),
                ):
                    measure_entries = sorted(measure_groups[measure_key], key=_entry_sort_key)
                    cards = "".join(
                        [
                            (
                                '<figure class="chart-card">'
                                f'<figcaption class="chart-caption">{_entry_caption(entry, measure_entries)}</figcaption>'
                                f'<div class="meta-subline">{entry.get("path_label", "")}</div>'
                                f'<div class="chart-load-state" data-loaded="false">Chart queued</div>'
                                f'<iframe data-src="../charts/{entry["file"]}" class="lazy-chart-frame is-unloaded" loading="lazy" title="{_entry_caption(entry, measure_entries)}"></iframe>'
                                "</figure>"
                            )
                            for entry in measure_entries
                        ]
                    )
                    measure_sections.append(
                        f'<section class="measure-group" style="margin:6px 0 14px 0;">'
                        + (
                            ""
                            if measure_key == "Energy balance (PJ)"
                            else f'<h3 style="margin:0 0 8px 0;font-size:14px;font-weight:600;color:#5b6470;">{measure_key}</h3>'
                        )
                        + f'<div class="dashboard-grid{_grid_expand_class(len(measure_entries))}">{cards}</div>'
                        + "</section>"
                    )
            heading = "" if len(grouped_chart_entries) == 1 and section_labels[section_id] == "Charts" else (
                f'<h2 style="margin:18px 0 8px 0;font-size:var(--section-title-size);color:#23384d;">'
                f'{section_labels[section_id]}</h2>'
            )
            section_blocks.append(
                f'<section id="{section_id}" style="scroll-margin-top:150px;">{heading}{"".join(measure_sections)}</section>'
            )
        chart_html = "".join(section_blocks)
    elif child_links:
        chart_html = ""
    else:
        chart_html = f'<section class="section-note">{empty_notice}</section>'

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{title}</title>
  <style>{_V2_STYLE_CSS}</style>
</head>
<body>
  <div class="page-shell">
    <header class="page-header" id="page-header">
      <div class="header-collapsible">
      <div class="header-main-row">
        <div style="min-width:220px;flex:1 1 320px;">
          <h1 style="margin:0;font-size:24px;line-height:1.15;"><a href="{title_href}" style="color:inherit;text-decoration:none;">{title}</a></h1>
          {page_measure_html}
        </div>
        <div class="header-side-controls">
          <div class="header-inline-controls">
            {_header_page_links_html()}
          </div>
        </div>
      </div>
      {section_jump_row}
      </div>
      <div class="header-toggle-row">
        <button id="header-toggle" class="header-toggle" type="button" aria-expanded="true" aria-label="Collapse header">▴</button>
      </div>
    </header>
    <main class="page-body">
      {children_section}
      {note_html}
      <section>{chart_html}</section>
    </main>
  </div>
  <script>{_HEADER_TOGGLE_SCRIPT}</script>
  <script>{_CHART_VIRTUALIZATION_SCRIPT}</script>
</body>
</html>
"""


def _build_about_page_html(
    *,
    title: str,
    top_links: list[tuple[str, str]],
    current_file: str,
    about_config: dict[str, Any],
) -> str:
    heading = _clean_token(about_config.get("title", "")) or "About This Dashboard"
    intro = _clean_token(about_config.get("intro", ""))
    sections = about_config.get("sections", [])
    if not isinstance(sections, list):
        sections = []

    def _header_page_links_html() -> str:
        chips: list[str] = []
        separator_after = {"Others", "Other transformation"}
        for label, href in top_links:
            current = "true" if href == current_file else "false"
            chips.append(f'<a href="{href}" class="header-chip" data-current="{current}">{label}</a>')
            if label in separator_after:
                chips.append('<span class="header-nav-separator" aria-hidden="true">|</span>')
        return "".join(chips)

    def _paragraph_html(text: object) -> str:
        content = _clean_token(text)
        if not content:
            return ""
        return f"<p>{content}</p>"

    section_html: list[str] = []
    for section in sections:
        if not isinstance(section, dict):
            continue
        section_title = _clean_token(section.get("title", ""))
        body = section.get("body", [])
        if isinstance(body, str):
            body_items = [body]
        elif isinstance(body, list):
            body_items = body
        else:
            body_items = []
        bullets = section.get("bullets", [])
        bullet_items = bullets if isinstance(bullets, list) else []
        paragraphs = "".join(_paragraph_html(item) for item in body_items)
        bullet_html = ""
        if bullet_items:
            bullet_html = "<ul>" + "".join(
                f"<li>{_clean_token(item)}</li>" for item in bullet_items if _clean_token(item)
            ) + "</ul>"
        if section_title or paragraphs or bullet_html:
            section_html.append(
                '<section class="about-section">'
                + (f"<h2>{section_title}</h2>" if section_title else "")
                + paragraphs
                + bullet_html
                + "</section>"
            )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width, initial-scale=1" />
  <title>{title}</title>
  <style>{_V2_STYLE_CSS}
.about-content {{
  max-width: 980px;
  background: #fff;
  border: 1px solid #d6dde5;
  border-radius: 8px;
  padding: 20px;
  box-sizing: border-box;
}}
.about-content p {{ margin: 0 0 12px 0; }}
.about-content ul {{ margin: 0 0 12px 22px; padding: 0; }}
.about-content li {{ margin: 5px 0; }}
.about-section {{ margin-top: 18px; }}
.about-section h2 {{ margin: 0 0 8px 0; font-size: var(--section-title-size); color: #23384d; }}
  </style>
</head>
<body>
  <div class="page-shell">
    <header class="page-header" id="page-header">
      <div class="header-collapsible">
        <div class="header-main-row">
          <div style="min-width:220px;flex:1 1 320px;">
            <h1 style="margin:0;font-size:24px;line-height:1.15;"><a href="{current_file}" style="color:inherit;text-decoration:none;">{heading}</a></h1>
          </div>
          <div class="header-side-controls">
            <div class="header-inline-controls">{_header_page_links_html()}</div>
          </div>
        </div>
      </div>
      <div class="header-toggle-row">
        <button id="header-toggle" class="header-toggle" type="button" aria-expanded="true" aria-label="Collapse header">â–´</button>
      </div>
    </header>
    <main class="page-body">
      <article class="about-content">
        {_paragraph_html(intro)}
        {"".join(section_html)}
      </article>
    </main>
  </div>
  <script>{_HEADER_TOGGLE_SCRIPT}</script>
</body>
</html>
"""


def render_balance_dashboards(
    *,
    comparison_long: pd.DataFrame,
    mapping_status: pd.DataFrame,
    structure_config: dict[str, Any],
    output_dir: Path | str,
    chart_backend: str = "plotly",
    hide_leap_only_charts: bool = False,
    chart_navigation_guide_path: Path | str | None = None,
) -> dict[str, Any]:
    """
    Render charts with existing utilities and dashboards from structure.json.
    """
    out_dir = _resolve(output_dir)
    charts_dir = out_dir / "charts"
    dashboards_dir = out_dir / "dashboards"
    charts_dir.mkdir(parents=True, exist_ok=True)
    dashboards_dir.mkdir(parents=True, exist_ok=True)
    for stale_chart in charts_dir.glob("*.html"):
        stale_chart.unlink()
    for stale_chart in charts_dir.glob("*.png"):
        stale_chart.unlink()

    template_allowlist = _load_dashboard_template_allowlist(chart_navigation_guide_path)
    page_tree = list(structure_config.get("page_tree", []) or [])
    sheet_catalog = dict(structure_config.get("sheet_catalog", {}) or {})
    esto_flow_to_sheet = {
        str(k).strip(): str(v).strip()
        for k, v in (structure_config.get("esto_flow_to_sheet", {}) or {}).items()
        if str(k).strip() and str(v).strip()
    }

    def _template_flow_paths() -> dict[str, tuple[str, ...]]:
        if not template_allowlist:
            return {}
        top_label_inverse = {v: k for k, v in BALANCE_DASHBOARD_TOP_LABELS.items()}
        flow_paths: dict[str, tuple[str, ...]] = {}
        measure_default = str((template_allowlist.get("defaults") or {}).get("measure", "")).strip()

        def _raw_path(display_path: tuple[str, ...]) -> tuple[str, ...]:
            if not display_path:
                return display_path
            return (top_label_inverse.get(display_path[0], display_path[0]), *display_path[1:])

        def _remember_flow(flow: object, display_path: tuple[str, ...]) -> None:
            flow_text = str(flow or "").strip()
            if flow_text and display_path:
                candidate_path = _raw_path(display_path)
                existing_path = flow_paths.get(flow_text)
                # Parent aggregates are walked before child graph nodes. Prefer the
                # deepest declared template path so direct leaf charts stay attached
                # to their intended subsector pages instead of collapsing upward.
                if existing_path is None or len(candidate_path) > len(existing_path):
                    flow_paths[flow_text] = candidate_path

        def _walk(node: dict[str, Any], display_path: tuple[str, ...]) -> None:
            for aggregate in _dashboard_template_aggregate_specs(node, default_measure=measure_default):
                for flow in list(aggregate.get("source_flows", []) or []):
                    _remember_flow(flow, display_path)
            for spec in _dashboard_template_graph_specs(node, default_measure=measure_default):
                for flow in list(spec.get("esto_flows", []) or [spec.get("esto_flow", "")]):
                    _remember_flow(flow, display_path)
            for key, child in node.items():
                if _dashboard_template_is_reserved_key(key) or not isinstance(child, dict):
                    continue
                label = str(key).strip()
                if label:
                    _walk(child, (*display_path, label))

        _walk(template_allowlist, ())
        return flow_paths

    template_flow_paths = _template_flow_paths()
    if template_flow_paths and mapping_status is not None and not mapping_status.empty:
        status_for_paths = mapping_status.copy()
        for col in ["sheet", "esto_flow"]:
            if col not in status_for_paths.columns:
                status_for_paths[col] = ""
            status_for_paths[col] = status_for_paths[col].fillna("").astype(str).str.strip()
        for row in status_for_paths[["sheet", "esto_flow"]].drop_duplicates().itertuples(index=False):
            sheet = str(row.sheet).strip()
            flow = str(row.esto_flow).strip()
            path = template_flow_paths.get(flow)
            if not sheet or not path:
                continue
            esto_flow_to_sheet[flow] = sheet
            existing_cfg = dict(sheet_catalog.get(sheet, {}) or {})
            existing_cfg["display_label"] = path[-1]
            existing_cfg["path"] = list(path)
            if not str(existing_cfg.get("measure", "")).strip():
                existing_cfg["measure"] = "Energy balance (PJ)"
            if "sort_order" not in existing_cfg:
                existing_cfg["sort_order"] = len(sheet_catalog)
            existing_cfg.update(
                _dashboard_hierarchy_from_path(
                    path,
                    entry_kind="direct",
                    measure=existing_cfg.get("measure", ""),
                    fallback_label=sheet,
                )
            )
            sheet_catalog[sheet] = existing_cfg
        structure_config = {
            **structure_config,
            "sheet_catalog": sheet_catalog,
            "esto_flow_to_sheet": esto_flow_to_sheet,
        }

    empty_notice = _clean_token(structure_config.get("empty_page_notice", "")) or DEFAULT_EMPTY_PAGE_NOTICE
    chart_comparison_long = _split_transformation_input_output_measures(comparison_long, sheet_catalog)
    chart_comparison_long = _backfill_dashboard_hierarchy(chart_comparison_long, sheet_catalog=sheet_catalog)
    mapping_status = _backfill_dashboard_hierarchy(mapping_status, sheet_catalog=sheet_catalog)

    written_charts = build_charts(
        chart_comparison_long,
        charts_dir=charts_dir,
        backend=chart_backend,
        hide_leap_only_charts=hide_leap_only_charts,
    )

    flat_nodes = _flatten_page_tree(page_tree)
    node_paths: list[tuple[str, ...]] = []
    node_path_set: set[tuple[str, ...]] = set()

    def _append_path_with_prefixes(path: Sequence[str]) -> None:
        clean_path = tuple(str(part).strip() for part in path if str(part).strip())
        for idx in range(1, len(clean_path) + 1):
            prefix = clean_path[:idx]
            if prefix in node_path_set:
                continue
            node_path_set.add(prefix)
            node_paths.append(prefix)

    for node in flat_nodes:
        _append_path_with_prefixes(node["path"])

    chart_lookup = _chart_file_lookup(chart_comparison_long, hide_leap_only_charts=hide_leap_only_charts)
    node_entries, _ = _collect_node_chart_entries(
        comparison_long=chart_comparison_long,
        sheet_catalog=sheet_catalog,
        chart_lookup=chart_lookup,
    )

    # Include any missing sheet paths from catalog as pages in case tree omitted them.
    for _, cfg in sheet_catalog.items():
        path = tuple(str(v).strip() for v in list((cfg or {}).get("path", []) or []) if str(v).strip())
        if path:
            _append_path_with_prefixes(path)

    render_long = _prepare_render_long(chart_comparison_long)
    if "measure" not in render_long.columns:
        render_long["measure"] = ""
    render_long["measure"] = render_long["measure"].fillna("").astype(str)
    render_long["sheet"] = render_long["sheet"].fillna("").astype(str)
    render_long["fuel_label"] = render_long["fuel_label"].fillna("").astype(str)
    render_long["source"] = render_long["source"].fillna("").astype(str)
    render_long["value"] = pd.to_numeric(render_long["value"], errors="coerce")

    sheet_paths: dict[str, tuple[str, ...]] = {}
    for sheet, cfg in sheet_catalog.items():
        path = tuple(str(v).strip() for v in list((cfg or {}).get("path", []) or []) if str(v).strip())
        if path:
            sheet_paths[str(sheet)] = path

    def _display_path(path: tuple[str, ...]) -> list[str]:
        return [
            _display_balance_path_part(part, is_top=idx == 0)
            for idx, part in enumerate(path)
        ]

    def _path_label(path: tuple[str, ...]) -> str:
        return " > ".join(_display_path(path))

    def _path_title(path: tuple[str, ...]) -> str:
        if not path:
            return ""
        return _display_balance_path_part(path[-1], is_top=len(path) == 1)

    def _is_aggregate_entry(entry: dict[str, str]) -> bool:
        return (
            str(entry.get("entry_kind", "")).strip() == "aggregate"
            or str(entry.get("file", "")).startswith("aggregate__")
        )

    def _as_sort_int(value: object, default: int = 9999) -> int:
        try:
            return int(str(value).strip())
        except (TypeError, ValueError):
            return default

    mapping_lookup: dict[tuple[str, str, str], dict[str, Any]] = {}
    if mapping_status is not None and not mapping_status.empty:
        status = mapping_status.copy()
        for col in ["sheet", "measure", "fuel_label", "esto_flow", "esto_product"]:
            if col not in status.columns:
                status[col] = ""
            status[col] = status[col].fillna("").astype(str).str.strip()
        for (sheet, measure, fuel), group in status.groupby(["sheet", "measure", "fuel_label"], dropna=False):
            pairs = (
                group[["esto_flow", "esto_product"]]
                .drop_duplicates()
                .sort_values(["esto_flow", "esto_product"], kind="mergesort")
                .to_dict("records")
            )
            mapping_lookup[(str(sheet), str(measure), str(fuel))] = {
                "esto_pairs": pairs,
                "esto_flows": sorted({str(row["esto_flow"]) for row in pairs if str(row["esto_flow"]).strip()}),
                "esto_products": sorted({str(row["esto_product"]) for row in pairs if str(row["esto_product"]).strip()}),
            }

    def _chart_mapping_info(entry: dict[str, str]) -> dict[str, Any]:
        exact_key = (
            str(entry.get("sheet", "")),
            str(entry.get("measure", "")),
            str(entry.get("fuel", "")),
        )
        if exact_key in mapping_lookup:
            return mapping_lookup[exact_key]
        default_measure_key = (
            str(entry.get("sheet", "")),
            "Energy balance (PJ)",
            str(entry.get("fuel", "")),
        )
        if default_measure_key in mapping_lookup:
            return mapping_lookup[default_measure_key]
        blank_measure_key = (
            str(entry.get("sheet", "")),
            "",
            str(entry.get("fuel", "")),
        )
        return mapping_lookup.get(blank_measure_key, {"esto_pairs": [], "esto_flows": [], "esto_products": []})

    allowed_template_paths: set[tuple[str, ...]] = set()
    allowed_template_nodes: dict[tuple[str, ...], dict[str, Any]] = {}
    allowed_template_graph_specs: dict[tuple[str, ...], list[dict[str, Any]]] = {}
    template_entry_order: dict[tuple[tuple[str, ...], str], int] = {}
    template_measure_default = str((template_allowlist.get("defaults") or {}).get("measure", "")).strip() or "Energy balance (PJ)"

    def _collect_template_allowlist(node: dict[str, Any], path: tuple[str, ...] = ()) -> None:
        if path:
            allowed_template_paths.add(path)
            allowed_template_nodes[path] = node
        order_idx = 0
        for key, child in node.items():
            if key in {"aggregate_graphs", "aggregate"} and isinstance(child, dict):
                template_entry_order.setdefault((path, "aggregate"), order_idx)
                order_idx += 1
            elif key in {"by_fuel_graphs", "graphs"} and isinstance(child, dict):
                template_entry_order.setdefault((path, "direct"), order_idx)
                order_idx += 1
            elif not _dashboard_template_is_reserved_key(key) and isinstance(child, dict):
                order_idx += 1
        graph_specs = _dashboard_template_graph_specs(node, default_measure=template_measure_default)
        if graph_specs:
            allowed_template_graph_specs[path] = graph_specs
        for key, child in node.items():
            if _dashboard_template_is_reserved_key(key) or not isinstance(child, dict):
                continue
            label = str(key).strip()
            if label:
                _collect_template_allowlist(child, (*path, label))

    if template_allowlist:
        _collect_template_allowlist(template_allowlist)

    def _display_path_tuple(path: tuple[str, ...]) -> tuple[str, ...]:
        return tuple(_display_path(path))

    def _path_allowed_by_template(path: tuple[str, ...]) -> bool:
        if not template_allowlist:
            return True
        display_path = _display_path_tuple(path)
        return display_path in allowed_template_paths

    def _template_node_has_aggregate(path: tuple[str, ...]) -> bool:
        if not template_allowlist:
            return False
        node = allowed_template_nodes.get(_display_path_tuple(path), {})
        return isinstance(node, dict) and bool(
            _dashboard_template_aggregate_specs(node, default_measure=template_measure_default)
        )

    def _template_order_for_entry(path: tuple[str, ...], entry: dict[str, str]) -> int:
        kind = "aggregate" if _is_aggregate_entry(entry) else "direct"
        fallback = 0 if kind == "aggregate" else 1
        if not template_allowlist:
            return fallback
        return template_entry_order.get((_display_path_tuple(path), kind), fallback)

    def _template_aggregate_specs_for_path(path: tuple[str, ...]) -> list[dict[str, Any]]:
        if not template_allowlist:
            return []
        node = allowed_template_nodes.get(_display_path_tuple(path), {})
        if not isinstance(node, dict):
            return []
        return _dashboard_template_aggregate_specs(node, default_measure=template_measure_default)

    def _component_sheets_for_aggregate_path(path: tuple[str, ...]) -> list[str]:
        aggregate_specs = _template_aggregate_specs_for_path(path)
        configured_flows = [
            str(flow).strip()
            for spec in aggregate_specs
            for flow in list(spec.get("esto_flows", spec.get("source_flows", [])) or [])
            if str(flow).strip()
        ]
        if not configured_flows:
            return _leaf_descendant_sheets(path)

        sheets: list[str] = []
        seen: set[str] = set()
        for flow in configured_flows:
            sheet = str(esto_flow_to_sheet.get(flow, "")).strip()
            if not sheet or sheet in seen:
                continue
            seen.add(sheet)
            sheets.append(sheet)
        return sorted(sheets, key=lambda sheet: tuple(part.lower() for part in sheet_paths.get(sheet, (sheet,))))

    def _entry_allowed_by_template(path: tuple[str, ...], entry: dict[str, str]) -> bool:
        if not template_allowlist:
            return True
        display_path = _display_path_tuple(path)
        node = allowed_template_nodes.get(display_path)
        if not isinstance(node, dict):
            return False
        measure = str(entry.get("measure", "")).strip()
        fuel = str(entry.get("fuel", "")).strip()
        if _is_aggregate_entry(entry):
            aggregate_specs = _dashboard_template_aggregate_specs(node, default_measure=template_measure_default)
            if not aggregate_specs:
                return False
            for aggregate in aggregate_specs:
                aggregate_fuel = str(aggregate.get("fuel", "Total")).strip() or "Total"
                if fuel and fuel != aggregate_fuel:
                    continue
                measures = {str(value).strip() for value in list(aggregate.get("measures", []) or []) if str(value).strip()}
                if not measures or not measure or measure in measures:
                    return True
            return False
        graph_specs = allowed_template_graph_specs.get(display_path, [])
        if not graph_specs:
            return False
        mapping_info = _chart_mapping_info(entry)
        pairs = list(mapping_info.get("esto_pairs", []) or [])
        if not pairs:
            return False
        for spec in graph_specs:
            spec_measure = str(spec.get("measure", "")).strip()
            spec_flow_norms = set(spec.get("esto_flows_norm", set()))
            if not spec_flow_norms:
                spec_flow_norms = {" ".join(str(spec.get("esto_flow", "")).strip().lower().split())}
            matching_pairs = [
                pair for pair in pairs
                if " ".join(str(pair.get("esto_flow", "")).strip().lower().split()) in spec_flow_norms
            ]
            if not matching_pairs:
                continue
            if spec_measure and measure and spec_measure != measure:
                continue
            if str(fuel).strip() == "Total":
                continue
            if spec.get("include_all_products", False):
                if any(
                    str(pair.get("esto_product", "")).strip()
                    and not _product_is_total(pair.get("esto_product", ""))
                    and " ".join(str(pair.get("esto_product", "")).strip().lower().split())
                    not in set(spec.get("exclude_products_norm", set()))
                    for pair in matching_pairs
                ):
                    return True
                continue
            products_norm = set(spec.get("products_norm", set()))
            exclude_norm = set(spec.get("exclude_products_norm", set()))
            if any(
                " ".join(str(pair.get("esto_product", "")).strip().lower().split()) in products_norm
                and " ".join(str(pair.get("esto_product", "")).strip().lower().split()) not in exclude_norm
                for pair in matching_pairs
            ):
                return True
        return False

    def _add_template_direct_chart_entries() -> None:
        if not template_allowlist:
            return
        existing_keys = {
            (
                path,
                str(entry.get("sheet", "")),
                str(entry.get("measure", "")),
                str(entry.get("fuel", "")),
                str(entry.get("file", "")),
                str(entry.get("entry_kind", "direct")),
            )
            for path, entries in node_entries.items()
            for entry in entries
        }
        top_label_inverse = {v: k for k, v in BALANCE_DASHBOARD_TOP_LABELS.items()}

        def _raw_path(display_path: tuple[str, ...]) -> tuple[str, ...]:
            if not display_path:
                return display_path
            return (top_label_inverse.get(display_path[0], display_path[0]), *display_path[1:])

        def _template_sheet_key(path: tuple[str, ...]) -> str:
            token = "__".join(_safe_token(str(part).replace("\\", "_")) for part in path if str(part).strip())
            return f"template__{token or 'root'}"

        template_sheet_paths = {
            _template_sheet_key(_raw_path(display_path)): _raw_path(display_path)
            for display_path in allowed_template_graph_specs
            if _raw_path(display_path)
        }

        for display_path, graph_specs in allowed_template_graph_specs.items():
            if not graph_specs:
                continue
            raw_path = _raw_path(display_path)
            if not raw_path:
                continue
            _append_path_with_prefixes(raw_path)
            for (sheet, measure, fuel_label), file_name in sorted(
                chart_lookup.items(),
                key=lambda item: (item[0][0], item[0][1], item[0][2]),
            ):
                sheet_text = str(sheet)
                if sheet_text.startswith("template__") and template_sheet_paths.get(sheet_text) != raw_path:
                    continue
                sheet_cfg = sheet_catalog.get(sheet_text, {}) or {}
                display_sheet = _clean_token(sheet_cfg.get("display_label", "")) or _path_title(raw_path)
                entry = {
                    "sheet": sheet_text,
                    "measure": str(measure),
                    "fuel": str(fuel_label),
                    "file": str(file_name),
                    "display_sheet": display_sheet,
                    "path_label": _path_label(raw_path),
                    "entry_kind": "direct",
                    "template_order": str(_template_order_for_entry(raw_path, {"entry_kind": "direct"})),
                    **_dashboard_hierarchy_from_path(
                        raw_path,
                        entry_kind="direct",
                        measure=measure,
                        fallback_label=sheet,
                    ),
                }
                if not _entry_allowed_by_template(raw_path, entry):
                    continue
                key = (
                    raw_path,
                    entry["sheet"],
                    entry["measure"],
                    entry["fuel"],
                    entry["file"],
                    entry["entry_kind"],
                )
                if key in existing_keys:
                    continue
                existing_keys.add(key)
                node_entries[raw_path].append(entry)

    def _apply_template_allowlist_to_paths_and_entries() -> None:
        if not template_allowlist:
            return
        configured_paths = {
            tuple(str(v).strip() for v in list((cfg or {}).get("path", []) or []) if str(v).strip())
            for cfg in sheet_catalog.values()
            if tuple(str(v).strip() for v in list((cfg or {}).get("path", []) or []) if str(v).strip())
        }
        allowed_entry_paths = {
            path
            for path, entries in node_entries.items()
            if _path_allowed_by_template(path)
            and any(_entry_allowed_by_template(path, entry) for entry in entries)
        }
        allowed_entry_paths.update(
            path
            for path in node_paths
            if _path_allowed_by_template(path) and _template_node_has_aggregate(path)
        )
        allowed_entry_paths.update(
            path
            for path in node_paths
            if _path_allowed_by_template(path) and path in configured_paths
        )
        allowed_with_prefixes: set[tuple[str, ...]] = set()
        for path in allowed_entry_paths:
            for idx in range(1, len(path) + 1):
                allowed_with_prefixes.add(path[:idx])
        filtered_paths = [
            path
            for path in node_paths
            if _path_allowed_by_template(path) and path in allowed_with_prefixes
        ]
        node_paths[:] = filtered_paths
        node_path_set.clear()
        node_path_set.update(filtered_paths)
        for path in list(node_entries):
            if not _path_allowed_by_template(path):
                node_entries.pop(path, None)
                continue
            kept = [entry for entry in node_entries.get(path, []) if _entry_allowed_by_template(path, entry)]
            if kept:
                node_entries[path] = kept
            else:
                node_entries.pop(path, None)

    _add_template_direct_chart_entries()
    _apply_template_allowlist_to_paths_and_entries()

    def _leaf_descendant_sheets(path: tuple[str, ...]) -> list[str]:
        descendants = [
            sheet
            for sheet, sheet_path in sheet_paths.items()
            if len(sheet_path) >= len(path) and sheet_path[: len(path)] == path
        ]
        leafs: list[str] = []
        for sheet in descendants:
            sheet_path = sheet_paths[sheet]
            has_child_sheet = any(
                other != sheet
                and len(other_path) > len(sheet_path)
                and other_path[: len(sheet_path)] == sheet_path
                for other, other_path in sheet_paths.items()
            )
            if not has_child_sheet:
                leafs.append(sheet)
        return sorted(set(leafs), key=lambda value: tuple(part.lower() for part in sheet_paths.get(value, (value,))))

    def _has_child_path(path: tuple[str, ...]) -> bool:
        return any(
            len(candidate) > len(path) and candidate[: len(path)] == path
            for candidate in node_path_set
        )

    def _has_required_aggregate_dataset_families(frame: pd.DataFrame) -> bool:
        if frame.empty or "source" not in frame.columns or "value" not in frame.columns:
            return False
        values = pd.to_numeric(frame["value"], errors="coerce")
        valid = frame[values.notna()].copy()
        if valid.empty:
            return False
        sources = set(valid["source"].fillna("").astype(str).str.strip())
        return bool(sources & {"base", "base_estimated", "base_mixed"}) and bool(
            sources & {"projection", "projection_estimated", "projection_mixed"}
        )

    def _has_nonzero_values(frame: pd.DataFrame) -> bool:
        if frame.empty or "value" not in frame.columns:
            return False
        values = pd.to_numeric(frame["value"], errors="coerce").dropna()
        return bool(not values.empty and values.abs().gt(1e-12).any())

    def _add_parent_aggregate_chart_entries() -> int:
        added = 0
        for path in sorted(node_paths, key=lambda item: (len(item), tuple(part.lower() for part in item))):
            if not path or not _has_child_path(path):
                if not _template_node_has_aggregate(path):
                    continue
            component_sheets = _component_sheets_for_aggregate_path(path)
            is_transformation_parent = path[0] in TRANSFORMATION_DASHBOARD_TOP_GROUPS
            if len(component_sheets) <= 1 and not is_transformation_parent and not _template_node_has_aggregate(path):
                continue
            subset = render_long[
                render_long["sheet"].isin(component_sheets)
                & render_long["fuel_label"].ne("Total")
            ].copy()
            if subset.empty:
                continue
            title = _path_title(path)
            path_slug = "__".join(_safe_token(part.replace("\\", "_")) for part in path)
            entries: list[dict[str, str]] = []
            for measure, measure_subset in subset.groupby("measure", dropna=False):
                measure_text = str(measure or "").strip() or "Energy balance (PJ)"
                total_rows = _aggregate_display_rows_to_total(
                    measure_subset,
                    title=title,
                    measure_value=measure_text,
                    collapse_base_family=True,
                    collapse_projection_family=True,
                )
                if total_rows.empty or not _has_required_aggregate_dataset_families(total_rows) or not _has_nonzero_values(total_rows):
                    continue
                chart_path = make_chart(
                    title,
                    "Total",
                    total_rows,
                    charts_dir,
                    backend=chart_backend,
                    display_sheet=f"{title} - {measure_text}",
                    file_sheet=f"aggregate__{path_slug}__{measure_text}",
                )
                if chart_path is None:
                    continue
                entries.append(
                    {
                        "sheet": title,
                        "measure": measure_text,
                        "fuel": "Total",
                        "file": chart_path.name,
                        "path_label": _path_label(path),
                        "entry_kind": "aggregate",
                        "template_order": str(_template_order_for_entry(path, {"entry_kind": "aggregate"})),
                        **_dashboard_hierarchy_from_path(
                            path,
                            entry_kind="aggregate",
                            measure=measure_text,
                            fallback_label=title,
                        ),
                    }
                )
            if entries:
                node_entries[path] = [*entries, *node_entries.get(path, [])]
                added += len(entries)
        return added

    aggregate_charts_written = _add_parent_aggregate_chart_entries()
    _apply_template_allowlist_to_paths_and_entries()
    for node_path, entries in node_entries.items():
        for entry in entries:
            entry["template_order"] = str(_template_order_for_entry(node_path, entry))
    all_node_entries_for_chart_groups = {
        path: [dict(entry) for entry in entries]
        for path, entries in node_entries.items()
    }

    def _filter_entries_for_display_policy() -> None:
        """
        Keep parent pages focused on one aggregate chart and leaf pages focused on fuels.

        Direct ``fuel == Total`` charts are useful as generated artifacts, but
        they make the dashboard ambiguous once parent aggregate totals exist.
        The chart navigation JSON mirrors this policy and is copied to config
        as the render guide for future review.
        """
        for node_path in list(node_entries):
            entries = list(node_entries.get(node_path, []))
            if _has_child_path(node_path):
                kept = [entry for entry in entries if _is_aggregate_entry(entry)]
            else:
                kept = [
                    entry
                    for entry in entries
                    if (
                        _is_aggregate_entry(entry)
                        and _template_node_has_aggregate(node_path)
                    )
                    or (
                        not _is_aggregate_entry(entry)
                        and str(entry.get("fuel", "")).strip() != "Total"
                    )
                ]
            if kept:
                node_entries[node_path] = kept
            else:
                node_entries.pop(node_path, None)

    _filter_entries_for_display_policy()

    configured_top_order = [
        _clean_token(node.get("label", ""))
        for node in page_tree
        if isinstance(node, dict) and _clean_token(node.get("label", ""))
    ]
    top_level_set = {path[0] for path in node_paths if path}
    top_level = [label for label in configured_top_order if label in top_level_set]
    top_level += sorted([label for label in top_level_set if label not in set(top_level)], key=lambda s: s.lower())
    top_page_paths = [(label,) for label in top_level if label in BALANCE_DASHBOARD_MAJOR_SECTOR_PAGES]
    if not top_page_paths:
        top_page_paths = [(label,) for label in top_level]
    top_links = [(_display_balance_path_part(label, is_top=True), _page_filename_from_path((label,))) for label, in top_page_paths]
    about_page_config = _dashboard_template_about_page(template_allowlist)
    if about_page_config:
        top_links = [("About", "about.html"), *top_links]

    path_to_filename = {path: _page_filename_from_path(path) for path in node_paths if path}

    def _chart_count_for_path(target_path: tuple[str, ...]) -> int:
        return sum(
            len(entries)
            for node_path, entries in node_entries.items()
            if node_path[: len(target_path)] == target_path
        )

    sheet_to_flow = {
        str(sheet).strip(): str(flow).strip()
        for flow, sheet in (structure_config.get("esto_flow_to_sheet", {}) or {}).items()
        if str(flow).strip() and str(sheet).strip()
    }

    def _compact_dashboard_template() -> dict[str, Any]:
        """
        Write the config-side dashboard template with only authoring inputs.

        Generated details such as chart file names, default measures, and
        duplicated esto_pairs are intentionally left out because the renderer
        can derive them from path, fuel, and literal ESTO flow/product keys.
        """
        default_measure = "Energy balance (PJ)"
        template: dict[str, Any] = {
            "defaults": {
                "measure": default_measure,
            }
        }

        def _node_for_display_path(display_path: list[str]) -> dict[str, Any]:
            node = template
            for part in display_path:
                node = node.setdefault(part, {})
            return node

        for node_path in node_paths:
            entries = node_entries.get(node_path, [])
            if not entries:
                continue
            display_path = [
                _display_balance_path_part(part, is_top=idx == 0)
                for idx, part in enumerate(node_path)
            ]
            node = _node_for_display_path(display_path)
            aggregate_entries = [entry for entry in entries if _is_aggregate_entry(entry)]
            if aggregate_entries:
                source_sheets = _component_sheets_for_aggregate_path(node_path)
                source_flows = sorted(
                    {
                        sheet_to_flow.get(str(sheet), "")
                        for sheet in source_sheets
                        if sheet_to_flow.get(str(sheet), "")
                    }
                )
                aggregate: dict[str, Any] = {
                    "fuels": "Total",
                    "esto_flows": source_flows,
                }
                measures = sorted(
                    {
                        str(entry.get("measure", "")).strip()
                        for entry in aggregate_entries
                        if str(entry.get("measure", "")).strip()
                        and str(entry.get("measure", "")).strip() != default_measure
                    }
                )
                if measures:
                    aggregate["measures"] = measures
                node["aggregate_graphs"] = aggregate
                continue

            graph_specs: list[dict[str, Any]] = []
            for entry in entries:
                mapping_info = _chart_mapping_info(entry)
                pairs = list(mapping_info.get("esto_pairs", []) or [])
                flows = sorted({str(pair.get("esto_flow", "")).strip() for pair in pairs if str(pair.get("esto_flow", "")).strip()})
                products = sorted(
                    {
                        str(pair.get("esto_product", "")).strip()
                        for pair in pairs
                        if str(pair.get("esto_product", "")).strip() and not _product_is_total(pair.get("esto_product", ""))
                    }
                )
                if not flows and not products:
                    continue
                measure = str(entry.get("measure", "")).strip()
                spec: dict[str, Any] = {}
                if flows:
                    spec["esto_flows"] = flows
                if measure and measure != default_measure:
                    spec["measure"] = measure
                spec["products"] = products or "all"
                graph_specs.append(spec)
            if graph_specs:
                unique_flows = sorted(
                    {
                        flow
                        for spec in graph_specs
                        for flow in list(spec.get("esto_flows", []) or [])
                        if str(flow).strip()
                    }
                )
                unique_products = sorted(
                    {
                        product
                        for spec in graph_specs
                        for product in list(spec.get("products", []) or [])
                        if str(product).strip() and str(product).strip().lower() != "all"
                    }
                )
                by_fuel: dict[str, Any] = {"products": unique_products or "All"}
                if unique_flows:
                    by_fuel["esto_flows"] = unique_flows
                measures = sorted(
                    {
                        str(spec.get("measure", "")).strip()
                        for spec in graph_specs
                        if str(spec.get("measure", "")).strip()
                        and str(spec.get("measure", "")).strip() != default_measure
                    }
                )
                if measures:
                    by_fuel["measures"] = measures
                node["by_fuel_graphs"] = by_fuel

        return template

    def _descendant_links_for_path(path: tuple[str, ...]) -> list[dict[str, object]]:
        descendant_paths = [p for p in node_paths if len(p) > len(path) and p[: len(path)] == path]
        leaf_label_counts: dict[str, int] = defaultdict(int)
        for descendant in descendant_paths:
            leaf_label_counts[descendant[-1]] += 1
        links: list[dict[str, object]] = []
        for child in descendant_paths:
            relative = child[len(path):]
            label = child[-1]
            if leaf_label_counts.get(label, 0) > 1 and len(relative) > 1:
                label = " > ".join(relative)
            links.append(
                {
                    "label": label,
                    "href": path_to_filename.get(child, "index.html"),
                    "count": _chart_count_for_path(child),
                    "level": min(4, max(1, len(relative))),
                }
            )
        return links

    def _section_links_for_top_path(path: tuple[str, ...]) -> list[dict[str, object]]:
        section_paths = [
            p
            for p in node_paths
            if len(path) < len(p) <= len(path) + 2 and p[: len(path)] == path
        ]
        if not section_paths and _chart_count_for_path(path):
            section_paths = [path]
        links: list[dict[str, object]] = []
        for section_path in section_paths:
            label = _display_balance_path_part(section_path[-1], is_top=False)
            links.append(
                {
                    "label": label,
                    "href": f"#{_section_id_from_path(section_path)}",
                    "count": _chart_count_for_path(section_path),
                    "level": max(1, min(4, len(section_path) - len(path))),
                }
            )
        return links

    def _top_level_links() -> list[dict[str, object]]:
        return [
            {
                "label": _display_balance_path_part(path[0], is_top=True),
                "href": href,
                "count": _chart_count_for_path(path),
                "level": 1,
            }
            for path, (_, href) in zip(top_page_paths, top_links)
        ]

    def _header_nav_links_for_path(path: tuple[str, ...] | None) -> list[dict[str, object]]:
        if not path:
            return []
        return _section_links_for_top_path((path[0],))

    def _chart_entries_for_path(path: tuple[str, ...]) -> list[dict[str, str]]:
        entries: list[dict[str, str]] = []
        for node_path in node_paths:
            if node_path[: len(path)] != path:
                continue
            if len(path) != 1 and node_path != path:
                continue
            section_depth = min(len(node_path), len(path) + 2)
            section_path = node_path[:section_depth] if len(node_path) >= section_depth else path
            section_label = _display_balance_path_part(section_path[-1], is_top=False)
            section_id = _section_id_from_path(section_path)
            for entry in node_entries.get(node_path, []):
                tagged = dict(entry)
                tagged["section_label"] = section_label
                tagged["section_id"] = section_id
                entries.append(tagged)
        return entries

    def _write_chart_navigation_hierarchy() -> tuple[Path, Path, Path]:
        hierarchy: dict[str, Any] = {}
        rows: list[dict[str, str]] = []

        for node_path in node_paths:
            entries = node_entries.get(node_path, [])
            if not entries:
                continue
            display_path = [
                _display_balance_path_part(part, is_top=idx == 0)
                for idx, part in enumerate(node_path)
            ]
            node = hierarchy
            for part in display_path:
                node = node.setdefault(part, {})
            fuels = node.setdefault("fuels", {})
            for entry in entries:
                mapping_info = _chart_mapping_info(entry)
                chart_record = {
                    "sheet": str(entry.get("sheet", "")),
                    "page_key": str(entry.get("page_key", "")),
                    "page_label": str(entry.get("page_label", "")),
                    "chart_group_key": str(entry.get("chart_group_key", "")),
                    "chart_group_label": str(entry.get("chart_group_label", "")),
                    "measure": str(entry.get("measure", "")),
                    "template_order": str(entry.get("template_order", "")),
                    "chart_file": f"charts/{entry.get('file', '')}",
                    "esto_flows": mapping_info.get("esto_flows", []),
                    "esto_products": mapping_info.get("esto_products", []),
                    "esto_pairs": mapping_info.get("esto_pairs", []),
                }
                fuel = str(entry.get("fuel", ""))
                fuels.setdefault(fuel, []).append(chart_record)
                pairs = list(mapping_info.get("esto_pairs", []) or [])
                if not pairs:
                    pairs = [{"esto_flow": "", "esto_product": ""}]
                for pair in pairs:
                    rows.append(
                        {
                            "dashboard_path": " > ".join(display_path),
                            "sheet": str(entry.get("sheet", "")),
                            "page_key": str(entry.get("page_key", "")),
                            "page_label": str(entry.get("page_label", "")),
                            "chart_group_key": str(entry.get("chart_group_key", "")),
                            "chart_group_label": str(entry.get("chart_group_label", "")),
                            "measure": str(entry.get("measure", "")),
                            "template_order": str(entry.get("template_order", "")),
                            "fuel": fuel,
                            "chart_file": f"charts/{entry.get('file', '')}",
                            "esto_flow": str(pair.get("esto_flow", "")),
                            "esto_product": str(pair.get("esto_product", "")),
                        }
                    )

        json_path = out_dir / "chart_navigation_hierarchy.json"
        csv_path = out_dir / "chart_navigation_hierarchy.csv"
        json_path.write_text(json.dumps(hierarchy, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
        pd.DataFrame(
            rows,
            columns=[
                "dashboard_path",
                "sheet",
                "page_key",
                "page_label",
                "chart_group_key",
                "chart_group_label",
                "measure",
                "template_order",
                "fuel",
                "chart_file",
                "esto_flow",
                "esto_product",
            ],
        ).sort_values(["dashboard_path", "template_order", "chart_group_key", "sheet", "measure", "fuel", "esto_flow", "esto_product"]).to_csv(csv_path, index=False)
        rendered_template_path = out_dir / "chart_navigation_rendered_template.json"
        rendered_template_path.write_text(
            json.dumps(_compact_dashboard_template(), ensure_ascii=True, indent=2) + "\n",
            encoding="utf-8",
        )
        return json_path, csv_path, rendered_template_path

    def _chart_group_rows_from_entries(
        entries_by_path: dict[tuple[str, ...], list[dict[str, str]]],
    ) -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        for node_path in sorted(entries_by_path, key=lambda item: (len(item), tuple(str(part).lower() for part in item))):
            entries = entries_by_path.get(node_path, [])
            if not entries:
                continue
            display_path = [
                _display_balance_path_part(part, is_top=idx == 0)
                for idx, part in enumerate(node_path)
            ]
            dashboard_path = " > ".join(display_path)
            section_id = _section_id_from_path(node_path)
            section_label = _display_balance_path_part(node_path[-1], is_top=False) if node_path else ""
            for entry in entries:
                chart_file = f"charts/{entry.get('file', '')}"
                hierarchy = _dashboard_hierarchy_from_path(
                    node_path,
                    entry_kind=str(entry.get("entry_kind", "direct") or "direct"),
                    measure=entry.get("measure", ""),
                    fallback_label=entry.get("sheet", ""),
                )
                rows.append(
                    {
                        "chart_group_id": f"chart::{chart_file}",
                        "dashboard_path": dashboard_path,
                        "sheet": str(entry.get("sheet", "")),
                        "page_key": str(entry.get("page_key", "") or hierarchy["page_key"]),
                        "page_label": str(entry.get("page_label", "") or hierarchy["page_label"]),
                        "chart_group_key": str(entry.get("chart_group_key", "") or hierarchy["chart_group_key"]),
                        "chart_group_label": str(entry.get("chart_group_label", "") or hierarchy["chart_group_label"]),
                        "measure": str(entry.get("measure", "")),
                        "fuel_label": str(entry.get("fuel", "")),
                        "chart_file": chart_file,
                        "section_id": section_id,
                        "section_label": section_label,
                        "entry_kind": str(entry.get("entry_kind", "direct") or "direct"),
                        "template_order": str(entry.get("template_order", "")),
                    }
                )
        return rows

    def _all_direct_chart_group_rows() -> list[dict[str, str]]:
        rows: list[dict[str, str]] = []
        for (sheet, measure, fuel), file_name in sorted(chart_lookup.items(), key=lambda item: (item[0][0], item[0][1], item[0][2])):
            path = tuple(str(v).strip() for v in list((sheet_catalog.get(str(sheet), {}) or {}).get("path", []) or []) if str(v).strip())
            if not path:
                path = (str(sheet),)
            display_path = [
                _display_balance_path_part(part, is_top=idx == 0)
                for idx, part in enumerate(path)
            ]
            chart_file = f"charts/{file_name}"
            rows.append(
                {
                    "chart_group_id": f"chart::{chart_file}",
                    "dashboard_path": " > ".join(display_path),
                    "sheet": str(sheet),
                    **_dashboard_hierarchy_from_path(
                        path,
                        entry_kind="direct",
                        measure=measure,
                        fallback_label=sheet,
                    ),
                    "measure": str(measure),
                    "fuel_label": str(fuel),
                    "chart_file": chart_file,
                    "section_id": _section_id_from_path(path),
                    "section_label": _display_balance_path_part(path[-1], is_top=False) if path else "",
                    "entry_kind": "direct",
                    "template_order": str(_template_order_for_entry(path, {"entry_kind": "direct"})),
                }
            )
        return rows

    def _template_chart_group_rows() -> list[dict[str, str]]:
        """Build chart group exposure rows directly from the template as the source of truth.

        The template owns chart exposure.  Each graph spec is matched directly against the
        rendered chart lookup table, using the ESTO flow/product declarations in the template
        and the actual chart-to-ESTO mapping metadata.  This intentionally avoids any
        candidate-sheet narrowing or rendered-node heuristics when deciding which direct
        chart groups are exposed.

        Aggregate entries are still appended separately from the rendered aggregate chart
        entries because those are declared at the template-node level rather than as direct
        per-fuel graph specs.
        """
        if not template_allowlist:
            return _chart_group_rows_from_entries(all_node_entries_for_chart_groups)

        def _is_total_fuel(fuel: str) -> bool:
            text = str(fuel).strip().lower()
            return text == "total" or bool(re.match(r"^\d+(?:[._]\d+)*\s+total$", text))

        def _spec_matches_chart(entry: dict[str, str], spec: dict[str, Any]) -> bool:
            mapping_info = _chart_mapping_info(entry)
            pairs = list(mapping_info.get("esto_pairs", []) or [])
            if not pairs:
                return False

            spec_flow_norms = set(spec.get("esto_flows_norm", set()))
            if not spec_flow_norms:
                spec_flow = str(spec.get("esto_flow", "")).strip()
                spec_flow_norms = {" ".join(spec_flow.lower().split())} if spec_flow else set()
            if not spec_flow_norms:
                return False

            spec_measure = str(spec.get("measure", "")).strip()
            entry_measure = str(entry.get("measure", "")).strip()
            if spec_measure and entry_measure and spec_measure != entry_measure:
                return False

            matching_pairs = [
                pair
                for pair in pairs
                if " ".join(str(pair.get("esto_flow", "")).strip().lower().split()) in spec_flow_norms
            ]
            if not matching_pairs:
                return False

            exclude_norm = set(spec.get("exclude_products_norm", set()))
            if spec.get("include_all_products", False):
                return any(
                    str(pair.get("esto_product", "")).strip()
                    and not _is_total_fuel(str(pair.get("esto_product", "")).strip())
                    and " ".join(str(pair.get("esto_product", "")).strip().lower().split()) not in exclude_norm
                    for pair in matching_pairs
                )

            products_norm = set(spec.get("products_norm", set()))
            return any(
                " ".join(str(pair.get("esto_product", "")).strip().lower().split()) in products_norm
                and " ".join(str(pair.get("esto_product", "")).strip().lower().split()) not in exclude_norm
                for pair in matching_pairs
            )

        rows: list[dict[str, str]] = []
        seen: set[tuple[str, str, str, str]] = set()  # (dashboard_path, sheet, measure, fuel)

        for tmpl_path, specs in sorted(
            allowed_template_graph_specs.items(),
            key=lambda item: (len(item[0]), item[0]),
        ):
            if not specs or not tmpl_path:
                continue

            display_parts = [
                _display_balance_path_part(part, is_top=(idx == 0))
                for idx, part in enumerate(tmpl_path)
            ]
            dashboard_path = " > ".join(display_parts)
            section_id = _section_id_from_path(tmpl_path)
            section_label = _display_balance_path_part(tmpl_path[-1], is_top=False)
            for spec in specs:
                for (sheet, measure, fuel_label), file_name in sorted(
                    chart_lookup.items(),
                    key=lambda item: (item[0][0], item[0][1], item[0][2]),
                ):
                    entry = {
                        "sheet": str(sheet),
                        "measure": str(measure),
                        "fuel": str(fuel_label),
                        "file": str(file_name),
                    }
                    if not _spec_matches_chart(entry, spec):
                        continue
                    if _is_total_fuel(fuel_label):
                        continue
                    key = (dashboard_path, str(sheet), str(measure), str(fuel_label))
                    if key in seen:
                        continue
                    seen.add(key)
                    chart_file = f"charts/{file_name}"
                    rows.append(
                        {
                            "chart_group_id": f"chart::{chart_file}",
                            "dashboard_path": dashboard_path,
                            "sheet": str(sheet),
                            **_dashboard_hierarchy_from_path(
                                tmpl_path,
                                entry_kind="direct",
                                measure=measure,
                                fallback_label=sheet,
                            ),
                            "measure": str(measure),
                            "fuel_label": str(fuel_label),
                            "chart_file": chart_file,
                            "section_id": section_id,
                            "section_label": section_label,
                            "entry_kind": "direct",
                            "template_order": str(_template_order_for_entry(tmpl_path, {"entry_kind": "direct"})),
                        }
                    )

        # Append aggregate entries from the template-filtered node_entries snapshot.
        for path, entries in all_node_entries_for_chart_groups.items():
            agg_display = [
                _display_balance_path_part(part, is_top=(idx == 0))
                for idx, part in enumerate(path)
            ]
            agg_dashboard_path = " > ".join(agg_display)
            agg_section_id = _section_id_from_path(path)
            agg_section_label = _display_balance_path_part(path[-1], is_top=False) if path else ""
            for entry in entries:
                if str(entry.get("entry_kind", "")).strip() != "aggregate":
                    continue
                sh = str(entry.get("sheet", ""))
                m = str(entry.get("measure", ""))
                fl = str(entry.get("fuel", ""))
                file_name = entry.get("file", "")
                if not file_name:
                    continue
                key = (agg_dashboard_path, sh, m, fl)
                if key in seen:
                    continue
                seen.add(key)
                chart_file = f"charts/{file_name}"
                rows.append(
                    {
                        "chart_group_id": f"chart::{chart_file}",
                        "dashboard_path": agg_dashboard_path,
                        "sheet": sh,
                        **_dashboard_hierarchy_from_path(
                            path,
                            entry_kind="aggregate",
                            measure=m,
                            fallback_label=sh,
                        ),
                        "measure": m,
                        "fuel_label": fl,
                        "chart_file": chart_file,
                        "section_id": agg_section_id,
                        "section_label": agg_section_label,
                        "entry_kind": "aggregate",
                        "template_order": str(_template_order_for_entry(path, {"entry_kind": "aggregate"})),
                    }
                )

        return sorted(
            rows,
            key=lambda row: (
                str(row.get("dashboard_path", "")),
                _as_sort_int(row.get("template_order", "")),
                str(row.get("chart_group_key", "")),
                str(row.get("sheet", "")),
                str(row.get("measure", "")),
                str(row.get("fuel_label", "")),
                str(row.get("chart_file", "")),
            ),
        )

    def _write_chart_group_exposure() -> tuple[Path, Path]:
        exposed_rows = _template_chart_group_rows()
        exposed_ids = {str(row["chart_group_id"]).strip() for row in exposed_rows if str(row.get("chart_group_id", "")).strip()}
        all_rows = _all_direct_chart_group_rows()
        aggregate_rows = [row for row in exposed_rows if str(row.get("entry_kind", "")).strip() == "aggregate"]
        all_rows.extend(aggregate_rows)
        deduped_all_rows: list[dict[str, str]] = []
        seen: set[tuple[str, str, str, str, str]] = set()
        for row in all_rows:
            key = (
                str(row.get("chart_group_id", "")),
                str(row.get("dashboard_path", "")),
                str(row.get("chart_group_key", "")),
                str(row.get("sheet", "")),
                str(row.get("measure", "")),
                str(row.get("fuel_label", "")),
            )
            if key in seen:
                continue
            seen.add(key)
            row = dict(row)
            row["exposed_in_dashboard"] = str(row.get("chart_group_id", "")).strip() in exposed_ids
            deduped_all_rows.append(row)

        chart_group_path = out_dir / "chart_group_exposure.csv"
        exposed_df = pd.DataFrame(
            exposed_rows,
            columns=[
                "chart_group_id",
                "dashboard_path",
                "sheet",
                "page_key",
                "page_label",
                "chart_group_key",
                "chart_group_label",
                "measure",
                "fuel_label",
                "chart_file",
                "section_id",
                "section_label",
                "entry_kind",
                "template_order",
            ],
        )
        if not exposed_df.empty:
            exposed_df["_template_order_sort"] = exposed_df["template_order"].map(_as_sort_int)
            exposed_df = exposed_df.sort_values(
                ["dashboard_path", "_template_order_sort", "chart_group_key", "sheet", "measure", "fuel_label", "chart_file", "chart_group_id"],
                kind="mergesort",
            ).drop(columns=["_template_order_sort"])
        exposed_df.to_csv(chart_group_path, index=False)

        all_chart_groups_path = out_dir / "all_chart_groups.csv"
        all_chart_groups_df = pd.DataFrame(
            deduped_all_rows,
            columns=[
                "chart_group_id",
                "dashboard_path",
                "sheet",
                "page_key",
                "page_label",
                "chart_group_key",
                "chart_group_label",
                "measure",
                "fuel_label",
                "chart_file",
                "section_id",
                "section_label",
                "entry_kind",
                "template_order",
                "exposed_in_dashboard",
            ],
        )
        if not all_chart_groups_df.empty:
            all_chart_groups_df["_template_order_sort"] = all_chart_groups_df["template_order"].map(_as_sort_int)
            all_chart_groups_df = all_chart_groups_df.sort_values(
                ["dashboard_path", "_template_order_sort", "chart_group_key", "sheet", "measure", "fuel_label", "chart_file", "chart_group_id"],
                kind="mergesort",
            ).drop(columns=["_template_order_sort"])
        all_chart_groups_df.to_csv(all_chart_groups_path, index=False)
        return chart_group_path, all_chart_groups_path

    chart_hierarchy_json_path, chart_hierarchy_csv_path, rendered_template_path = _write_chart_navigation_hierarchy()
    chart_group_exposure_path, all_chart_groups_path = _write_chart_group_exposure()

    empty_pages: list[dict[str, Any]] = []
    for stale_page in dashboards_dir.glob("node__*.html"):
        stale_page.unlink()
    stale_index = dashboards_dir / "index.html"
    if stale_index.exists():
        stale_index.unlink()

    render_page_paths = [path for path in node_paths if path]
    for path in render_page_paths:
        if not path:
            continue
        title = _display_balance_path_part(path[-1], is_top=len(path) == 1)
        filename = path_to_filename[path]

        child_links = _section_links_for_top_path(path)

        chart_entries = _chart_entries_for_path(path)
        fallback_note = ""
        if not chart_entries and not child_links:
            empty_pages.append({"path": " > ".join(path), "level": len(path)})

        html = _build_page_html(
            title=title,
            top_links=top_links,
            header_nav_links=_header_nav_links_for_path(path),
            child_links=child_links,
            current_file=filename,
            page_measure="Energy balance (PJ)",
            chart_entries=chart_entries,
            empty_notice=empty_notice,
            fallback_note=fallback_note,
        )
        (dashboards_dir / filename).write_text(html, encoding="utf-8")

    if about_page_config:
        about_html = _build_about_page_html(
            title=_clean_token(about_page_config.get("title", "")) or "About This Dashboard",
            top_links=top_links,
            current_file="about.html",
            about_config=about_page_config,
        )
        (dashboards_dir / "about.html").write_text(about_html, encoding="utf-8")

    empty_pages_path = dashboards_dir / "empty_pages.csv"
    empty_pages_df = pd.DataFrame(empty_pages, columns=["path", "level"])
    if not empty_pages_df.empty:
        empty_pages_df = empty_pages_df.sort_values(["path", "level"], kind="mergesort").reset_index(drop=True)
    empty_pages_df.to_csv(empty_pages_path, index=False)
    dashboard_index = str(dashboards_dir / path_to_filename[top_page_paths[0]]) if top_page_paths else ""

    graph_fuel_coverage_csv: str | None = None
    if template_allowlist:
        graph_fuel_coverage_csv = write_dashboard_graph_fuel_coverage(
            template=template_allowlist,
            mapping_status=mapping_status,
            output_path=out_dir / "dashboard_graph_fuel_coverage.csv",
            chart_comparison_long=chart_comparison_long,
        )

    return {
        "charts_written": len(written_charts) + aggregate_charts_written,
        "dashboard_index": dashboard_index,
        "charts_dir": str(charts_dir),
        "dashboards_dir": str(dashboards_dir),
        "empty_pages_csv": str(empty_pages_path),
        "chart_navigation_hierarchy": str(chart_hierarchy_json_path),
        "chart_navigation_hierarchy_flat": str(chart_hierarchy_csv_path),
        "chart_navigation_rendered_template": str(rendered_template_path),
        "chart_group_exposure": str(chart_group_exposure_path),
        "all_chart_groups": str(all_chart_groups_path),
        "graph_fuel_coverage_csv": graph_fuel_coverage_csv,
    }


def clone_v2_structure_from_paths(sheet_paths: dict[str, list[str]]) -> dict[str, Any]:
    """
    Build a structure-config object from a V2 sheet->path mapping.

    This is used for one-time generation of config/leap_results_balance_dashboard_structure.json.
    """
    sheet_catalog: dict[str, Any] = {}
    for order, (sheet, path) in enumerate(sorted(sheet_paths.items(), key=lambda kv: tuple([p.lower() for p in kv[1]] + [kv[0].lower()]))):
        clean_path = [str(token).strip() for token in path if str(token).strip()]
        if not clean_path:
            clean_path = [str(sheet).strip()]
        sheet_catalog[str(sheet)] = {
            "display_label": str(sheet),
            "path": clean_path,
            "measure": "Energy balance (PJ)",
            "sort_order": order,
        }

    # Build a nested tree from all unique path prefixes.
    root: dict[str, Any] = {"children": {}}
    for cfg in sheet_catalog.values():
        path = cfg["path"]
        node = root
        for token in path:
            node = node["children"].setdefault(token, {"children": {}})

    def _to_nodes(node: dict[str, Any], prefix: list[str]) -> list[dict[str, Any]]:
        out_nodes: list[dict[str, Any]] = []
        for label, child in sorted(node.get("children", {}).items(), key=lambda kv: kv[0].lower()):
            full = prefix + [label]
            out_nodes.append(
                {
                    "id": "__".join(_safe_token(part) for part in full),
                    "label": label,
                    "children": _to_nodes(child, full),
                }
            )
        return out_nodes

    page_tree = _to_nodes(root, [])

    return {
        "page_tree": page_tree,
        "sheet_catalog": sheet_catalog,
        "flow_to_sheet": {},
        "empty_page_notice": DEFAULT_EMPTY_PAGE_NOTICE,
    }


def _strip_esto_code_prefix(label: object) -> str:
    text = str(label or "").strip()
    return re.sub(r"^\d{2}(?:\.\d{2})*\s*", "", text).strip()


def _esto_flow_code(flow: object) -> str:
    match = re.match(r"^(\d{2}(?:\.\d{2})*)\b", str(flow or "").strip())
    return match.group(1) if match else ""


def _esto_flow_sort_key(flow: object) -> tuple[int, ...]:
    code = _esto_flow_code(flow)
    if not code:
        return (999,)
    parts: list[int] = []
    for part in code.split("."):
        try:
            parts.append(int(part))
        except ValueError:
            parts.append(999)
    return tuple(parts)


def _esto_parent_codes(code: str) -> list[str]:
    if not code:
        return []
    parts = code.split(".")
    return [".".join(parts[:idx]) for idx in range(1, len(parts))]


def _esto_flow_label_lookup(flows: Iterable[object]) -> dict[str, str]:
    lookup = dict(ESTO_FLOW_LABEL_FALLBACKS)
    for flow in flows:
        code = _esto_flow_code(flow)
        label = _strip_esto_code_prefix(flow)
        if code and label:
            lookup[code] = label
    return lookup


def _esto_label_for_code(code: str, label_lookup: dict[str, str], fallback: str = "") -> str:
    return label_lookup.get(code, "").strip() or fallback or code


def _esto_flow_path(flow: object, label_lookup: dict[str, str]) -> list[str]:
    code = _esto_flow_code(flow)
    label = _strip_esto_code_prefix(flow) or str(flow or "").strip() or "Unknown flow"
    if not code:
        return ["Other sector", label]

    if code == "01":
        return ["Supply", "Production"]
    if code == "02":
        return ["Supply", "Imports"]
    if code == "03":
        return ["Supply", "Exports"]
    if code == "04":
        return ["Bunkers", "International marine bunkers"]
    if code == "05":
        return ["Bunkers", "International aviation bunkers"]
    if code == "07":
        return ["Supply", "Total primary energy supply"]

    if code == "08" or code.startswith("08."):
        path = ["Other transformation", "Transfers"]
        if code != "08":
            path.append(label)
        return path

    if code == "09":
        return ["Other transformation", "Total transformation sector"]
    if code in {"09.01.01", "09.02.01"}:
        return ["Power", "Electricity plants"]
    if code in {"09.01.02", "09.02.02"}:
        return ["Power", "CHP plants"]
    if code in {"09.01.03", "09.02.03"}:
        return ["Power", "Heat plants"]
    if code.startswith("09.06"):
        path = ["Other transformation", "Gas processing plants"]
        if code != "09.06":
            path.append(label)
        return path
    if code == "09.07" or code.startswith("09.07."):
        if code == "09.07":
            return ["Refining", "Oil refineries"]
        return ["Refining", "Oil refineries", label]
    if code.startswith("09.08"):
        path = ["Other transformation", "Coal transformation"]
        if code != "09.08":
            path.append(label)
        return path
    if code.startswith("09.12"):
        if code == "09.12":
            return ["Other transformation", "Non-specified transformation"]
        return ["Other transformation", "Non-specified transformation", label]
    if code.startswith("09.13"):
        path = ["Other transformation", "Hydrogen transformation"]
        if code != "09.13":
            path.append(label)
        return path
    if code.startswith("09."):
        parent_code = ".".join(code.split(".")[:2])
        parent_label = _esto_label_for_code(parent_code, label_lookup, "Other transformation")
        if code == parent_code:
            return ["Other transformation", parent_label]
        return ["Other transformation", parent_label, label]

    if code == "10":
        return ["Losses & own use"]
    if code.startswith("10.01"):
        if code == "10.01":
            return ["Losses & own use", "Own use"]
        return ["Losses & own use", "Own use", label]
    if code.startswith("10.02"):
        return ["Losses & own use", "Transmission and distribution losses"]

    if code in {"12", "13"}:
        return ["Demand", label]

    if code == "14":
        return ["Industry sector"]
    if code == "14.03":
        return ["Industry sector", "Manufacturing"]
    if code.startswith("14.03."):
        return ["Industry sector", "Manufacturing", label]
    if code.startswith("14."):
        return ["Industry sector", label]

    if code == "15":
        return ["Transport sector"]
    if code.startswith("15."):
        return ["Transport sector", label]

    if code == "16.01":
        return ["Buildings", "Commercial and public services"]
    if code == "16.01.01":
        return ["Buildings", "Datacentres"]
    if code.startswith("16.01."):
        return ["Buildings", "Commercial and public services", label]
    if code == "16.02":
        return ["Buildings", "Residential"]
    if code.startswith("16.02."):
        return ["Buildings", "Residential", label]
    if code == "16.03":
        return ["Other sector", "Agriculture and fishing", "Agriculture"]
    if code == "16.04":
        return ["Other sector", "Agriculture and fishing", "Fishing"]
    if code == "16.05":
        return ["Other sector", "Non-specified others"]
    if code == "16":
        return ["Other sector"]
    if code.startswith("16."):
        return ["Other sector", label]

    if code == "17" or code.startswith("17."):
        if code == "17":
            return ["Other sector", "Non-energy use"]
        return ["Other sector", "Non-energy use", label]
    if code == "18" or code.startswith("18."):
        if code == "18":
            return ["Power", "Electricity output in GWh"]
        return ["Power", "Electricity output in GWh", label]
    if code == "19" or code.startswith("19."):
        if code == "19":
            return ["Power", "Heat output in PJ"]
        return ["Power", "Heat output in PJ", label]

    return ["Other sector", label]


def _esto_dashboard_path_sort_key(path: Sequence[str], flow: object = "") -> tuple[object, ...]:
    clean_path = [str(part).strip() for part in path if str(part).strip()]
    top = clean_path[0] if clean_path else ""
    return (
        ESTO_DASHBOARD_GROUP_ORDER.get(top, 999),
        tuple(part.lower() for part in clean_path),
        _esto_flow_sort_key(flow),
    )


def _sheet_key_from_esto_flow(flow: object) -> str:
    base = _strip_esto_code_prefix(flow) or str(flow or "").strip() or "Unknown flow"
    code = _esto_flow_code(flow)
    if code:
        code_token = _safe_token(code.replace(".", "_"))
        return f"esto__{code_token}__{_safe_token(base)}"
    return f"esto__{_safe_token(base)}"


def _esto_flow_key(flow: object) -> str:
    """Stable internal key for one ESTO flow."""
    text = _clean_token(flow)
    if not text:
        return ""
    code = _esto_flow_code(text)
    if code:
        return f"esto_flow__{_safe_token(code.replace('.', '_'))}__{_safe_token(_strip_esto_code_prefix(text) or text)}"
    return f"esto_flow__{_safe_token(text)}"


def _esto_flow_group_key(flows: Sequence[object]) -> str:
    """Stable internal key for a declared group of one or more ESTO flows."""
    clean_flows = sorted(
        {_clean_token(flow) for flow in flows if _clean_token(flow)},
        key=lambda flow: (_esto_flow_sort_key(flow), flow.lower()),
    )
    if not clean_flows:
        return ""
    flow_keys = [_esto_flow_key(flow).replace("esto_flow__", "", 1) for flow in clean_flows]
    return "esto_flow_group__" + "__".join(flow_keys)


def _esto_flow_group_label(flows: Sequence[object], fallback_label: object = "") -> str:
    """Human label for a declared ESTO flow group."""
    fallback = _clean_token(fallback_label)
    if fallback:
        return fallback
    clean_flows = [_clean_token(flow) for flow in flows if _clean_token(flow)]
    if len(clean_flows) == 1:
        return _strip_esto_code_prefix(clean_flows[0]) or clean_flows[0]
    return " + ".join(_strip_esto_code_prefix(flow) or flow for flow in clean_flows)


def _add_dashboard_context_aliases(frame: pd.DataFrame) -> pd.DataFrame:
    """Add clearer dashboard/context names while preserving legacy columns."""
    out = pd.DataFrame() if frame is None else frame.copy()
    for col in ["page_key", "page_label", "chart_group_key", "chart_group_label"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()
    alias_pairs = {
        "dashboard_page_key": "page_key",
        "dashboard_page_label": "page_label",
        "dashboard_section_key": "chart_group_key",
        "dashboard_section_label": "chart_group_label",
    }
    for alias_col, source_col in alias_pairs.items():
        if alias_col not in out.columns:
            out[alias_col] = out[source_col]
        else:
            out[alias_col] = out[alias_col].fillna("").astype(str).str.strip()
            out.loc[out[alias_col].eq(""), alias_col] = out.loc[out[alias_col].eq(""), source_col]
    if "chart_kind" not in out.columns:
        out["chart_kind"] = "by_fuel"
    out["chart_kind"] = out["chart_kind"].fillna("").astype(str).str.strip().replace("", "by_fuel")
    return out


def _add_esto_flow_context_columns(frame: pd.DataFrame) -> pd.DataFrame:
    """Add stable ESTO flow and flow-group keys used for comparison/dedupe."""
    out = pd.DataFrame() if frame is None else frame.copy()
    for col in ["esto_flow", "chart_group_label"]:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()
    if "esto_flow_key" not in out.columns:
        out["esto_flow_key"] = out["esto_flow"].map(_esto_flow_key)
    else:
        out["esto_flow_key"] = out["esto_flow_key"].fillna("").astype(str).str.strip()
        missing = out["esto_flow_key"].eq("")
        out.loc[missing, "esto_flow_key"] = out.loc[missing, "esto_flow"].map(_esto_flow_key)
    if "esto_flow_group_key" not in out.columns:
        out["esto_flow_group_key"] = out["esto_flow"].map(lambda flow: _esto_flow_group_key([flow]))
    else:
        out["esto_flow_group_key"] = out["esto_flow_group_key"].fillna("").astype(str).str.strip()
        missing = out["esto_flow_group_key"].eq("")
        out.loc[missing, "esto_flow_group_key"] = out.loc[missing, "esto_flow"].map(lambda flow: _esto_flow_group_key([flow]))
    if "esto_flow_group_label" not in out.columns:
        out["esto_flow_group_label"] = [
            _esto_flow_group_label([flow], fallback_label=label)
            for flow, label in zip(out["esto_flow"], out["chart_group_label"], strict=False)
        ]
    else:
        out["esto_flow_group_label"] = out["esto_flow_group_label"].fillna("").astype(str).str.strip()
        missing = out["esto_flow_group_label"].eq("")
        out.loc[missing, "esto_flow_group_label"] = [
            _esto_flow_group_label([flow], fallback_label=label)
            for flow, label in zip(out.loc[missing, "esto_flow"], out.loc[missing, "chart_group_label"], strict=False)
        ]
    return out


def _collapse_template_multi_flow_comparison_rows(
    comparison_long: pd.DataFrame,
    template_groups: pd.DataFrame,
) -> pd.DataFrame:
    """Make template multi-flow rows aggregate at the declared flow-group level."""
    if comparison_long.empty or template_groups.empty or "template_is_multi_flow" not in template_groups.columns:
        return comparison_long.copy()

    multi_template_groups = template_groups[template_groups["template_is_multi_flow"].fillna(False).astype(bool)].copy()
    if multi_template_groups.empty:
        return comparison_long.copy()

    out = comparison_long.copy()
    match_cols = ["scenario", "sheet", "chart_group_key", "measure", "fuel_label"]
    needed_cols = match_cols + ["esto_flow_key", "esto_flow_group_key"]
    for col in needed_cols:
        if col not in out.columns:
            out[col] = ""
        out[col] = out[col].fillna("").astype(str).str.strip()
    for col in match_cols:
        if col not in multi_template_groups.columns:
            multi_template_groups[col] = ""
        multi_template_groups[col] = multi_template_groups[col].fillna("").astype(str).str.strip()

    covered = set(
        tuple(row)
        for row in multi_template_groups[match_cols]
        .drop_duplicates()
        .itertuples(index=False, name=None)
    )
    if not covered:
        return out

    template_mask = out["sheet"].fillna("").astype(str).str.startswith("template__")
    covered_mask = out[match_cols].apply(lambda row: tuple(row) in covered, axis=1)
    collapse_mask = template_mask & covered_mask & out["esto_flow_group_key"].ne("")
    if collapse_mask.any():
        out.loc[collapse_mask, "esto_flow_key"] = out.loc[collapse_mask, "esto_flow_group_key"]
    return out


def _add_leap_parent_transfer_rows_to_template(comparison_long: pd.DataFrame) -> pd.DataFrame:
    """
    Add LEAP parent 08 transfer rows to the template transfer chart only.

    ESTO and 9th treat parent 08 Transfers as a subtotal, so the dashboard template
    is authored against the active 08.xx transfer rows. The LEAP export, however,
    carries real non-subtotal transfer values on the parent 08 row too. Without
    this bridge, the template transfer charts show only the 08.xx LEAP portion.
    """
    if comparison_long.empty:
        return comparison_long.copy()

    required_cols = {"sheet", "source", "measure", "fuel_label", "scenario", "year", "value"}
    if not required_cols.issubset(comparison_long.columns):
        return comparison_long.copy()

    frame = comparison_long.copy()
    parent_mask = frame["sheet"].fillna("").astype(str).str.strip().eq("esto__08__Transfers")
    parent_mask &= frame["source"].fillna("").astype(str).str.strip().eq("leap")
    parent_rows = frame.loc[parent_mask].copy()
    if parent_rows.empty:
        return frame

    template_mask = frame["sheet"].fillna("").astype(str).str.strip().eq("template__Other_transformation__Transfers")
    template_rows = frame.loc[template_mask].copy()
    if template_rows.empty:
        return frame

    meta_cols = [
        "scenario",
        "measure",
        "fuel_label",
        "sheet",
        "page_key",
        "page_label",
        "chart_group_key",
        "chart_group_label",
        "dashboard_page_key",
        "dashboard_page_label",
        "dashboard_section_key",
        "dashboard_section_label",
        "chart_kind",
        "esto_flow_key",
        "esto_flow_group_key",
        "esto_flow_group_label",
        "ninth_pairs_label",
    ]
    for col in meta_cols:
        if col not in template_rows.columns:
            template_rows[col] = ""
        if col not in parent_rows.columns:
            parent_rows[col] = ""

    template_meta = (
        template_rows[meta_cols]
        .drop_duplicates(subset=["scenario", "measure", "fuel_label"])
        .reset_index(drop=True)
    )
    if template_meta.empty:
        return frame

    value_cols = ["economy", "scenario", "measure", "fuel_label", "source", "year", "value"]
    parent_for_template = parent_rows[value_cols].merge(
        template_meta,
        on=["scenario", "measure", "fuel_label"],
        how="inner",
        suffixes=("", "_template"),
    )
    if parent_for_template.empty:
        return frame

    parent_for_template = parent_for_template.reindex(columns=frame.columns, fill_value="")
    return pd.concat([frame, parent_for_template], ignore_index=True, sort=False)


def _fill_template_transfer_base_values(
    comparison_long: pd.DataFrame,
    *,
    base_df: pd.DataFrame,
    base_year: int,
    base_economy: str,
) -> pd.DataFrame:
    """Fill template transfer ESTO base rows from their explicit 08.xx pairs."""
    if comparison_long.empty or base_df is None or base_df.empty:
        return comparison_long.copy()
    required_cols = {"sheet", "source", "fuel_label", "value"}
    if not required_cols.issubset(comparison_long.columns):
        return comparison_long.copy()

    out = comparison_long.copy()
    out["value"] = pd.to_numeric(out["value"], errors="coerce")
    transfer_template = out["sheet"].fillna("").astype(str).str.strip().eq("template__Other_transformation__Transfers")
    base_source = out["source"].fillna("").astype(str).str.strip().eq("base")
    target = transfer_template & base_source
    if not target.any():
        return out

    transfer_child_flows = [
        "08.01 Recycled products",
        "08.02 Interproduct transfers",
        "08.03 Products transferred",
        "08.04 Gas separation",
        "08.99 Transfers nonspecified",
    ]
    product_lookup: dict[str, str] = {}
    if {"flows", "products"}.issubset(base_df.columns):
        base_products = base_df[
            base_df["flows"].fillna("").astype(str).str.strip().isin(transfer_child_flows)
        ].copy()
        if "is_subtotal" in base_products.columns:
            base_products = base_products[~base_products["is_subtotal"].map(_to_bool)].copy()
        for product in base_products["products"].dropna().astype(str).str.strip().unique():
            if product and not _product_is_total(product):
                product_lookup.setdefault((_strip_esto_code_prefix(product) or product).strip(), product)

    cache: dict[tuple[str, str], float] = {}
    for idx, row in out.loc[target].iterrows():
        flow = _clean_token(row.get("esto_flow", "")) if "esto_flow" in out.columns else ""
        product = _clean_token(row.get("esto_product", "")) if "esto_product" in out.columns else ""
        if flow.startswith("08.") and product:
            flow_product_pairs = [(flow, product)]
        else:
            fuel_label = _clean_token(row.get("fuel_label", ""))
            product = product_lookup.get(fuel_label, "")
            flow_product_pairs = [(flow_item, product) for flow_item in transfer_child_flows if product]
        values: list[float] = []
        for flow_item, product_item in flow_product_pairs:
            key = (flow_item, product_item)
            if key not in cache:
                cache[key] = pull_base_year_value(
                    base_df,
                    base_year=base_year,
                    economy_code=base_economy,
                    esto_flow=flow_item,
                    esto_product=product_item,
                    value_sign_role="",
                )
            values.append(cache[key])
        out.at[idx, "value"] = sum(value for value in values if pd.notna(value)) if values else float("nan")
    return out


def _build_page_tree_from_paths(paths: list[list[str]]) -> list[dict[str, Any]]:
    root: dict[str, Any] = {"children": {}}
    for path in paths:
        node = root
        for token in path:
            clean = str(token).strip()
            if not clean:
                continue
            node = node["children"].setdefault(clean, {"children": {}})

    def _to_nodes(node: dict[str, Any], prefix: list[str]) -> list[dict[str, Any]]:
        out: list[dict[str, Any]] = []
        for label, child in node.get("children", {}).items():
            full = prefix + [label]
            out.append(
                {
                    "id": "__".join(_safe_token(part) for part in full),
                    "label": label,
                    "children": _to_nodes(child, full),
                }
            )
        return out

    return _to_nodes(root, [])


def _resolve_esto_structure(
    *,
    grouped: pd.DataFrame,
    structure_config: dict[str, Any],
) -> tuple[pd.DataFrame, dict[str, Any], pd.DataFrame]:
    structure = dict(structure_config or {})
    sheet_catalog = dict(structure.get("sheet_catalog", {}) or {})
    esto_flow_to_sheet = {
        str(k).strip(): str(v).strip()
        for k, v in (structure.get("esto_flow_to_sheet", {}) or {}).items()
        if str(k).strip() and str(v).strip()
    }

    out = grouped.copy()
    out["sheet"] = out["esto_flow"].map(esto_flow_to_sheet).fillna("")

    auto_rows: list[dict[str, Any]] = []
    if out["sheet"].eq("").any():
        missing_flows = (
            out.loc[out["sheet"].eq(""), "esto_flow"]
            .dropna()
            .astype(str)
            .str.strip()
            .drop_duplicates()
            .tolist()
        )
        label_lookup = _esto_flow_label_lookup([*out["esto_flow"].dropna().astype(str).tolist(), *missing_flows])
        for flow in sorted(missing_flows, key=lambda value: _esto_dashboard_path_sort_key(_esto_flow_path(value, label_lookup), value)):
            sheet_key = _sheet_key_from_esto_flow(flow)
            esto_flow_to_sheet[flow] = sheet_key
            flow_label = _strip_esto_code_prefix(flow) or flow
            flow_path = _esto_flow_path(flow, label_lookup)
            if sheet_key not in sheet_catalog:
                sheet_catalog[sheet_key] = {
                    "display_label": flow_label,
                    "path": flow_path,
                    "measure": "Energy balance (PJ)",
                    "sort_order": len(sheet_catalog),
                }
                auto_rows.append({"esto_flow": flow, "sheet": sheet_key, "path": " > ".join(sheet_catalog[sheet_key]["path"])})
        out["sheet"] = out["esto_flow"].map(esto_flow_to_sheet).fillna("")

    sheet_to_flow = {sheet: flow for flow, sheet in esto_flow_to_sheet.items()}
    catalog_items = sorted(
        sheet_catalog.items(),
        key=lambda item: (
            _esto_dashboard_path_sort_key((item[1] or {}).get("path", []) or [], sheet_to_flow.get(str(item[0]), "")),
            str(item[0]).lower(),
        ),
    )
    all_paths = [
        [str(token).strip() for token in list((cfg or {}).get("path", []) or []) if str(token).strip()]
        for _, cfg in catalog_items
    ]
    all_paths = [p for p in all_paths if p]
    page_tree = _build_page_tree_from_paths(all_paths) if all_paths else structure.get("page_tree")

    resolved = {
        **structure,
        "page_tree": page_tree,
        "sheet_catalog": sheet_catalog,
        "esto_flow_to_sheet": esto_flow_to_sheet,
        "empty_page_notice": _clean_token(structure.get("empty_page_notice", "")) or DEFAULT_EMPTY_PAGE_NOTICE,
    }
    auto_df = pd.DataFrame(auto_rows, columns=["esto_flow", "sheet", "path"])
    return out, resolved, auto_df


def load_balance_leap_long_esto_axis(
    *,
    ref_workbook_path: Path | str = DEFAULT_REF_WORKBOOK_PATH,
    tgt_workbook_path: Path | str = DEFAULT_TGT_WORKBOOK_PATH,
    template_sheet: str = "EBal|2060",
    mapping_pairs_path: ConfigTableRef = DEFAULT_MAPPING_PAIRS_PATH,
    codebook_path: Path | str = DEFAULT_CODEBOOK_PATH,
    structure_config: dict[str, Any] | None = None,
    known_issues: dict[str, Any] | None = None,
    projection_economy: str = "20_USA",
    explicit_pair_mappings_only: bool = False,
    allow_descendant_mapping_expansion: bool = True,
) -> dict[str, Any]:
    structure = structure_config or {}
    issues_cfg = known_issues or {}

    ref_path = _resolve(ref_workbook_path)
    tgt_path = _resolve(tgt_workbook_path)
    mapping_pairs = _resolve_config_table_ref(mapping_pairs_path)
    codebook = _resolve(codebook_path)
    for candidate in [ref_path, tgt_path]:
        if not candidate.exists():
            raise FileNotFoundError(f"Missing required input: {candidate}")
    mapping_pairs_file, mapping_pairs_sheet = split_config_table_ref(mapping_pairs)
    if not config_table_exists(mapping_pairs_file, sheet_name=mapping_pairs_sheet):
        raise FileNotFoundError(f"Missing required input: {mapping_pairs}")
    if not config_table_exists(codebook, sheet_name="code_to_name"):
        raise FileNotFoundError(f"Missing required input: {codebook}")

    extracted_ref = _extract_balance_workbook(
        ref_path,
        template_sheet=template_sheet,
        mapping_pairs_path=mapping_pairs,
        codebook_path=codebook,
        explicit_pair_mappings_only=explicit_pair_mappings_only,
        allow_descendant_mapping_expansion=allow_descendant_mapping_expansion,
    )
    extracted_tgt = _extract_balance_workbook(
        tgt_path,
        template_sheet=template_sheet,
        mapping_pairs_path=mapping_pairs,
        codebook_path=codebook,
        explicit_pair_mappings_only=explicit_pair_mappings_only,
        allow_descendant_mapping_expansion=allow_descendant_mapping_expansion,
    )

    combined = pd.concat([extracted_ref["mapped_long"], extracted_tgt["mapped_long"]], ignore_index=True, sort=False)
    if combined.empty:
        raise RuntimeError("Balance extraction produced no rows.")

    combined["scenario"] = combined.get("scenario", "").map(_normalize_scenario)
    combined["year"] = pd.to_numeric(combined.get("year", pd.NA), errors="coerce").astype("Int64")
    combined["value_petajoule"] = pd.to_numeric(
        combined.get("value_petajoule", combined.get("value", pd.NA)),
        errors="coerce",
    )

    for col in ["leap_sector", "leap_fuel", "leap_sector_name", "leap_fuel_name", "esto_flow", "esto_product"]:
        if col not in combined.columns:
            combined[col] = ""
        combined[col] = combined[col].fillna("").astype(str).str.strip()

    combined, override_report = _apply_mapping_overrides(
        combined,
        list(issues_cfg.get("mapping_overrides", []) or []),
    )
    for subtotal_col in ["leap_is_subtotal", "esto_is_subtotal", "ninth_is_subtotal"]:
        if subtotal_col not in combined.columns:
            combined[subtotal_col] = False
        combined[subtotal_col] = combined[subtotal_col].fillna(False).map(_to_bool)
    inferred_subtotal = combined.apply(_infer_subtotal_flag, axis=1)
    combined["leap_is_subtotal"] = combined["leap_is_subtotal"] | inferred_subtotal
    combined["is_subtotal"] = combined["leap_is_subtotal"]

    row_filters = issues_cfg.get("row_filters", {}) or {}
    combined = _apply_balance_row_filters(combined, row_filters)

    required_esto_mask = combined["esto_flow"].ne("") & combined["esto_product"].ne("")
    incomplete_rows = combined[~required_esto_mask].copy()
    mapped = combined[required_esto_mask].copy()

    for col in ["esto_is_subtotal", "ninth_is_subtotal"]:
        if col not in mapped.columns:
            mapped[col] = False
        mapped[col] = mapped[col].fillna(False).map(_to_bool)

    pre_group_leap_mapped = mapped.copy()

    grouped = (
        mapped.groupby(["scenario", "year", "esto_flow", "esto_product"], as_index=False)
        .agg(
            leap_value=("value_petajoule", "sum"),
            leap_sector=("leap_sector", _coalesce_pipe_tokens_unique),
            leap_fuel=("leap_fuel", _coalesce_pipe_tokens_unique),
            leap_sector_name=("leap_sector_name", _coalesce_unique),
            leap_fuel_name=("leap_fuel_name", _coalesce_unique),
            is_subtotal=("is_subtotal", "max"),
            esto_is_subtotal=("esto_is_subtotal", "max"),
            ninth_is_subtotal=("ninth_is_subtotal", "max"),
            source_sheet=("source_sheet", lambda s: "|".join(sorted(set([v for v in s.astype(str) if v])))),
            source_workbook=("source_workbook", lambda s: "|".join(sorted(set([v for v in s.astype(str) if v])))),
        )
        .reset_index(drop=True)
    )

    grouped["sector_name"] = grouped["esto_flow"].map(_strip_esto_code_prefix)
    grouped["fuel_label"] = grouped["esto_product"].map(_strip_esto_code_prefix)
    grouped["measure"] = "Energy balance (PJ)"

    grouped, resolved_structure, auto_sheet_rows = _resolve_esto_structure(
        grouped=grouped,
        structure_config=structure,
    )

    grouped["economy"] = projection_economy
    grouped["region"] = ""
    grouped["sector_code_9th"] = grouped["leap_sector"]
    grouped["ninth_fuel_code"] = grouped["leap_fuel"]
    grouped["leap_variable"] = "Energy Balance"
    grouped["leap_units"] = "Petajoule"
    grouped["leap_scale_note"] = ""

    leap_long = grouped[
        [
            "economy",
            "scenario",
            "region",
            "sheet",
            "sector_code_9th",
            "sector_name",
            "fuel_label",
            "year",
            "leap_value",
            "leap_variable",
            "leap_units",
            "measure",
            "leap_scale_note",
            "ninth_fuel_code",
            "esto_flow",
            "esto_product",
            "source_sheet",
            "source_workbook",
            "leap_sector",
            "leap_fuel",
            "leap_sector_name",
            "leap_fuel_name",
            "is_subtotal",
            "esto_is_subtotal",
            "ninth_is_subtotal",
        ]
    ].rename(columns={"sheet": "sheet_name"})

    leap_long["year"] = pd.to_numeric(leap_long["year"], errors="coerce").astype("Int64")
    leap_long["leap_value"] = pd.to_numeric(leap_long["leap_value"], errors="coerce")
    leap_long = leap_long.sort_values(["scenario", "sheet_name", "fuel_label", "year"], kind="mergesort").reset_index(drop=True)

    mapping_status = leap_long[
        [
            "sheet_name",
            "measure",
            "fuel_label",
            "sector_code_9th",
            "ninth_fuel_code",
            "esto_flow",
            "esto_product",
        ]
    ].drop_duplicates().rename(columns={"sheet_name": "sheet"})
    mapping_status["mapped"] = True
    mapping_status["has_any_mapping"] = True
    mapping_status["base_mapping_complete"] = True
    mapping_status["projection_mapping_complete"] = True
    mapping_status["partially_mapped"] = False
    mapping_status["missing_ninth_fuel"] = False
    mapping_status["missing_esto_flow"] = False
    mapping_status["missing_esto_product"] = False
    mapping_status["has_mapping_note"] = False
    mapping_status["mapping_source"] = "balance_table_esto_axis"
    mapping_status["flow_source"] = "esto_flow"
    mapping_status["fuel_source"] = "esto_product"
    mapping_status["sector_match_method"] = "esto_pair_axis"
    mapping_status["mapping_note"] = ""
    mapping_status["projection_parent_fallback"] = False
    mapping_status["projection_parent_sector_code"] = ""
    mapping_status["comparator_scope"] = "esto_pair"

    issues_df = _rows_to_issue_records(
        incomplete_rows,
        reason="missing_esto_pair",
        details="No explicit LEAP-to-ESTO pair mapping was found.",
    )

    return {
        "leap_long": leap_long,
        "mapping_status": mapping_status,
        "issues": issues_df,
        "override_report": override_report,
        "pre_group_leap_mapped": pre_group_leap_mapped,
        "pre_group_incomplete_rows": incomplete_rows.copy(),
        "unit_diagnostics": pd.concat([extracted_ref["unit_diag"], extracted_tgt["unit_diag"]], ignore_index=True, sort=False),
        "coverage": pd.concat([extracted_ref["coverage"], extracted_tgt["coverage"]], ignore_index=True, sort=False),
        "matching_diagnostics": pd.concat(
            [
                extracted_ref["report"].get("matching_diagnostics", pd.DataFrame()),
                extracted_tgt["report"].get("matching_diagnostics", pd.DataFrame()),
            ],
            ignore_index=True,
            sort=False,
        ),
        "resolved_structure": resolved_structure,
        "auto_sheet_rows": auto_sheet_rows,
        "extraction_summary": {
            "ref": extracted_ref["report"].get("summary", {}),
            "tgt": extracted_tgt["report"].get("summary", {}),
            "selected_template_ref": extracted_ref["template_sheet"],
            "selected_template_tgt": extracted_tgt["template_sheet"],
            "leap_rows_after_filters": int(len(leap_long)),
            "mapping_rows": int(len(mapping_status)),
            "issue_rows": int(len(issues_df)),
            "auto_sheet_rows": int(len(auto_sheet_rows)),
        },
    }



def _dashboard_template_esto_axis_records(
    chart_navigation_guide_path: Path | str | None,
    *,
    scenario_names: Sequence[str],
    default_measure: str = "Energy balance (PJ)",
    leap_working: pd.DataFrame | None = None,
    base_df: pd.DataFrame | None = None,
    ninth_df: pd.DataFrame | None = None,
    esto_to_ninth: dict[tuple[str, str], list[tuple[str, str]]] | None = None,
    base_year: int | None = None,
    base_economy: str = "",
    projection_economy: str = "",
    projection_years: Sequence[int] = (),
    scenario_to_projection: dict[str, str] | None = None,
) -> pd.DataFrame:
    """Return template-declared ESTO-axis rows that may not exist in LEAP exports."""
    template = _load_dashboard_template_allowlist(chart_navigation_guide_path)
    columns = [
        "scenario",
        "sheet",
        "page_key",
        "page_label",
        "chart_group_key",
        "chart_group_label",
        "measure",
        "fuel_label",
        "esto_flow",
        "esto_flow_key",
        "esto_flow_group_key",
        "esto_flow_group_label",
        "esto_product",
        "dashboard_page_key",
        "dashboard_page_label",
        "dashboard_section_key",
        "dashboard_section_label",
        "chart_kind",
        "template_is_multi_flow",
        "use_esto_to_ninth_mapping",
    ]
    if not template:
        return pd.DataFrame(columns=columns)

    rows: list[dict[str, str]] = []
    measure_default = str((template.get("defaults") or {}).get("measure", default_measure)).strip() or default_measure
    scenarios = [_normalize_scenario(name) for name in scenario_names if _clean_token(name)]
    scenario_projection = dict(scenario_to_projection or {})

    def _norm(value: object) -> str:
        return " ".join(str(value or "").strip().lower().split())

    def _nonzero_series(series: pd.Series) -> bool:
        values = pd.to_numeric(series, errors="coerce").dropna()
        return bool(not values.empty and values.abs().gt(1e-12).any())

    def _virtual_sheet_for_path(path: tuple[str, ...]) -> str:
        token = "__".join(_safe_token(str(part).replace("\\", "_")) for part in path if str(part).strip())
        return f"template__{token or 'root'}"

    projection_cache: dict[tuple[str, str, str], pd.Series] = {}

    def _projection_pair_has_nonzero(esto_flow: str, product: str) -> bool:
        if ninth_df is None or ninth_df.empty or not projection_years:
            return False
        targets = list((esto_to_ninth or {}).get((esto_flow, product), []))
        if not targets:
            return False
        projection_scenarios = sorted(set(scenario_projection.values()) or {str(name).lower() for name in scenarios})
        for projection_scenario in projection_scenarios:
            for sector_code, fuel_code in targets:
                cache_key = (sector_code, fuel_code, projection_scenario)
                if cache_key not in projection_cache:
                    projection_cache[cache_key] = pull_projection_series(
                        ninth_df,
                        sector_code=sector_code,
                        fuel_code=fuel_code,
                        economy_code=projection_economy,
                        scenario=projection_scenario,
                        projection_years=projection_years,
                        value_sign_role="",
                    )
                if _nonzero_series(projection_cache[cache_key]):
                    return True
        return False

    def _all_products_for_flows(flows: list[str], exclude_norm: set[str]) -> list[str]:
        flow_norms = {_norm(flow) for flow in flows if _norm(flow)}
        products: set[str] = set()

        if leap_working is not None and not leap_working.empty:
            leap = leap_working.copy()
            for col in ["esto_flow", "esto_product", "leap_value", "esto_is_subtotal"]:
                if col not in leap.columns:
                    leap[col] = False if col == "esto_is_subtotal" else ""
            leap["esto_flow"] = leap["esto_flow"].fillna("").astype(str).str.strip()
            leap["esto_product"] = leap["esto_product"].fillna("").astype(str).str.strip()
            leap["leap_value"] = pd.to_numeric(leap["leap_value"], errors="coerce").fillna(0.0)
            leap["esto_is_subtotal"] = leap["esto_is_subtotal"].fillna(False).astype(bool)
            for row in leap.itertuples(index=False):
                flow = str(getattr(row, "esto_flow", "")).strip()
                product = str(getattr(row, "esto_product", "")).strip()
                if (
                    _norm(flow) in flow_norms
                    and product
                    and not _product_is_total(product)
                    and _norm(product) not in exclude_norm
                    and not bool(getattr(row, "esto_is_subtotal", False))
                    and abs(float(getattr(row, "leap_value", 0.0) or 0.0)) > 1e-12
                ):
                    products.add(product)

        if base_df is not None and not base_df.empty and base_year is not None:
            base = base_df.copy()
            year_col = str(base_year)
            for col in ["economy", "flows", "products", year_col, "is_subtotal"]:
                if col not in base.columns:
                    base[col] = False if col == "is_subtotal" else ""
            base = base[base["economy"].fillna("").astype(str).str.strip().eq(str(base_economy))].copy()
            base["value"] = pd.to_numeric(base[year_col], errors="coerce").fillna(0.0)
            base["is_subtotal"] = base["is_subtotal"].fillna(False).astype(bool)
            for row in base.itertuples(index=False):
                flow = str(getattr(row, "flows", "")).strip()
                product = str(getattr(row, "products", "")).strip()
                if (
                    _norm(flow) in flow_norms
                    and product
                    and not _product_is_total(product)
                    and _norm(product) not in exclude_norm
                    and not bool(getattr(row, "is_subtotal", False))
                    and abs(float(getattr(row, "value", 0.0) or 0.0)) > 1e-12
                ):
                    products.add(product)

        for flow, product in sorted((esto_to_ninth or {}).keys()):
            if (
                _norm(flow) in flow_norms
                and product
                and not _product_is_total(product)
                and _norm(product) not in exclude_norm
                and _projection_pair_has_nonzero(flow, product)
            ):
                products.add(product)

        return sorted(products, key=lambda item: (_esto_flow_sort_key(item), item.lower()))

    def _walk(node: dict[str, Any], path: tuple[str, ...]) -> None:
        for spec in _dashboard_template_graph_specs(node, default_measure=measure_default):
            measure = str(spec.get("measure", "")).strip() or measure_default
            flows = [str(flow).strip() for flow in list(spec.get("esto_flows", []) or [spec.get("esto_flow", "")]) if str(flow).strip()]
            if not flows:
                continue
            if bool(spec.get("include_all_products", False)):
                products = _all_products_for_flows(flows, set(spec.get("exclude_products_norm", set())))
            else:
                products = list(spec.get("products", []) or [])
            group_label = _esto_flow_group_label(flows, fallback_label=path[-1] if path else "")
            flow_group_key = _esto_flow_group_key(flows)
            for product in products:
                product_text = str(product).strip()
                if not product_text:
                    continue
                sheet = _virtual_sheet_for_path(path) if len(flows) > 1 else _sheet_key_from_esto_flow(flows[0])
                fuel_label_text = _strip_esto_code_prefix(product_text) or product_text
                hierarchy = _dashboard_hierarchy_from_path(
                    path,
                    entry_kind="direct",
                    measure=measure,
                    fallback_label=sheet,
                )
                for flow in flows:
                    for scenario in scenarios:
                        rows.append(
                            {
                                "scenario": scenario,
                                "sheet": sheet,
                                **hierarchy,
                                "measure": measure,
                                "fuel_label": fuel_label_text,
                                "esto_flow": flow,
                                "esto_flow_key": _esto_flow_key(flow),
                                "esto_flow_group_key": flow_group_key,
                                "esto_flow_group_label": group_label,
                                "esto_product": product_text,
                                "dashboard_page_key": hierarchy["page_key"],
                                "dashboard_page_label": hierarchy["page_label"],
                                "dashboard_section_key": hierarchy["chart_group_key"],
                                "dashboard_section_label": hierarchy["chart_group_label"],
                                "chart_kind": "by_fuel",
                                "template_is_multi_flow": len(flows) > 1,
                                "use_esto_to_ninth_mapping": bool(spec.get("use_esto_to_ninth_mapping", False)),
                            }
                        )
        for key, child in node.items():
            if _dashboard_template_is_reserved_key(key) or not isinstance(child, dict):
                continue
            label = str(key).strip()
            if label:
                _walk(child, (*path, label))

    _walk(template, ())
    if not rows:
        return pd.DataFrame(columns=columns)
    return _add_dashboard_context_aliases(pd.DataFrame(rows, columns=columns)).drop_duplicates().reset_index(drop=True)

def build_balance_comparison_esto_axis(
    *,
    leap_long: pd.DataFrame,
    mapping_status: pd.DataFrame,
    base_year: int,
    projection_years: Sequence[int],
    base_economy: str,
    projection_economy: str,
    scenario_map: dict[str, str],
    sheet_map_path: Path | str = DEFAULT_SHEET_MAP_PATH,
    backup_mappings_path: Path | str = DEFAULT_BACKUP_MAPPINGS_PATH,
    codebook_path: Path | str = DEFAULT_CODEBOOK_PATH,
    canonical_pairs_path: ConfigTableRef = DEFAULT_MAPPING_PAIRS_PATH,
    explicit_mappings_path: Path | str = DEFAULT_EXPLICIT_MAPPINGS_PATH,
    explicit_reassignments_path: Path | str = DEFAULT_EXPLICIT_REASSIGNMENTS_PATH,
    synthetic_reference_rows_path: Path | str = DEFAULT_SYNTHETIC_REFERENCE_ROWS_PATH,
    esto_table_path: Path | str = DEFAULT_BASE_TABLE_PATH,
    projection_table_path: Path | str = DEFAULT_PROJECTION_TABLE_PATH,
    chart_navigation_guide_path: Path | str | None = None,
    balance_mapping_workbook_path: Path | str | None = None,
    known_issues: dict[str, Any] | None = None,
    base_df: pd.DataFrame | None = None,
    ninth_df: pd.DataFrame | None = None,
) -> dict[str, Any]:
    if leap_long.empty:
        raise RuntimeError("leap_long is empty; cannot build ESTO-axis comparison outputs.")

    mapping_inputs: dict[str, Any] | None = None
    reassignment_status = pd.DataFrame()
    synthetic_reference_status = pd.DataFrame()
    if base_df is None or ninth_df is None:
        mapping_inputs = load_mapping_inputs(
            sheet_map_path=_resolve(sheet_map_path),
            backup_mappings_path=_resolve(backup_mappings_path),
            codebook_path=_resolve(codebook_path),
            canonical_pairs_path=_resolve_config_table_ref(canonical_pairs_path),
            explicit_mappings_path=_resolve(explicit_mappings_path),
            explicit_reassignments_path=_resolve(explicit_reassignments_path),
        )
        base_df, ninth_df, reassignment_status, synthetic_reference_status = load_reference_tables(
            esto_table_path=_resolve(esto_table_path),
            projection_table_path=_resolve(projection_table_path),
            explicit_reassignments=mapping_inputs["explicit_reassignments"],
            explicit_mappings=mapping_inputs["explicit_mappings"],
            canonical_pairs=mapping_inputs["canonical_pairs"],
            synthetic_reference_rows_path=_resolve(synthetic_reference_rows_path),
            drop_all_zero_base_rows=True,
            drop_all_zero_projection_rows=False,
        )
    else:
        mapping_inputs = load_mapping_inputs(
            sheet_map_path=_resolve(sheet_map_path),
            backup_mappings_path=_resolve(backup_mappings_path),
            codebook_path=_resolve(codebook_path),
            canonical_pairs_path=_resolve_config_table_ref(canonical_pairs_path),
            explicit_mappings_path=_resolve(explicit_mappings_path),
            explicit_reassignments_path=_resolve(explicit_reassignments_path),
        )

    scenario_to_projection = {
        _normalize_scenario(k): str(v).strip().lower()
        for k, v in (scenario_map or {}).items()
        if _clean_token(k) and _clean_token(v)
    }
    projection_scenario_values = sorted(set(scenario_to_projection.values()) or set(scenario_map.values() or []))
    if ninth_df is not None and not ninth_df.empty:
        ninth_df = _prepare_ninth_projection_frame(
            ninth_df,
            economy_code=projection_economy,
            scenario_values={str(value).strip().lower() for value in projection_scenario_values if str(value).strip()},
        )

    esto_to_ninth: dict[tuple[str, str], list[tuple[str, str]]] = {}
    def _add_esto_to_ninth_target(esto_flow: object, esto_product: object, ninth_sector: object, ninth_fuel: object) -> None:
        flow = _clean_token(esto_flow)
        product = _clean_token(esto_product)
        sector = _clean_token(ninth_sector)
        fuel = _clean_token(ninth_fuel)
        if not (flow and product and sector and fuel):
            return
        pairs = esto_to_ninth.setdefault((flow, product), [])
        pair = (sector, fuel)
        if pair not in pairs:
            pairs.append(pair)

    def _split_mapping_tokens(value: object) -> list[str]:
        return [token.strip() for token in str(value or "").split("|") if token.strip()]

    def _iter_ninth_mapping_pairs(sectors: object, fuels: object) -> list[tuple[str, str]]:
        sector_tokens = _split_mapping_tokens(sectors)
        fuel_tokens = _split_mapping_tokens(fuels)
        if not sector_tokens or not fuel_tokens:
            return []
        if len(sector_tokens) == len(fuel_tokens):
            return list(dict.fromkeys(zip(sector_tokens, fuel_tokens)))
        if len(fuel_tokens) == 1:
            return [(sector, fuel_tokens[0]) for sector in sector_tokens]
        if len(sector_tokens) == 1:
            return [(sector_tokens[0], fuel) for fuel in fuel_tokens]
        return [(sector, fuel) for sector in sector_tokens for fuel in fuel_tokens]

    if not mapping_status.empty:
        _ms = mapping_status.copy()
        for _col in ("esto_flow", "esto_product", "sector_code_9th", "ninth_fuel_code"):
            if _col not in _ms.columns:
                _ms[_col] = ""
            _ms[_col] = _ms[_col].fillna("").astype(str).str.strip()
        _ms = _ms[_ms["esto_flow"].ne("") & _ms["esto_product"].ne("")].copy()
        for _, _row in _ms[["esto_flow", "esto_product", "sector_code_9th", "ninth_fuel_code"]].drop_duplicates().iterrows():
            for _s, _f in _iter_ninth_mapping_pairs(_row["sector_code_9th"], _row["ninth_fuel_code"]):
                _add_esto_to_ninth_target(_row["esto_flow"], _row["esto_product"], _s, _f)

    nonzero_ninth_pairs = _nonzero_ninth_projection_pairs(
        ninth_df,
        projection_economy=projection_economy,
        projection_scenarios=projection_scenario_values,
        projection_years=projection_years,
    )
    mapping_workbook_crosswalk = _load_active_balance_mapping_crosswalk(balance_mapping_workbook_path)
    if not mapping_workbook_crosswalk.empty and nonzero_ninth_pairs:
        mapping_workbook_crosswalk = mapping_workbook_crosswalk[
            mapping_workbook_crosswalk.apply(
                lambda row: (
                    _clean_token(row.get("ninth_sector", "")),
                    _clean_token(row.get("ninth_fuel", "")),
                )
                in nonzero_ninth_pairs,
                axis=1,
            )
        ].copy()
    elif not mapping_workbook_crosswalk.empty:
        mapping_workbook_crosswalk = mapping_workbook_crosswalk.iloc[0:0].copy()
    if not mapping_workbook_crosswalk.empty:
        for _row in mapping_workbook_crosswalk[
            ["esto_flow", "esto_product", "ninth_sector", "ninth_fuel"]
        ].drop_duplicates().itertuples(index=False):
            _add_esto_to_ninth_target(_row.esto_flow, _row.esto_product, _row.ninth_sector, _row.ninth_fuel)

    def _template_enabled_esto_to_ninth_pairs(canonical_pairs_frame: pd.DataFrame) -> pd.DataFrame:
        if chart_navigation_guide_path is None or canonical_pairs_frame.empty:
            return canonical_pairs_frame.iloc[0:0].copy()
        template = _load_dashboard_template_allowlist(chart_navigation_guide_path)
        if not template:
            return canonical_pairs_frame.iloc[0:0].copy()

        enabled_all_product_flows: set[str] = set()
        enabled_pairs: set[tuple[str, str]] = set()

        def _norm_text(value: object) -> str:
            return " ".join(str(value or "").strip().lower().split())

        def _walk_enabled_specs(node: dict[str, Any]) -> None:
            for aggregate in _dashboard_template_aggregate_specs(node):
                if not bool(aggregate.get("use_esto_to_ninth_mapping", False)):
                    continue
                for flow in list(aggregate.get("esto_flows", aggregate.get("source_flows", [])) or []):
                    flow_norm = _norm_text(flow)
                    if flow_norm:
                        # Aggregate totals need all mapped products for the flow.
                        enabled_all_product_flows.add(flow_norm)

            for spec in _dashboard_template_graph_specs(node):
                if not bool(spec.get("use_esto_to_ninth_mapping", False)):
                    continue
                flow_norms = {_norm_text(flow) for flow in list(spec.get("esto_flows", []) or []) if _norm_text(flow)}
                if not flow_norms:
                    continue
                if bool(spec.get("include_all_products", False)):
                    enabled_all_product_flows.update(flow_norms)
                    continue
                for product in list(spec.get("products", []) or []):
                    product_norm = _norm_text(product)
                    if product_norm:
                        for flow_norm in flow_norms:
                            enabled_pairs.add((flow_norm, product_norm))

            for key, child in node.items():
                if _dashboard_template_is_reserved_key(key) or not isinstance(child, dict):
                    continue
                _walk_enabled_specs(child)

        _walk_enabled_specs(template)
        if not enabled_all_product_flows and not enabled_pairs:
            return canonical_pairs_frame.iloc[0:0].copy()

        work = canonical_pairs_frame.copy()
        work["_esto_flow_norm"] = work["esto_flow"].map(_norm_text)
        work["_esto_product_norm"] = work["esto_product"].map(_norm_text)
        allowed = work["_esto_flow_norm"].isin(enabled_all_product_flows) | work.apply(
            lambda row: (row["_esto_flow_norm"], row["_esto_product_norm"]) in enabled_pairs,
            axis=1,
        )
        return work.loc[allowed].drop(columns=["_esto_flow_norm", "_esto_product_norm"]).copy()

    canonical_pairs = (mapping_inputs or {}).get("canonical_pairs", pd.DataFrame())
    if (
        ninth_df is not None
        and not ninth_df.empty
        and isinstance(canonical_pairs, pd.DataFrame)
        and not canonical_pairs.empty
    ):
        _cp = canonical_pairs.copy()
        for _col in ("esto_flow", "esto_product", "ninth_sector", "ninth_fuel"):
            if _col not in _cp.columns:
                _cp[_col] = ""
            _cp[_col] = _cp[_col].fillna("").astype(str).str.strip()
        _cp = _cp[
            _cp["esto_flow"].ne("")
            & _cp["esto_product"].ne("")
            & _cp["ninth_sector"].ne("")
            & _cp["ninth_fuel"].ne("")
        ].copy()
        _cp = _template_enabled_esto_to_ninth_pairs(_cp)
        for _row in _cp[["esto_flow", "esto_product", "ninth_sector", "ninth_fuel"]].drop_duplicates().itertuples(index=False):
            _add_esto_to_ninth_target(_row.esto_flow, _row.esto_product, _row.ninth_sector, _row.ninth_fuel)

    if esto_to_ninth:
        for _key in esto_to_ninth:
            esto_to_ninth[_key] = sorted(esto_to_ninth[_key])

    def _ninth_hierarchy_tokens(code: object) -> tuple[str, ...]:
        tokens: list[str] = []
        for token in str(code or "").strip().lower().split("_"):
            token = token.strip()
            if not token:
                break
            if token.isdigit() or token == "x":
                tokens.append(token)
                continue
            break
        return tuple(tokens)

    def _is_ninth_parent_code(parent_code: object, child_code: object) -> bool:
        parent_tokens = _ninth_hierarchy_tokens(parent_code)
        child_tokens = _ninth_hierarchy_tokens(child_code)
        if not parent_tokens or len(parent_tokens) >= len(child_tokens):
            return False
        if "x" in parent_tokens:
            # Treat explicit x-buckets as catch-all leaves unless subtotal flags
            # identify them as always-subtotal below.
            return False
        return child_tokens[: len(parent_tokens)] == parent_tokens

    def _drop_parent_ninth_targets(targets: list[tuple[str, str]]) -> list[tuple[str, str]]:
        """Remove parent 9th sector/fuel targets when child targets are present."""
        deduped = sorted(set(targets))
        if len(deduped) <= 1:
            return deduped
        keep: list[tuple[str, str]] = []
        for sector, fuel in deduped:
            has_child_sector_same_fuel = any(
                other_fuel == fuel and _is_ninth_parent_code(sector, other_sector)
                for other_sector, other_fuel in deduped
                if (other_sector, other_fuel) != (sector, fuel)
            )
            has_child_fuel_same_sector = any(
                other_sector == sector and _is_ninth_parent_code(fuel, other_fuel)
                for other_sector, other_fuel in deduped
                if (other_sector, other_fuel) != (sector, fuel)
            )
            if has_child_sector_same_fuel or has_child_fuel_same_sector:
                continue
            keep.append((sector, fuel))
        return keep or deduped

    if esto_to_ninth:
        for _key in list(esto_to_ninth):
            esto_to_ninth[_key] = _drop_parent_ninth_targets(esto_to_ninth[_key])

    # When an ESTO pair maps to a mix of always-subtotal and non-subtotal ninth pairs,
    # drop the always-subtotal ones. Otherwise a subtotal row and its component rows are
    # both included in the sum, producing double-counted projection values.
    if ninth_df is not None and not ninth_df.empty:
        _ninth_sector_cols = [c for c in ["sectors", "sub1sectors", "sub2sectors", "sub3sectors", "sub4sectors"] if c in ninth_df.columns]
        _ninth_fuel_cols = [c for c in ["fuels", "subfuels"] if c in ninth_df.columns]
        if _ninth_sector_cols and _ninth_fuel_cols:
            _sub_layout = ninth_df["subtotal_layout"].fillna(False).astype(bool) if "subtotal_layout" in ninth_df.columns else pd.Series(False, index=ninth_df.index)
            _sub_results = ninth_df["subtotal_results"].fillna(False).astype(bool) if "subtotal_results" in ninth_df.columns else pd.Series(False, index=ninth_df.index)
            _is_sub_arr = (_sub_layout | _sub_results).values
            _sf_records: list[pd.DataFrame] = []
            for _sc in _ninth_sector_cols:
                for _fc in _ninth_fuel_cols:
                    _tmp = pd.DataFrame({
                        "_sector": ninth_df[_sc].fillna("").astype(str).str.strip().str.lower().values,
                        "_fuel": ninth_df[_fc].fillna("").astype(str).str.strip().str.lower().values,
                        "_is_sub": _is_sub_arr,
                    })
                    _sf_records.append(_tmp[_tmp["_sector"].ne("") & _tmp["_fuel"].ne("")])
            if _sf_records:
                _sf_long = pd.concat(_sf_records, ignore_index=True)
                _pair_subtotal_agg = (
                    _sf_long.groupby(["_sector", "_fuel"])["_is_sub"]
                    .agg(all_sub="all")
                    .reset_index()
                )
                _ninth_always_subtotal: set[tuple[str, str]] = set(
                    zip(
                        _pair_subtotal_agg.loc[_pair_subtotal_agg["all_sub"], "_sector"],
                        _pair_subtotal_agg.loc[_pair_subtotal_agg["all_sub"], "_fuel"],
                    )
                )
                if _ninth_always_subtotal:
                    for _esto_key in list(esto_to_ninth.keys()):
                        _all_ninth = esto_to_ninth[_esto_key]
                        _non_sub_ninth = [
                            (s, f) for s, f in _all_ninth
                            if (s.lower(), f.lower()) not in _ninth_always_subtotal
                        ]
                        if _non_sub_ninth and len(_non_sub_ninth) < len(_all_ninth):
                            esto_to_ninth[_esto_key] = _non_sub_ninth

    leap_working = leap_long.copy()
    leap_working["scenario"] = leap_working["scenario"].map(_normalize_scenario)
    leap_working["year"] = pd.to_numeric(leap_working["year"], errors="coerce").astype("Int64")
    leap_working["leap_value"] = pd.to_numeric(leap_working["leap_value"], errors="coerce")
    for col in ["sheet_name", "measure", "fuel_label", "esto_flow", "esto_product"]:
        leap_working[col] = leap_working.get(col, "").fillna("").astype(str).str.strip()
    hierarchy_sheet_catalog = (
        build_esto_axis_structure_from_dashboard_template(chart_navigation_guide_path).get("sheet_catalog", {})
        if chart_navigation_guide_path
        else {}
    )
    leap_working = _backfill_dashboard_hierarchy(
        leap_working,
        sheet_catalog=hierarchy_sheet_catalog,
        sheet_col="sheet_name",
    )
    leap_working = _add_dashboard_context_aliases(_add_esto_flow_context_columns(leap_working))

    rows: list[dict[str, Any]] = []
    base_cache: dict[tuple[str, str], float] = {}
    projection_series_cache: dict[tuple[str, str, str], pd.Series] = {}
    ninth_projection_component_rows: list[dict[str, Any]] = []
    # Per-(ESTO flow group, scenario): set of (ninth_sector, ninth_fuel) pairs already used in projection sums.
    # Prevents counting the same 9th series multiple times when several ESTO products
    # share a 9th pair, even if dashboard template aliases assign different rendering keys.
    used_ninth_pairs: dict[tuple[str, str], set[tuple[str, str]]] = {}
    # (sheet, measure, fuel_label) -> compact annotation string for chart rendering
    ninth_pairs_label_map: dict[tuple[str, str, str], str] = {}

    leap_rows = leap_working[
        [
            "scenario",
            "sheet_name",
            "page_key",
            "page_label",
            "chart_group_key",
            "chart_group_label",
            "dashboard_page_key",
            "dashboard_page_label",
            "dashboard_section_key",
            "dashboard_section_label",
            "chart_kind",
            "esto_flow_key",
            "esto_flow_group_key",
            "esto_flow_group_label",
            "measure",
            "fuel_label",
            "year",
            "leap_value",
        ]
    ].rename(columns={"sheet_name": "sheet", "leap_value": "value"})
    leap_rows["economy"] = projection_economy
    leap_rows["source"] = "leap"
    rows.extend(
        leap_rows[
            [
                "economy",
                "scenario",
                "sheet",
                "page_key",
                "page_label",
                "chart_group_key",
                "chart_group_label",
                "dashboard_page_key",
                "dashboard_page_label",
                "dashboard_section_key",
                "dashboard_section_label",
                "chart_kind",
                "esto_flow_key",
                "esto_flow_group_key",
                "esto_flow_group_label",
                "measure",
                "fuel_label",
                "source",
                "year",
                "value",
            ]
        ].to_dict("records")
    )

    for col in ["esto_is_subtotal", "ninth_is_subtotal"]:
        if col not in leap_working.columns:
            leap_working[col] = False
        leap_working[col] = leap_working[col].fillna(False).astype(bool)

    groups = (
        leap_working[
            [
                "scenario",
                "sheet_name",
                "page_key",
                "page_label",
                "chart_group_key",
                "chart_group_label",
                "dashboard_page_key",
                "dashboard_page_label",
                "dashboard_section_key",
                "dashboard_section_label",
                "chart_kind",
                "esto_flow_key",
                "esto_flow_group_key",
                "esto_flow_group_label",
                "measure",
                "fuel_label",
                "esto_flow",
                "esto_product",
                "esto_is_subtotal",
                "ninth_is_subtotal",
            ]
        ]
        .drop_duplicates(subset=["scenario", "sheet_name", "measure", "fuel_label", "esto_flow", "esto_product"])
        .rename(columns={"sheet_name": "sheet"})
    )
    template_groups = _dashboard_template_esto_axis_records(
        chart_navigation_guide_path,
        scenario_names=list(scenario_to_projection) or list(leap_working["scenario"].dropna().astype(str).unique()),
        leap_working=leap_working,
        base_df=base_df,
        ninth_df=ninth_df,
        esto_to_ninth=esto_to_ninth,
        base_year=base_year,
        base_economy=base_economy,
        projection_economy=projection_economy,
        projection_years=projection_years,
        scenario_to_projection=scenario_to_projection,
    )
    if not template_groups.empty and "template_is_multi_flow" in template_groups.columns:
        multi_template_groups = template_groups[template_groups["template_is_multi_flow"].fillna(False).astype(bool)].copy()
        if not multi_template_groups.empty:
            multi_keys = multi_template_groups[
                [
                    "scenario",
                    "sheet",
                    "page_key",
                    "page_label",
                    "chart_group_key",
                    "chart_group_label",
                    "dashboard_page_key",
                    "dashboard_page_label",
                    "dashboard_section_key",
                    "dashboard_section_label",
                    "chart_kind",
                    "esto_flow_key",
                    "esto_flow_group_key",
                    "esto_flow_group_label",
                    "measure",
                    "fuel_label",
                    "esto_flow",
                    "esto_product",
                ]
            ].drop_duplicates()
            leap_multi = leap_working.merge(
                multi_keys,
                left_on=["scenario", "esto_flow", "esto_product"],
                right_on=["scenario", "esto_flow", "esto_product"],
                how="inner",
                suffixes=("", "_template"),
            )
            if not leap_multi.empty:
                leap_multi_rows = pd.DataFrame(
                    {
                        "economy": projection_economy,
                        "scenario": leap_multi["scenario"],
                        "sheet": leap_multi["sheet"],
                        "page_key": leap_multi["page_key_template"],
                        "page_label": leap_multi["page_label_template"],
                        "chart_group_key": leap_multi["chart_group_key_template"],
                        "chart_group_label": leap_multi["chart_group_label_template"],
                        "dashboard_page_key": leap_multi["dashboard_page_key_template"],
                        "dashboard_page_label": leap_multi["dashboard_page_label_template"],
                        "dashboard_section_key": leap_multi["dashboard_section_key_template"],
                        "dashboard_section_label": leap_multi["dashboard_section_label_template"],
                        "chart_kind": leap_multi["chart_kind_template"],
                        "esto_flow_key": leap_multi["esto_flow_key_template"],
                        "esto_flow_group_key": leap_multi["esto_flow_group_key_template"],
                        "esto_flow_group_label": leap_multi["esto_flow_group_label_template"],
                        "measure": leap_multi["measure_template"],
                        "fuel_label": leap_multi["fuel_label_template"],
                        "source": "leap",
                        "year": leap_multi["year"],
                        "value": leap_multi["leap_value"],
                    }
                )
                rows.extend(leap_multi_rows.to_dict("records"))
    if not template_groups.empty:
        groups = (
            pd.concat([groups, template_groups], ignore_index=True, sort=False)
            .drop_duplicates(subset=["scenario", "sheet", "measure", "fuel_label", "esto_flow", "esto_product"])
            .reset_index(drop=True)
        )
    leap_backed_projection_keys = {
        (
            _normalize_scenario(row.scenario),
            _clean_token(row.measure),
            _clean_token(row.fuel_label),
            _clean_token(row.esto_flow),
            _clean_token(row.esto_product),
        )
        for row in leap_working[["scenario", "measure", "fuel_label", "esto_flow", "esto_product"]]
        .drop_duplicates()
        .itertuples(index=False)
    }
    if leap_backed_projection_keys:
        groups["_projection_priority"] = groups.apply(
            lambda row: 0
            if (
                _normalize_scenario(row.get("scenario", "")),
                _clean_token(row.get("measure", "")),
                _clean_token(row.get("fuel_label", "")),
                _clean_token(row.get("esto_flow", "")),
                _clean_token(row.get("esto_product", "")),
            )
            in leap_backed_projection_keys
            else 1,
            axis=1,
        )
        groups = groups.sort_values(
            ["_projection_priority", "scenario", "sheet", "measure", "fuel_label", "esto_flow", "esto_product"],
            kind="mergesort",
        ).reset_index(drop=True)
    for _col in ("esto_is_subtotal", "ninth_is_subtotal"):
        if _col not in groups.columns:
            groups[_col] = False
        groups[_col] = groups[_col].fillna(False).astype(bool)
    groups = _backfill_dashboard_hierarchy(groups, sheet_catalog=hierarchy_sheet_catalog)
    groups = _add_dashboard_context_aliases(_add_esto_flow_context_columns(groups))

    # Precompute which (ESTO flow group, measure, ninth_sector, ninth_fuel) are claimed by more than one
    # (esto_flow, esto_product) group. Used to decide whether to show annotation and to deduplicate.
    _ninth_pair_claimants: dict[tuple[str, str, str, str], list[str]] = {}
    for _grow in groups.itertuples(index=False):
        _grd = _grow._asdict()
        _gflow_group_key = _clean_token(_grd.get("esto_flow_group_key", ""))
        if not _gflow_group_key:
            _gflow_group_key = _esto_flow_group_key([_grd.get("esto_flow", "")]) or _clean_token(_grd.get("sheet", ""))
        _gmeasure = _clean_token(_grd.get("measure", ""))
        _gfuel = _clean_token(_grd.get("fuel_label", ""))
        _gflow = _clean_token(_grd.get("esto_flow", ""))
        _gproduct = _clean_token(_grd.get("esto_product", ""))
        for _ns, _nf in esto_to_ninth.get((_gflow, _gproduct), []):
            _key = (_gflow_group_key, _gmeasure, _ns, _nf)
            if _key not in _ninth_pair_claimants:
                _ninth_pair_claimants[_key] = []
            if _gfuel not in _ninth_pair_claimants[_key]:
                _ninth_pair_claimants[_key].append(_gfuel)

    for row in groups.itertuples(index=False):
        rd = row._asdict()
        scenario = _normalize_scenario(rd.get("scenario", ""))
        sheet = _clean_token(rd.get("sheet", ""))
        page_key = _clean_token(rd.get("page_key", ""))
        page_label = _clean_token(rd.get("page_label", ""))
        chart_group_key = _clean_token(rd.get("chart_group_key", "")) or sheet
        chart_group_label = _clean_token(rd.get("chart_group_label", "")) or sheet
        dashboard_page_key = _clean_token(rd.get("dashboard_page_key", "")) or page_key
        dashboard_page_label = _clean_token(rd.get("dashboard_page_label", "")) or page_label
        dashboard_section_key = _clean_token(rd.get("dashboard_section_key", "")) or chart_group_key
        dashboard_section_label = _clean_token(rd.get("dashboard_section_label", "")) or chart_group_label
        chart_kind = _clean_token(rd.get("chart_kind", "")) or "by_fuel"
        measure = _clean_token(rd.get("measure", ""))
        fuel_label = _clean_token(rd.get("fuel_label", ""))
        esto_flow = _clean_token(rd.get("esto_flow", ""))
        esto_flow_key = _clean_token(rd.get("esto_flow_key", "")) or _esto_flow_key(esto_flow)
        esto_flow_group_key = _clean_token(rd.get("esto_flow_group_key", "")) or _esto_flow_group_key([esto_flow])
        esto_flow_group_label = _clean_token(rd.get("esto_flow_group_label", "")) or _esto_flow_group_label([esto_flow], chart_group_label)
        esto_product = _clean_token(rd.get("esto_product", ""))
        esto_is_subtotal = bool(rd.get("esto_is_subtotal", False))
        ninth_is_subtotal = bool(rd.get("ninth_is_subtotal", False))

        base_key = (esto_flow, esto_product)
        if base_key not in base_cache:
            if esto_is_subtotal:
                base_cache[base_key] = float("nan")
            else:
                base_cache[base_key] = pull_base_year_value(
                    base_df,
                    base_year=base_year,
                    economy_code=base_economy,
                    esto_flow=esto_flow,
                    esto_product=esto_product,
                    value_sign_role="",
                )
        rows.append(
            {
                "economy": base_economy,
                "scenario": scenario,
                "sheet": sheet,
                "page_key": page_key,
                "page_label": page_label,
                "chart_group_key": chart_group_key,
                "chart_group_label": chart_group_label,
                "dashboard_page_key": dashboard_page_key,
                "dashboard_page_label": dashboard_page_label,
                "dashboard_section_key": dashboard_section_key,
                "dashboard_section_label": dashboard_section_label,
                "chart_kind": chart_kind,
                "esto_flow_key": esto_flow_key,
                "esto_flow_group_key": esto_flow_group_key,
                "esto_flow_group_label": esto_flow_group_label,
                "measure": measure,
                "fuel_label": fuel_label,
                "source": "base",
                "year": int(base_year),
                "value": float(base_cache[base_key]) if pd.notna(base_cache[base_key]) else float("nan"),
            }
        )

        proj_scenario = scenario_to_projection.get(scenario, scenario.lower())
        sheet_scenario_key = (esto_flow_group_key or sheet, scenario)
        if sheet_scenario_key not in used_ninth_pairs:
            used_ninth_pairs[sheet_scenario_key] = set()

        if ninth_is_subtotal:
            proj_total_series: pd.Series = pd.Series(index=list(projection_years), dtype="float64")
        else:
            all_targets = list(esto_to_ninth.get((esto_flow, esto_product), []))
            unused_targets = [(ns, nf) for ns, nf in all_targets if (ns, nf) not in used_ninth_pairs[sheet_scenario_key]]
            total_series: pd.Series | None = None
            for sector_code, fuel_code in unused_targets:
                cache_key = (sector_code, fuel_code, proj_scenario)
                if cache_key not in projection_series_cache:
                    projection_series_cache[cache_key] = pull_projection_series(
                        ninth_df,
                        sector_code=sector_code,
                        fuel_code=fuel_code,
                        economy_code=projection_economy,
                        scenario=proj_scenario,
                        projection_years=projection_years,
                        value_sign_role="",
                    )
                series = projection_series_cache[cache_key]
                for component_year, component_value in series.items():
                    ninth_projection_component_rows.append(
                        {
                            "scenario": scenario,
                            "year": int(component_year),
                            "ninth_sector": sector_code,
                            "esto_flow": esto_flow,
                            "ninth_fuel": fuel_code,
                            "esto_product": esto_product,
                            "value_pj": float(component_value) if pd.notna(component_value) else float("nan"),
                            "subtotal": False,
                            "sheet": sheet,
                            "measure": measure,
                            "fuel_label": fuel_label,
                            "chart_group_key": chart_group_key,
                        }
                    )
                if total_series is None:
                    total_series = series.copy()
                else:
                    total_series = total_series.add(series, fill_value=0.0)
            # Mark newly used pairs so later fuel rows in the same ESTO flow group/scenario skip them.
            for pair in unused_targets:
                used_ninth_pairs[sheet_scenario_key].add(pair)
            if total_series is None:
                total_series = pd.Series(index=list(projection_years), dtype="float64")
            proj_total_series = total_series

            # Build annotation label: show 9th pairs when cardinality is not one-to-one
            ann_key = (chart_group_key, measure, fuel_label)
            if ann_key not in ninth_pairs_label_map:
                is_shared = any(
                    len(_ninth_pair_claimants.get((esto_flow_group_key, measure, ns, nf), [])) > 1
                    for ns, nf in all_targets
                )
                if len(all_targets) != 1 or is_shared:
                    parts: list[str] = []
                    for ns, nf in sorted(set(all_targets)):
                        claimants = _ninth_pair_claimants.get((esto_flow_group_key, measure, ns, nf), [])
                        shared_with = [c for c in claimants if c != fuel_label]
                        tag = f"{nf}/{ns}" if ns else nf
                        if shared_with:
                            others = ", ".join(shared_with[:3])
                            tag = f"{tag} (shared w/ {others})"
                        parts.append(tag)
                    if parts:
                        ninth_pairs_label_map[ann_key] = "; ".join(parts)

        for proj_year, proj_value in proj_total_series.items():
            rows.append(
                {
                    "economy": projection_economy,
                    "scenario": scenario,
                    "sheet": sheet,
                    "page_key": page_key,
                    "page_label": page_label,
                    "chart_group_key": chart_group_key,
                    "chart_group_label": chart_group_label,
                    "dashboard_page_key": dashboard_page_key,
                    "dashboard_page_label": dashboard_page_label,
                    "dashboard_section_key": dashboard_section_key,
                    "dashboard_section_label": dashboard_section_label,
                    "chart_kind": chart_kind,
                    "esto_flow_key": esto_flow_key,
                    "esto_flow_group_key": esto_flow_group_key,
                    "esto_flow_group_label": esto_flow_group_label,
                    "measure": measure,
                    "fuel_label": fuel_label,
                    "source": "projection",
                    "year": int(proj_year),
                    "value": float(proj_value) if pd.notna(proj_value) else float("nan"),
                }
        )

    comparison_long = pd.DataFrame(rows)
    comparison_long["year"] = pd.to_numeric(comparison_long["year"], errors="coerce").astype("Int64")
    comparison_long["value"] = pd.to_numeric(comparison_long["value"], errors="coerce")
    if not template_groups.empty and "template_is_multi_flow" in template_groups.columns:
        multi_template_groups = template_groups[template_groups["template_is_multi_flow"].fillna(False).astype(bool)].copy()
        if not multi_template_groups.empty:
            covered_cols = [
                "chart_group_key",
                "measure",
                "fuel_label",
                "esto_flow_key",
            ]
            for col in covered_cols:
                if col not in multi_template_groups.columns:
                    multi_template_groups[col] = ""
                if col not in comparison_long.columns:
                    comparison_long[col] = ""
                multi_template_groups[col] = multi_template_groups[col].fillna("").astype(str).str.strip()
                comparison_long[col] = comparison_long[col].fillna("").astype(str).str.strip()
            covered = set(
                tuple(row)
                for row in multi_template_groups[covered_cols]
                .drop_duplicates()
                .itertuples(index=False, name=None)
            )
            if covered:
                direct_duplicate_mask = (
                    ~comparison_long["sheet"].fillna("").astype(str).str.startswith("template__")
                    & comparison_long[covered_cols].apply(lambda row: tuple(row) in covered, axis=1)
                )
                comparison_long = comparison_long.loc[~direct_duplicate_mask].copy()
    group_cols = [
        "economy",
        "scenario",
        "sheet",
        "page_key",
        "page_label",
        "chart_group_key",
        "chart_group_label",
        "dashboard_page_key",
        "dashboard_page_label",
        "dashboard_section_key",
        "dashboard_section_label",
        "chart_kind",
        "esto_flow_key",
        "esto_flow_group_key",
        "esto_flow_group_label",
        "measure",
        "fuel_label",
        "source",
        "year",
    ]
    comparison_long = _backfill_dashboard_hierarchy(comparison_long, sheet_catalog=hierarchy_sheet_catalog)
    comparison_long = _add_dashboard_context_aliases(_add_esto_flow_context_columns(comparison_long))
    comparison_long = _collapse_template_multi_flow_comparison_rows(comparison_long, template_groups)
    comparison_long = _add_leap_parent_transfer_rows_to_template(comparison_long)
    comparison_long = _fill_template_transfer_base_values(
        comparison_long,
        base_df=base_df,
        base_year=base_year,
        base_economy=base_economy,
    )
    comparison_long = (
        comparison_long.groupby(group_cols, as_index=False)["value"]
        .sum(min_count=1)
        .sort_values(group_cols, kind="mergesort")
        .reset_index(drop=True)
    )
    comparison_long = _fill_template_transfer_base_values(
        comparison_long,
        base_df=base_df,
        base_year=base_year,
        base_economy=base_economy,
    )
    # Attach compact 9th-pairs annotation for chart rendering (only when cardinality is not one-to-one)
    if ninth_pairs_label_map:
        label_series = comparison_long.apply(
            lambda r: ninth_pairs_label_map.get(
                (
                    str(r.get("chart_group_key", "") or r["sheet"]),
                    str(r["measure"]),
                    str(r["fuel_label"]),
                ),
                "",
            ),
            axis=1,
        )
        comparison_long["ninth_pairs_label"] = label_series
    else:
        comparison_long["ninth_pairs_label"] = ""

    comparison_wide = (
        comparison_long.pivot_table(
            index=[
                "economy",
                "scenario",
                "sheet",
                "page_key",
                "page_label",
                "chart_group_key",
                "chart_group_label",
                "dashboard_page_key",
                "dashboard_page_label",
                "dashboard_section_key",
                "dashboard_section_label",
                "chart_kind",
                "esto_flow_key",
                "esto_flow_group_key",
                "esto_flow_group_label",
                "measure",
                "fuel_label",
                "year",
            ],
            columns="source",
            values="value",
            aggfunc="first",
        )
        .reset_index()
    )
    if hasattr(comparison_wide.columns, "name"):
        comparison_wide.columns.name = None

    mapping_status_for_availability = mapping_status.copy()
    mapping_status_for_availability = _backfill_dashboard_hierarchy(
        mapping_status_for_availability,
        sheet_catalog=hierarchy_sheet_catalog,
    )
    mapping_status_for_availability = _add_dashboard_context_aliases(
        _add_esto_flow_context_columns(mapping_status_for_availability)
    )
    for col in ["esto_flow", "esto_product", "sector_code_9th", "ninth_fuel_code", "mapping_note"]:
        if col not in mapping_status_for_availability.columns:
            mapping_status_for_availability[col] = ""
        mapping_status_for_availability[col] = mapping_status_for_availability[col].fillna("").astype(str).str.strip()
    if esto_to_ninth and not mapping_status_for_availability.empty:
        def _joined_pruned_targets(row: pd.Series, item_index: int) -> str:
            targets = esto_to_ninth.get(
                (_clean_token(row.get("esto_flow", "")), _clean_token(row.get("esto_product", ""))),
                [],
            )
            return "|".join(sorted({pair[item_index] for pair in targets if pair[item_index]}))

        mapping_status_for_availability["sector_code_9th"] = mapping_status_for_availability.apply(
            lambda row: _joined_pruned_targets(row, 0) or str(row.get("sector_code_9th", "")),
            axis=1,
        )
        mapping_status_for_availability["ninth_fuel_code"] = mapping_status_for_availability.apply(
            lambda row: _joined_pruned_targets(row, 1) or str(row.get("ninth_fuel_code", "")),
            axis=1,
        )
    if not template_groups.empty:
        template_status = template_groups.copy()
        for col in [
            "sheet",
            "page_key",
            "page_label",
            "chart_group_key",
            "chart_group_label",
            "dashboard_page_key",
            "dashboard_page_label",
            "dashboard_section_key",
            "dashboard_section_label",
            "chart_kind",
            "esto_flow",
            "esto_flow_key",
            "esto_flow_group_key",
            "esto_flow_group_label",
            "esto_product",
            "measure",
        ]:
            if col not in mapping_status_for_availability.columns:
                mapping_status_for_availability[col] = ""
            mapping_status_for_availability[col] = mapping_status_for_availability[col].fillna("").astype(str).str.strip()
        targets_by_pair = {
            (str(flow), str(product)): targets
            for (flow, product), targets in esto_to_ninth.items()
        }
        expanded_rows: list[dict[str, Any]] = []
        for record in template_status.to_dict("records"):
            targets = targets_by_pair.get((record.get("esto_flow", ""), record.get("esto_product", "")), [])
            if not targets:
                targets = [("", "")]
            for sector_code, fuel_code in targets:
                row = dict(record)
                row["sector_code_9th"] = sector_code
                row["ninth_fuel_code"] = fuel_code
                row["mapped"] = bool(sector_code or fuel_code)
                row["has_any_mapping"] = bool(sector_code or fuel_code)
                row["base_mapping_complete"] = True
                row["projection_mapping_complete"] = bool(sector_code and fuel_code)
                row["partially_mapped"] = False
                row["missing_ninth_fuel"] = not bool(fuel_code)
                row["missing_esto_flow"] = False
                row["missing_esto_product"] = False
                row["has_mapping_note"] = True
                row["mapping_source"] = "dashboard_template_esto_axis"
                row["flow_source"] = "dashboard_template"
                row["fuel_source"] = "dashboard_template"
                row["sector_match_method"] = "esto_pair_axis"
                row["mapping_note"] = "Template-declared ESTO-axis row with no LEAP export row."
                row["projection_parent_fallback"] = False
                row["projection_parent_sector_code"] = ""
                row["comparator_scope"] = "esto_pair"
                expanded_rows.append(row)
        if expanded_rows:
            template_status = pd.DataFrame(expanded_rows)
            for col in mapping_status_for_availability.columns:
                if col not in template_status.columns:
                    template_status[col] = ""
            for col in template_status.columns:
                if col not in mapping_status_for_availability.columns:
                    mapping_status_for_availability[col] = ""
            mapping_status_for_availability = (
                pd.concat([mapping_status_for_availability, template_status[mapping_status_for_availability.columns]], ignore_index=True, sort=False)
                .drop_duplicates(subset=["sheet", "measure", "fuel_label", "sector_code_9th", "ninth_fuel_code", "esto_flow", "esto_product"])
                .reset_index(drop=True)
            )

    mapping_status_out = _build_mapping_status_with_availability(mapping_status_for_availability, comparison_long)
    mapping_status_out = _backfill_dashboard_hierarchy(mapping_status_out, sheet_catalog=hierarchy_sheet_catalog)
    mapping_status_out = _add_dashboard_context_aliases(_add_esto_flow_context_columns(mapping_status_out))
    mapping_status_out["projection_targets_count"] = mapping_status_out.apply(
        lambda r: len(esto_to_ninth.get((_clean_token(r.get("esto_flow", "")), _clean_token(r.get("esto_product", ""))), [])),
        axis=1,
    )
    mapping_status_out["mapping_note"] = mapping_status_out["mapping_note"].fillna("").astype(str)
    no_proj = mapping_status_out["projection_targets_count"].fillna(0).eq(0)
    mapping_status_out.loc[no_proj, "mapping_note"] = mapping_status_out.loc[no_proj, "mapping_note"].apply(
        lambda text: text if text else "No 9th pairs found for this ESTO flow/product."
    )

    return {
        "comparison_long": comparison_long,
        "comparison_wide": comparison_wide,
        "mapping_status": mapping_status_out,
        "ninth_projection_components": pd.DataFrame(ninth_projection_component_rows),
        "base_df": base_df,
        "ninth_df": ninth_df,
        "reassignment_status": reassignment_status,
        "synthetic_reference_status": synthetic_reference_status,
        "mapping_inputs": mapping_inputs,
    }
